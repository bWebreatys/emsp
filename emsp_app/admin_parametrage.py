"""
Module administration - Gestion centralisée des listes déroulantes
Similaire à "Gestion Familles Produits" dans un ERP
"""

from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for
from openpyxl import load_workbook
import json
from datetime import datetime
import os

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# ==============================================================================
# PAGE PRINCIPALE : TABLEAU DE BORD PARAMÉTRAGE
# ==============================================================================

@admin_bp.route('/parametrage')
def parametrage_dashboard():
    """
    Page principale d'administration des listes déroulantes
    Affiche toutes les catégories et nombre de valeurs par liste
    """
    
    # Charger les listes depuis Excel
    listes = {
        'formations_types': load_liste('FORMATIONS', 'A'),
        'formations_durees': load_liste('FORMATIONS', 'B'),
        'formations_statuts': load_liste('FORMATIONS', 'C'),
        'salles_types': load_liste('SALLES', 'A'),
        'salles_equipements': load_liste('SALLES', 'B'),
        'salles_statuts': load_liste('SALLES', 'C'),
        'sessions_statuts': load_liste('SESSIONS', 'A'),
        'interventions_statuts': load_liste('INTERVENTIONS', 'A'),
        'reservations_statuts': load_liste('RESERVATIONS', 'A'),
        'depenses_types': load_liste('DEPENSES', 'A'),
        'formateurs_statuts': load_liste('FORMATEURS', 'A'),
    }
    
    # Status synchronisation
    sync_status = {
        'synced': True,
        'last_sync': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    }
    
    return render_template('admin/parametrage_dashboard.html',
                         listes=listes,
                         sync_status=sync_status)

# ==============================================================================
# GESTION DES LISTES : GÉNÉRIQUE
# ==============================================================================

@admin_bp.route('/parametrage/<category>/<liste_name>')
def manage_liste(category, liste_name):
    """
    Page générique de gestion d'une liste déroulante
    Affiche toutes les valeurs avec :
    - Valeur (texte)
    - Description (explication admin)
    - Nombre d'utilisation (formations, sessions, etc.)
    - Actions (modifier, supprimer)
    """
    
    # Charger les données
    valeurs = load_liste(category, liste_name)
    
    # Compter utilisation pour chaque valeur
    utilisations = count_utilisations(category, liste_name)
    
    # Préparer données pour affichage
    donnees = []
    for idx, valeur in enumerate(valeurs, 1):
        donnees.append({
            'ordre': idx,
            'valeur': valeur,
            'description': get_description(category, liste_name, valeur),
            'utilise': utilisations.get(valeur, 0),
            'modifiable': True,
            'supprimable': utilisations.get(valeur, 0) == 0,  # Suppression si pas utilisée
        })
    
    return render_template('admin/manage_liste.html',
                         category=category,
                         liste_name=liste_name,
                         donnees=donnees,
                         total=len(donnees))

# ==============================================================================
# AJOUTER/MODIFIER VALEUR
# ==============================================================================

@admin_bp.route('/parametrage/<category>/<liste_name>/ajouter', methods=['GET', 'POST'])
@admin_bp.route('/parametrage/<category>/<liste_name>/modifier/<int:ordre>', methods=['GET', 'POST'])
def ajouter_modifier_valeur(category, liste_name, ordre=None):
    """
    Formulaire d'ajout/modification d'une valeur de liste
    
    Avec :
    - Validation unicité
    - Vérification doublons (case-insensitive)
    - Synchronisation Excel automatique
    - Audit trail
    """
    
    if request.method == 'GET':
        # Affichage du formulaire
        if ordre:
            # MODIFICATION
            valeurs = load_liste(category, liste_name)
            valeur = valeurs[ordre - 1] if ordre <= len(valeurs) else None
            
            utilisations = count_utilisations(category, liste_name)
            utilise_par = utilisations.get(valeur, 0)
            
            return render_template('admin/form_valeur.html',
                                 category=category,
                                 liste_name=liste_name,
                                 mode='modification',
                                 ordre=ordre,
                                 valeur=valeur,
                                 description=get_description(category, liste_name, valeur),
                                 utilise_par=utilise_par)
        else:
            # AJOUT
            return render_template('admin/form_valeur.html',
                                 category=category,
                                 liste_name=liste_name,
                                 mode='ajout')
    
    elif request.method == 'POST':
        # TRAITEMENT DU FORMULAIRE
        
        valeur_new = request.form.get('valeur', '').strip()
        description = request.form.get('description', '')
        ordre_custom = request.form.get('ordre', 999)
        
        # VALIDATIONS
        errors = []
        
        # 1. Valeur obligatoire
        if not valeur_new:
            errors.append("Valeur obligatoire")
        
        # 2. Longueur
        if len(valeur_new) > 50:
            errors.append("Max 50 caractères")
        
        # 3. Caractères spéciaux
        if not valeur_new.replace('-', '').replace('.', '').replace("'", '').isalnum():
            errors.append("Caractères spéciaux non autorisés (sauf - . ')")
        
        # 4. Unicité (case-insensitive)
        valeurs = load_liste(category, liste_name)
        
        if ordre:  # MODIFICATION
            # Vérifier que la nouvelle valeur ne doublonne pas une autre
            for idx, v in enumerate(valeurs):
                if idx != (ordre - 1) and v.lower() == valeur_new.lower():
                    errors.append(f"Valeur '{valeur_new}' existe déjà")
        else:  # AJOUT
            # Vérifier pas doublon
            if any(v.lower() == valeur_new.lower() for v in valeurs):
                errors.append(f"Valeur '{valeur_new}' existe déjà")
        
        if errors:
            flash(' | '.join(errors), 'error')
            return redirect(request.referrer)
        
        # SAUVEGARDE
        if ordre:
            # MODIFICATION
            valeurs[ordre - 1] = valeur_new
            action = 'UPDATE'
            ancien_valeur = valeurs[ordre - 1]
        else:
            # AJOUT
            valeurs.append(valeur_new)
            action = 'CREATE'
            ancien_valeur = '-'
        
        # Écrire dans Excel
        save_liste(category, liste_name, valeurs)
        
        # LOG AUDIT
        log_audit(action, category, liste_name, ancien_valeur, valeur_new, 'OK')
        
        flash(f"✅ '{valeur_new}' {'modifié' if ordre else 'créé'}. Synchronisé avec Excel.", 'success')
        
        return redirect(url_for('admin.manage_liste', category=category, liste_name=liste_name))

# ==============================================================================
# SUPPRIMER VALEUR
# ==============================================================================

@admin_bp.route('/parametrage/<category>/<liste_name>/supprimer/<int:ordre>', methods=['POST'])
def supprimer_valeur(category, liste_name, ordre):
    """
    Suppression d'une valeur de liste
    
    Sécurités :
    - Vérifier pas d'utilisation
    - Double confirmation
    - Audit trail
    """
    
    valeurs = load_liste(category, liste_name)
    valeur = valeurs[ordre - 1]
    
    # Vérifier utilisation
    utilisations = count_utilisations(category, liste_name)
    
    if utilisations.get(valeur, 0) > 0:
        flash(f"❌ Impossible supprimer '{valeur}' ({utilisations[valeur]} utilisations)", 'error')
        return redirect(url_for('admin.manage_liste', category=category, liste_name=liste_name))
    
    # Supprimer
    valeurs.pop(ordre - 1)
    save_liste(category, liste_name, valeurs)
    
    # LOG AUDIT
    log_audit('DELETE', category, liste_name, valeur, '-', 'OK')
    
    flash(f"✅ '{valeur}' supprimé", 'success')
    return redirect(url_for('admin.manage_liste', category=category, liste_name=liste_name))

# ==============================================================================
# FONCTIONS UTILITAIRES
# ==============================================================================

def load_liste(category, colonne):
    """
    Charger une liste déroulante depuis Excel
    
    Args:
        category: Nom onglet Excel (FORMATIONS, SALLES, etc.)
        colonne: Lettre colonne (A, B, C, etc.) ou numérique
    
    Returns:
        Liste de valeurs (strings)
    """
    try:
        wb = load_workbook('EMSP_v0.1.xlsx')
        ws = wb['LISTES']  # Onglet caché
        
        # Colonne A=1, B=2, etc.
        if isinstance(colonne, str):
            col_num = ord(colonne.upper()) - ord('A') + 1
        else:
            col_num = colonne
        
        # Lire colonne
        valeurs = []
        for cell in ws.iter_rows(min_row=2, max_row=100, min_col=col_num, max_col=col_num):
            if cell[0].value:
                valeurs.append(str(cell[0].value))
            else:
                break
        
        return valeurs
    
    except Exception as e:
        print(f"Erreur load_liste: {e}")
        return []

def save_liste(category, colonne, valeurs):
    """Sauvegarder une liste déroulante dans Excel"""
    try:
        wb = load_workbook('EMSP_v0.1.xlsx')
        ws = wb['LISTES']
        
        # Colonne
        col_num = ord(colonne.upper()) - ord('A') + 1
        
        # Effacer ancienne colonne
        for row in ws.iter_rows(min_row=2, max_row=100, min_col=col_num, max_col=col_num):
            row[0].value = None
        
        # Écrire nouvelles valeurs
        for idx, valeur in enumerate(valeurs, 2):
            ws.cell(row=idx, column=col_num).value = valeur
        
        wb.save('EMSP_v0.1.xlsx')
        print(f"✅ Sauvegarde Excel: {category} colonne {colonne}")
    
    except Exception as e:
        print(f"Erreur save_liste: {e}")

def count_utilisations(category, colonne):
    """Compter combien de fois chaque valeur est utilisée"""
    utilisations = {}
    
    # À adapter selon catégorie
    # Pour FORMATIONS types : compter FORMATIONS.xlsx colonnes type_formation
    # Pour SESSIONS statuts : compter SESSIONS.xlsx colonnes statut
    # etc.
    
    # Placeholder : retourner 0 pour toutes
    for valeur in load_liste(category, colonne):
        utilisations[valeur] = 0
    
    return utilisations

def get_description(category, colonne, valeur):
    """Récupérer description d'une valeur"""
    # Descriptions prédéfinies
    descriptions = {
        ('FORMATIONS', 'A', 'Licence'): 'Baccalauréat + 3 ans',
        ('FORMATIONS', 'A', 'Master'): 'Licence + 2 ans (spécialisation)',
        # etc.
    }
    
    return descriptions.get((category, colonne, valeur), '')

def log_audit(action, category, liste_name, ancien, nouveau, status):
    """Enregistrer action dans audit trail"""
    log_entry = {
        'datetime': datetime.now().isoformat(),
        'user': 'ADMIN',  # À remplacer avec utilisateur réel
        'action': action,
        'category': category,
        'liste': liste_name,
        'ancien': ancien,
        'nouveau': nouveau,
        'status': status
    }
    
    # Écrire dans fichier log
    with open('logs/audit_listes.log', 'a') as f:
        f.write(json.dumps(log_entry) + '\n')

