"""
EMSP - Gestion Formation
Application Flask locale - saisie multi-onglets Excel
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os, copy

app = Flask(__name__)
app.secret_key = "emsp-local-2026"

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "data.xlsx")

# ─── Couleurs cohérentes avec le fichier Excel ───────────────────────────────
BLACK = "FF000000"
RED   = "FFCC0000"

# ─── Sheets protégées (contiennent des formules à ne pas toucher) ────────────
FORMULA_SHEETS = {"BILANS", "ACCUEIL"}

# ─── Colonnes avec formules dans chaque onglet (index 0-based) ──────────────
FORMULA_COLS = {
    "SESSIONS":      [10],   # col K = Taux_Réalisation
    "RESERVATIONS":  [10, 11],  # cols K, L
    "BILANS":        list(range(20)),
}

# ─── Définition des onglets ──────────────────────────────────────────────────
SHEETS_CONFIG = {
    "PARAM_FORMATIONS": {
        "label":   "📚 Formations",
        "icon":    "📚",
        "color":   "#2E75B6",
        "columns": [
            {"key": "Code_Formation",           "label": "Code formation",            "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. FC-HYG-01"},
            {"key": "Intitulé",                 "label": "Intitulé",                  "type": "text",   "required": True,  "spec": True,  "placeholder": "Nom complet de la formation"},
            {"key": "Type_Formation",           "label": "Type de formation",         "type": "select", "required": True,  "spec": True,  "options": ["Initiale","Continue","Recyclage"]},
            {"key": "Filière",                  "label": "Filière",                   "type": "select", "required": False, "spec": True,  "options": ["Infirmier","Aide-soignant","Médecin","Pharmacien","Transversal","Formateurs"]},
            {"key": "Durée (h)",                "label": "Durée (heures)",            "type": "number", "required": True,  "spec": True,  "placeholder": "ex. 24"},
            {"key": "Responsable_Pédagogique",  "label": "Responsable pédagogique",   "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. Dr. Dupont"},
            {"key": "Coût_Prévisionnel (FCFA)", "label": "Coût prévisionnel (FCFA)",  "type": "number", "required": False, "spec": True,  "placeholder": "ex. 250000"},
            {"key": "Financement_Principal",    "label": "Financement principal",     "type": "select", "required": False, "spec": True,  "options": ["ODS","Budget EMSP","Projet","Partenariat","Autofinancement"]},
            {"key": "Nb_Max_Participants",      "label": "Nb max participants",       "type": "number", "required": False, "spec": False, "placeholder": "ex. 30"},
            {"key": "Objectifs_Généraux",       "label": "Objectifs généraux",        "type": "textarea","required": False,"spec": False, "placeholder": "Décrire les objectifs..."},
            {"key": "Pré-requis",               "label": "Pré-requis",                "type": "text",   "required": False, "spec": False, "placeholder": "ex. BAC ou équivalent"},
        ]
    },
    "PARAM_SALLES": {
        "label":   "🏫 Salles",
        "icon":    "🏫",
        "color":   "#2E75B6",
        "columns": [
            {"key": "Code_Salle",    "label": "Code salle",    "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. SAL-TP-01"},
            {"key": "Nom_Salle",     "label": "Nom de la salle","type": "text",   "required": True,  "spec": True,  "placeholder": "ex. Salle TP 1"},
            {"key": "Type_Salle",    "label": "Type de salle",  "type": "select", "required": True,  "spec": True,  "options": ["Amphi","TP","Simulation","Cours","Réunion","Laboratoire"]},
            {"key": "Capacité",      "label": "Capacité",       "type": "number", "required": True,  "spec": True,  "placeholder": "Nb de places"},
            {"key": "Bâtiment",      "label": "Bâtiment",       "type": "text",   "required": False, "spec": True,  "placeholder": "ex. Bât. A"},
            {"key": "Étage",         "label": "Étage",          "type": "text",   "required": False, "spec": True,  "placeholder": "ex. RDC, R+1"},
            {"key": "Équipements",   "label": "Équipements",    "type": "text",   "required": False, "spec": True,  "placeholder": "ex. Vidéo, Tableau blanc"},
            {"key": "Statut",        "label": "Statut",         "type": "select", "required": True,  "spec": True,  "options": ["En service","Hors service","En rénovation","Réservée"]},
            {"key": "Resp. Salle",   "label": "Responsable salle","type": "text", "required": False, "spec": False, "placeholder": "Nom du responsable"},
            {"key": "Observations",  "label": "Observations",   "type": "textarea","required": False,"spec": False, "placeholder": "Notes diverses..."},
        ]
    },
    "FORMATEURS_RH": {
        "label":   "👩‍🏫 Formateurs",
        "icon":    "👩‍🏫",
        "color":   "#70AD47",
        "columns": [
            {"key": "Code_Formateur",              "label": "Code formateur",            "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. F001"},
            {"key": "Nom",                         "label": "Nom",                       "type": "text",   "required": True,  "spec": True,  "placeholder": "NOM en majuscules"},
            {"key": "Prénom",                      "label": "Prénom",                    "type": "text",   "required": True,  "spec": True,  "placeholder": "Prénom"},
            {"key": "Statut",                      "label": "Statut",                    "type": "select", "required": True,  "spec": True,  "options": ["Interne EMSP","Intervenant externe","Expert projet","Enseignant UDC","Vacataire"]},
            {"key": "Discipline_Principale",       "label": "Discipline principale",     "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. Soins infirmiers"},
            {"key": "Niveau_Qualification",        "label": "Niveau qualification",      "type": "select", "required": False, "spec": True,  "options": ["Licence","Master","Doctorat","Spécialiste","DES","DESC"]},
            {"key": "Type_Intervention",           "label": "Type d'intervention",       "type": "text",   "required": False, "spec": True,  "placeholder": "ex. Cours magistral, TP"},
            {"key": "Exp. Ens. (ans)",             "label": "Expérience enseignement (ans)","type": "number","required": False,"spec": True, "placeholder": "Années"},
            {"key": "Exp. Clinique (ans)",         "label": "Expérience clinique (ans)", "type": "number", "required": False, "spec": True,  "placeholder": "Années"},
            {"key": "Taux_Disponibilité",          "label": "Taux disponibilité",        "type": "select", "required": False, "spec": True,  "options": ["Plein temps","Temps partiel","Occasionnel","Vacataire"]},
            {"key": "Rattachement_Institutionnel", "label": "Rattachement institutionnel","type": "text",  "required": False, "spec": True,  "placeholder": "ex. EMSP, Hôpital Central"},
            {"key": "Formations_Suivies_Projet",   "label": "Formations suivies (projet)","type": "text",  "required": False, "spec": True,  "placeholder": "ex. Simulation – 2026"},
            {"key": "Date_Dernière_MAJ",           "label": "Date dernière MAJ",         "type": "date",   "required": False, "spec": True,  "placeholder": "JJ/MM/AAAA"},
            {"key": "Email",                       "label": "Email",                     "type": "email",  "required": False, "spec": False, "placeholder": "prenom.nom@emsp.org"},
            {"key": "Téléphone",                   "label": "Téléphone",                 "type": "text",   "required": False, "spec": False, "placeholder": "ex. +241 XX XX XX XX"},
            {"key": "Heures_Max_S1",               "label": "Heures max S1",             "type": "number", "required": False, "spec": False, "placeholder": "ex. 300"},
        ]
    },
    "SESSIONS": {
        "label":   "📅 Sessions",
        "icon":    "📅",
        "color":   "#ED7D31",
        "columns": [
            {"key": "Code_Session",           "label": "Code session",          "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. FC-HYG-01-2026-01"},
            {"key": "Code_Formation",         "label": "Code formation",        "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. FC-HYG-01"},
            {"key": "Intitulé_Session",       "label": "Intitulé session",      "type": "text",   "required": True,  "spec": True,  "placeholder": "Nom de la session"},
            {"key": "Date_Début",             "label": "Date début",            "type": "date",   "required": True,  "spec": True,  "placeholder": "JJ/MM/AAAA"},
            {"key": "Date_Fin",               "label": "Date fin",              "type": "date",   "required": True,  "spec": True,  "placeholder": "JJ/MM/AAAA"},
            {"key": "Responsable_Session",    "label": "Responsable session",   "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. Dr. Dupont"},
            {"key": "Nb_Participants_Prévus", "label": "Nb participants prévus","type": "number", "required": True,  "spec": True,  "placeholder": "ex. 30"},
            {"key": "Nb_Participants_Réels",  "label": "Nb participants réels", "type": "number", "required": False, "spec": True,  "placeholder": "ex. 28"},
            {"key": "Type_Public",            "label": "Type public",           "type": "select", "required": False, "spec": True,  "options": ["Étudiants EMSP","Agents hospitaliers","Formateurs EMSP","Mixte","Externe"]},
            {"key": "Statut_Session",         "label": "Statut session",        "type": "select", "required": True,  "spec": True,  "options": ["Planifiée","En cours","Terminée","Annulée","Reportée"]},
            # col 11 (index 10) = Taux_Réalisation : FORMULE - pas de saisie
            {"key": "Lieu_Principal",         "label": "Lieu principal",        "type": "text",   "required": False, "spec": False, "placeholder": "ex. SAL-AMP-01"},
            {"key": "Observations",           "label": "Observations",          "type": "textarea","required": False,"spec": False, "placeholder": "Notes..."},
        ]
    },
    "INTERVENTIONS": {
        "label":   "🔗 Interventions",
        "icon":    "🔗",
        "color":   "#ED7D31",
        "columns": [
            {"key": "Code_Intervention",  "label": "Code intervention",   "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. INT_013"},
            {"key": "Code_Session",       "label": "Code session",        "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. FC-HYG-01-2026-01"},
            {"key": "Code_Formateur",     "label": "Code formateur",      "type": "text",   "required": True,  "spec": True,  "placeholder": "ex. F001"},
            {"key": "Rôle",               "label": "Rôle",                "type": "select", "required": True,  "spec": True,  "options": ["Responsable","Co-formateur","Intervenant externe","Tuteur stage","Observateur"]},
            {"key": "Nb_Heures",          "label": "Nb heures",           "type": "number", "required": True,  "spec": True,  "placeholder": "ex. 16"},
            {"key": "Date_Début_Session", "label": "Date début session",  "type": "date",   "required": False, "spec": True,  "placeholder": "JJ/MM/AAAA"},
            # cols 7,8 = lookups auto (Intitulé_Formateur, Intitulé_Session) - pas de saisie
            {"key": "Observations",       "label": "Observations",        "type": "textarea","required": False,"spec": False, "placeholder": "Notes..."},
        ]
    },
    "RESERVATIONS": {
        "label":   "🗓 Réservations",
        "icon":    "🗓",
        "color":   "#7030A0",
        "columns": [
            {"key": "Date",              "label": "Date",                "type": "date",   "required": True,  "spec": True, "placeholder": "JJ/MM/AAAA"},
            {"key": "Heure_Début",       "label": "Heure début",         "type": "time",   "required": True,  "spec": True, "placeholder": "HH:MM"},
            {"key": "Heure_Fin",         "label": "Heure fin",           "type": "time",   "required": True,  "spec": True, "placeholder": "HH:MM"},
            {"key": "Code_Salle",        "label": "Code salle",          "type": "text",   "required": True,  "spec": True, "placeholder": "ex. SAL-SIM-01"},
            {"key": "Type_Activité",     "label": "Type activité",       "type": "select", "required": True,  "spec": True, "options": ["Cours initial","TP","Simulation","Formation continue","Réunion","Examen"]},
            {"key": "Intitulé_Activité", "label": "Intitulé activité",   "type": "text",   "required": True,  "spec": True, "placeholder": "ex. Anatomie S1"},
            {"key": "Responsable",       "label": "Responsable",         "type": "text",   "required": True,  "spec": True, "placeholder": "ex. Dr. Dupont"},
            {"key": "Lien_Session",      "label": "Session liée",        "type": "text",   "required": False, "spec": True, "placeholder": "ex. FI-INF-01-2026-S1"},
            {"key": "Statut",            "label": "Statut",              "type": "select", "required": True,  "spec": True, "options": ["Prévu","Confirmé","Réalisé","Annulé","Reporté"]},
            {"key": "Commentaire",       "label": "Commentaire",         "type": "textarea","required": False,"spec": True, "placeholder": "Notes..."},
            # cols 11,12 = formules Durée et Capacité - pas de saisie
        ]
    },
    "DEPENSES": {
        "label":   "💶 Dépenses",
        "icon":    "💶",
        "color":   "#C00000",
        "columns": [
            {"key": "Date_Dépense",              "label": "Date dépense",              "type": "date",   "required": True,  "spec": True, "placeholder": "JJ/MM/AAAA"},
            {"key": "Code_Session",              "label": "Code session",              "type": "text",   "required": True,  "spec": True, "placeholder": "ex. FC-HYG-01-2026-01 ou HORS_SESSION"},
            {"key": "Nature_Dépense",            "label": "Nature dépense",            "type": "select", "required": True,  "spec": True, "options": ["Intervenants","Supports pédagogiques","Matériel TP","Hébergement","Transport","Collations","Location matériel","Fournitures administratives","Autre"]},
            {"key": "Montant (FCFA)",            "label": "Montant (FCFA)",            "type": "number", "required": True,  "spec": True, "placeholder": "ex. 45000"},
            {"key": "Fournisseur/Bénéficiaire",  "label": "Fournisseur / Bénéficiaire","type": "text",   "required": True,  "spec": True, "placeholder": "ex. Traiteur Express"},
            {"key": "Mode_Paiement",             "label": "Mode paiement",             "type": "select", "required": True,  "spec": True, "options": ["Virement","Chèque","Espèces","Prélèvement","Bon de commande"]},
            {"key": "Justificatif",              "label": "N° justificatif",           "type": "text",   "required": False, "spec": True, "placeholder": "ex. FAC-2026-042"},
            {"key": "Source_Financement",        "label": "Source financement",        "type": "select", "required": True,  "spec": True, "options": ["ODS","Budget EMSP","Projet","Partenariat","Autofinancement"]},
            {"key": "Exercice_Budgétaire",       "label": "Exercice budgétaire",       "type": "text",   "required": False, "spec": False, "placeholder": "ex. 2026"},
            {"key": "Validé_par",                "label": "Validé par",                "type": "text",   "required": False, "spec": False, "placeholder": "Nom du valideur"},
        ]
    },
}


# ─── Utilitaires Excel ────────────────────────────────────────────────────────

def load_wb():
    return load_workbook(EXCEL_PATH)


def get_sheet_data(sheet_name):
    """Retourne (headers, rows) pour un onglet donné."""
    wb = load_wb()
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            data.append([format_cell(v) for v in row])
    return headers, data


def format_cell(val):
    """Formate une valeur cellule pour affichage."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(round(val, 4))
    return str(val)


def parse_date(s):
    """Parse JJ/MM/AAAA → datetime.date, ou None."""
    s = s.strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return s  # laisser tel quel si on ne sait pas parser


def parse_value(raw, col_type):
    """Convertit la valeur brute du formulaire selon le type de champ."""
    if raw is None or raw.strip() == "":
        return None
    raw = raw.strip()
    if col_type == "date":
        return parse_date(raw)
    if col_type == "number":
        try:
            f = float(raw.replace(" ", "").replace(",", "."))
            return int(f) if f == int(f) else f
        except ValueError:
            return raw
    return raw


def append_row_to_sheet(sheet_name, values_dict, config_cols):
    """
    Ajoute une ligne à l'onglet Excel sans toucher aux formules existantes.
    values_dict = {col_key: valeur_brute, ...}
    """
    wb = load_wb()
    ws = wb[sheet_name]

    # Récupérer les headers de la feuille
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Construire la ligne à insérer (None par défaut pour les cols formule)
    new_row = [None] * len(headers)
    for col_cfg in config_cols:
        key = col_cfg["key"]
        col_type = col_cfg["type"]
        if key in values_dict:
            val = parse_value(values_dict[key], col_type)
            # Trouver la colonne correspondante dans Excel
            try:
                idx = headers.index(key)
                new_row[idx] = val
            except ValueError:
                pass  # colonne pas trouvée → on ignore

    # Ajouter la ligne
    next_row = ws.max_row + 1
    for ci, val in enumerate(new_row, 1):
        cell = ws.cell(row=next_row, column=ci)
        if val is not None:
            cell.value = val
            # Appliquer le format date si nécessaire
            if hasattr(val, 'strftime'):
                cell.number_format = "DD/MM/YYYY"

    wb.save(EXCEL_PATH)


def update_row_in_sheet(sheet_name, row_idx, values_dict, config_cols):
    """
    Met à jour une ligne existante (1-based, row 1 = header).
    row_idx : numéro de ligne Excel (2 = première ligne de données).
    NE TOUCHE PAS aux colonnes formule.
    """
    wb = load_wb()
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    formula_col_indices = set(FORMULA_COLS.get(sheet_name, []))  # 0-based

    for col_cfg in config_cols:
        key = col_cfg["key"]
        col_type = col_cfg["type"]
        try:
            idx = headers.index(key)  # 0-based
        except ValueError:
            continue
        if idx in formula_col_indices:
            continue  # ne pas toucher aux formules
        val = parse_value(values_dict.get(key, ""), col_type)
        cell = ws.cell(row=row_idx, column=idx + 1)
        cell.value = val
        if hasattr(val, 'strftime'):
            cell.number_format = "DD/MM/YYYY"

    wb.save(EXCEL_PATH)


def delete_row_in_sheet(sheet_name, row_idx):
    """Supprime une ligne (row_idx = numéro Excel, 2-based pour données)."""
    wb = load_wb()
    ws = wb[sheet_name]
    ws.delete_rows(row_idx)
    wb.save(EXCEL_PATH)


# ─── Routes principales ───────────────────────────────────────────────────────

@app.route("/")
def index():
    """Page d'accueil — tableau de bord navigation."""
    stats = {}
    wb = load_wb()
    for sname, scfg in SHEETS_CONFIG.items():
        try:
            ws = wb[sname]
            stats[sname] = max(0, ws.max_row - 1)
        except Exception:
            stats[sname] = "?"
    return render_template("index.html", sheets=SHEETS_CONFIG, stats=stats)


@app.route("/sheet/<sheet_name>")
def sheet_view(sheet_name):
    """Liste des enregistrements d'un onglet."""
    if sheet_name not in SHEETS_CONFIG:
        flash(f"Onglet inconnu : {sheet_name}", "error")
        return redirect(url_for("index"))
    cfg = SHEETS_CONFIG[sheet_name]
    headers, rows = get_sheet_data(sheet_name)
    return render_template(
        "sheet_list.html",
        sheet_name=sheet_name,
        cfg=cfg,
        headers=headers,
        rows=rows,
        sheets=SHEETS_CONFIG,
    )


@app.route("/sheet/<sheet_name>/new", methods=["GET", "POST"])
def sheet_new(sheet_name):
    """Formulaire d'ajout d'un enregistrement."""
    if sheet_name not in SHEETS_CONFIG:
        return redirect(url_for("index"))
    cfg = SHEETS_CONFIG[sheet_name]

    if request.method == "POST":
        values = {k: v for k, v in request.form.items()}
        # Validation des champs obligatoires
        errors = []
        for col in cfg["columns"]:
            if col["required"] and not values.get(col["key"], "").strip():
                errors.append(f"Le champ « {col['label']} » est obligatoire.")
        if errors:
            for e in errors:
                flash(e, "error")
            return render_template("sheet_form.html", sheet_name=sheet_name,
                                   cfg=cfg, values=values, mode="new", sheets=SHEETS_CONFIG)
        try:
            append_row_to_sheet(sheet_name, values, cfg["columns"])
            flash("✅ Enregistrement ajouté avec succès.", "success")
        except Exception as ex:
            flash(f"❌ Erreur lors de l'écriture : {ex}", "error")
        return redirect(url_for("sheet_view", sheet_name=sheet_name))

    return render_template("sheet_form.html", sheet_name=sheet_name,
                           cfg=cfg, values={}, mode="new", sheets=SHEETS_CONFIG)


@app.route("/sheet/<sheet_name>/edit/<int:row_idx>", methods=["GET", "POST"])
def sheet_edit(sheet_name, row_idx):
    """Formulaire de modification d'un enregistrement (row_idx = ligne Excel)."""
    if sheet_name not in SHEETS_CONFIG:
        return redirect(url_for("index"))
    cfg = SHEETS_CONFIG[sheet_name]
    headers, rows = get_sheet_data(sheet_name)
    # row_idx 2-based dans Excel, donc index dans rows = row_idx - 2
    data_index = row_idx - 2
    if data_index < 0 or data_index >= len(rows):
        flash("Ligne introuvable.", "error")
        return redirect(url_for("sheet_view", sheet_name=sheet_name))

    row_data = rows[data_index]
    current_values = {h: (row_data[i] if i < len(row_data) else "") for i, h in enumerate(headers)}

    if request.method == "POST":
        values = {k: v for k, v in request.form.items()}
        errors = []
        for col in cfg["columns"]:
            if col["required"] and not values.get(col["key"], "").strip():
                errors.append(f"Le champ « {col['label']} » est obligatoire.")
        if errors:
            for e in errors:
                flash(e, "error")
            return render_template("sheet_form.html", sheet_name=sheet_name,
                                   cfg=cfg, values=values, mode="edit",
                                   row_idx=row_idx, sheets=SHEETS_CONFIG)
        try:
            update_row_in_sheet(sheet_name, row_idx, values, cfg["columns"])
            flash("✅ Enregistrement mis à jour.", "success")
        except Exception as ex:
            flash(f"❌ Erreur : {ex}", "error")
        return redirect(url_for("sheet_view", sheet_name=sheet_name))

    return render_template("sheet_form.html", sheet_name=sheet_name,
                           cfg=cfg, values=current_values, mode="edit",
                           row_idx=row_idx, sheets=SHEETS_CONFIG)


@app.route("/sheet/<sheet_name>/delete/<int:row_idx>", methods=["POST"])
def sheet_delete(sheet_name, row_idx):
    """Supprime une ligne."""
    if sheet_name not in SHEETS_CONFIG:
        return redirect(url_for("index"))
    try:
        delete_row_in_sheet(sheet_name, row_idx)
        flash("🗑 Ligne supprimée.", "success")
    except Exception as ex:
        flash(f"❌ Erreur suppression : {ex}", "error")
    return redirect(url_for("sheet_view", sheet_name=sheet_name))


@app.route("/formules")
def formules():
    """Page de documentation : formules et nomenclatures."""
    FORMULAS_DOC = {
        "SESSIONS – Taux_Réalisation (colonne K)": {
            "formule": "=IF(G{n}=0, \"–\", H{n}/G{n})",
            "description": "Calcule le taux de réalisation d'une session : participants réels / participants prévus.",
            "colonnes": {"G": "Nb_Participants_Prévus", "H": "Nb_Participants_Réels", "K": "Taux_Réalisation (résultat)"},
            "format": "Pourcentage (0%)",
            "spec": False,
        },
        "RESERVATIONS – Durée (colonne K)": {
            "formule": "=IF(AND(C{n}<>\"\",B{n}<>\"\"), TIMEVALUE(C{n})-TIMEVALUE(B{n}), \"–\")",
            "description": "Calcule la durée d'occupation d'une salle en heures.",
            "colonnes": {"B": "Heure_Début", "C": "Heure_Fin", "K": "Durée (résultat)"},
            "format": "[h]:mm",
            "spec": False,
        },
        "RESERVATIONS – Capacité_Salle (colonne L)": {
            "formule": "Valeur statique issue de PARAM_SALLES",
            "description": "Capacité de la salle réservée, copiée depuis le référentiel PARAM_SALLES.",
            "colonnes": {"D": "Code_Salle", "L": "Capacité_Salle (résultat)"},
            "format": "Nombre entier",
            "spec": False,
        },
        "BILANS – Nb total de sessions (D5)": {
            "formule": "=COUNTA('📅 SESSIONS'!A:A)-1",
            "description": "Compte toutes les sessions enregistrées (hors en-tête).",
            "colonnes": {},
            "format": "Nombre entier",
            "spec": True,
        },
        "BILANS – Sessions terminées (D7)": {
            "formule": "=COUNTIF('📅 SESSIONS'!J:J, \"Terminée\")",
            "description": "Nombre de sessions avec statut = Terminée.",
            "colonnes": {"J": "Statut_Session"},
            "format": "Nombre entier",
            "spec": True,
        },
        "BILANS – Total participants prévus (D10)": {
            "formule": "=SUM('📅 SESSIONS'!G:G)",
            "description": "Somme de tous les participants prévus sur l'ensemble des sessions.",
            "colonnes": {"G": "Nb_Participants_Prévus"},
            "format": "Nombre entier",
            "spec": True,
        },
        "BILANS – Taux global réalisation (D12)": {
            "formule": "=IF(SUM(SESSIONS!G:G)=0, \"–\", SUM(SESSIONS!H:H)/SUM(SESSIONS!G:G))",
            "description": "Taux global = total réels / total prévus sur toutes sessions.",
            "colonnes": {"G": "Nb_Participants_Prévus", "H": "Nb_Participants_Réels"},
            "format": "0%",
            "spec": False,
        },
        "BILANS – Dépenses par nature (D18:D26)": {
            "formule": "=SUMIF(DEPENSES!C:C, B{n}, DEPENSES!D:D)",
            "description": "Somme des dépenses filtrées par Nature_Dépense.",
            "colonnes": {"B": "Nature_Dépense (libellé)", "C": "Nature_Dépense (DEPENSES)", "D": "Montant (FCFA)"},
            "format": "#,##0",
            "spec": True,
        },
        "BILANS – Heures par formateur (G{n})": {
            "formule": "=SUMIF(INTERVENTIONS!C:C, B{n}, INTERVENTIONS!E:E)",
            "description": "Total heures d'intervention du formateur sur la période.",
            "colonnes": {"B": "Code_Formateur", "C": "Code_Formateur (INTERVENTIONS)", "E": "Nb_Heures"},
            "format": "#,##0",
            "spec": True,
        },
    }
    NOMENCLATURES = {
        "Code_Formation":  {"pattern": "TYPE-FILIERE-NN", "exemple": "FC-HYG-01, FI-INF-02"},
        "Code_Session":    {"pattern": "CODE_FORMATION-AAAA-NN", "exemple": "FC-HYG-01-2026-01"},
        "Code_Formateur":  {"pattern": "F + numéro 3 chiffres", "exemple": "F001, F008"},
        "Code_Salle":      {"pattern": "SAL-TYPE-NN", "exemple": "SAL-TP-01, SAL-AMP-01"},
        "Code_Intervention":{"pattern": "INT_NNN", "exemple": "INT_001, INT_012"},
        "Code_Pièce":      {"pattern": "PIE-CATEG-NN", "exemple": "PIE-FILTRE-01"},
        "Date":            {"pattern": "JJ/MM/AAAA", "exemple": "09/05/2026"},
        "Justificatif":    {"pattern": "FAC-AAAA-NNN ou BC-AAAA-NNN ou RECU-NNN", "exemple": "FAC-2026-042"},
    }
    return render_template("formules.html", formulas=FORMULAS_DOC,
                           nomenclatures=NOMENCLATURES, sheets=SHEETS_CONFIG)


if __name__ == "__main__":
    print("\n" + "="*60)
    print("  EMSP – Gestion Formation  |  Serveur local")
    print("  Ouvrez votre navigateur : http://127.0.0.1:5000")
    print("="*60 + "\n")
    app.run(debug=False, host="127.0.0.1", port=5000)
