# -*- coding: utf-8 -*-
"""Couche d'ACCES AUX DONNEES — lecture seule du classeur EMSP_V1.xlsx.
Cette couche ne contient AUCUNE logique metier : elle ne fait que lire le
classeur et renvoyer des structures brutes (en-tetes, lignes, listes).
La logique metier (calculs, provenance, tableau de bord) est dans metier.py.
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from datetime import datetime
import os
import re
import json
import config

# Cache de lecture partage : {chemin: (mtime, wb)}. Le classeur n'est recharge
# que si le fichier a change sur disque (hot reload preserve). Evite de re-parser
# le fichier a chaque acces (la cause de la lenteur en V1.2 : ~90 chargements/page).
_CACHE_LECTURE = {}


def _invalider_cache(chemin):
    _CACHE_LECTURE.pop(chemin, None)


class AccesDonnees:
    def __init__(self, chemin=config.WORKBOOK):
        self.chemin = chemin
        self._wb = None

    @property
    def wb(self):
        # Recharge uniquement si le fichier a change (mtime). data_only=True pour
        # lire les valeurs calculees ; read_only=False => chargement complet en
        # memoire, pas de handle de fichier persistant (pas de verrou sous Windows).
        mt = os.path.getmtime(self.chemin)
        cache = _CACHE_LECTURE.get(self.chemin)
        if cache and cache[0] == mt:
            return cache[1]
        wb = load_workbook(self.chemin, data_only=True, read_only=False)
        _CACHE_LECTURE[self.chemin] = (mt, wb)
        return wb

    def onglets(self):
        return self.wb.sheetnames

    def titre(self, onglet):
        ws = self.wb[onglet]
        c = ws.cell(row=1, column=1).value
        return c if c else onglet

    def entetes(self, onglet):
        ws = self.wb[onglet]
        vals = [ws.cell(row=config.LIGNE_ENTETES, column=col).value
                for col in range(1, ws.max_column + 1)]
        # on coupe les colonnes finales vides
        while vals and vals[-1] in (None, ""):
            vals.pop()
        return [v if v is not None else "" for v in vals]

    def lignes(self, onglet, limite=None):
        """Renvoie les lignes de donnees (a partir de la ligne 3), non vides."""
        ws = self.wb[onglet]
        ncol = len(self.entetes(onglet)) or ws.max_column
        out = []
        for r in ws.iter_rows(min_row=config.LIGNE_DONNEES, values_only=True):
            cells = list(r[:ncol])
            if all(c in (None, "") for c in cells):
                continue
            out.append([("" if c is None else c) for c in cells])
            if limite and len(out) >= limite:
                break
        return out

    def nb_lignes(self, onglet):
        return len(self.lignes(onglet))

    def colonne(self, onglet, nom_entete):
        """Valeurs d'une colonne identifiee par son libelle d'en-tete exact."""
        ent = self.entetes(onglet)
        if nom_entete not in ent:
            return []
        i = ent.index(nom_entete)
        return [lig[i] for lig in self.lignes(onglet)]

    def listes_parametres(self):
        """Renvoie {libelle_liste: [valeurs...]} depuis P0_Parametres."""
        ent = self.entetes("P0_Parametres")
        lignes = self.lignes("P0_Parametres")
        res = {}
        for i, nom in enumerate(ent):
            vals = [lig[i] for lig in lignes if i < len(lig) and lig[i] not in (None, "")]
            res[nom] = vals
        return res

    def dictionnaire(self):
        """Renvoie les lignes du Dictionnaire sous forme de dicts."""
        ws = self.wb["Dictionnaire"]
        cols = ["Onglet", "Champ", "Type", "Obligatoire", "Provenance", "Liste", "Description"]
        out = []
        for r in ws.iter_rows(min_row=config.LIGNE_DONNEES, values_only=True):
            if all(c in (None, "") for c in r[:7]):
                continue
            out.append({cols[i]: ("" if r[i] is None else r[i]) for i in range(7)})
        return out

    def lignes_libres(self, onglet, ncol=2):
        """Lecture cle/valeur (Guide, Legende) : liste de tuples non vides."""
        ws = self.wb[onglet]
        out = []
        for r in ws.iter_rows(values_only=True):
            cells = list(r[:ncol])
            if all(c in (None, "") for c in cells):
                continue
            out.append([("" if c is None else c) for c in cells])
        return out

    # =======================================================================
    # ECRITURE (V1.2) — seules primitives autorisees a modifier le classeur.
    # Garde-fous integres (defense en profondeur, en plus de la regle de role
    # appliquee dans metier.py) : jamais d'ecriture sur un onglet en lecture
    # seule ni sur une colonne calcul ; les formules sont preservees ; le motif
    # de formule est recopie dans chaque nouvelle ligne (aucune limite par ligne).
    # =======================================================================
    def _ouvrir_ecriture(self):
        # data_only=False : on PRESERVE les formules. read_only=False : ecriture.
        return load_workbook(self.chemin, data_only=False)

    def _index_colonne(self, ws, entete_brut):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=config.LIGNE_ENTETES, column=col).value == entete_brut:
                return col
        return None

    def _premiere_ligne_vide(self, ws, ncol):
        r = config.LIGNE_DONNEES
        while True:
            if all(ws.cell(row=r, column=c).value in (None, "")
                   for c in range(1, ncol + 1)):
                return r
            r += 1

    def _propager_formule(self, ws, col, ligne):
        """Recopie (en translatant les references) la derniere formule de la
        colonne `col` au-dessus de `ligne`, dans la cellule (ligne, col)."""
        rr = ligne - 1
        while rr >= config.LIGNE_DONNEES:
            v = ws.cell(row=rr, column=col).value
            if isinstance(v, str) and v.startswith("="):
                origine = "%s%d" % (get_column_letter(col), rr)
                dest = "%s%d" % (get_column_letter(col), ligne)
                ws.cell(row=ligne, column=col).value = \
                    Translator(v, origin=origine).translate_formula(dest)
                return True
            rr -= 1
        return False

    def ajouter_ligne(self, onglet, valeurs):
        """Ajoute une ligne de donnees. `valeurs` = {entete_brut: valeur}.
        Refuse les onglets en lecture seule ; ignore toute tentative d'ecrire
        une colonne calcul (la formule y est recopiee a la place)."""
        if onglet in config.READONLY_TABS:
            raise PermissionError("Onglet en lecture seule : %s" % onglet)
        ent = self.entetes(onglet)
        ro = set(config.READONLY_COLS.get(onglet, []))
        wb = self._ouvrir_ecriture()
        ws = wb[onglet]
        ncol = len(ent)
        ligne = self._premiere_ligne_vide(ws, ncol)
        if (ligne - config.LIGNE_DONNEES) >= config.CAPACITE:
            wb.close()
            raise OverflowError("Capacite atteinte (%d lignes) pour %s"
                                % (config.CAPACITE, onglet))
        for i, brut in enumerate(ent, start=1):
            if brut in ro:                       # colonne calcul : on (re)pose la formule
                self._propager_formule(ws, i, ligne)
            elif brut in valeurs:
                ws.cell(row=ligne, column=i, value=valeurs[brut])
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return ligne

    def modifier_ligne(self, onglet, index, valeurs):
        """Modifie une ligne EXISTANTE. `index` = position 0-based parmi les lignes
        NON VIDES (meme logique que `lignes()`, donc aligne sur l'affichage).
        `valeurs` = {entete_brut: valeur} ; une valeur vide ECRASE (effacement).
        Refuse les onglets en lecture seule ; ne touche JAMAIS une colonne calcul
        (sa formule est preservee telle quelle)."""
        if onglet in config.READONLY_TABS:
            raise PermissionError("Onglet en lecture seule : %s" % onglet)
        ent = self.entetes(onglet)
        ncol = len(ent)
        ro = set(config.READONLY_COLS.get(onglet, []))
        wb = self._ouvrir_ecriture()
        ws = wb[onglet]
        cible, k, r = None, 0, config.LIGNE_DONNEES
        while r <= ws.max_row:
            if not all(ws.cell(row=r, column=c).value in (None, "")
                       for c in range(1, ncol + 1)):
                if k == index:
                    cible = r
                    break
                k += 1
            r += 1
        if cible is None:
            wb.close()
            raise IndexError("Ligne introuvable (index %s)" % index)
        for i, brut in enumerate(ent, start=1):
            if brut in ro:                       # colonne calcul : formule preservee
                continue
            if brut in valeurs:
                ws.cell(row=cible, column=i, value=valeurs[brut])
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return cible

    # --- AJOUT PAR LOT (append pur) — V1.20 ---------------------------------
    # Ajoute PLUSIEURS lignes en UNE ouverture du classeur (perf : saisie en
    # grille du registre de tresorerie). Append uniquement (pas d'upsert) : chaque
    # element de `lignes` cree une nouvelle ligne. Memes garde-fous que ajouter_ligne
    # (onglet en lecture seule refuse ; colonnes calcul recoivent la formule recopiee).
    def ajouter_lignes(self, onglet, lignes):
        if onglet in config.READONLY_TABS:
            raise PermissionError("Onglet en lecture seule : %s" % onglet)
        if not lignes:
            return 0
        ent = self.entetes(onglet)
        ro = set(config.READONLY_COLS.get(onglet, []))
        wb = self._ouvrir_ecriture()
        ws = wb[onglet]
        ncol = len(ent)
        ligne = self._premiere_ligne_vide(ws, ncol)
        n = 0
        for valeurs in lignes:
            if (ligne - config.LIGNE_DONNEES) >= config.CAPACITE:
                wb.close()
                raise OverflowError("Capacite atteinte (%d) pour %s"
                                    % (config.CAPACITE, onglet))
            for i, brut in enumerate(ent, start=1):
                if brut in ro:
                    self._propager_formule(ws, i, ligne)
                elif brut in valeurs:
                    ws.cell(row=ligne, column=i, value=valeurs[brut])
            ligne += 1
            n += 1
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return n

    # --- ECRITURE PAR LOT (upsert) — V1.4 -----------------------------------
    # Ecrit/MET A JOUR plusieurs lignes en UNE ouverture du classeur (perf).
    # Chaque ligne est identifiee par un jeu de COLONNES-CLE : si une ligne avec
    # la meme cle existe deja, ses cellules sont mises a jour (pas de doublon) ;
    # sinon la ligne est ajoutee. Memes garde-fous que ajouter_ligne : onglet en
    # lecture seule refuse, colonnes calcul jamais saisies (formule recopiee),
    # cles jamais ecrasees lors d'une mise a jour. Utilise par la saisie des
    # presences par seance (A2_Presences).
    def ecrire_lignes_lot(self, onglet, lignes, cles):
        if onglet in config.READONLY_TABS:
            raise PermissionError("Onglet en lecture seule : %s" % onglet)
        ent = self.entetes(onglet)
        ro = set(config.READONLY_COLS.get(onglet, []))
        cles = [k for k in cles if k in ent]
        wb = self._ouvrir_ecriture()
        ws = wb[onglet]
        ncol = len(ent)
        col_de = {brut: i for i, brut in enumerate(ent, start=1)}
        # Index des lignes existantes par tuple de cle.
        existant = {}
        r = config.LIGNE_DONNEES
        while True:
            if all(ws.cell(row=r, column=c).value in (None, "")
                   for c in range(1, ncol + 1)):
                break
            tup = tuple(str(ws.cell(row=r, column=col_de[k]).value or "").strip()
                        for k in cles)
            existant.setdefault(tup, r)
            r += 1
        fin = r  # premiere ligne libre
        n_ajout = n_maj = 0
        for valeurs in lignes:
            tup = tuple(str(valeurs.get(k, "") or "").strip() for k in cles)
            ligne = existant.get(tup)
            if ligne is None:                         # AJOUT
                ligne = fin
                if (ligne - config.LIGNE_DONNEES) >= config.CAPACITE:
                    wb.close()
                    raise OverflowError("Capacite atteinte (%d) pour %s"
                                        % (config.CAPACITE, onglet))
                for brut, c in col_de.items():
                    if brut in ro:
                        self._propager_formule(ws, c, ligne)
                    elif brut in valeurs:
                        ws.cell(row=ligne, column=c, value=valeurs[brut])
                existant[tup] = ligne
                fin += 1
                n_ajout += 1
            else:                                     # MISE A JOUR (cles preservees)
                for brut, c in col_de.items():
                    if brut in ro or brut in cles:
                        continue
                    if brut in valeurs:
                        ws.cell(row=ligne, column=c, value=valeurs[brut])
                n_maj += 1
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return {"ajout": n_ajout, "maj": n_maj}

    # --- Remplacement complet des donnees d'un onglet (vide puis reecrit) — V1.16 ---
    # Utilise pour la zone d'import (IMPORT_zone) : on remplace tout le contenu a
    # chaque import. Les colonnes calcul (READONLY_COLS) recoivent la formule
    # (recopiee depuis un modele capture AVANT le vidage). Refuse lecture seule.
    def remplacer_donnees(self, onglet, lignes):
        if onglet in config.READONLY_TABS:
            raise PermissionError("Onglet en lecture seule : %s" % onglet)
        ent = self.entetes(onglet)
        ncol = len(ent)
        ro = set(config.READONLY_COLS.get(onglet, []))
        if len(lignes) > config.CAPACITE:
            raise OverflowError("Capacite atteinte (%d) pour %s" % (config.CAPACITE, onglet))
        wb = self._ouvrir_ecriture()
        ws = wb[onglet]
        # 1) capturer un modele de formule par colonne calcul (avant vidage)
        modele = {}
        for i, brut in enumerate(ent, start=1):
            if brut in ro:
                rr = config.LIGNE_DONNEES
                while rr <= ws.max_row:
                    v = ws.cell(row=rr, column=i).value
                    if isinstance(v, str) and v.startswith("="):
                        modele[i] = (v, rr)
                        break
                    rr += 1
        # 2) vider toute la zone de donnees existante
        for rr in range(config.LIGNE_DONNEES, ws.max_row + 1):
            for c in range(1, ncol + 1):
                ws.cell(row=rr, column=c).value = None
        # 3) reecrire
        for idx, valeurs in enumerate(lignes):
            ligne = config.LIGNE_DONNEES + idx
            for i, brut in enumerate(ent, start=1):
                if brut in ro:
                    if i in modele:
                        src, src_row = modele[i]
                        origine = "%s%d" % (get_column_letter(i), src_row)
                        dest = "%s%d" % (get_column_letter(i), ligne)
                        ws.cell(row=ligne, column=i,
                                value=Translator(src, origin=origine).translate_formula(dest))
                elif brut in valeurs:
                    ws.cell(row=ligne, column=i, value=valeurs[brut])
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return len(lignes)

    # --- Suppression d'une ligne par CLE (onglets sans colonne calcul) — V1.13 ---
    # Utilise pour l'administration des droits (P1_Roles). Recompacte les lignes.
    # Refuse les onglets a colonne calcul (reecriture des valeurs incompatible avec
    # les formules) et les onglets en lecture seule.
    def supprimer_ligne_par_cle(self, onglet, cle_entete, valeur):
        if onglet in config.READONLY_TABS:
            raise PermissionError("Onglet en lecture seule : %s" % onglet)
        if config.READONLY_COLS.get(onglet):
            raise PermissionError("Suppression non supportee (colonnes calcul) : %s" % onglet)
        valeur = str(valeur).strip()
        ent = self.entetes(onglet)
        ncol = len(ent)
        wb = self._ouvrir_ecriture()
        ws = wb[onglet]
        col = self._index_colonne(ws, cle_entete)
        if col is None:
            wb.close()
            raise KeyError("Colonne inconnue : %s" % cle_entete)
        r, rows = config.LIGNE_DONNEES, []
        while True:
            if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ncol + 1)):
                break
            rows.append([ws.cell(row=r, column=c).value for c in range(1, ncol + 1)])
            r += 1
        fin = r
        restantes = [row for row in rows if str(row[col - 1] or "").strip() != valeur]
        if len(restantes) == len(rows):
            wb.close()
            return False
        for idx, row in enumerate(restantes):           # reecriture compactee
            rr = config.LIGNE_DONNEES + idx
            for c in range(1, ncol + 1):
                ws.cell(row=rr, column=c, value=row[c - 1])
        for rr in range(config.LIGNE_DONNEES + len(restantes), fin):  # purge du reste
            for c in range(1, ncol + 1):
                ws.cell(row=rr, column=c).value = None
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return True

    def supprimer_ligne_par_index(self, onglet, index):
        """Supprime la `index`-ieme ligne NON VIDE (indexation alignee sur lignes() et
        modifier_ligne, donc sur l'affichage). Compactage par reecriture (pas de
        delete_rows : preserve la structure). Sert aux onglets SANS cle unique
        (ex. F5_Budget_Prev). Renvoie True si une ligne a ete supprimee."""
        if onglet in config.READONLY_TABS:
            raise PermissionError("Onglet en lecture seule : %s" % onglet)
        try:
            index = int(index)
        except (TypeError, ValueError):
            return False
        ent = self.entetes(onglet)
        ncol = len(ent)
        wb = self._ouvrir_ecriture()
        ws = wb[onglet]
        r, rows = config.LIGNE_DONNEES, []
        while True:
            if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, ncol + 1)):
                break
            rows.append([ws.cell(row=r, column=c).value for c in range(1, ncol + 1)])
            r += 1
        fin = r
        if index < 0 or index >= len(rows):
            wb.close()
            return False
        restantes = [row for i, row in enumerate(rows) if i != index]
        for idx, row in enumerate(restantes):           # reecriture compactee
            rr = config.LIGNE_DONNEES + idx
            for c in range(1, ncol + 1):
                ws.cell(row=rr, column=c, value=row[c - 1])
        for rr in range(config.LIGNE_DONNEES + len(restantes), fin):  # purge du reste
            for c in range(1, ncol + 1):
                ws.cell(row=rr, column=c).value = None
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return True

    # --- P0_Parametres : colonnes-listes independantes (ajout / suppression) ---
    def ajouter_valeur_liste(self, colonne_entete, valeur):
        """Ajoute une valeur en bas d'une colonne-liste de P0_Parametres.
        Renvoie True si ajoutee, False si doublon."""
        valeur = str(valeur).strip()
        if not valeur:
            return False
        wb = self._ouvrir_ecriture()
        ws = wb["P0_Parametres"]
        col = self._index_colonne(ws, colonne_entete)
        if col is None:
            wb.close()
            raise KeyError("Colonne inconnue : %s" % colonne_entete)
        r, existantes = config.LIGNE_DONNEES, []
        while True:
            v = ws.cell(row=r, column=col).value
            if v in (None, ""):
                break
            existantes.append(str(v).strip())
            r += 1
        if valeur in existantes:
            wb.close()
            return False
        if (r - config.LIGNE_DONNEES) >= config.CAPACITE:
            wb.close()
            raise OverflowError("Capacite atteinte pour la liste %s" % colonne_entete)
        ws.cell(row=r, column=col, value=valeur)
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return True

    def supprimer_valeur_liste(self, colonne_entete, valeur):
        """Retire une valeur d'une colonne-liste de P0_Parametres et recompacte.
        Renvoie True si retiree, False si absente."""
        valeur = str(valeur).strip()
        wb = self._ouvrir_ecriture()
        ws = wb["P0_Parametres"]
        col = self._index_colonne(ws, colonne_entete)
        if col is None:
            wb.close()
            raise KeyError("Colonne inconnue : %s" % colonne_entete)
        r, vals = config.LIGNE_DONNEES, []
        while True:
            v = ws.cell(row=r, column=col).value
            if v in (None, ""):
                break
            vals.append(v)
            r += 1
        restantes = [v for v in vals if str(v).strip() != valeur]
        if len(restantes) == len(vals):
            wb.close()
            return False
        for i in range(len(vals)):
            ws.cell(row=config.LIGNE_DONNEES + i, column=col).value = \
                (restantes[i] if i < len(restantes) else None)
        wb.save(self.chemin)
        _invalider_cache(self.chemin)
        wb.close()
        return True

    # --- Documents lies aux bailleurs (3b) : I/O fichiers + index.json --------
    # Couche d'acces pure : aucune verification de role ni de type ici (assurees
    # en amont par metier/auth). Le dossier fait foi ; pas de suppression (un
    # document obsolete est marque via son statut, jamais efface).
    def _slug_bailleur(self, id_bailleur):
        """Nom de dossier sur a partir de l'ID bailleur (retire separateurs et
        caracteres de controle, conserve un libelle lisible)."""
        s = str(id_bailleur).strip().replace("/", "_").replace("\\", "_")
        s = s.replace(os.sep, "_")
        s = re.sub(r"[\x00-\x1f]", "", s).strip(". ")
        return s or "_"

    def _slug_fichier(self, nom):
        nom = os.path.basename(str(nom)).strip()
        nom = re.sub(r"[^\w.\- ]", "_", nom, flags=re.UNICODE).replace(" ", "_")
        return nom or "fichier"

    def dossier_documents_bailleur(self, id_bailleur, creer=False):
        d = os.path.join(config.DOCS_BAILLEURS_DIR, self._slug_bailleur(id_bailleur))
        if creer:
            os.makedirs(d, exist_ok=True)
        return d

    def _chemin_index_bailleur(self, id_bailleur, creer=False):
        return os.path.join(
            self.dossier_documents_bailleur(id_bailleur, creer), "index.json")

    def lire_index_documents_bailleur(self, id_bailleur):
        """Liste de dicts (metadonnees) ; [] si aucun document ou index illisible."""
        p = self._chemin_index_bailleur(id_bailleur)
        if not os.path.exists(p):
            return []
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (ValueError, OSError):
            return []
        docs = data.get("documents", []) if isinstance(data, dict) else []
        return docs if isinstance(docs, list) else []

    def _ecrire_index_bailleur(self, id_bailleur, docs):
        p = self._chemin_index_bailleur(id_bailleur, creer=True)
        tmp = p + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({"documents": docs}, f, ensure_ascii=False, indent=2)
        os.replace(tmp, p)  # ecriture atomique (pas d'index a moitie ecrit)

    def _nom_stocke_unique(self, dossier, nom_original):
        horo = datetime.now().strftime("%Y%m%d_%H%M%S")
        nom = "%s_%s" % (horo, self._slug_fichier(nom_original))
        cible = os.path.join(dossier, nom)
        i = 1
        while os.path.exists(cible):
            racine, ext = os.path.splitext(nom)
            cible = os.path.join(dossier, "%s_%d%s" % (racine, i, ext))
            i += 1
        return os.path.basename(cible)

    def ajouter_document_bailleur(self, id_bailleur, donnees_octets,
                                  nom_original, type_doc, ajoute_par):
        """Ecrit le binaire dans le dossier du bailleur et ajoute une entree a
        l'index. Renvoie le dict de metadonnees cree."""
        dossier = self.dossier_documents_bailleur(id_bailleur, creer=True)
        nom_stocke = self._nom_stocke_unique(dossier, nom_original)
        with open(os.path.join(dossier, nom_stocke), "wb") as f:
            f.write(donnees_octets)
        meta = {
            "nom_stocke": nom_stocke,
            "nom_original": os.path.basename(str(nom_original)),
            "type": str(type_doc or ""),
            "date_ajout": datetime.now().strftime("%d/%m/%Y"),
            "taille": len(donnees_octets),
            "ajoute_par": str(ajoute_par or ""),
            "statut": "",
        }
        docs = self.lire_index_documents_bailleur(id_bailleur)
        docs.append(meta)
        self._ecrire_index_bailleur(id_bailleur, docs)
        return meta

    def definir_statut_document_bailleur(self, id_bailleur, nom_stocke, statut):
        """Met a jour le statut/libelle d'un document (jamais de suppression).
        Renvoie True si le document a ete trouve."""
        docs = self.lire_index_documents_bailleur(id_bailleur)
        for d in docs:
            if d.get("nom_stocke") == nom_stocke:
                d["statut"] = str(statut or "")
                self._ecrire_index_bailleur(id_bailleur, docs)
                return True
        return False

    def chemin_document_bailleur(self, id_bailleur, nom_stocke):
        """Chemin disque d'un document pour telechargement, ou None si absent
        ou tentative de traversee de repertoire."""
        if not nom_stocke or self._slug_fichier(nom_stocke) != nom_stocke:
            return None
        if nom_stocke == "index.json" or nom_stocke.endswith(".tmp"):
            return None  # fichiers internes : jamais servis
        dossier = os.path.realpath(self.dossier_documents_bailleur(id_bailleur))
        chemin = os.path.realpath(os.path.join(dossier, nom_stocke))
        if not (chemin == dossier or chemin.startswith(dossier + os.sep)):
            return None
        return chemin if os.path.exists(chemin) else None

    def definir_nb_documents_bailleur(self, id_bailleur, n):
        """Ecrit le nombre de documents dans la cellule 'Documents lies' de la
        ligne F4 du bailleur (compteur calcule, 3c). Ecriture interne : ne passe
        pas par le garde-fou de saisie IHM (colonne en lecture seule cote editeur).
        Renvoie True si la ligne du bailleur a ete trouvee."""
        idb = str(id_bailleur or "").strip()
        if not idb:
            return False
        wb = self._ouvrir_ecriture()
        ws = wb["F4_Bailleurs"]
        c_id = self._index_colonne(ws, "ID bailleur (*)")
        c_doc = self._index_colonne(ws, "Documents lies (*)")
        if not c_id or not c_doc:
            wb.close()
            return False
        trouve = False
        r = config.LIGNE_DONNEES
        while True:
            v = ws.cell(row=r, column=c_id).value
            if v in (None, ""):
                break
            if str(v).strip().lower() == idb.lower():
                ws.cell(row=r, column=c_doc).value = int(n)
                trouve = True
                break
            r += 1
        if trouve:
            wb.save(self.chemin)
            _invalider_cache(self.chemin)
        wb.close()
        return trouve
