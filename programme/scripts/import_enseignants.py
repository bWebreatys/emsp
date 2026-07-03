# -*- coding: utf-8 -*-
"""
EMSP — Chargement des ENSEIGNANTS (E1_Enseignants) dans le classeur de
PRODUCTION (copie deployee de EMSP_V1.xlsx).

Source : scripts/enseignants_data.json, fige depuis Enseignants_EMSP.xlsx
(onglet « Enseignants »), 52 enseignants.

Regles appliquees (validees Bernard) :
  - Matricule provisoire : ENS-001 ... ENS-052 (numero du fichier source).
  - Nom : le libelle complet de la source est conserve TEL QUEL dans « Nom ».
    « Prenom » laisse VIDE (a separer ensuite a la main via la fiche si besoin).
  - Statut : « Permanent » -> titulaire ; « Vacataire » -> vacataire.
  - Matieres enseignees : liste texte regroupee de la source (champ unique).
    L'onglet « Associations » (prof x matiere x niveau) N'EST PAS charge : il
    reste matiere premiere du planning previsionnel (candidat hors V1).
  - Colonnes laissees VIDES (completees ensuite via la fiche) : Genre,
    Qualifications, Departement, Chef dept validant, Taux horaire,
    Mode de remuneration, Cout mensuel.

ATTENTION DESSINS : ce script ecrit via openpyxl, donc sur une COPIE DEPLOYEE
(donnees/data), jamais sur le master fige (qui conserve ses 16 dessins). C'est la
meme voie d'ecriture que l'IHM (ajouter_ligne). Le master template n'est pas touche.

Idempotent : purge les lignes de donnees de E1 puis reecrit ; rejouable sans
doublon.

Lancement (depuis la racine du programme) :  python scripts/import_enseignants.py
Chemin surchargeable :                       EMSP_DB=...\\EMSP_V1.xlsx python scripts/import_enseignants.py
"""
import os, json, sys
from openpyxl import load_workbook

ICI = os.path.dirname(os.path.abspath(__file__))
RACINE = os.path.dirname(ICI)
sys.path.insert(0, RACINE)
try:
    import config
    DEFAUT = config.WORKBOOK
    L_ENT, L_DON = config.LIGNE_ENTETES, config.LIGNE_DONNEES
except Exception:
    DEFAUT = os.path.join(RACINE, "donnees", "data", "EMSP_V1.xlsx")
    L_ENT, L_DON = 2, 3

XLSX = os.environ.get("EMSP_DB") or DEFAUT
DATA = os.path.join(ICI, "enseignants_data.json")
ONGLET = "E1_Enseignants"


def _propre(h):
    h = "" if h is None else str(h).strip()
    for suf in (" (**)", " (*)", "(**)", "(*)"):
        if h.endswith(suf):
            return h[: -len(suf)].strip()
    return h


def _index(ws):
    """Libelle propre -> numero de colonne (1-based), depuis la ligne d'en-tetes."""
    idx = {}
    for c in range(1, ws.max_column + 1):
        lib = _propre(ws.cell(L_ENT, c).value)
        if lib:
            idx[lib] = c
    return idx


def _purge(ws):
    if ws.max_row >= L_DON:
        ws.delete_rows(L_DON, ws.max_row - L_DON + 1)


def main():
    if not os.path.exists(XLSX):
        raise SystemExit("Classeur introuvable : %s" % XLSX)
    ens = json.load(open(DATA, encoding="utf-8"))
    wb = load_workbook(XLSX)
    if ONGLET not in wb.sheetnames:
        raise SystemExit("Onglet %s absent du classeur." % ONGLET)

    ws = wb[ONGLET]
    ix = _index(ws)
    _purge(ws)

    r = L_DON
    for e in ens:
        valeurs = {
            "Matricule ens.": "ENS-%03d" % int(e.get("n", r - L_DON + 1)),
            "Genre": "",
            "Nom": e.get("nom", ""),
            "Prenom": "",
            "Statut (titulaire/vacataire)": e.get("statut", ""),
            "Matieres enseignees": e.get("matieres", ""),
            "Qualifications": "",
            "Departement": "",
            "Chef dept validant": "",
            "Taux horaire (KMF/h)": "",
            "Mode de remuneration": "",
            "Cout mensuel (KMF)": "",
        }
        for lib, val in valeurs.items():
            c = ix.get(lib)
            if c:
                ws.cell(r, c).value = val
        r += 1

    wb.save(XLSX)
    from collections import Counter
    rep = Counter(e.get("statut", "") for e in ens)
    print("%s : %d enseignants charges (titulaires %d, vacataires %d)."
          % (ONGLET, len(ens), rep.get("titulaire", 0), rep.get("vacataire", 0)))


if __name__ == "__main__":
    main()
