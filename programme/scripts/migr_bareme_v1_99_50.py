#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Migration V1.99.50 — barème des UE (EMSP_Notes.xlsx / N1_Bareme_UE).

IDEMPOTENT : peut etre relance sans effet de bord. Agit EN PLACE sur le
classeur des notes (PAS le maitre : EMSP_Notes.xlsx ne contient aucun dessin,
openpyxl.save y est autorise).

Actions :
  1. Colonne « Coef matiere (*) » inseree apres « Matiere » si absente ;
     valeur 1 sur toutes les lignes ou elle est vide (defaut officiel :
     moyenne d'UE arithmetique, conforme au releve officiel L2 SI).
  2. « Coef confirme » = Non sur TOUTES les lignes, L2 SI compris : le bareme
     present dans N1 suit la MAQUETTE REVISEE (UE16..UE24 en S3), alors que le
     releve officiel 2024-2025 suit l'ancienne numerotation (UE10..UE16). Aucun
     bareme de N1 n'est donc couvert ligne a ligne par un document officiel :
     tout est provisoire (mention reactivee a l'ecran et a l'impression), la
     scolarite confirmera filiere par filiere via l'ecran Bareme des UE. Les
     REGLES du moteur, elles, restent validees contre le releve officiel
     (banc d'essai : S3 = 9,89 ; sess. 2 = 11,87) sur un bareme temoin jetable.
  3. Soins infirmiers / Soins obstetricaux : Coef UE = ECTS/2 quand l'ECTS est
     renseigne (regle du programme de formation revise : Coef = Credit/2),
     sinon valeur existante conservee.
  4. Aides-soignants : bareme cree s'il est absent — 11 modules (programme
     revise), 1 ligne par module, semestre 1, niveau vide (aligne sur
     A1_Etudiants), coef 1 partout, Coef confirme = Non.

Usage :
    python scripts/migr_bareme_v1_99_50.py [chemin_EMSP_Notes.xlsx]
Sans argument : donnees/data/EMSP_Notes.xlsx relatif a la racine du kit,
sinon config.WORKBOOK_NOTES si importable.
"""
import os
import sys

import openpyxl

LIGNE_ENTETES = 2
LIGNE_DONNEES = 3

# Modules Aides-soignants (Programme de formation revise 2, onglet AS).
MODULES_AS = [
    ("M1", "Accompagnement et soins de la personne dans les activités de la vie quotidienne"),
    ("M2", "Repérage et prévention des situations à risque"),
    ("M3", "Evaluation de l'état clinique et mise en œuvre de soins adaptés"),
    ("M4", "Mise en œuvre des soins adaptés, évaluation et réajustement"),
    ("M5", "Accompagnement de la mobilité de la personne aidée"),
    ("M6", "Relation et communication avec les personnes et leur entourage"),
    ("M7", "Accompagnement des personnes en formation et communication avec les pairs"),
    ("M8", "Entretien des locaux et des matériels et prévention des risques associés"),
    ("M9", "Traitement des informations"),
    ("M10", "Travail en équipe pluri professionnelle, qualité et gestion des risques"),
    ("M11", "Adaptation aux valeurs locales"),
]


def _chemin_defaut():
    ici = os.path.dirname(os.path.abspath(__file__))
    for cand in (
        os.path.join(ici, "..", "..", "donnees", "data", "EMSP_Notes.xlsx"),
        os.path.join(ici, "..", "donnees", "data", "EMSP_Notes.xlsx"),
    ):
        cand = os.path.normpath(cand)
        if os.path.exists(cand):
            return cand
    try:
        sys.path.insert(0, os.path.normpath(os.path.join(ici, "..")))
        import config  # noqa: E402
        return config.WORKBOOK_NOTES
    except Exception:
        raise SystemExit("Classeur des notes introuvable : passer le chemin en argument.")


def _propre(lib):
    """Libelle sans marqueur (*) / (**)."""
    s = str(lib or "").strip()
    for m in (" (*)", " (**)"):
        if s.endswith(m):
            s = s[: -len(m)]
    return s.strip()


def _cols(ws):
    ent = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(LIGNE_ENTETES, c).value
        if v is not None and str(v).strip():
            ent[_propre(v)] = c
    return ent


def _num(x):
    try:
        return float(str(x).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


def main():
    chemin = sys.argv[1] if len(sys.argv) > 1 else _chemin_defaut()
    print("Classeur :", chemin)
    wb = openpyxl.load_workbook(chemin)
    ws = wb["N1_Bareme_UE"]
    cols = _cols(ws)

    # --- 1. Colonne Coef matiere -------------------------------------------
    if "Coef matiere" not in cols:
        pos = cols["Matiere"] + 1
        ws.insert_cols(pos)
        ws.cell(LIGNE_ENTETES, pos).value = "Coef matiere (*)"
        # style d'entete copie de la colonne Matiere
        src = ws.cell(LIGNE_ENTETES, cols["Matiere"])
        dst = ws.cell(LIGNE_ENTETES, pos)
        try:
            import copy
            dst.font = copy.copy(src.font); dst.fill = copy.copy(src.fill)
            dst.border = copy.copy(src.border); dst.alignment = copy.copy(src.alignment)
        except Exception:
            pass
        ws.column_dimensions[dst.column_letter].width = 8
        print("Colonne 'Coef matiere (*)' inseree en position", pos)
        cols = _cols(ws)
    else:
        print("Colonne 'Coef matiere' deja presente — rien a inserer.")

    cF, cN, cS = cols["Filiere"], cols["Niveau"], cols["Semestre"]
    cU, cI, cM = cols["N° UE"], cols["Intitule UE"], cols["Matiere"]
    cCM, cCU, cE, cK = cols["Coef matiere"], cols["Coef UE"], cols["ECTS UE"], cols["Coef confirme"]

    n_cm = n_conf_non = n_siso = 0
    a_as = False
    derniere = LIGNE_ENTETES
    for r in range(LIGNE_DONNEES, ws.max_row + 1):
        fil = str(ws.cell(r, cF).value or "").strip()
        if not fil:
            continue
        derniere = r
        filn = fil.lower()
        if filn.startswith("aides"):
            a_as = True
        # 1b. Coef matiere par defaut = 1
        if str(ws.cell(r, cCM).value or "").strip() in ("", "None"):
            ws.cell(r, cCM).value = 1
            n_cm += 1
        # 2. Tout provisoire : le bareme N1 (maquette revisee) n'est couvert par
        # aucun document officiel ligne a ligne — la scolarite confirmera.
        if str(ws.cell(r, cK).value or "").strip() != "Non":
            ws.cell(r, cK).value = "Non"
        n_conf_non += 1
        # 3. Regle Credit/2 du programme revise pour SI / SO
        if filn in ("soins infirmiers", "soins obstétricaux", "soins obstetricaux"):
            ects = _num(ws.cell(r, cE).value)
            if ects and ects > 0:
                cible = ects / 2.0
                if _num(ws.cell(r, cCU).value) != cible:
                    ws.cell(r, cCU).value = (int(cible) if cible == int(cible) else cible)
                    n_siso += 1

    # --- 4. Bareme Aides-soignants -----------------------------------------
    n_as = 0
    if not a_as:
        r = derniere + 1
        for num, mod in MODULES_AS:
            ws.cell(r, cF).value = "Aides-soignants"
            # Niveau volontairement vide (aligne sur A1_Etudiants)
            ws.cell(r, cS).value = "1"
            ws.cell(r, cU).value = num
            ws.cell(r, cI).value = mod
            ws.cell(r, cM).value = mod
            ws.cell(r, cCM).value = 1
            ws.cell(r, cCU).value = 1
            ws.cell(r, cK).value = "Non"
            r += 1
            n_as += 1
        print("Bareme Aides-soignants cree :", n_as, "modules.")
    else:
        print("Bareme Aides-soignants deja present — rien a creer.")

    wb.save(chemin)
    print("Coef matiere initialises a 1 :", n_cm)
    print("Coefs UE SI/SO recalcules ECTS/2 :", n_siso)
    print("Lignes en 'Coef confirme = Non' (bareme provisoire) :", n_conf_non)
    print("OK — migration V1.99.50 terminee.")


if __name__ == "__main__":
    main()
