# -*- coding: utf-8 -*-
"""Migration R2 (V1.99.36) — Statut enseignant en liste editable a 4 valeurs.

A executer UNE fois sur un deploiement existant pour mettre sa base
donnees/data/EMSP_V1.xlsx a niveau (la base n'est jamais ecrasee par une MAJ).
IDEMPOTENT : relançable sans effet de bord.

Operations (base active uniquement — 0 dessin, openpyxl sûr) :
  1) E1_Enseignants : renomme l'en-tete 'Statut (titulaire/vacataire) (*)' -> 'Statut (*)'
  2) E1_Enseignants : migre les valeurs   titulaire -> Permanent,  vacataire -> Vacataire
  3) P0_Parametres  : cree la colonne-liste 'Statuts_enseignant'
                      amorcee avec  Permanent / Contractuel / Vacataire / Benevole
  4) Dictionnaire   : champ 'Statut (...)' -> 'Statut', source 'Titulaire/Vacataire'
                      -> 'Statuts_enseignant'

Le classeur MAITRE (16 dessins) n'est PAS concerne ici : il a deja la structure
cible (chirurgie ZIP livree avec la version). Ne JAMAIS openpyxl.save le maitre.

Usage :
    python scripts/migr_statut_enseignant_v1_99_36.py [chemin_EMSP_V1.xlsx]
Sans argument : utilise config.WORKBOOK.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402
import openpyxl  # noqa: E402

SEEDS = ["Permanent", "Contractuel", "Vacataire", "Bénévole"]
MIGRATION = {"titulaire": "Permanent", "vacataire": "Vacataire"}
LE, LD = config.LIGNE_ENTETES, config.LIGNE_DONNEES


def _col_par_entete(ws, prefixe=None, exact=None):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=LE, column=c).value
        if exact is not None and v == exact:
            return c
        if prefixe is not None and v and str(v).startswith(prefixe):
            return c
    return None


def migrer(chemin):
    wb = openpyxl.load_workbook(chemin)  # garde les formules (pas data_only)
    rapport = []

    # 1 + 2) E1_Enseignants : en-tete + valeurs
    e1 = wb["E1_Enseignants"]
    col = _col_par_entete(e1, prefixe="Statut")
    if col is None:
        raise RuntimeError("Colonne statut introuvable dans E1_Enseignants.")
    ancien = e1.cell(row=LE, column=col).value
    if ancien != "Statut (*)":
        e1.cell(row=LE, column=col).value = "Statut (*)"
        rapport.append("E1 en-tete '%s' -> 'Statut (*)'" % ancien)
    else:
        rapport.append("E1 en-tete deja 'Statut (*)' (rien a faire)")
    nmig = 0
    for r in range(LD, e1.max_row + 1):
        v = e1.cell(row=r, column=col).value
        if v is None:
            continue
        k = str(v).strip().lower()
        if k in MIGRATION and str(v).strip() != MIGRATION[k]:
            e1.cell(row=r, column=col).value = MIGRATION[k]
            nmig += 1
    rapport.append("E1 valeurs migrees : %d" % nmig)

    # 3) P0_Parametres : colonne-liste Statuts_enseignant
    p0 = wb["P0_Parametres"]
    col0 = _col_par_entete(p0, exact="Statuts_enseignant")
    if col0 is None:
        last = 0
        for c in range(1, p0.max_column + 1):
            if p0.cell(row=LE, column=c).value not in (None, ""):
                last = c
        col0 = last + 1
        p0.cell(row=LE, column=col0).value = "Statuts_enseignant"
        for i, val in enumerate(SEEDS):
            p0.cell(row=LD + i, column=col0).value = val
        rapport.append("P0 colonne 'Statuts_enseignant' creee (%d valeurs)" % len(SEEDS))
    else:
        # complete les valeurs par defaut manquantes (sans doublon, sans ecraser)
        existantes = []
        r = LD
        while True:
            v = p0.cell(row=r, column=col0).value
            if v in (None, ""):
                break
            existantes.append(str(v).strip())
            r += 1
        ajoutes = 0
        for val in SEEDS:
            if val not in existantes:
                p0.cell(row=r, column=col0).value = val
                existantes.append(val)
                r += 1
                ajoutes += 1
        rapport.append("P0 'Statuts_enseignant' deja presente (%d valeur(s) defaut ajoutee(s))" % ajoutes)

    # 4) Dictionnaire : champ + source
    dic = wb["Dictionnaire"]
    hdr = {dic.cell(row=LE, column=c).value: c for c in range(1, dic.max_column + 1)}
    cc, ct, cl = hdr["Onglet"], hdr["Champ"], hdr["Liste / source"]
    maj = False
    for r in range(LD, dic.max_row + 1):
        if dic.cell(row=r, column=cc).value == "E1_Enseignants" \
                and str(dic.cell(row=r, column=ct).value or "").startswith("Statut"):
            if dic.cell(row=r, column=ct).value != "Statut":
                dic.cell(row=r, column=ct).value = "Statut"
                maj = True
            if dic.cell(row=r, column=cl).value != "Statuts_enseignant":
                dic.cell(row=r, column=cl).value = "Statuts_enseignant"
                maj = True
            break
    rapport.append("Dictionnaire ligne statut maj : %s" % ("oui" if maj else "deja a jour"))

    wb.save(chemin)
    return rapport


if __name__ == "__main__":
    chemin = sys.argv[1] if len(sys.argv) > 1 else config.WORKBOOK
    if not os.path.exists(chemin):
        sys.exit("Classeur introuvable : %s" % chemin)
    print("Migration R2 (statut enseignant) sur : %s" % chemin)
    for ligne in migrer(chemin):
        print("  - " + ligne)
    print("Termine.")
