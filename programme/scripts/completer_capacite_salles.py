# -*- coding: utf-8 -*-
"""
EMSP — Complete la colonne « Capacite » de L1_Salles depuis le nombre de chaises
de l'inventaire des salles, sur les LIGNES EXISTANTES uniquement (match par ID).

NE CREE AUCUNE LIGNE, NE PURGE RIEN, NE TOUCHE PAS aux autres colonnes : le
referentiel des salles reste celui de import_patrimoine_bloc1.py (cours, TP,
salles specialisees, bureaux, divers) et son rattachement a M1_Equipements.

Source (validee Bernard) : nb de chaises releve dans INVENTAIRE_DES_SALLES_DES_CLASSES.
  SC02=37  SC04=86  SC06=87  SC07=19  SC08=25  SC09=50  SC10=74  SINFO=42
TP / bureaux / divers : pas de comptage chaises -> laisses vides.

ATTENTION DESSINS : ecrit via openpyxl sur la COPIE DEPLOYEE (donnees/data),
jamais sur le master. Idempotent : reapplique les memes valeurs, rejouable.

Lancement (depuis la racine du programme) :  python scripts/completer_capacite_salles.py
Chemin surchargeable :                       EMSP_DB=...\\EMSP_V1.xlsx python scripts/completer_capacite_salles.py
"""
import os, sys
from openpyxl import load_workbook

ICI = os.path.dirname(os.path.abspath(__file__))
RACINE = os.path.dirname(ICI)
sys.path.insert(0, RACINE)
try:
    import config
    DEFAUT = config.WORKBOOK
    L_ENT, L_DON = config.LIGNE_ENTETES, config.LIGNE_DONNEES
except Exception:
    DEFAUT = os.path.join(RACINE, "donnees", "data", "EMSP_V1.xlsx")
    L_ENT, L_DON = 2, 3

XLSX = os.environ.get("EMSP_DB") or DEFAUT
ONGLET = "L1_Salles"

CAPACITES = {
    "SC02": 37, "SC04": 86, "SC06": 87, "SC07": 19,
    "SC08": 25, "SC09": 50, "SC10": 74, "SINFO": 42,
}


def _propre(h):
    h = "" if h is None else str(h).strip()
    for suf in (" (**)", " (*)", "(**)", "(*)"):
        if h.endswith(suf):
            return h[: -len(suf)].strip()
    return h


def _index(ws):
    idx = {}
    for c in range(1, ws.max_column + 1):
        lib = _propre(ws.cell(L_ENT, c).value)
        if lib:
            idx[lib] = c
    return idx


def main():
    if not os.path.exists(XLSX):
        raise SystemExit("Classeur introuvable : %s" % XLSX)
    wb = load_workbook(XLSX)
    if ONGLET not in wb.sheetnames:
        raise SystemExit("Onglet %s absent du classeur." % ONGLET)
    ws = wb[ONGLET]
    ix = _index(ws)
    c_id = ix.get("ID salle")
    c_cap = ix.get("Capacite")
    if not c_id or not c_cap:
        raise SystemExit("Colonnes 'ID salle' / 'Capacite' introuvables.")

    maj = 0
    for r in range(L_DON, ws.max_row + 1):
        sid = ws.cell(r, c_id).value
        sid = "" if sid is None else str(sid).strip()
        if sid in CAPACITES:
            ws.cell(r, c_cap).value = CAPACITES[sid]
            maj += 1

    wb.save(XLSX)
    print("%s : capacite renseignee sur %d salle(s) existante(s) (sur %d prevues)."
          % (ONGLET, maj, len(CAPACITES)))


if __name__ == "__main__":
    main()
