# -*- coding: utf-8 -*-
"""LOGIQUE METIER — calculs et interpretations a partir de la couche d'acces.
Separee de data.py (acces) et de app.py (presentation / routage).
"""
import re
import os
import json
import datetime
import unicodedata
import shutil
import config
import auth
from data import AccesDonnees

_db = AccesDonnees()
# Fichier des notes SEPARE (confidentialite). Couche d'acces dediee ; le routage par
# onglet se fait via _db_pour(). Les onglets de notes sont declares dans config.ONGLETS_NOTES.
_db_notes = AccesDonnees(config.WORKBOOK_NOTES)



def couleur_login(login):
    """Couleur d'identite d'un login (poste partage). PRIORITE a la couleur CHOISIE
    par l'informatique (instance/comptes.json) ; a defaut, derivee du login par
    hachage sur la palette config.COULEURS_UTILISATEUR. Login vide -> bleu charte."""
    choisie = auth.couleur(login)
    if choisie:
        return choisie
    pal = getattr(config, "COULEURS_UTILISATEUR", ["#1F4E79"]) or ["#1F4E79"]
    s = str(login or "").strip().lower()
    if not s:
        return "#1F4E79"
    h = 0
    for ch in s:
        h = (h * 31 + ord(ch)) & 0xFFFFFFFF
    return pal[h % len(pal)]

def _db_pour(onglet):
    """Retourne la couche d'acces du bon classeur selon l'onglet (notes vs principal)."""
    return _db_notes if onglet in getattr(config, "ONGLETS_NOTES", ()) else _db


# ---------------------------------------------------------------------------
# Provenance des champs : (sans marqueur)=TDR, (*)=initiative, (**)=hors TDR
# ---------------------------------------------------------------------------
def decoupe_provenance(entete):
    """Renvoie (libelle_propre, code_provenance) pour un en-tete brut."""
    if entete is None:
        return "", "TDR"
    txt = str(entete).strip()
    if txt.endswith("(**)"):
        return txt[:-4].strip(), "**"
    if txt.endswith("(*)"):
        return txt[:-3].strip(), "*"
    return txt, "TDR"


def entetes_meta(onglet):
    """Liste de dicts {libelle, brut, prov, prov_info, readonly} pour un onglet."""
    bruts = _db_pour(onglet).entetes(onglet)
    ro_cols = set(config.READONLY_COLS.get(onglet, []))
    tab_ro = onglet in config.READONLY_TABS
    meta = []
    for b in bruts:
        lib, prov = decoupe_provenance(b)
        meta.append({
            "libelle": lib,
            "brut": b,
            "prov": prov,
            "prov_info": config.PROVENANCE[prov],
            "readonly": tab_ro or (b in ro_cols),
        })
    return meta


def table(onglet, limite=None):
    """Donnees pretes pour affichage : en-tetes meta + lignes.
    Les colonnes calcul (formules) sont remplies en Python pour l'affichage IHM
    (Decision (a) V1.9) : openpyxl n'evalue pas les formules des lignes ajoutees,
    donc la cellule serait vide tant qu'Excel n'a pas recalcule. La formule reste
    dans le classeur pour Excel ; seul l'affichage est calcule ici."""
    db = _db_pour(onglet)
    entetes = entetes_meta(onglet)
    lignes = _appliquer_calc_affichage(onglet, entetes, db.lignes(onglet, limite))
    # Colonnes d'affichage SUPPLEMENTAIRES (calculees a la volee, non stockees).
    extra = COLONNES_AFFICHAGE_EXTRA.get(onglet)
    if extra:
        entetes = list(entetes)
        libelles = [m["libelle"] for m in entetes]
        lignes = [list(l) for l in lignes]
        for spec in extra:
            pinfo = next((m["prov_info"] for m in entetes
                          if m["libelle"] == spec.get("prov_ref")),
                         entetes[-1]["prov_info"] if entetes else None)
            entetes.append({"libelle": spec["libelle"], "brut": spec["libelle"],
                            "prov": None, "prov_info": pinfo, "readonly": True})
            fn = spec["calc"]
            for lig in lignes:
                rowd = {libelles[j]: (lig[j] if j < len(lig) else "")
                        for j in range(len(libelles))}
                try:
                    lig.append(fn(rowd))
                except Exception:
                    lig.append("")
    return {
        "titre": db.titre(onglet),
        "entetes": entetes,
        "lignes": lignes,
        "nb": db.nb_lignes(onglet),
    }


# Affichage des COLONNES CALCUL (Decision (a) V1.9). Pour chaque onglet a colonne
# calcul, une regle {libelle_colonne_calcul: fonction(row_dict)->valeur}. Le row_dict
# est indexe par libelle propre (sans marqueur). On ne remplit QUE si la cellule lue
# est vide (formule non encore evaluee) : si Excel a deja calcule la valeur, on la garde.
# Reutilisable pour F2_Comptes / IMPORT_zone lors de leur activation.
def _num(x):
    """Convertit une valeur de cellule en nombre (tolere espaces, virgule decimale)."""
    try:
        return float(str(x).replace(" ", "").replace(",", ".") or 0)
    except (ValueError, TypeError):
        return 0.0


def _fmt_kmf(n):
    """Montant sans decimales inutiles, separateur de milliers ESPACE (charte).
    Entier -> '799 423' (coherent avec _kmf_aff) ; valeur fractionnaire (ex. heures)
    -> decimales conservees. Tolere les chaines (espaces, virgule decimale)."""
    try:
        n = round(float(_num(n)), 2)
    except (TypeError, ValueError):
        return str(n)
    if abs(n - round(n)) < 1e-9:
        return "{:,.0f}".format(round(n)).replace(",", " ")
    ent, dec = ("%.2f" % n).split(".")
    ent = "{:,d}".format(int(ent)).replace(",", " ")
    return (ent + "." + dec).rstrip("0").rstrip(".")


def _solde_courant_f2(r):
    """Solde courant d'un compte (F2) = solde initial + recettes - depenses du compte
    (lus dans F1). Calcule en Python pour un affichage LIVE (les saisies F1 dans l'IHM
    ne declenchent pas le recalcul Excel). r = ligne F2 {libelle: valeur}."""
    compte = str(r.get("Nom du compte / caisse", "")).strip()
    if not compte:
        return ""
    net = _num(r.get("Solde initial (KMF)", 0))
    comptes = _db.colonne("F1_Mouvements", _brut("F1_Mouvements", "Compte / caisse"))
    rec = _db.colonne("F1_Mouvements", _brut("F1_Mouvements", "Montant Recette (KMF)"))
    dep = _db.colonne("F1_Mouvements", _brut("F1_Mouvements", "Montant Depense (KMF)"))
    for i, c in enumerate(comptes):
        if str(c).strip() == compte:
            net += _num(rec[i] if i < len(rec) else 0) - _num(dep[i] if i < len(dep) else 0)
    return _fmt_kmf(net)


# Affichage des COLONNES CALCUL (V1.9, etendu V1.15). Calcule en Python pour l'IHM.
# E2 = total trivial ; F2 = solde live (depend de F1, donc TOUJOURS recalcule a
# l'affichage car la valeur en cache Excel serait perimee apres une saisie F1).
def _statut_vs_base_import(r):
    """IMPORT_zone : NOUVEAU si le matricule n'est pas dans A1_Etudiants, sinon
    EXISTANT (calcul Python live, comme la formule H du classeur)."""
    mat = str(r.get("Matricule", "")).strip()
    if not mat:
        return ""
    base = set(str(x).strip() for x in _db.colonne("A1_Etudiants",
               _brut("A1_Etudiants", "Matricule")) if str(x).strip())
    return "NOUVEAU" if mat not in base else "EXISTANT"


# Affichage des COLONNES CALCUL (V1.9, etendu V1.15/V1.16). Calcule en Python pour l'IHM.
# E2 = total trivial ; F2 = solde live (depend de F1) ; IMPORT = statut vs base (depend de A1).
# TOUJOURS recalcule (la valeur en cache Excel serait perimee apres une saisie liee).
CALC_AFFICHAGE = {
    "E2_Releve_heures": {
        "Total heures a payer": lambda r: r.get("Vol. horaire constate", ""),
    },
    "F2_Comptes": {
        "Solde courant (KMF)": _solde_courant_f2,
    },
    "IMPORT_zone": {
        "Statut vs base": _statut_vs_base_import,
    },
}


def _fmt_ecart(n):
    """Ecart d'heures : sans decimale inutile, signe explicite (+ si programme > constate)."""
    n = round(n, 2)
    if n == int(n):
        n = int(n)
    return ("+" + str(n)) if n > 0 else str(n)


def _num_h(x):
    """Conversion robuste d'une valeur d'heures (tolere virgule decimale et espaces)."""
    try:
        return float(str(x).replace(" ", "").replace(",", ".") or 0)
    except (ValueError, TypeError):
        return 0.0


def _ecart_prog_constate(rowd, lib_prog="Vol. horaire prog.", lib_cons="Vol. horaire constate"):
    """Ecart = programme - constate. Vide tant que les DEUX valeurs ne sont pas
    saisies (sinon un programme seul afficherait un faux ecart avant le constate)."""
    sp = str(rowd.get(lib_prog, "")).strip()
    sc = str(rowd.get(lib_cons, "")).strip()
    if sp == "" or sc == "":
        return ""
    return _fmt_ecart(_num_h(rowd.get(lib_prog, 0)) - _num_h(rowd.get(lib_cons, 0)))


# Colonnes AJOUTEES A L'AFFICHAGE (calculees, NON stockees au classeur). A distinguer
# de CALC_AFFICHAGE (qui remplit une colonne formule deja presente) : ici on ajoute une
# colonne VIRTUELLE en fin de table, pour la consultation seulement. Le classeur reste
# inchange. 'prov_ref' = libelle d'une colonne existante dont on reprend la provenance
# (couleur/etoile) ; la colonne est marquee 'calcul' (lecture seule).
COLONNES_AFFICHAGE_EXTRA = {
    "E2_Releve_heures": [
        {"libelle": "Ecart (prog. - constate)", "calc": _ecart_prog_constate,
         "prov_ref": "Vol. horaire prog."},
    ],
}


def _appliquer_calc_affichage(onglet, entetes, lignes):
    regles = CALC_AFFICHAGE.get(onglet)
    if not regles:
        return lignes
    libelles = [m["libelle"] for m in entetes]
    idx = {lib: i for i, lib in enumerate(libelles)}
    out = []
    for lig in lignes:
        lig = list(lig)
        rowd = {libelles[i]: (lig[i] if i < len(lig) else "")
                for i in range(len(libelles))}
        for col_lib, fn in regles.items():
            i = idx.get(col_lib)
            if i is None or i >= len(lig):
                continue
            try:                      # TOUJOURS recalcule (valeur Excel possiblement perimee)
                lig[i] = fn(rowd)
            except Exception:
                pass
        out.append(lig)
    return out


def dictionnaire_par_onglet():
    """Dictionnaire regroupe par onglet, pour les pages d'aide et la saisie.
    Complete par config.DICTIONNAIRE_SUPPLEMENT (onglets ajoutes par chirurgie, non
    presents dans l'onglet Dictionnaire) et corrige par config.DICTIONNAIRE_SURCHARGE
    (remplacement de cles d'entrees existantes) — sans modifier le classeur."""
    groupes = {}
    for r in _db.dictionnaire():
        groupes.setdefault(r["Onglet"], []).append(dict(r))
    # Surcharges (ex. S1 'Lieu de stage' -> source S2)
    for onglet, champs in getattr(config, "DICTIONNAIRE_SURCHARGE", {}).items():
        for d in groupes.get(onglet, []):
            maj = champs.get(d.get("Champ"))
            if maj:
                d.update(maj)
    # Supplements (ex. S2_Lieux_stage)
    for onglet, lignes in getattr(config, "DICTIONNAIRE_SUPPLEMENT", {}).items():
        if onglet not in groupes:
            groupes[onglet] = [dict(x) for x in lignes]
    # Champs ajoutes a un onglet DEJA present (ex. G1 'Type d'écart') : on append
    # les champs absents (rapproches par 'Champ'), sans modifier le classeur.
    for onglet, lignes in getattr(config, "DICTIONNAIRE_CHAMPS_SUP", {}).items():
        existants = {d.get("Champ") for d in groupes.get(onglet, [])}
        for x in lignes:
            if x.get("Champ") not in existants:
                groupes.setdefault(onglet, []).append(dict(x))
    return groupes


# ---------------------------------------------------------------------------
# Nomenclature budgetaire — ecran de curation (C-2)
# ---------------------------------------------------------------------------
def donnees_nomenclature(filtres=None):
    """Table P3_Nomenclature filtrable (sens / source / actif / recherche) +
    compteurs globaux. Lecture seule (la bascule passe par basculer_actif_codes)."""
    f = filtres or {}
    sens = (f.get("sens") or "").strip()
    source = (f.get("source") or "").strip()
    actif = (f.get("actif") or "").strip().lower()
    q = (f.get("q") or "").strip().lower()
    rows = []
    for r in _lignes_dict("P3_Nomenclature"):
        code = str(r.get("Code", "")).strip()
        if not code:
            continue
        rows.append({
            "code": code,
            "intitule": str(r.get("Intitule", "")).strip(),
            "sens": str(r.get("Sens", "")).strip(),
            "niveau": str(r.get("Niveau", "")).strip(),
            "source": str(r.get("Source", "")).strip(),
            "actif": str(r.get("Actif", "")).strip(),
        })
    total = len(rows)
    nb_actifs = sum(1 for r in rows if r["actif"].lower() == "oui")

    def garde(r):
        if sens and r["sens"] != sens:
            return False
        if source and r["source"] != source:
            return False
        if actif and r["actif"].lower() != actif:
            return False
        if q and q not in r["code"].lower() and q not in r["intitule"].lower():
            return False
        return True

    filtrees = [r for r in rows if garde(r)]
    filtrees.sort(key=lambda r: (r["sens"], r["code"]))
    return {"lignes": filtrees, "total": total, "nb_actifs": nb_actifs,
            "nb_reserve": total - nb_actifs, "nb_filtre": len(filtrees),
            "sens_opts": ["Recette", "Depense", "Investissement"],
            "source_opts": ["OHADA", "EMSP"]}


def basculer_actif_codes(codes, actif):
    """Bascule la colonne Actif (Oui/Non) pour une liste de codes EXISTANTS de
    P3_Nomenclature. Upsert cle = Code ; ne touche qu'a la colonne Actif. Les
    codes inconnus sont ignores (jamais de ligne partielle creee)."""
    val = "Oui" if str(actif).strip().lower() in ("oui", "1", "true", "on") else "Non"
    existants = {str(r.get("Code", "")).strip()
                 for r in _lignes_dict("P3_Nomenclature")}
    cibles = [c for c in {str(x).strip() for x in (codes or [])}
              if c and c in existants]
    if not cibles:
        return 0
    lignes = [{"Code": c, "Actif": val} for c in cibles]
    _db.ecrire_lignes_lot("P3_Nomenclature", lignes, cles=["Code"])
    return len(cibles)


def kpis(filtres=None):
    """Indicateurs du tableau de bord, filtrables (V1.71).
    Etudiants et presences reagissent a filiere/niveau/annee/periode.
    Finances et heures ne reagissent qu'a annee/periode (F1/E2 ne portent pas
    filiere/niveau) : si un filtre filiere/niveau est actif, on les marque
    non_filtrable=True pour que l'IHM grise/masque ces cartes (option B)."""
    f = filtres or {}
    fn_actif = bool(f.get("filiere") or f.get("niveau"))
    et_rows = _lignes_filtrees("A1_Etudiants", f)
    pr_rows = _lignes_filtrees("A2_Presences", f)
    mv_rows = _lignes_filtrees("F1_Mouvements", f)
    he_rows = _lignes_filtrees("E2_Releve_heures", f)
    statuts = [str(r.get("Statut", "")).strip().lower() for r in et_rows]
    pres = [str(r.get("Present (O/N)", "")).strip() for r in pr_rows]
    nb_pres = len([p for p in pres if p != ""])
    taux = (len([p for p in pres if p.upper() == "O"]) / nb_pres) if nb_pres else 0.0
    rec = sum(_num(r.get("Montant Recette (KMF)", 0)) for r in mv_rows)
    dep = sum(_num(r.get("Montant Depense (KMF)", 0)) for r in mv_rows)
    heures = sum(_num(r.get("Vol. horaire constate", 0)) for r in he_rows)
    nb_ens = len([r for r in _lignes_filtrees("E1_Enseignants", f)
                  if str(r.get("Matricule ens.", "")).strip()])
    return {
        "etudiants": len([r for r in et_rows if str(r.get("Matricule", "")).strip()]),
        "actifs": len([s for s in statuts if s == "actif"]),
        "diplomes": len([s for s in statuts if s == "diplome"]),
        "recettes": rec, "depenses": dep, "solde": rec - dep,
        "taux_presence": round(taux * 100, 1),
        "heures": heures, "enseignants": nb_ens,
        "reste_du": reste_du_par_filiere(f)["total_reste"],
        "finances_non_filtrable": fn_actif,
        "heures_non_filtrable": fn_actif,
    }


def _compte_par(valeurs, ordre=None):
    c = {}
    for v in valeurs:
        k = str(v).strip()
        if k == "":
            continue
        c[k] = c.get(k, 0) + 1
    if ordre:
        labels = [o for o in ordre if o in c] + [k for k in c if k not in ordre]
    else:
        labels = sorted(c)
    return labels, [c[l] for l in labels]


def donnees_graphiques(filtres=None):
    """Jeux de donnees des graphiques, filtrables (V1.71).
    Finances et heures ne reagissent qu'a annee/periode ; si un filtre
    filiere/niveau est actif, ils sont marques non_filtrable (l'IHM les grise)."""
    f = filtres or {}
    fn_actif = bool(f.get("filiere") or f.get("niveau"))
    par = _db.listes_parametres()
    et_rows = _lignes_filtrees("A1_Etudiants", f)
    pr_rows = _lignes_filtrees("A2_Presences", f)
    mv_rows = _lignes_filtrees("F1_Mouvements", f)
    he_rows = _lignes_filtrees("E2_Releve_heures", f)

    # 1) Effectif par filiere (toutes les filieres parametrees apparaissent)
    fcount = {}
    for r in et_rows:
        k = str(r.get("Filiere", "")).strip()
        if k:
            fcount[k] = fcount.get(k, 0) + 1
    fil_labels = [str(x).strip() for x in par.get("Filieres", []) if str(x).strip()]
    for k in fcount:
        if k not in fil_labels:
            fil_labels.append(k)
    filieres = {"labels": fil_labels, "valeurs": [fcount.get(l, 0) for l in fil_labels]}

    # 2) Statuts
    sl, sv = _compte_par([r.get("Statut", "") for r in et_rows],
                         ordre=["Actif", "Diplome", "Abandonne", "Radie"])
    statuts = {"labels": sl, "valeurs": sv}

    # 3) Finances par categorie (annee/periode uniquement)
    catset, r_by, d_by = [], {}, {}
    for r in mv_rows:
        k = str(r.get("Categorie", "")).strip()
        if not k:
            continue
        if k not in catset:
            catset.append(k)
        r_by[k] = r_by.get(k, 0) + _num(r.get("Montant Recette (KMF)", 0))
        d_by[k] = d_by.get(k, 0) + _num(r.get("Montant Depense (KMF)", 0))
    finances = {"labels": catset, "non_filtrable": fn_actif,
                "series": [
                    {"nom": "Recettes", "valeurs": [r_by.get(k, 0) for k in catset]},
                    {"nom": "Depenses", "valeurs": [d_by.get(k, 0) for k in catset]},
                ]}

    # 4) Presence par creneau
    tot, oui = {}, {}
    for r in pr_rows:
        k = str(r.get("Creneau", "")).strip()
        if not k:
            continue
        tot[k] = tot.get(k, 0) + 1
        if str(r.get("Present (O/N)", "")).strip().upper() == "O":
            oui[k] = oui.get(k, 0) + 1
    cr_labels = ["10h", "12h", "15h", "17h"]
    cr_labels = [c for c in cr_labels if c in tot] + [c for c in tot if c not in cr_labels]
    presence = {"labels": cr_labels,
                "valeurs": [round(100 * oui.get(k, 0) / tot[k], 1) if tot.get(k) else 0
                            for k in cr_labels]}

    # 5) Heures par enseignant (annee/periode uniquement)
    hby = {}
    for r in he_rows:
        k = str(r.get("Matricule ens.", "")).strip()
        if not k:
            continue
        hby[k] = hby.get(k, 0) + _num(r.get("Vol. horaire constate", 0))
    heures = {"labels": list(hby.keys()), "valeurs": list(hby.values()),
              "non_filtrable": fn_actif}

    return {"filieres": filieres, "statuts": statuts, "finances": finances,
            "presence": presence, "heures": heures,
            "reste_du_filiere": reste_du_par_filiere(f)}


# ===========================================================================
# CALENDRIER & SALLES (vues facon Outlook) — V1.1
# ---------------------------------------------------------------------------
# Rappel perimetre : A3_Sessions est un emploi du temps HEBDOMADAIRE recurrent
# (colonne Jour = Lun..Sam, pas de date). Les vues mois/semaine/jour PROJETTENT
# cette grille. Affichage seul (Option 1) : aucune ecriture dans le classeur.
# ===========================================================================
import datetime as _dt

_JOURS = config.JOURS_PLEINS  # index 0 = Lundi ... 6 = Dimanche


def _lignes_dict(onglet):
    """Lignes d'un onglet sous forme de dicts {libelle_propre: valeur}."""
    bruts = _db.entetes(onglet)
    cles = [decoupe_provenance(b)[0] for b in bruts]
    out = []
    for lig in _db.lignes(onglet):
        d = {}
        for i, c in enumerate(cles):
            d[c] = lig[i] if i < len(lig) else ""
        out.append(d)
    return out


def _parse_heure(txt):
    """'8h00' / '8h' / '10h05' / '08:00' -> minutes depuis minuit (ou None)."""
    if txt in (None, ""):
        return None
    s = str(txt).strip().lower().replace(" ", "")
    m = re.match(r"^(\d{1,2})[h:](\d{0,2})$", s)
    if not m:
        m = re.match(r"^(\d{1,2})$", s)
        if m:
            return int(m.group(1)) * 60
        return None
    h = int(m.group(1))
    mn = int(m.group(2)) if m.group(2) else 0
    return h * 60 + mn


def fmt_date(d):
    """date -> 'JJ/MM/AAAA' (charte)."""
    return d.strftime("%d/%m/%Y")


def normaliser_date_saisie(s=""):
    """Normalise une date de saisie (FR ou ISO) -> (JJ/MM/AAAA, AAAA-MM-JJ).
    Vide ou invalide -> jour courant : pratique pour la saisie des presences,
    ou la date du jour est proposee par defaut (calendrier cliquable)."""
    d = _parse_date_fr(s) if s else None
    if d is None:
        d = _dt.date.today()
    return fmt_date(d), d.isoformat()


# --- Donnees de demonstration (visualiser le rendu tant que le classeur est vide) ---
def _salles_demo():
    return [
        {"id": "SAL-01", "nom": "Amphi A", "type": "Amphitheatre", "capacite": "120",
         "equipements": "Videoprojecteur; Sonorisation; Tableau", "batiment": "Batiment principal"},
        {"id": "SAL-02", "nom": "Salle 1", "type": "Cours", "capacite": "40",
         "equipements": "Tableau; Videoprojecteur", "batiment": "Batiment principal"},
        {"id": "SAL-03", "nom": "Salle 2", "type": "Cours", "capacite": "40",
         "equipements": "Tableau", "batiment": "Batiment principal"},
        {"id": "SAL-04", "nom": "Labo TP", "type": "TP", "capacite": "24",
         "equipements": "Paillasses; Microscopes; Point d'eau", "batiment": "Annexe laboratoire"},
    ]


def _seances_demo():
    g = "L1 / S.I"
    base = [
        ("Lundi", "8h00", "10h00", "Liquide et electrolytes", "Dr Mohamed Moundhirou", "Amphi A", "Cours"),
        ("Lundi", "10h05", "12h05", "Communication", "M. Abdou Mroimana", "Salle 1", "Cours"),
        ("Lundi", "13h00", "15h00", "Anatomie appareil locomoteur", "Dr Mbae Toyib", "Salle 1", "Cours"),
        ("Mardi", "8h00", "10h00", "Liquide et electrolytes", "Dr Mohamed Moundhirou", "Amphi A", "Cours"),
        ("Mardi", "10h05", "12h05", "Biochimie", "Dr El Habib", "Salle 2", "Cours"),
        ("Mercredi", "8h00", "10h00", "Systeme neurologique", "Dr Moussa Mohamed", "Amphi A", "Cours"),
        ("Mercredi", "10h05", "12h05", "Microbiologie", "Dr El Habib", "Salle 2", "Cours"),
        ("Jeudi", "8h00", "10h00", "Expression francaise", "Mr Ameldine", "Salle 1", "Cours"),
        ("Jeudi", "13h00", "15h00", "Examens de laboratoire", "Dr Ahamada Fazul", "Labo TP", "TP"),
        ("Vendredi", "8h00", "10h00", "Langue chinoise", "M. Mreha Abdou", "Amphi A", "Cours"),
    ]
    out = []
    for i, (j, d, f, mat, ens, sal, typ) in enumerate(base, 1):
        out.append({
            "id": "DEMO-%02d" % i, "filiere": "Soins infirmiers", "niveau": "L1",
            "section": "S.I", "matiere": mat, "enseignant": ens, "salle": sal,
            "jour": j, "debut": d, "fin": f, "type": typ, "groupe": g,
            "deb_min": _parse_heure(d), "fin_min": _parse_heure(f),
        })
    return out


def _presences_demo(date_iso):
    # repartition fictive stable par creneau
    return {"10h": {"present": 44, "total": 48}, "12h": {"present": 41, "total": 48},
            "15h": {"present": 39, "total": 47}, "17h": {"present": 35, "total": 45}}


# --- Lecture reelle (depuis le classeur) ---
def seances(demo=False):
    if demo:
        return _seances_demo()
    out = []
    for r in _lignes_dict("A3_Sessions"):
        jour = str(r.get("Jour", "")).strip()
        if not jour and not str(r.get("Matiere", "")).strip():
            continue
        grp = " / ".join([x for x in [str(r.get("Filiere", "")).strip(),
                                       str(r.get("Niveau", "")).strip(),
                                       str(r.get("Section", "")).strip()] if x])
        out.append({
            "id": r.get("ID session", ""), "filiere": r.get("Filiere", ""),
            "niveau": r.get("Niveau", ""), "section": r.get("Section", ""),
            "matiere": r.get("Matiere", ""), "enseignant": r.get("Enseignant", ""),
            "salle": str(r.get("Salle", "")).strip(), "jour": jour,
            "debut": r.get("Heure debut", ""), "fin": r.get("Heure fin", ""),
            "type": r.get("Type", ""), "groupe": grp,
            "deb_min": _parse_heure(r.get("Heure debut", "")),
            "fin_min": _parse_heure(r.get("Heure fin", "")),
        })
    return out


def salles(demo=False):
    if demo:
        return _salles_demo()
    out = []
    for r in _lignes_dict("L1_Salles"):
        if not str(r.get("ID salle", "")).strip() and not str(r.get("Nom / libelle", "")).strip():
            continue
        out.append({
            "id": str(r.get("ID salle", "")).strip(),
            "nom": str(r.get("Nom / libelle", "")).strip(),
            "type": r.get("Type", ""), "capacite": r.get("Capacite", ""),
            "equipements": r.get("Equipements", ""),
            "batiment": r.get("Batiment / localisation", ""),
        })
    if not out:
        # V1.89 : L1_Salles vide -> on derive les salles des noms presents dans
        # l'inventaire materiel (M1 « Salle / localisation »), en attendant une
        # numerotation officielle. Lecture seule, aucune ecriture.
        compte = {}
        for r in _lignes_dict("M1_Equipements"):
            nom = str(r.get("Salle / localisation", "")).strip()
            if nom:
                compte[nom] = compte.get(nom, 0) + 1
        for nom in sorted(compte):
            out.append({"id": "", "nom": nom, "type": "", "capacite": "",
                        "equipements": "%d equipement(s) inventorie(s)" % compte[nom],
                        "batiment": "", "depuis_materiel": True})
    return out


def _seance_dans_salle(seance, salle):
    """Une seance occupe la salle si Salle reference son ID ou son nom."""
    ref = str(seance.get("salle", "")).strip().lower()
    return ref and ref in (str(salle["id"]).strip().lower(),
                           str(salle["nom"]).strip().lower())


def presences_du_jour(date_iso, demo=False):
    """Synthese des presences pour une date (JJ/MM/AAAA), par creneau."""
    if demo:
        return _presences_demo(date_iso)
    try:
        d = _dt.date.fromisoformat(date_iso)
    except (TypeError, ValueError):
        return {}
    cible = fmt_date(d)
    res = {c: {"present": 0, "total": 0} for c in config.CRENEAUX}
    for r in _lignes_dict("A2_Presences"):
        dt = str(r.get("Date", "")).strip()
        if dt != cible and dt != date_iso:
            continue
        cr = str(r.get("Creneau", "")).strip()
        if cr not in res:
            res[cr] = {"present": 0, "total": 0}
        res[cr]["total"] += 1
        if str(r.get("Present (O/N)", "")).strip().upper() == "O":
            res[cr]["present"] += 1
    return {k: v for k, v in res.items() if v["total"] > 0}


# --- Vues calendrier ---
def _aujourdhui():
    return _dt.date.today()


def _date_ou_defaut(date_iso):
    try:
        return _dt.date.fromisoformat(date_iso)
    except (TypeError, ValueError):
        return _aujourdhui()


def calendrier_jour(date_iso, demo=False):
    d = _date_ou_defaut(date_iso)
    jour_nom = _JOURS[d.weekday()]
    sc = [s for s in seances(demo) if str(s["jour"]).strip().lower() == jour_nom.lower()]
    sc.sort(key=lambda s: (s["deb_min"] if s["deb_min"] is not None else 9999))
    salles_occ = sorted({s["salle"] for s in sc if s["salle"]})
    return {
        "date": d, "date_iso": d.isoformat(), "date_fr": fmt_date(d),
        "jour_nom": jour_nom, "seances": sc, "salles_occupees": salles_occ,
        "presences": presences_du_jour(d.isoformat(), demo),
        "prev": (d - _dt.timedelta(days=1)).isoformat(),
        "next": (d + _dt.timedelta(days=1)).isoformat(),
    }


def calendrier_semaine(date_iso, demo=False):
    d = _date_ou_defaut(date_iso)
    lundi = d - _dt.timedelta(days=d.weekday())
    sc_all = seances(demo)
    jours = []
    for i, nom in enumerate(config.JOURS_SEMAINE):   # Lundi..Samedi
        jd = lundi + _dt.timedelta(days=i)
        sc = [s for s in sc_all if str(s["jour"]).strip().lower() == nom.lower()]
        sc.sort(key=lambda s: (s["deb_min"] if s["deb_min"] is not None else 9999))
        jours.append({"nom": nom, "date": jd, "date_iso": jd.isoformat(),
                      "date_fr": fmt_date(jd), "seances": sc,
                      "est_aujourdhui": jd == _aujourdhui()})
    return {
        "lundi": lundi, "jours": jours, "ancre": lundi.isoformat(),
        "libelle": "Semaine du %s au %s" % (fmt_date(lundi), fmt_date(lundi + _dt.timedelta(days=5))),
        "heure_min": config.CAL_HEURE_MIN, "heure_max": config.CAL_HEURE_MAX,
        "prev": (lundi - _dt.timedelta(days=7)).isoformat(),
        "next": (lundi + _dt.timedelta(days=7)).isoformat(),
    }


_MOIS_FR = ["", "Janvier", "Fevrier", "Mars", "Avril", "Mai", "Juin",
            "Juillet", "Aout", "Septembre", "Octobre", "Novembre", "Decembre"]


def _periode_libelle(mois_annee):
    """'MM/AAAA' -> 'Mois AAAA' lisible (ex. '06/2026' -> 'Juin 2026').
    Renvoie la valeur telle quelle si elle n'est pas au format MM/AAAA."""
    s = str(mois_annee or "").strip()
    m = re.match(r"^(\d{1,2})/(\d{4})$", s)
    if not m:
        return s
    mois = int(m.group(1))
    if 1 <= mois <= 12:
        return "%s %s" % (_MOIS_FR[mois], m.group(2))
    return s


def calendrier_mois(date_iso, demo=False):
    d = _date_ou_defaut(date_iso)
    premier = d.replace(day=1)
    # nb seances par jour de semaine (projection)
    sc_all = seances(demo)
    par_jour = {nom: 0 for nom in _JOURS}
    for s in sc_all:
        j = str(s["jour"]).strip()
        for nom in _JOURS:
            if j.lower() == nom.lower():
                par_jour[nom] += 1
    debut = premier - _dt.timedelta(days=premier.weekday())  # lundi de la 1re semaine
    semaines, cur = [], debut
    for _ in range(6):
        semaine = []
        for _ in range(7):
            nom = _JOURS[cur.weekday()]
            semaine.append({
                "date": cur, "date_iso": cur.isoformat(), "num": cur.day,
                "dans_mois": cur.month == premier.month,
                "est_aujourdhui": cur == _aujourdhui(),
                "nb_seances": par_jour.get(nom, 0) if cur.month == premier.month else 0,
            })
            cur += _dt.timedelta(days=1)
        semaines.append(semaine)
        if cur.month != premier.month and cur > premier:
            break
    mois_prec = (premier - _dt.timedelta(days=1)).replace(day=1)
    j2 = premier.replace(day=28) + _dt.timedelta(days=4)
    mois_suiv = j2.replace(day=1)
    return {
        "annee": premier.year, "mois": premier.month, "ancre": premier.isoformat(),
        "libelle": "%s %d" % (_MOIS_FR[premier.month], premier.year),
        "semaines": semaines, "entetes": _JOURS,
        "prev": mois_prec.isoformat(), "next": mois_suiv.isoformat(),
    }


# --- Occupation d'une salle (detail) ---
def salle_occupation(salle_id, demo=False):
    cible = None
    for s in salles(demo):
        if str(s["id"]).strip().lower() == str(salle_id).strip().lower() \
           or str(s["nom"]).strip().lower() == str(salle_id).strip().lower():
            cible = s
            break
    if cible is None:
        return None
    occ = [s for s in seances(demo) if _seance_dans_salle(s, cible)]
    par_jour = []
    for nom in config.JOURS_SEMAINE:
        jl = [s for s in occ if str(s["jour"]).strip().lower() == nom.lower()]
        jl.sort(key=lambda s: (s["deb_min"] if s["deb_min"] is not None else 9999))
        if jl:
            par_jour.append({"jour": nom, "seances": jl})
    equip = [e.strip() for e in str(cible.get("equipements", "")).replace(",", ";").split(";") if e.strip()]
    nom_salle = cible.get("nom", "")
    return {"salle": cible, "occupation": par_jour, "nb": len(occ),
            "equipements": equip,
            "materiel": materiel_de_salle(nom_salle),
            "index": salle_index(salle_id),
            "reservations": reservations_salle(nom_salle),
            "seances_a3": seances_salle(nom_salle),
            "enseignants": enseignants_choix(),
            "types_reservation": list(config.TYPES_RESERVATION)}


def materiel_de_salle(nom_salle):
    """Liste (lecture seule) du materiel inventorie pour une salle, depuis
    M1_Equipements (source unique). Filtre sur la colonne « Salle / localisation »
    par nom de salle (insensible a la casse/espaces). Aucune ecriture ici :
    l'ajout/retrait de materiel se fait sur l'ecran Logistique > Materiel."""
    cible = str(nom_salle or "").strip().lower()
    if not cible:
        return []
    out = []
    for r in _lignes_dict("M1_Equipements"):
        if str(r.get("Salle / localisation", "")).strip().lower() == cible:
            out.append({
                "designation": str(r.get("Designation", "")).strip(),
                "categorie": str(r.get("Categorie", "")).strip(),
                "quantite": r.get("Quantite", ""),
                "etat": str(r.get("Etat", "")).strip(),
            })
    return out


def inventaire_salle(nom_salle):
    """Inventaire IMPRIMABLE du materiel d'UNE salle (M1 filtre par
    'Salle / localisation'). Meme structure 'table' que inventaire_equipements
    (R5, V1.99.41). Colonnes : ID, Designation, Categorie, Etat, Bailleur,
    Quantite, Montant (KMF) + total."""
    cible = str(nom_salle or "").strip().lower()
    cols = ["ID", "Designation", "Categorie", "Etat", "Bailleur",
            "Quantite", "Montant (KMF)"]
    items = []
    for r in _lignes_dict("M1_Equipements"):
        if str(r.get("Salle / localisation", "")).strip().lower() != cible:
            continue
        idv = str(r.get("ID equipement", "")).strip()
        des = str(r.get("Designation", "")).strip()
        if not idv and not des:
            continue
        items.append({
            "id": idv, "des": des,
            "cat": str(r.get("Categorie", "")).strip(),
            "etat": str(r.get("Etat", "")).strip(),
            "bailleur": str(r.get("Source de financement / Bailleur", "")).strip(),
            "qte": _num(r.get("Quantite", 0)) or 0,
            "mtt": _num(r.get("Montant (KMF)", 0)) or 0,
        })
    items.sort(key=lambda x: (x["des"].lower(), x["id"]))
    lignes, tq, tm = [], 0.0, 0.0
    for it in items:
        lignes.append([it["id"], it["des"], it["cat"], it["etat"], it["bailleur"],
                       _kmf_aff(it["qte"]) if it["qte"] else "",
                       _kmf_aff(it["mtt"]) if it["mtt"] else ""])
        tq += it["qte"]
        tm += it["mtt"]
    total = ["TOTAL", "", "", "", "", _kmf_aff(tq), _kmf_aff(tm)]
    leg = "Inventaire du materiel — salle %s · %d equipement(s) · Edite le %s" % (
        str(nom_salle).strip(), len(items), fmt_date(_dt.date.today()))
    return {"colonnes": cols, "lignes": lignes, "total": total,
            "contexte": {"legende": leg}}


def salle_index(salle_id):
    """Position 0-based d'une salle parmi les lignes de donnees de L1_Salles
    (meme convention que data.modifier_ligne / lignes()). Match par ID salle ou
    Nom, insensible a la casse. Retourne None si introuvable."""
    cible = str(salle_id or "").strip().lower()
    for i, r in enumerate(_lignes_dict("L1_Salles")):
        if str(r.get("ID salle", "")).strip().lower() == cible \
           or str(r.get("Nom / libelle", "")).strip().lower() == cible:
            return i
    return None


# ===========================================================================
# DROITS D'ACCES, CAPACITE & PARAMETRES EDITABLES (ecriture — V1.2)
# ===========================================================================
def roles():
    """Liste des roles depuis P1_Roles : login, role, lecture, ecriture, financier,
    admin, superuser. Les SUPERUTILISATEURS (config.SUPERUSER_LOGINS) sont garantis :
    leurs pouvoirs sont forces a l'acces total + admin quoi que dise P1_Roles, et
    tout superutilisateur absent de P1_Roles est ajoute (selectionnable, anti-blocage)."""
    out, vus = [], set()
    for r in _lignes_dict("P1_Roles"):
        login = str(r.get("Utilisateur / Login", "")).strip()
        if not login:
            continue
        sup = login in config.SUPERUSER_LOGINS
        out.append({
            "login": login,
            "role": str(r.get("Role", "")).strip(),
            "lecture": "Tous" if sup else str(r.get("Modules lecture", "")).strip(),
            "ecriture": "Tous" if sup else str(r.get("Modules ecriture", "")).strip(),
            "financier": "O" if sup else str(r.get("Acces financier (O/N)", "")).strip().upper(),
            "admin": "O" if sup else str(r.get("Admin droits (O/N)", "")).strip().upper(),
            "superuser": sup,
        })
        vus.add(login)
    # superutilisateurs absents de P1_Roles -> synthetises (toujours disponibles)
    for login in config.SUPERUSER_LOGINS:
        if login not in vus:
            out.insert(0, {"login": login, "role": "Super-administrateur",
                           "lecture": "Tous", "ecriture": "Tous", "financier": "O",
                           "admin": "O", "superuser": True})
    return out


def role_par_login(login):
    for r in roles():
        if r["login"] == login:
            return r
    # login inconnu : si c'est un superutilisateur, on le garantit malgre tout
    if login in config.SUPERUSER_LOGINS:
        return {"login": login, "role": "Super-administrateur", "lecture": "Tous",
                "ecriture": "Tous", "financier": "O", "admin": "O", "superuser": True}
    rs = roles()
    return rs[0] if rs else {"login": "", "role": "", "lecture": "",
                             "ecriture": "Tous", "financier": "O", "admin": "O",
                             "superuser": False}


def _expanse_groupes(libelle_groupes):
    """'Academique, Stages' -> set d'onglets ; 'Tous' -> tous les onglets editables."""
    txt = str(libelle_groupes or "").strip()
    if not txt:
        return set()
    if txt.lower() == "tous":
        tous = set(config.ONGLETS_DIRECTION)
        for ongs in config.MODULES_ONGLETS.values():
            tous.update(ongs)
        return tous
    res = set()
    for g in [x.strip() for x in txt.split(",") if x.strip()]:
        res.update(config.MODULES_ONGLETS.get(g, []))
    return res


def peut_ecrire(role_row, onglet):
    """Vrai si le role peut ECRIRE sur cet onglet (matrice P1_Roles)."""
    if onglet in config.READONLY_TABS:
        return False
    tabs = _expanse_groupes(role_row.get("ecriture", ""))
    if onglet not in tabs:
        return False
    if onglet in config.ONGLETS_FINANCIERS and role_row.get("financier", "") != "O":
        return False
    return True


def peut_lire(role_row, onglet):
    """Vrai si le role DECLARE la lecture de cet onglet (colonne Modules lecture).
    NB : la lecture n'est pas encore restreinte dans l'IHM (seule l'ecriture l'est) ;
    cette fonction reflete le droit DECLARE dans P1_Roles, pour la matrice."""
    if onglet in _expanse_groupes(role_row.get("ecriture", "")):
        return True  # qui peut ecrire peut lire
    return onglet in _expanse_groupes(role_row.get("lecture", ""))


def matrice_autorisations():
    """Matrice r:roles x modules pour la page de gouvernance des droits.
    Lecture = droit DECLARE (P1_Roles) ; Ecriture = droit EFFECTIF (peut_ecrire,
    inclut le verrou Acces financier et les onglets en lecture seule).
    Renvoie aussi la definition des modules (decoupage module -> onglets)."""
    groupes = list(config.MODULES_ONGLETS.keys())
    modules_def = [{"module": g, "onglets": list(config.MODULES_ONGLETS[g])}
                   for g in groupes]
    modules_def.append({"module": "Direction (parametrage)",
                        "onglets": list(config.ONGLETS_DIRECTION)})
    colonnes = groupes + ["Direction (parametrage)"]

    def _statut(role, tabs):
        if not tabs:
            return {"lecture": False, "ecriture": False, "etat": "—"}
        ecr = all(peut_ecrire(role, t) for t in tabs)
        lec = all(peut_lire(role, t) for t in tabs)
        if ecr and lec:
            etat = "Lecture + Ecriture"
        elif ecr:
            etat = "Ecriture"
        elif lec:
            etat = "Lecture"
        else:
            etat = "—"
        return {"lecture": lec, "ecriture": ecr, "etat": etat}

    lignes = []
    for role in roles():
        cells = [_statut(role, config.MODULES_ONGLETS[g]) for g in groupes]
        cells.append(_statut(role, config.ONGLETS_DIRECTION))
        lignes.append({
            "login": role["login"], "role": role["role"],
            "financier": role["financier"] or "N", "cells": cells,
            "admin": role.get("admin", "") or "N", "superuser": role.get("superuser", False),
        })
    return {"colonnes": colonnes, "lignes": lignes, "modules_def": modules_def,
            "financiers": list(config.ONGLETS_FINANCIERS)}


# =========================================================================
# ADMINISTRATION DES DROITS (V1.13) — edition de P1_Roles depuis l'IHM,
# reservee aux administrateurs (Admin droits=O) et au superutilisateur.
# Garde-fous anti-blocage : superutilisateur jamais modifiable/supprimable ;
# impossible de retirer le dernier administrateur.
# =========================================================================
def est_admin(role_row):
    """Vrai si le role peut administrer les droits (superutilisateur ou Admin droits=O)."""
    return bool(role_row.get("superuser")) or str(role_row.get("admin", "")).upper() == "O"


def est_superuser(login):
    return str(login).strip() in config.SUPERUSER_LOGINS


def _nb_admins():
    """Nombre d'administrateurs effectifs presents dans P1_Roles (superutilisateurs
    inclus s'ils y figurent ; le superutilisateur reste de toute facon garanti
    par le code, donc le recours existe meme si ce nombre tombe a 0)."""
    return sum(1 for r in roles() if est_admin(r))


def _compose_groupes(groupes, tous):
    """Construit la chaine 'Modules ...' : 'Tous' si tous=True, sinon la liste
    des groupes valides separes par ', '. Filtre sur les groupes connus."""
    if tous:
        return config.GROUPE_TOUS
    valides = [g for g in (groupes or []) if g in config.MODULES_ONGLETS]
    return ", ".join(valides)


def enregistrer_utilisateur(acteur, login, role, lecture_groupes, ecriture_groupes,
                            lecture_tous=False, ecriture_tous=False,
                            financier=False, admin=False):
    """Ajoute ou met a jour un utilisateur dans P1_Roles (upsert par login).
    `acteur` = role courant (doit etre admin). Renvoie (ok, message)."""
    if not est_admin(acteur):
        return False, "Action reservee aux administrateurs des droits."
    login = str(login or "").strip()
    if not login:
        return False, "Le login est obligatoire."
    if est_superuser(login):
        return False, ("Le superutilisateur est garanti par le systeme : "
                       "ses droits ne se modifient pas ici.")
    valeurs = {
        "Utilisateur / Login (**)": login,
        "Role (**)": str(role or "").strip() or login,
        "Modules lecture (**)": _compose_groupes(lecture_groupes, lecture_tous),
        "Modules ecriture (**)": _compose_groupes(ecriture_groupes, ecriture_tous),
        "Acces financier (O/N) (**)": "O" if financier else "N",
        "Admin droits (O/N) (**)": "O" if admin else "N",
    }
    # Garde-fou : ne pas retirer le droit admin du DERNIER administrateur.
    if not admin:
        deja = next((r for r in roles() if r["login"] == login), None)
        if deja and est_admin(deja) and _nb_admins() <= 1:
            return False, ("Impossible de retirer le dernier administrateur : "
                           "designez d'abord un autre administrateur.")
    res = _db.ecrire_lignes_lot("P1_Roles", [valeurs], cles=["Utilisateur / Login (**)"])
    if res.get("ajout"):
        return True, "Utilisateur ajoute : %s." % login
    return True, "Utilisateur mis a jour : %s." % login


def rubriques():
    """Liste des grandes rubriques d'appartenance (config, editable)."""
    return list(getattr(config, "RUBRIQUES", []))


def enregistrer_compte_it(acteur, login, role):
    """Console informatique : cree/met a jour la LIGNE P1_Roles d'un compte avec
    UNIQUEMENT login + role. Les droits par module (lecture/ecriture/financier/
    admin) ne sont PAS touches ici : ils restent geres dans P1_Roles par la
    Direction (ecrire_lignes_lot preserve les colonnes non fournies en MAJ).
    Renvoie (ok, message, nouveau)."""
    if not est_admin(acteur):
        return False, "Action reservee au responsable informatique.", False
    login = str(login or "").strip()
    if not login:
        return False, "Le login est obligatoire.", False
    if est_superuser(login):
        return False, ("Le superutilisateur est garanti par le systeme : "
                       "il ne se gere pas ici."), False
    nouveau = not any(r["login"] == login for r in roles())
    valeurs = {
        "Utilisateur / Login (**)": login,
        "Role (**)": str(role or "").strip() or login,
    }
    _db.ecrire_lignes_lot("P1_Roles", [valeurs], cles=["Utilisateur / Login (**)"])
    return True, ("Compte cree : %s." if nouveau else "Compte mis a jour : %s.") % login, nouveau


def supprimer_utilisateur(acteur, login):
    """Supprime un utilisateur de P1_Roles. Renvoie (ok, message)."""
    if not est_admin(acteur):
        return False, "Action reservee aux administrateurs des droits."
    login = str(login or "").strip()
    if est_superuser(login):
        return False, "Le superutilisateur ne peut pas etre supprime."
    cible = next((r for r in roles() if r["login"] == login), None)
    if not cible:
        return False, "Utilisateur introuvable."
    if est_admin(cible) and _nb_admins() <= 1:
        return False, ("Impossible de supprimer le dernier administrateur : "
                       "designez d'abord un autre administrateur.")
    ok = _db.supprimer_ligne_par_cle("P1_Roles", "Utilisateur / Login (**)", login)
    return (True, "Utilisateur supprime : %s." % login) if ok \
        else (False, "Utilisateur introuvable.")


def utilisateurs_admin():
    """Donnees pour la console comptes & acces : un dict par utilisateur. Les droits
    par module (lecture/ecriture/financier/admin) sont AFFICHES en lecture seule
    (geres dans P1_Roles). Les attributs operationnels (rubrique, couleur, validite,
    etat du mot de passe) viennent de instance/comptes.json (hors depot)."""
    out = []
    for r in roles():
        login = r["login"]
        att = auth.attributs(login)
        lec = str(r.get("lecture", "")).strip()
        ecr = str(r.get("ecriture", "")).strip()
        out.append({
            "login": login, "role": r["role"],
            "superuser": r.get("superuser", False),
            "admin": str(r.get("admin", "")).upper() == "O",
            "financier": str(r.get("financier", "")).upper() == "O",
            "lecture_tous": lec.lower() == "tous",
            "ecriture_tous": ecr.lower() == "tous",
            "lecture": [g.strip() for g in lec.split(",") if g.strip()
                        and g.strip().lower() != "tous"],
            "ecriture": [g.strip() for g in ecr.split(",") if g.strip()
                         and g.strip().lower() != "tous"],
            # Attributs operationnels (comptes.json, hors depot)
            "compte_actif": bool(att.get("hash")),
            "rubrique": att.get("rubrique", ""),
            "couleur": couleur_login(login),
            "valide_jusqu": att.get("valide_jusqu", ""),
            "expire": auth.est_expire(login),
            "doit_changer": bool(att.get("doit_changer")),
        })
    return out


def capacite_onglet(onglet):
    """Indicateur de remplissage d'un onglet : nb lignes, max, ratio, proche."""
    try:
        nb = _db_pour(onglet).nb_lignes(onglet)
    except Exception:
        nb = 0
    maxi = config.CAPACITE
    ratio = (nb / maxi) if maxi else 0
    return {"onglet": onglet, "nb": nb, "max": maxi, "ratio": ratio,
            "proche": ratio >= config.ALERTE_CAPACITE_RATIO}


def alertes_capacite():
    """Onglets qui approchent la capacite (pour le bandeau d'alerte IHM)."""
    vus, out = set(), []
    for sec in config.GUIDE_STRUCTURE:
        for cle, lib, _ in sec["modules"]:
            if (cle in config.SPECIAL_ROUTES or cle in config.READONLY_TABS
                    or cle in vus):
                continue
            vus.add(cle)
            c = capacite_onglet(cle)
            if c["proche"]:
                c["libelle"] = lib
                out.append(c)
    return out


def parametres_editables():
    """Structure de P0_Parametres pour l'ecran d'edition : une entree par
    colonne-liste, avec libelle propre, provenance et valeurs actuelles."""
    bruts = _db.entetes("P0_Parametres")
    out = []
    for b in bruts:
        lib, prov = decoupe_provenance(b)
        vals = [v for v in _db.colonne("P0_Parametres", b)
                if str(v).strip() != ""]
        out.append({"brut": b, "libelle": lib, "prov": prov,
                    "prov_info": config.PROVENANCE[prov], "valeurs": vals})
    return out


# ---------------------------------------------------------------------------
# Formulaires de SAISIE generiques (pilotes par le Dictionnaire) — V1.3
# ---------------------------------------------------------------------------
def options_liste(source):
    """Renvoie la liste d'options pour un champ 'Liste' a partir du libelle de
    source du Dictionnaire, ou None si saisie libre.
    Source = soit une liste inline (config.LISTES_INLINE), soit une colonne-liste
    de P0_Parametres (rapprochee par libelle propre)."""
    src = str(source or "").strip()
    if not src or src == "-":
        return None
    if src in config.LISTES_INLINE:
        return list(config.LISTES_INLINE[src])
    # liste alimentee par une colonne d'un autre onglet de donnees (lien A1/A3...)
    if src in getattr(config, "LISTES_ONGLET", {}):
        onglet_src, champ_src = config.LISTES_ONGLET[src]
        vus, vals = set(), []
        for v in _db.colonne(onglet_src, champ_src):
            s = str(v).strip()
            if s and s not in vus:
                vus.add(s)
                vals.append(s)
        return sorted(vals)
    # liste composite : plusieurs colonnes d'un autre onglet -> libelle lisible
    # (ex. A3 'Enseignant' <- E1 "Nom Prenom"). Lien Seances <-> formateurs (V1.8).
    if src in getattr(config, "LISTES_ONGLET_COMPOSITE", {}):
        onglet_src, champs_src, sep = config.LISTES_ONGLET_COMPOSITE[src]
        cols = [_db.colonne(onglet_src, c) for c in champs_src]
        n = max((len(c) for c in cols), default=0)
        vus, vals = set(), []
        for i in range(n):
            parts = [str(c[i]).strip() for c in cols
                     if i < len(c) and str(c[i]).strip()]
            lab = sep.join(parts)
            if lab and lab not in vus:
                vus.add(lab)
                vals.append(lab)
        return sorted(vals)
    # liste VALEUR != LIBELLE : on stocke une cle (colonne valeur) mais on affiche un
    # libelle lisible "<valeur> — <colonnes complementaires>". Renvoie des {value,label}.
    # Lien E2 'Matricule ens.' <- E1 (V1.9, choix (b2)) : stocke le matricule, affiche
    # "Matricule — Nom Prenom".
    if src in getattr(config, "LISTES_ONGLET_VALLABEL_FILTRE", {}):
        spec = config.LISTES_ONGLET_VALLABEL_FILTRE[src]
        onglet_src, col_val, cols_lab = spec[0], spec[1], spec[2]
        # 5-uplet : (col_filtre, val_filtre) -> 1 critere ; 4-uplet : [(col, val), ...] -> N (ET)
        if len(spec) == 5:
            filtres = [(spec[3], spec[4])]
        else:
            filtres = list(spec[3])
        valcol = _db.colonne(onglet_src, col_val)
        labcols = [_db.colonne(onglet_src, c) for c in cols_lab]
        filcols = [(_db.colonne(onglet_src, c), str(v).strip().lower()) for c, v in filtres]
        vus, out = set(), []
        for i in range(len(valcol)):
            if any(i >= len(fc) or str(fc[i]).strip().lower() != cible
                   for fc, cible in filcols):
                continue
            v = str(valcol[i]).strip()
            if not v or v in vus:
                continue
            vus.add(v)
            compl = " ".join(str(labcols[j][i]).strip() for j in range(len(labcols))
                             if i < len(labcols[j]) and str(labcols[j][i]).strip())
            out.append({"value": v, "label": (v + " — " + compl) if compl else v})
        return sorted(out, key=lambda o: o["label"])
    if src in getattr(config, "LISTES_ONGLET_VALLABEL", {}):
        onglet_src, col_val, cols_lab = config.LISTES_ONGLET_VALLABEL[src]
        valcol = _db.colonne(onglet_src, col_val)
        labcols = [_db.colonne(onglet_src, c) for c in cols_lab]
        vus, out = set(), []
        for i in range(len(valcol)):
            v = str(valcol[i]).strip()
            if not v or v in vus:
                continue
            vus.add(v)
            compl = " ".join(str(labcols[j][i]).strip() for j in range(len(labcols))
                             if i < len(labcols[j]) and str(labcols[j][i]).strip())
            out.append({"value": v, "label": (v + " — " + compl) if compl else v})
        return sorted(out, key=lambda o: o["label"])
    # liste COMBINEE "X OU Y" : concatene plusieurs colonnes P0 (distinct, ordre
    # conserve). Ex. F1 'Categorie' = "Cat_Recettes OU Cat_Depenses" (V1.14, choix a).
    if " OU " in src:
        propre = {decoupe_provenance(b)[0]: v for b, v in _db.listes_parametres().items()}
        out, vus = [], set()
        for part in [p.strip() for p in src.split(" OU ")]:
            for v in propre.get(part, []):
                s = str(v).strip()
                if s and s not in vus:
                    vus.add(s)
                    out.append(s)
        return out or None
    # rapprochement avec une colonne de P0_Parametres (libelle sans marqueur)
    par = _db.listes_parametres()   # {entete_brut: [valeurs]}
    for brut, vals in par.items():
        if decoupe_provenance(brut)[0] == src:
            return [v for v in vals if str(v).strip() != ""]
    return None   # saisie libre


def _prochain_code(onglet, colonne, prefixe):
    """Prochain code libre "<prefixe>-<n>" pour une colonne (max des n + 1, 1 si
    aucun). Insensible a la casse, ignore les valeurs hors format. Sert aux
    identifiants pre-suggeres : matricule provisoire "NC-<n>" (E1), ID salle
    "SAL-<n>" (L1). Le jour ou la vraie valeur existe, on remplace le code."""
    rx = re.compile(r"^%s-0*(\d+)$" % re.escape(prefixe), re.IGNORECASE)
    maxi = 0
    for v in _db.colonne(onglet, _brut(onglet, colonne)):
        m = rx.match(str(v).strip())
        if m:
            n = int(m.group(1))
            if n > maxi:
                maxi = n
    return "%s-%d" % (prefixe, maxi + 1)


def _prochain_nc():
    return _prochain_code("E1_Enseignants", "Matricule ens.", "NC")


def _prochain_sal():
    return _prochain_code("L1_Salles", "ID salle", "SAL")


def _prochain_res():
    return _prochain_code("L2_Reservations", "ID reservation", "RES")


def _prochain_eq():
    return _prochain_code("M1_Equipements", "ID equipement", "EQ")


def _prochain_bes():
    return _prochain_code("L3_Besoins", "ID besoin", "BES")


def _prochain_ordre():
    """Prochain n° d'ordre libre pour A1 (max numerique + 1, 1 si vide)."""
    maxi = 0
    for v in _db.colonne("A1_Etudiants", _brut("A1_Etudiants", "N ordre")):
        s = str(v).strip()
        if s.isdigit() and int(s) > maxi:
            maxi = int(s)
    return str(maxi + 1)


def _dernier_saisi(onglet, libelle, token):
    """Resout le token \"@last\" / \"@last|<repli>\" : derniere valeur NON VIDE saisie
    dans la colonne <libelle> de l'onglet (report des dernieres valeurs en saisie
    serie). Si la colonne est encore vide, on prend le <repli> (texte litteral, ou
    \"@today\" => date du jour). Sans repli : chaine vide. Lecture via _db_pour pour
    router le bon classeur (notes/principal)."""
    repli = token.split("|", 1)[1].strip() if "|" in token else ""
    brut = _brut(onglet, libelle)
    derniere = ""
    if brut:
        for v in _db_pour(onglet).colonne(onglet, brut):
            if v is not None and str(v).strip() != "":
                derniere = str(v).strip()
    if derniere:
        return derniere
    if repli == "@today":
        return fmt_date(_dt.date.today())
    return repli


# Etats de materiel qui justifient une expression de besoin (panne / indisponibilite).
ETATS_BESOIN = {"En panne", "Hors service"}


def materiels_en_panne():
    """Equipements M1 dont l'etat justifie une expression de besoin (En panne /
    Hors service). Sert au panneau de declenchement sur l'ecran Equipements."""
    iID, iDes, iEtat, iSalle = _idx_map(_db, "M1_Equipements",
        ["ID equipement", "Designation", "Etat", "Salle / localisation"])
    out = []
    for r in _db.lignes("M1_Equipements"):
        etat = (str(r[iEtat]).strip() if iEtat >= 0 else "")
        if etat not in ETATS_BESOIN:
            continue
        out.append({
            "id": str(r[iID]).strip() if iID >= 0 else "",
            "designation": str(r[iDes]).strip() if iDes >= 0 else "",
            "etat": etat,
            "salle": str(r[iSalle]).strip() if iSalle >= 0 else "",
        })
    return sorted(out, key=lambda d: d["designation"].lower())


def champs_saisie(onglet):
    """Specs des champs saisissables d'un onglet (pour le formulaire d'ajout).
    Construit a partir du Dictionnaire (type, obligatoire, liste) rapproche des
    en-tetes reels. Les colonnes calcul (readonly) sont exclues."""
    dico = {d["Champ"]: d for d in dictionnaire_par_onglet().get(onglet, [])}
    auto = set(config.CHAMPS_AUTO_LOGIN.get(onglet, []))
    champs = []
    for meta in entetes_meta(onglet):
        if meta["readonly"] or meta["libelle"] in auto:
            continue
        d = dico.get(meta["libelle"], {})
        typ = str(d.get("Type", "")).strip().lower()
        obl = str(d.get("Obligatoire", "")).strip().lower().startswith("o")
        opts = options_liste(d.get("Liste", "")) if typ == "liste" else None
        # Normalisation : options toujours en liste de {value, label}. Une source
        # peut renvoyer des chaines (value=label) ou deja des {value, label}
        # (cas valeur != libelle, ex. matricule affiche "Matricule — Nom Prenom").
        if opts is not None:
            opts = [o if isinstance(o, dict) else {"value": o, "label": o} for o in opts]
        if typ in ("nombre", "kmf"):
            html = "number"
        elif typ == "date":
            html = "date_fr"   # saisie texte JJ/MM/AAAA (charte)
        else:
            html = "texte"
        brut_def = config.SAISIE_DEFAUTS.get(onglet, {}).get(meta["libelle"], "")
        if brut_def == "@today":
            defaut = fmt_date(_dt.date.today())
        elif brut_def == "@next_nc":
            defaut = _prochain_nc()
        elif brut_def == "@next_sal":
            defaut = _prochain_sal()
        elif brut_def == "@next_res":
            defaut = _prochain_res()
        elif brut_def == "@next_eq":
            defaut = _prochain_eq()
        elif brut_def == "@next_bes":
            defaut = _prochain_bes()
        elif brut_def == "@next_ordre":
            defaut = _prochain_ordre()
        elif brut_def.startswith("@last"):
            defaut = _dernier_saisi(onglet, meta["libelle"], brut_def)
        else:
            defaut = brut_def
        # Suggestions maquette (datalist) : seul le champ 'matiere' de l'onglet
        # concerne porte l'id ; les autres champs n'en ont pas. La saisie reste
        # libre (input texte + list=...), pas une liste stricte.
        dl_spec = config.MAQUETTE_DATALIST.get(onglet)
        datalist = None
        if dl_spec and meta["libelle"] == dl_spec.get("matiere"):
            datalist = "dl_maquette_" + onglet
        elif onglet in config.CHAMPS_DATALIST and meta["libelle"] in config.CHAMPS_DATALIST[onglet]:
            datalist = config.CHAMPS_DATALIST[onglet][meta["libelle"]]
        # Largeur de saisie (caracteres) : valeur dediee, sinon calcul auto.
        largeur = config.LARGEURS_CHAMPS.get(meta["libelle"])
        if largeur is None:
            if opts:
                largeur = min(max((len(str(o["label"])) for o in opts), default=8) + 5, 32)
            elif html == "date_fr":
                largeur = 12
            elif html == "number":
                largeur = 9
            else:
                largeur = 18
        # Mode de saisie clavier (V1.99.4) : classe les champs LISTE pour le moteur
        # static/js/saisie_clavier.js. Override explicite par nom de liste, sinon
        # regle auto (value==label + courts + peu nombreux -> "touche" ; sinon "auto").
        saisie_mode = "normal"
        if opts:
            forced = config.MODE_SAISIE_LISTE.get(str(d.get("Liste", "")).strip())
            if forced in ("touche", "auto", "normal"):
                saisie_mode = forced
            else:
                diff_val_label = any(str(o["value"]) != str(o["label"]) for o in opts)
                labels_courts = all(len(str(o["label"])) <= 18 for o in opts)
                if diff_val_label or len(opts) > config.SEUIL_AUTOCOMPLETE:
                    saisie_mode = "auto"
                elif len(opts) <= config.MAX_OPTS_TOUCHE and labels_courts:
                    saisie_mode = "touche"
                else:
                    saisie_mode = "auto"
        champs.append({"brut": meta["brut"], "libelle": meta["libelle"],
                       "obligatoire": obl, "type": html, "options": opts,
                       "defaut": defaut, "prov_info": meta["prov_info"],
                       "datalist": datalist, "largeur": largeur,
                       "saisie_mode": saisie_mode})
    # V1.83 : champs obligatoires regroupes en debut de ligne (tri stable :
    # l'ordre d'origine est conserve a l'interieur de chaque groupe), pour ne
    # pas tabuler jusqu'au bout de la ligne avant d'atteindre un champ requis.
    champs.sort(key=lambda c: not c["obligatoire"])
    return champs


def valide_saisie(onglet, valeurs):
    """Verifie les champs obligatoires + regles specifiques. Renvoie (ok, message)."""
    manquants = [c["libelle"] for c in champs_saisie(onglet)
                 if c["obligatoire"] and not str(valeurs.get(c["brut"], "")).strip()]
    if manquants:
        return False, "Champs obligatoires manquants : " + ", ".join(manquants) + "."
    fn = _VALIDATIONS_SPECIFIQUES.get(onglet)
    if fn:
        ok, msg = fn(valeurs)
        if not ok:
            return False, msg
    return True, ""


def _valide_f1_mouvements(valeurs):
    """Regle conditionnelle F1 (choix B-b, V1.14) : le montant doit correspondre au
    sens. valeurs = {entete_brut: valeur}."""
    g = lambda lib: str(valeurs.get(_brut("F1_Mouvements", lib), "")).strip()
    sens = g("Sens").lower()
    rec = g("Montant Recette (KMF)")
    dep = g("Montant Depense (KMF)")
    if sens.startswith("recette"):
        if not rec:
            return False, "Sens = Recette : indiquez le Montant Recette (KMF)."
        if dep:
            return False, "Sens = Recette : laissez le Montant Depense (KMF) vide."
    elif sens.startswith("dep") or sens.startswith("dép"):
        if not dep:
            return False, "Sens = Depense : indiquez le Montant Depense (KMF)."
        if rec:
            return False, "Sens = Depense : laissez le Montant Recette (KMF) vide."
    return True, ""


_VALIDATIONS_SPECIFIQUES = {
    "F1_Mouvements": _valide_f1_mouvements,
}


# ===========================================================================
# SAISIE DES PRESENCES PAR SEANCE (par lot) — V1.4
# ---------------------------------------------------------------------------
# Perimetre assume HORS-TDR (champs A2 en (**), besoin issu du CR du 11/06).
# Flux : choisir une SEANCE (A3) + une DATE + un CRENEAU (10/12/15/17) ;
# l'effectif de la classe est tire de A1 (par Filiere+Niveau+Section) ; on coche
# les presents ; l'enregistrement ecrit une ligne A2 par etudiant via un UPSERT
# (cle Date+Matricule+Session+Creneau) => re-saisir une seance ne cree PAS de
# doublon, elle corrige les valeurs. "Saisi par" = login du role courant.
# ---------------------------------------------------------------------------

# En-tetes BRUTS de A2 (avec marqueurs) resolus depuis le classeur, pour ne pas
# coder les "(**)"/"(*)" en dur (robustesse si le libelle evolue).
def _brut(onglet, libelle):
    for m in entetes_meta(onglet):
        if m["libelle"] == libelle:
            return m["brut"]
    return libelle


def _creneau_defaut(deb_min):
    """Checkpoint de presence (10h/12h/15h/17h) deduit de l'heure de debut."""
    if deb_min is None:
        return config.CRENEAUX[0]
    if deb_min < 11 * 60:
        return "10h"
    if deb_min < 13 * 60:
        return "12h"
    if deb_min < 16 * 60:
        return "15h"
    return "17h"


def seances_saisie():
    """Seances de A3_Sessions presentables dans le selecteur (id + libelle riche)."""
    out = []
    for s in seances(demo=False):
        sid = str(s.get("id", "")).strip()
        mat = str(s.get("matiere", "")).strip()
        if not sid and not mat:
            continue
        grp = s.get("groupe", "")
        horaire = " ".join(x for x in [str(s.get("jour", "")).strip(),
                                       str(s.get("debut", "")).strip()
                                       + ("-" + str(s.get("fin", "")).strip()
                                          if s.get("fin") else "")] if x).strip()
        bits = [b for b in [horaire, mat, grp,
                            str(s.get("salle", "")).strip()] if b]
        label = (("[" + sid + "] ") if sid else "") + " · ".join(bits)
        out.append({
            "id": sid, "label": label, "matiere": mat,
            "filiere": str(s.get("filiere", "")).strip(),
            "niveau": str(s.get("niveau", "")).strip(),
            "section": str(s.get("section", "")).strip(),
            "jour": str(s.get("jour", "")).strip(),
            "debut": str(s.get("debut", "")).strip(),
            "fin": str(s.get("fin", "")).strip(),
            "salle": str(s.get("salle", "")).strip(),
            "creneau_defaut": _creneau_defaut(s.get("deb_min")),
        })
    return out


def _seance_par_id(seance_id):
    sid = str(seance_id).strip()
    for s in seances_saisie():
        if s["id"] == sid:
            return s
    return None


def cle_session(seance):
    """Valeur stockee dans A2 'Session / Matiere' : l'ID seance (lien A3 stable),
    a defaut la matiere."""
    if not seance:
        return ""
    return seance["id"] or seance["matiere"]


def etudiants_de_seance(seance_id):
    """Effectif (depuis A1_Etudiants) correspondant a la classe de la seance :
    meme Filiere + Niveau, et meme Section si la seance en precise une."""
    s = _seance_par_id(seance_id)
    if s is None:
        return []
    fil, niv, sec = s["filiere"].lower(), s["niveau"].lower(), s["section"].lower()
    out = []
    for e in _lignes_dict("A1_Etudiants"):
        mat = str(e.get("Matricule", "")).strip()
        if not mat:
            continue
        ef = str(e.get("Filiere", "")).strip().lower()
        en = str(e.get("Niveau", "")).strip().lower()
        es = str(e.get("Section", "")).strip().lower()
        if fil and ef != fil:
            continue
        if niv and en != niv:
            continue
        if sec and es and es != sec:
            continue
        out.append({
            "matricule": mat,
            "nom": str(e.get("Nom", "")).strip(),
            "prenom": str(e.get("Prenom", "")).strip(),
            "section": str(e.get("Section", "")).strip(),
        })
    out.sort(key=lambda x: (x["nom"].lower(), x["prenom"].lower(), x["matricule"]))
    return out


def presences_existantes(date_fr, session_val, creneau):
    """Renvoie {matricule: 'O'/'N'} deja enregistres pour (date, session, creneau).
    Sert a precocher les cases et a montrer l'etat courant avant correction."""
    res = {}
    df = str(date_fr).strip()
    sv = str(session_val).strip()
    cr = str(creneau).strip()
    if not (df and sv and cr):
        return res
    for r in _lignes_dict("A2_Presences"):
        if str(r.get("Date", "")).strip() != df:
            continue
        if str(r.get("Session / Matiere", "")).strip() != sv:
            continue
        if str(r.get("Creneau", "")).strip() != cr:
            continue
        res[str(r.get("Matricule", "")).strip()] = \
            str(r.get("Present (O/N)", "")).strip().upper()
    return res


_RE_DATE_FR = re.compile(r"^\d{2}/\d{2}/\d{4}$")


def enregistrer_presences_lot(seance_id, date_fr, creneau, presents, saisi_par):
    """Ecrit une ligne A2 par etudiant de la classe (upsert). `presents` = set de
    matricules coches present. Renvoie (ok, message)."""
    s = _seance_par_id(seance_id)
    if s is None:
        return False, "Seance introuvable."
    date_fr = str(date_fr).strip()
    if not _RE_DATE_FR.match(date_fr):
        return False, "Date invalide : format attendu JJ/MM/AAAA."
    creneau = str(creneau).strip()
    if creneau not in config.CRENEAUX:
        return False, "Creneau invalide."
    roster = etudiants_de_seance(seance_id)
    if not roster:
        return False, ("Aucun etudiant pour cette classe "
                       "(verifier Filiere/Niveau/Section dans A1_Etudiants).")
    presents = set(str(m).strip() for m in presents)
    session_val = cle_session(s)
    bd, bm, bs = (_brut("A2_Presences", "Date"),
                  _brut("A2_Presences", "Matricule"),
                  _brut("A2_Presences", "Session / Matiere"))
    bc, bp, bsp = (_brut("A2_Presences", "Creneau"),
                   _brut("A2_Presences", "Present (O/N)"),
                   _brut("A2_Presences", "Saisi par"))
    lignes = []
    for e in roster:
        lignes.append({
            bd: date_fr, bm: e["matricule"], bs: session_val, bc: creneau,
            bp: "O" if e["matricule"] in presents else "N",
            bsp: str(saisi_par or "").strip(),
        })
    res = _db.ecrire_lignes_lot("A2_Presences", lignes, cles=[bd, bm, bs, bc])
    return True, ("Presences enregistrees : %d ajout(s), %d mise(s) a jour."
                  % (res["ajout"], res["maj"]))



# ===========================================================================
# PRESENCES - SAISIE EN LISTE, SEANCE AD HOC (option B, V1.64)
# ---------------------------------------------------------------------------
# Saisie d'une feuille de presence SANS pre-remplir A3_Sessions : la seance est
# definie a la volee (classe + date + plage horaire + matiere + enseignant). Le
# roster vient de A1 par Filiere + Niveau (+ Section si renseignee) ; l'ecriture
# se fait dans A2_Presences (upsert). Le creneau A2 = plage "HH:MM-HH:MM". Si
# l'agent coche "recurrente", la seance est d'abord creee dans A3_Sessions (ID
# genere) et cet ID sert de cle "Session / Matiere" ; sinon cle composee lisible.
# ---------------------------------------------------------------------------
_RE_HEURE = re.compile(r"^([01]?\d|2[0-3]):[0-5]\d$")
_JOURS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


def _jour_de_date(date_fr):
    """Nom du jour (Lundi..Dimanche) pour une date JJ/MM/AAAA, ou ''."""
    import datetime as _dt
    s = str(date_fr).strip()
    if not _RE_DATE_FR.match(s):
        return ""
    try:
        d = _dt.datetime.strptime(s, "%d/%m/%Y").date()
    except ValueError:
        return ""
    return _JOURS_FR[d.weekday()]


_REGLAGES_FILE = os.path.join(config.INSTANCE_DIR, "reglages.json")


def _reglage_get(cle, defaut=""):
    """Lit un reglage local (instance/reglages.json), hors depot/zip. defaut si absent."""
    try:
        with open(_REGLAGES_FILE, encoding="utf-8") as f:
            return json.load(f).get(cle, defaut)
    except Exception:
        return defaut


def _reglage_set(cle, valeur):
    """Ecrit/maj un reglage local (merge). Cree instance/ au besoin. Ne bloque jamais."""
    try:
        os.makedirs(config.INSTANCE_DIR, exist_ok=True)
        data = {}
        if os.path.exists(_REGLAGES_FILE):
            with open(_REGLAGES_FILE, encoding="utf-8") as f:
                data = json.load(f) or {}
        data[cle] = valeur
        with open(_REGLAGES_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def seuil_passage():
    """Seuil de passage conditionnel = ecart de credits ECTS tolere (annuel).
    Reglage editable 'seuil_passage_conditionnel' (instance/reglages.json) ;
    sinon defaut config.SEUIL_PASSAGE_CONDITIONNEL. Renvoie un nombre >= 0."""
    brut = str(_reglage_get("seuil_passage_conditionnel", "")).strip()
    if brut != "":
        v = _num_h(brut)
        if v is not None and v >= 0:
            return v
    return float(getattr(config, "SEUIL_PASSAGE_CONDITIONNEL", 5))


def definir_seuil_passage(acteur, valeur, login):
    """Definit le seuil de passage conditionnel (reglage). Reserve au droit
    d'ecriture du bareme (Notes : scolarite / direction). Renvoie (ok, msg)."""
    if not peut_ecrire(acteur, "N1_Bareme_UE"):
        return False, "Action non autorisee."
    v = _num_h(valeur)
    if v is None or v < 0:
        return False, "Seuil invalide (attendu un nombre de credits >= 0)."
    _reglage_set("seuil_passage_conditionnel", v)
    auth.journal(login, "Seuil de passage defini", str(v), "")
    return True, ("Seuil de passage conditionnel defini : %g credit(s) "
                  "d'ecart tolere(s)." % v)


# ---------------------------------------------------------------------------
# Selection des indicateurs du TABLEAU DE BORD DIRECTION — V1.99.15 (#20 A+B)
# Selection GLOBALE etablissement (pas de preference par utilisateur en V1),
# persistee dans instance/reglages.json sous la cle "tdb_selection" =
# {"kpis": [...ids...], "charts": [...ids...]}. Absente/invalide => tous.
# Une selection vide explicite est respectee (l'utilisateur peut tout decocher ;
# le bouton "reinitialiser" retablit l'ensemble complet).
# ---------------------------------------------------------------------------
def tdb_selection():
    """Selection effective des indicateurs du TDB direction (filtree au catalogue)."""
    tous_k = config.tdb_kpi_ids() + config.tdb_kpi_budget_ids()
    tous_c = config.tdb_chart_ids() + config.tdb_chart_budget_ids()
    brut = _reglage_get("tdb_selection", None)
    if not isinstance(brut, dict):
        return {"kpis": list(tous_k), "charts": list(tous_c)}

    def _garde(vals, ref):
        if not isinstance(vals, list):
            return list(ref)
        return [v for v in ref if v in vals]  # conserve l'ordre du catalogue

    return {
        "kpis": _garde(brut.get("kpis"), tous_k),
        "charts": _garde(brut.get("charts"), tous_c),
    }


def tdb_selection_set(kpis, charts):
    """Enregistre la selection TDB (globale). Ne garde que des ids connus."""
    tous_k = config.tdb_kpi_ids() + config.tdb_kpi_budget_ids()
    tous_c = config.tdb_chart_ids() + config.tdb_chart_budget_ids()
    sel = {
        "kpis": [k for k in tous_k if k in set(kpis or [])],
        "charts": [c for c in tous_c if c in set(charts or [])],
    }
    _reglage_set("tdb_selection", sel)
    return sel


def tdb_selection_reset():
    """Reinitialise la selection TDB = tous les indicateurs (bouton reinitialiser)."""
    return tdb_selection_set(config.tdb_kpi_ids() + config.tdb_kpi_budget_ids(),
                             config.tdb_chart_ids() + config.tdb_chart_budget_ids())


def _annee_acad_defaut():
    """Annee academique COURANTE. V1.95 : pivot explicite stocke dans
    instance/reglages.json ('annee_acad_courante') s'il est defini ; sinon repli sur
    l'heuristique historique (annee la plus frequente dans A1)."""
    fixe = str(_reglage_get("annee_acad_courante", "")).strip()
    if fixe:
        return fixe
    from collections import Counter
    c = Counter(str(e.get("Annee acad.", "")).strip()
                for e in _lignes_dict("A1_Etudiants"))
    c.pop("", None)
    return c.most_common(1)[0][0] if c else ""


def _annee_suivante_label(label):
    """'2024-2025' -> '2025-2026'. None si label non parsable."""
    fin = _annee_fin_label(label)
    return ("%d-%d" % (fin, fin + 1)) if fin else None


def definir_annee_courante(acteur, label, login):
    """Definit manuellement l'annee academique courante (Direction). Renvoie (ok, msg)."""
    if not est_admin(acteur):
        return False, "Action reservee a la Direction."
    label = str(label or "").strip()
    if not re.match(r"^\d{4}\s*-\s*\d{4}$", label):
        return False, "Format d'annee invalide (attendu AAAA-AAAA)."
    label = label.replace(" ", "")
    _reglage_set("annee_acad_courante", label)
    auth.journal(login, "Annee courante definie", label, "")
    return True, "Annee academique courante definie : %s." % label


def passer_annee_suivante(acteur, login):
    """Avance l'annee academique courante a la suivante (Direction). Renvoie (ok, msg)."""
    if not est_admin(acteur):
        return False, "Action reservee a la Direction."
    courante = _annee_acad_defaut()
    suivante = _annee_suivante_label(courante)
    if not suivante:
        return False, ("Annee courante non determinee ou illisible : definissez-la "
                       "d'abord manuellement (format AAAA-AAAA).")
    _reglage_set("annee_acad_courante", suivante)
    auth.journal(login, "Passage annee suivante", "%s -> %s" % (courante, suivante), "")
    return True, "Annee academique courante : %s (precedente : %s)." % (suivante, courante)


# ===========================================================================
# C-4 (V1.99.12) — Budget previsionnel par formation (onglet F5_Budget_Prev)
# ===========================================================================
_F5_TAB = "F5_Budget_Prev"


def _taux_eur():
    """Taux EUR (KMF pour 1 EUR). Priorite : reglage 'taux_eur' (instance/reglages.json)
    -> sinon onglet P2_Taux (ligne Devise == EUR) -> sinon config.TAUX_EUR_DEFAUT."""
    brut = str(_reglage_get("taux_eur", "")).strip()
    if brut:
        v = _num(brut)
        if v > 0:
            return v
    for r in _lignes_dict("P2_Taux"):
        if str(r.get("Devise", "")).strip().upper() == "EUR":
            v = _num(r.get("Taux en KMF", ""))
            if v > 0:
                return v
    return float(config.TAUX_EUR_DEFAUT)


def _frais_admin_pct():
    """% de frais administratifs EMSP. reglage 'frais_admin_pct' -> sinon defaut config."""
    brut = str(_reglage_get("frais_admin_pct", "")).strip()
    if brut:
        v = _num(brut)
        if v >= 0:
            return v
    return float(config.FRAIS_ADMIN_PCT_DEFAUT)


def _q_budget(x):
    """Quantite budgetaire : vide -> 1 (permet de saisir un cout forfaitaire avec le
    seul Cout unitaire) ; sinon la valeur saisie (0 reste 0)."""
    s = str(x).strip()
    if s == "":
        return 1.0
    return _num(s)


def _montant_ligne_kmf(q1, q2, cu):
    """Montant d'une ligne = Qte1 x Qte2 x Cout unitaire (KMF), arrondi a l'entier."""
    return int(round(_q_budget(q1) * _q_budget(q2) * _num(cu)))


def _ordre_niveau(niv):
    return {"M1": 0, "M2": 1, "L1": 2, "L2": 3, "L3": 4}.get(str(niv).strip().upper(), 99)


def _eur(montant_kmf, taux):
    return round(montant_kmf / taux, 2) if taux else 0.0


def donnees_budget_prev(formation="", session=""):
    """Budget previsionnel structure pour l'ecran et l'impression. Filtre optionnel par
    Formation et/ou Session. Regroupe par Formation -> Niveau, calcule par ligne
    (Montant KMF/EUR), et par niveau Total / frais admin (%) / sous-total, puis le total
    par formation. Tous les calculs en Python (jamais de lecture de cellule formule)."""
    taux = _taux_eur()
    pct = _frais_admin_pct()
    f_cible = str(formation).strip().lower()
    s_cible = str(session).strip().lower()

    parnf = {}
    for idx, r in enumerate(_lignes_dict(_F5_TAB)):
        form = str(r.get("Formation", "")).strip()
        if not form:
            continue
        if f_cible and form.lower() != f_cible:
            continue
        sess = str(r.get("Session", "")).strip()
        if s_cible and sess.lower() != s_cible:
            continue
        niv = str(r.get("Niveau", "")).strip()
        mkmf = _montant_ligne_kmf(r.get("Qte1"), r.get("Qte2"), r.get("Cout unitaire (KMF)"))
        parnf.setdefault(form, {}).setdefault(niv, []).append({
            "index": idx, "niveau": niv,
            "rubrique": str(r.get("Rubrique", "")).strip(),
            "designation": str(r.get("Designation", "")).strip(),
            "unite1": str(r.get("Unite1", "")).strip(), "qte1": str(r.get("Qte1", "")).strip(),
            "unite2": str(r.get("Unite2", "")).strip(), "qte2": str(r.get("Qte2", "")).strip(),
            "cout": _num(r.get("Cout unitaire (KMF)", "")),
            "poste": str(r.get("Poste budgetaire", "")).strip(),
            "bailleur": str(r.get("Source de financement / Bailleur", "")).strip(),
            "session": sess, "montant_kmf": mkmf, "montant_eur": _eur(mkmf, taux),
        })

    formations = []
    total_general_kmf = 0
    for form in sorted(parnf):
        niveaux, total_form = [], 0
        for niv in sorted(parnf[form], key=_ordre_niveau):
            lignes = parnf[form][niv]
            total_niv = sum(l["montant_kmf"] for l in lignes)
            frais = int(round(total_niv * pct / 100.0))
            sous_total = total_niv + frais
            total_form += sous_total
            niveaux.append({
                "niveau": niv, "lignes": lignes,
                "total_kmf": total_niv, "total_eur": _eur(total_niv, taux),
                "frais_kmf": frais, "frais_eur": _eur(frais, taux),
                "sous_total_kmf": sous_total, "sous_total_eur": _eur(sous_total, taux),
            })
        total_general_kmf += total_form
        formations.append({
            "formation": form, "niveaux": niveaux,
            "total_kmf": total_form, "total_eur": _eur(total_form, taux),
        })

    return {
        "taux_eur": taux, "frais_admin_pct": pct,
        "formation": formation, "session": session,
        "formations": formations,
        "total_general_kmf": total_general_kmf,
        "total_general_eur": _eur(total_general_kmf, taux),
    }


def _ligne_brute_f5(d):
    """{entete_physique: valeur} d'une ligne F5 a partir d'un dict de champs propres.
    Montant (KMF) et Montant (EUR) calcules et ecrits comme VALEUR (precedent E4)."""
    mkmf = _montant_ligne_kmf(d.get("Qte1"), d.get("Qte2"), d.get("Cout unitaire (KMF)"))
    meur = _eur(mkmf, _taux_eur())
    champs = {
        "Formation": d.get("Formation", ""), "Niveau": d.get("Niveau", ""),
        "Rubrique": d.get("Rubrique", ""), "Designation": d.get("Designation", ""),
        "Unite1": d.get("Unite1", ""), "Qte1": d.get("Qte1", ""),
        "Unite2": d.get("Unite2", ""), "Qte2": d.get("Qte2", ""),
        "Cout unitaire (KMF)": d.get("Cout unitaire (KMF)", ""),
        "Poste budgetaire": d.get("Poste budgetaire", ""),
        "Source de financement / Bailleur": d.get("Source de financement / Bailleur", ""),
        "Session": d.get("Session", ""),
        "Montant (KMF)": mkmf, "Montant (EUR)": meur,
    }
    return {_brut(_F5_TAB, k): v for k, v in champs.items()}


def _peut_ecrire_budget(acteur):
    return est_admin(acteur) or peut_ecrire(acteur, _F5_TAB)


def ajouter_ligne_budget_prev(acteur, d, login):
    """Ajoute une ligne de budget previsionnel (Montant calcule). Renvoie (ok, msg)."""
    if not _peut_ecrire_budget(acteur):
        return False, "Action non autorisee (acces financier requis)."
    if not str(d.get("Formation", "")).strip():
        return False, "Formation obligatoire."
    if not str(d.get("Designation", "")).strip():
        return False, "Designation obligatoire."
    _db.ajouter_ligne(_F5_TAB, _ligne_brute_f5(d))
    auth.journal(login, "Budget prev. ligne ajoutee",
                 "%s / %s / %s" % (d.get("Formation", ""), d.get("Niveau", ""),
                                   d.get("Designation", "")), "")
    return True, "Ligne ajoutee."


def modifier_ligne_budget_prev(acteur, index, d, login):
    """Modifie une ligne existante (par index d'affichage). Renvoie (ok, msg)."""
    if not _peut_ecrire_budget(acteur):
        return False, "Action non autorisee (acces financier requis)."
    if not str(d.get("Designation", "")).strip():
        return False, "Designation obligatoire."
    try:
        _db.modifier_ligne(_F5_TAB, int(index), _ligne_brute_f5(d))
    except (IndexError, ValueError):
        return False, "Ligne introuvable."
    auth.journal(login, "Budget prev. ligne modifiee", "index %s" % index, "")
    return True, "Ligne modifiee."


def supprimer_ligne_budget_prev(acteur, index, login):
    """Supprime une ligne de budget previsionnel (par index d'affichage). (ok, msg)."""
    if not _peut_ecrire_budget(acteur):
        return False, "Action non autorisee (acces financier requis)."
    if _db.supprimer_ligne_par_index(_F5_TAB, index):
        auth.journal(login, "Budget prev. ligne supprimee", "index %s" % index, "")
        return True, "Ligne supprimee."
    return False, "Ligne introuvable."


def definir_taux_eur(acteur, valeur, login):
    """Definit le taux EUR (reglage). Reserve a la Direction / acces financier. (ok, msg)."""
    if not _peut_ecrire_budget(acteur):
        return False, "Action non autorisee."
    v = _num(valeur)
    if v <= 0:
        return False, "Taux invalide (attendu un nombre > 0)."
    _reglage_set("taux_eur", v)
    auth.journal(login, "Taux EUR defini", str(v), "")
    return True, "Taux EUR defini : %s KMF." % _fmt_kmf(v)


def definir_frais_admin_pct(acteur, valeur, login):
    """Definit le % de frais administratifs (reglage). (ok, msg)."""
    if not _peut_ecrire_budget(acteur):
        return False, "Action non autorisee."
    v = _num(valeur)
    if v < 0:
        return False, "Pourcentage invalide (>= 0)."
    _reglage_set("frais_admin_pct", v)
    auth.journal(login, "Frais admin % defini", str(v), "")
    return True, "Frais administratifs definis : %s %%." % v


def roster_classe(filiere, niveau, section=""):
    """Effectif de A1 pour une classe : meme Filiere + Niveau, et meme Section
    seulement si une section est demandee ET renseignee cote etudiant. Trie par
    Nom, Prenom, Matricule. Renvoie [{matricule, nom, prenom}]."""
    fil = str(filiere or "").strip().lower()
    niv = str(niveau or "").strip().lower()
    sec = str(section or "").strip().lower()
    out = []
    for e in _lignes_dict("A1_Etudiants"):
        mat = str(e.get("Matricule", "")).strip()
        if not mat:
            continue
        if fil and str(e.get("Filiere", "")).strip().lower() != fil:
            continue
        if niv and str(e.get("Niveau", "")).strip().lower() != niv:
            continue
        es = str(e.get("Section", "")).strip().lower()
        if sec and es and es != sec:
            continue
        out.append({"matricule": mat,
                    "nom": str(e.get("Nom", "")).strip(),
                    "prenom": str(e.get("Prenom", "")).strip()})
    out.sort(key=lambda x: (x["nom"].lower(), x["prenom"].lower(), x["matricule"]))
    return out


def classes_en_service():
    """Classes reellement peuplees (depuis A1), pour les listes de l'ecran :
    [{filiere, niveaux:[{niveau, sections:[...]}]}], dans l'ordre d'apparition."""
    from collections import OrderedDict
    arbre = OrderedDict()
    for e in _lignes_dict("A1_Etudiants"):
        f = str(e.get("Filiere", "")).strip()
        n = str(e.get("Niveau", "")).strip()
        s = str(e.get("Section", "")).strip()
        if not f and not n:
            continue
        arbre.setdefault(f, OrderedDict())
        arbre[f].setdefault(n, [])
        if s and s not in arbre[f][n]:
            arbre[f][n].append(s)
    out = []
    for f, nivs in arbre.items():
        out.append({"filiere": f,
                    "niveaux": [{"niveau": n, "sections": sorted(secs)}
                                for n, secs in nivs.items()]})
    return out


def cle_session_libre(filiere, niveau, matiere, enseignant):
    """Cle composee lisible pour A2 'Session / Matiere' (seance ad hoc sans A3)."""
    classe = (str(filiere or "").strip() + " " + str(niveau or "").strip()).strip()
    bits = [b for b in [classe, str(matiere or "").strip(),
                        str(enseignant or "").strip()] if b]
    return " - ".join(bits)


def _prochain_id_session():
    """Prochain 'ID session' libre pour A3 : S001, S002, ... (max existant + 1)."""
    mx = 0
    for r in _lignes_dict("A3_Sessions"):
        sid = str(r.get("ID session", "")).strip()
        if sid[:1] in ("S", "s") and sid[1:].isdigit():
            mx = max(mx, int(sid[1:]))
        elif sid.isdigit():
            mx = max(mx, int(sid))
    return "S%03d" % (mx + 1)


def creer_session_recurrente(filiere, niveau, section, matiere, enseignant,
                             salle, jour, debut, fin):
    """Cree une ligne A3_Sessions (seance recurrente) et renvoie l'ID session."""
    sid = _prochain_id_session()
    valeurs = {
        _brut("A3_Sessions", "ID session"): sid,
        _brut("A3_Sessions", "Annee acad."): _annee_acad_defaut(),
        _brut("A3_Sessions", "Filiere"): str(filiere or "").strip(),
        _brut("A3_Sessions", "Niveau"): str(niveau or "").strip(),
        _brut("A3_Sessions", "Section"): str(section or "").strip(),
        _brut("A3_Sessions", "Matiere"): str(matiere or "").strip(),
        _brut("A3_Sessions", "Enseignant"): str(enseignant or "").strip(),
        _brut("A3_Sessions", "Salle"): str(salle or "").strip(),
        _brut("A3_Sessions", "Jour"): str(jour or "").strip(),
        _brut("A3_Sessions", "Heure debut"): str(debut or "").strip(),
        _brut("A3_Sessions", "Heure fin"): str(fin or "").strip(),
    }
    _db.ajouter_ligne("A3_Sessions", valeurs)
    return sid


def presences_existantes_libre(date_fr, filiere, niveau, matiere, enseignant,
                               debut, fin):
    """Etat deja enregistre pour une seance ad hoc (cle composee), {mat: 'O'/'N'}."""
    sv = cle_session_libre(filiere, niveau, matiere, enseignant)
    cr = "%s-%s" % (str(debut).strip(), str(fin).strip())
    return presences_existantes(date_fr, sv, cr)


CRENEAU_DEBUT = {"10h": "10:00", "12h": "12:00", "15h": "15:00", "17h": "17:00"}


def creneau_duree_heures(creneau, duree):
    """Convertit un creneau (10h/12h/15h/17h) + une duree (1 ou 2 heures) en
    (debut, fin) au format HH:MM. Renvoie ('', '') si l'entree est invalide.
    L'IHM des presences n'expose que creneau + duree ; le moteur de comptage
    des heures continue de recevoir debut/fin sans changement."""
    deb = CRENEAU_DEBUT.get(str(creneau).strip())
    try:
        d = int(str(duree).strip().lower().rstrip("h") or "0")
    except ValueError:
        d = 0
    if not deb or d not in (1, 2):
        return "", ""
    return deb, "%02d:00" % (int(deb[:2]) + d)


def enregistrer_presences_libre(filiere, niveau, section, date_fr, debut, fin,
                                matiere, enseignant, salle, presents,
                                recurrente, saisi_par, peut_creer_a3=True):
    """Saisie en liste d'une seance ad hoc. Ecrit A2 (upsert) ; cree A3 si
    'recurrente'. Renvoie (ok, message). N'ecrit jamais de formule."""
    date_fr = str(date_fr).strip()
    if not _RE_DATE_FR.match(date_fr):
        return False, "Date invalide : format attendu JJ/MM/AAAA."
    debut = str(debut).strip()
    fin = str(fin).strip()
    if not (_RE_HEURE.match(debut) and _RE_HEURE.match(fin)):
        return False, "Heures invalides : format attendu HH:MM (ex. 10:00)."
    if debut >= fin:
        return False, "L'heure de fin doit etre posterieure a l'heure de debut."
    if not str(matiere).strip():
        return False, "Choisissez une matiere."
    roster = roster_classe(filiere, niveau, section)
    if not roster:
        return False, ("Aucun etudiant pour cette classe "
                       "(verifier Filiere / Niveau / Section dans A1_Etudiants).")
    creneau = "%s-%s" % (debut, fin)
    msg_a3 = ""
    if recurrente:
        if not peut_creer_a3:
            return False, ("Creation de seance recurrente refusee : droit "
                           "d'ecriture sur A3_Sessions requis.")
        jour = _jour_de_date(date_fr)
        sid = creer_session_recurrente(filiere, niveau, section, matiere,
                                       enseignant, salle, jour, debut, fin)
        session_val = sid
        msg_a3 = " Seance creee dans A3_Sessions (%s, %s)." % (sid, jour or "?")
    else:
        session_val = cle_session_libre(filiere, niveau, matiere, enseignant)
    presents = set(str(m).strip() for m in presents)
    bd, bm, bs = (_brut("A2_Presences", "Date"),
                  _brut("A2_Presences", "Matricule"),
                  _brut("A2_Presences", "Session / Matiere"))
    bc, bp, bsp = (_brut("A2_Presences", "Creneau"),
                   _brut("A2_Presences", "Present (O/N)"),
                   _brut("A2_Presences", "Saisi par"))
    lignes = [{bd: date_fr, bm: e["matricule"], bs: session_val, bc: creneau,
               bp: "O" if e["matricule"] in presents else "N",
               bsp: str(saisi_par or "").strip()} for e in roster]
    res = _db.ecrire_lignes_lot("A2_Presences", lignes, cles=[bd, bm, bs, bc])
    nb_pres = len(presents & set(e["matricule"] for e in roster))
    return True, ("Presences enregistrees : %d present(s) / %d ; "
                  "%d ajout(s), %d mise(s) a jour.%s"
                  % (nb_pres, len(roster), res["ajout"], res["maj"], msg_a3))


# ===========================================================================
# IMPORT CSV NATIONAL -> zone de staging IMPORT_zone (V1.16)
# ---------------------------------------------------------------------------
# Flux : la Direction COLLE le CSV national ; les lignes remplissent IMPORT_zone
# (colonnes A-G). La colonne "Statut vs base" (calculee en Python) indique
# NOUVEAU / EXISTANT par rapport a A1_Etudiants. La copie des NOUVEAU vers A1
# reste MANUELLE (aucune ecriture dans A1 ici). "Retour en arriere" : avant chaque
# import / vidage, un instantane de la zone est sauvegarde sur disque ; il peut
# etre restaure en un clic.
# ---------------------------------------------------------------------------

# Colonnes saisissables d'IMPORT_zone (A-G), dans l'ordre, hors colonne calcul.
def _colonnes_import():
    return [m for m in entetes_meta("IMPORT_zone") if not m["readonly"]]


def parser_csv(texte):
    """Parse un CSV colle : auto-detection du separateur (tabulation / ; / ,),
    saut d'une eventuelle ligne d'en-tete, 7 colonnes (A-G), valeurs nettoyees.
    Renvoie une liste de listes (max 7 valeurs par ligne)."""
    lignes_txt = [l for l in (texte or "").replace("\r", "").split("\n") if l.strip()]
    if not lignes_txt:
        return []
    prem = lignes_txt[0]
    sep = "\t" if "\t" in prem else (";" if prem.count(";") >= prem.count(",") and ";" in prem
                                     else ("," if "," in prem else "\t"))
    n = len(_colonnes_import())
    rows = []
    for i, l in enumerate(lignes_txt):
        champs = [c.strip() for c in l.split(sep)]
        # ligne d'en-tete ignoree : 1re ligne dont la 1re cellule ne contient aucun
        # chiffre (un matricule est numerique ; un en-tete est un libelle).
        if i == 0 and champs and not any(ch.isdigit() for ch in champs[0]):
            continue
        champs = (champs + [""] * n)[:n]
        if any(c for c in champs):
            rows.append(champs)
    return rows


def import_zone_brut():
    """Lignes actuelles d'IMPORT_zone (colonnes A-G) en liste de {brut: valeur}.
    Sert d'instantane pour l'annulation et de base de relecture."""
    cols = _colonnes_import()
    donnees = {m["brut"]: _db.colonne("IMPORT_zone", m["brut"]) for m in cols}
    n = max((len(v) for v in donnees.values()), default=0)
    out = []
    for i in range(n):
        row = {b: (vals[i] if i < len(vals) else "") for b, vals in donnees.items()}
        if any(str(v).strip() for v in row.values()):
            out.append({b: ("" if v is None else v) for b, v in row.items()})
    return out


def _undo_path():
    return os.path.join(os.path.dirname(config.WORKBOOK), "import_undo.json")


def _sauver_undo(rows):
    try:
        with open(_undo_path(), "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False)
    except OSError:
        pass


def import_undo_existe():
    return os.path.exists(_undo_path())


def remplacer_zone_import(rows_valeurs):
    """rows_valeurs = liste de {brut: valeur}. Remplace tout le contenu de la zone."""
    return _db.remplacer_donnees("IMPORT_zone", rows_valeurs)


def importer_csv(texte):
    """Parse + remplace la zone d'import. Sauvegarde l'instantane precedent (undo).
    Renvoie (nb_lignes, message)."""
    rows = parser_csv(texte)
    if not rows:
        return 0, "Aucune ligne exploitable dans le texte colle."
    _sauver_undo(import_zone_brut())          # instantane avant remplacement
    cols = _colonnes_import()
    valeurs = [{cols[j]["brut"]: row[j] for j in range(len(cols))} for row in rows]
    n = remplacer_zone_import(valeurs)
    return n, "%d ligne(s) importee(s) dans la zone. Verifiez la colonne Statut vs base." % n


def vider_zone_import():
    """Vide la zone d'import (avec instantane pour annulation). Renvoie le message."""
    _sauver_undo(import_zone_brut())
    remplacer_zone_import([])
    return "Zone d'import videe."


def annuler_import():
    """Restaure le dernier instantane sauvegarde. Renvoie (ok, message)."""
    if not import_undo_existe():
        return False, "Aucune version precedente a restaurer."
    try:
        with open(_undo_path(), "r", encoding="utf-8") as f:
            rows = json.load(f)
    except (OSError, ValueError):
        return False, "Instantane illisible."
    remplacer_zone_import(rows)
    try:
        os.remove(_undo_path())               # annulation a un seul niveau
    except OSError:
        pass
    return True, "Version precedente restauree (%d ligne(s))." % len(rows)


def import_resume():
    """Table IMPORT_zone (avec Statut vs base calcule) + compteurs NOUVEAU/EXISTANT."""
    t = table("IMPORT_zone")
    libs = [m["libelle"] for m in t["entetes"]]
    try:
        j = libs.index("Statut vs base")
    except ValueError:
        j = None
    nouveaux = existants = 0
    if j is not None:
        for lig in t["lignes"]:
            v = str(lig[j]).strip().upper() if j < len(lig) else ""
            if v == "NOUVEAU":
                nouveaux += 1
            elif v == "EXISTANT":
                existants += 1
    t["nouveaux"] = nouveaux
    t["existants"] = existants
    t["undo"] = import_undo_existe()
    return t


# ===========================================================================
# MODULE IMPRESSIONS & EDITIONS — V1.18
# ---------------------------------------------------------------------------
# 6 documents : liste etudiants, feuille de presence vierge, releve d'heures
# individuel (paie), recapitulatif mensuel des heures, recu de paiement,
# attestation de passage. Modeles persistants dans D1_Modeles_docs (parties
# fixes editables + jetons {…}). Rendu HTML + @media print (window.print).
# L'export du tableau de bord = vrai .xlsx via openpyxl.
# ===========================================================================
import io as _io


def _kmf_aff(n):
    """Montant KMF en chiffres, separateur de milliers espace, sans decimales."""
    try:
        return "{:,.0f}".format(round(float(_num(n)))).replace(",", " ")
    except Exception:
        return str(n)


def _civilite(genre):
    """Civilite deduite du Genre (parametrable, config.CIVILITES)."""
    g = str(genre or "").strip().upper()
    if not g:
        return config.CIVILITE_DEFAUT
    if g.startswith("F"):              # Feminin / Femme / F
        return config.CIVILITES.get("F", "Madame")
    if g.startswith("M") or g.startswith("H"):   # Masculin / Homme / M
        return config.CIVILITES.get("M", "Monsieur")
    return config.CIVILITE_DEFAUT


# --- Modeles de documents (D1_Modeles_docs) --------------------------------
def _modeles_stockes():
    """Modeles enregistres dans D1, indexes par Cle doc."""
    out = {}
    if config.MODELE_TAB not in _db.onglets():
        return out
    for r in _lignes_dict(config.MODELE_TAB):
        cle = str(r.get("Cle doc", "")).strip()
        if cle:
            out[cle] = r
    return out


def modele_doc(cle):
    """Modele d'un document : valeurs stockees completees par les defauts config."""
    spec = config.MODELES_DOCS.get(cle)
    if not spec:
        return None
    stock = _modeles_stockes().get(cle, {})
    d = spec["defauts"]
    def val(col):
        v = str(stock.get(col, "")).strip()
        return v if v != "" else d.get(col, "")
    try:
        nbc = int(float(str(val("Nb copies")).strip() or "1"))
    except Exception:
        nbc = 1
    nbc = min(max(nbc, 1), 5)
    return {
        "cle": cle, "libelle": spec["libelle"], "source": spec["source"],
        "tabulaire": spec["tabulaire"], "icone": spec.get("icone", "ti-file"),
        "jetons": spec["jetons"],
        "entete": val("En-tete"), "titre": val("Titre"), "corps": val("Corps"),
        "mentions": val("Mentions / pied"), "signataire": val("Libelle signataire"),
        "nb_copies": nbc,
    }


def modeles_docs():
    """Liste ordonnee des modeles (pour l'ecran Modeles et le hub Impressions)."""
    return [modele_doc(c) for c in config.MODELES_ORDRE]


def enregistrer_modele(cle, valeurs):
    """Met a jour (upsert par Cle doc) un modele dans D1_Modeles_docs.
    valeurs = {En-tete, Titre, Corps, Mentions / pied, Libelle signataire, Nb copies}."""
    spec = config.MODELES_DOCS.get(cle)
    if not spec:
        return False, "Modele inconnu."
    ligne = {"Cle doc": cle, "Libelle": spec["libelle"]}
    for col in ["En-tete", "Titre", "Corps", "Mentions / pied",
                "Libelle signataire", "Nb copies"]:
        ligne[col] = str(valeurs.get(col, "")).strip()
    if not ligne["Titre"]:
        return False, "Le titre est obligatoire."
    try:
        nbc = int(float(ligne["Nb copies"] or "1"))
    except Exception:
        return False, "Nombre de copies invalide."
    ligne["Nb copies"] = str(min(max(nbc, 1), 5))
    try:
        _db.ecrire_lignes_lot(config.MODELE_TAB, [ligne], cles=["Cle doc"])
    except Exception:
        return False, "Echec de l'enregistrement du modele."
    return True, "Modele enregistre."


def rendre_modele(cle, jetons=None):
    """Modele avec jetons {…} remplaces (titre, corps, mentions)."""
    m = modele_doc(cle)
    if m is None:
        return None
    jetons = jetons or {}
    def remplir(txt):
        for k, v in jetons.items():
            txt = txt.replace("{" + k + "}", "" if v is None else str(v))
        return txt
    m = dict(m)
    m["titre"] = remplir(m["titre"])
    m["corps"] = remplir(m["corps"])
    m["mentions"] = remplir(m["mentions"])
    return m


# --- Numero de recu : scan-based (max F1 annee courante + 1), reset annuel ---
def _prochain_recu(annee=None):
    annee = annee or _dt.date.today().year
    rx = re.compile(r"^REC-%d-0*(\d+)$" % annee, re.IGNORECASE)
    maxi = 0
    for v in _db.colonne("F1_Mouvements", _brut("F1_Mouvements", "Reference / N piece")):
        m = rx.match(str(v).strip())
        if m:
            maxi = max(maxi, int(m.group(1)))
    return "REC-%d-%04d" % (annee, maxi + 1)


# --- Liste d'etudiants (A1) -------------------------------------------------
_LISTE_COLS = ["N ordre", "Matricule", "Genre", "Nom", "Prenom",
               "Date naissance", "Lieu naissance", "Origine / lieu actuel"]


def liste_filtres():
    """Valeurs distinctes des criteres de filtrage de la liste etudiants."""
    rows = _lignes_dict("A1_Etudiants")
    def distinct(champ):
        vus, out = set(), []
        for r in rows:
            v = str(r.get(champ, "")).strip()
            if v and v not in vus:
                vus.add(v)
                out.append(v)
        return sorted(out)
    return {"filiere": distinct("Filiere"), "niveau": distinct("Niveau"),
            "section": distinct("Section"), "annee": distinct("Annee acad.")}


def liste_etudiants(filiere="", niveau="", section="", annee=""):
    """Etudiants filtres + en-tete + colonnes pour impression."""
    rows = _lignes_dict("A1_Etudiants")
    def garde(r):
        return ((not filiere or str(r.get("Filiere", "")).strip() == filiere)
                and (not niveau or str(r.get("Niveau", "")).strip() == niveau)
                and (not section or str(r.get("Section", "")).strip() == section)
                and (not annee or str(r.get("Annee acad.", "")).strip() == annee))
    sel = [r for r in rows if garde(r)]
    lignes = []
    for i, r in enumerate(sel, start=1):
        lignes.append([str(i), r.get("Matricule", ""), r.get("Genre", ""),
                       r.get("Nom", ""), r.get("Prenom", ""),
                       r.get("Date naissance", ""), r.get("Lieu naissance", ""),
                       r.get("Origine / lieu actuel", "")])
    contexte = {"filiere": filiere or "Toutes", "niveau": niveau or "Tous",
                "section": section or "Toutes", "annee": annee or "Toutes",
                "effectif": len(sel), "date_jour": fmt_date(_dt.date.today())}
    return {"colonnes": _LISTE_COLS, "lignes": lignes, "contexte": contexte}


# --- Attestation : jetons d'un etudiant (A1) --------------------------------
def etudiants_dispo():
    """Etudiants selectionnables (matricule -> libelle Nom Prenom)."""
    out = []
    for r in _lignes_dict("A1_Etudiants"):
        mat = str(r.get("Matricule", "")).strip()
        if not mat:
            continue
        nom = ("%s %s" % (r.get("Nom", ""), r.get("Prenom", ""))).strip()
        out.append({"matricule": mat, "libelle": "%s — %s" % (mat, nom)})
    return sorted(out, key=lambda o: o["libelle"])


def attestation_jetons(matricule):
    matricule = str(matricule).strip()
    for r in _lignes_dict("A1_Etudiants"):
        if str(r.get("Matricule", "")).strip() == matricule:
            return {
                "civilite": _civilite(r.get("Genre", "")),
                "nom": r.get("Nom", ""), "prenom": r.get("Prenom", ""),
                "matricule": matricule,
                "date_naissance": r.get("Date naissance", ""),
                "lieu_naissance": r.get("Lieu naissance", ""),
                "niveau": r.get("Niveau", ""), "filiere": r.get("Filiere", ""),
                "section": r.get("Section", ""), "annee": r.get("Annee acad.", ""),
                "date_jour": fmt_date(_dt.date.today()),
            }
    return None


# --- Recu de paiement : recettes de F1 --------------------------------------
def recettes_dispo():
    """Recettes (Sens=Recette) de F1, selectionnables par index de ligne."""
    out = []
    for i, r in enumerate(_lignes_dict("F1_Mouvements")):
        if str(r.get("Sens", "")).strip().lower().startswith("recette"):
            out.append({
                "idx": i,
                "reference": str(r.get("Reference / N piece", "")).strip(),
                "tiers": r.get("Tiers", ""),
                "montant": _kmf_aff(r.get("Montant Recette (KMF)", 0)),
                "libelle": r.get("Libelle / description", ""),
                "date": r.get("Date operation", ""),
            })
    return out


def recu_jetons(idx):
    """Jetons du recu pour la recette F1 a l'index donne."""
    rows = _lignes_dict("F1_Mouvements")
    try:
        r = rows[int(idx)]
    except (ValueError, IndexError, TypeError):
        return None
    if not str(r.get("Sens", "")).strip().lower().startswith("recette"):
        return None
    return {
        "reference": str(r.get("Reference / N piece", "")).strip() or _prochain_recu(),
        "date_operation": r.get("Date operation", ""),
        "tiers": r.get("Tiers", ""),
        "montant": _kmf_aff(r.get("Montant Recette (KMF)", 0)),
        "categorie": r.get("Categorie", ""),
        "compte": r.get("Compte / caisse", ""),
        "mode_paiement": r.get("Mode paiement", ""),
        "libelle": r.get("Libelle / description", ""),
        "date_jour": fmt_date(_dt.date.today()),
    }


# --- Releve d'heures : individuel (paie) + recapitulatif mensuel ------------
def _ens_index():
    """Matricule ens. -> {nom, prenom, statut, departement} depuis E1."""
    idx = {}
    for r in _lignes_dict("E1_Enseignants"):
        mat = str(r.get("Matricule ens.", "")).strip()
        if mat:
            idx[mat] = {
                "nom": r.get("Nom", ""), "prenom": r.get("Prenom", ""),
                "statut": r.get("Statut", ""),
                "departement": r.get("Departement", ""),
            }
    return idx


def mois_dispo():
    """Periodes Mois/Annee distinctes presentes dans E2."""
    vus, out = set(), []
    for r in _lignes_dict("E2_Releve_heures"):
        v = str(r.get("Mois / Annee", "")).strip()
        if v and v not in vus:
            vus.add(v)
            out.append(v)
    return sorted(out)


def releve_individuel(mois_annee, matricule):
    """Jetons du releve individuel (E2 pour la periode + identite E1)."""
    mois_annee = str(mois_annee).strip()
    matricule = str(matricule).strip()
    ens = _ens_index().get(matricule, {})
    for r in _lignes_dict("E2_Releve_heures"):
        if (str(r.get("Mois / Annee", "")).strip() == mois_annee
                and str(r.get("Matricule ens.", "")).strip() == matricule):
            constate = r.get("Vol. horaire constate", "")
            return {
                "matricule": matricule, "nom": ens.get("nom", ""),
                "prenom": ens.get("prenom", ""), "statut": ens.get("statut", ""),
                "departement": ens.get("departement", ""),
                "mois_annee": _periode_libelle(mois_annee),
                "vol_prog": r.get("Vol. horaire prog.", ""),
                "vol_constate": constate,
                "total_heures": constate,   # Total a payer = constate (CALC_AFFICHAGE)
                "date_jour": fmt_date(_dt.date.today()),
            }
    return None


def releves_individuels_dispo(mois_annee):
    """Enseignants ayant un releve pour la periode (selection du releve individuel)."""
    mois_annee = str(mois_annee).strip()
    ens = _ens_index()
    out = []
    for r in _lignes_dict("E2_Releve_heures"):
        if str(r.get("Mois / Annee", "")).strip() != mois_annee:
            continue
        mat = str(r.get("Matricule ens.", "")).strip()
        if not mat:
            continue
        e = ens.get(mat, {})
        nom = ("%s %s" % (e.get("nom", ""), e.get("prenom", ""))).strip()
        out.append({"matricule": mat, "libelle": "%s — %s" % (mat, nom or "?")})
    return sorted(out, key=lambda o: o["libelle"])


def releve_recap(mois_annee):
    """Recapitulatif mensuel : une ligne par enseignant + total general."""
    mois_annee = str(mois_annee).strip()
    ens = _ens_index()
    lignes, tot_prog, tot_const = [], 0.0, 0.0
    for r in _lignes_dict("E2_Releve_heures"):
        if str(r.get("Mois / Annee", "")).strip() != mois_annee:
            continue
        mat = str(r.get("Matricule ens.", "")).strip()
        e = ens.get(mat, {})
        nom = ("%s %s" % (e.get("nom", ""), e.get("prenom", ""))).strip()
        prog = _num(r.get("Vol. horaire prog.", 0))
        const = _num(r.get("Vol. horaire constate", 0))
        tot_prog += prog
        tot_const += const
        lignes.append([mat, nom, _fmt_kmf(prog), _fmt_kmf(const), _fmt_kmf(const)])
    lignes.sort(key=lambda x: x[1])
    return {
        "colonnes": ["Matricule", "Nom Prenom", "Vol. programme",
                     "Vol. constate", "Total a payer (h)"],
        "lignes": lignes,
        "total": ["", "TOTAL GENERAL", _fmt_kmf(tot_prog),
                  _fmt_kmf(tot_const), _fmt_kmf(tot_const)],
        "contexte": {"mois_annee": _periode_libelle(mois_annee), "nb": len(lignes),
                     "date_jour": fmt_date(_dt.date.today())},
    }



# ===========================================================================
# HEURES CONSTATEES DES ENSEIGNANTS (V1.55)
# ---------------------------------------------------------------------------
# Principe : l'appel des eleves (A2_Presences) atteste qu'une seance a eu lieu.
# La seance du planning (A3_Sessions) donne l'enseignant programme et la duree
# (Heure fin - Heure debut, a defaut Vol. horaire prog.). Les heures constatees
# d'un enseignant pour un mois = somme des durees des seances appelees ce mois.
# L'onglet E3_Seances_faites ne stocke que les EXCEPTIONS (remplacant, cours
# annule, duree differente) qui surchargent ce calcul par defaut. Le resultat
# alimente le Vol. horaire constate d'E2 (report avec garde-fou anti-ecrasement) ;
# la chaine paie (releve_individuel / releve_recap) lit E2, inchangee.
# ---------------------------------------------------------------------------

def _mois_de_date(date_fr):
    """'JJ/MM/AAAA' -> 'MM/AAAA' (cle de periode E2). '' si format invalide."""
    m = re.match(r"^\d{2}/(\d{2})/(\d{4})$", str(date_fr or "").strip())
    return ("%s/%s" % (m.group(1), m.group(2))) if m else ""


def _a3_index():
    """ID session -> {enseignant(label), matiere, filiere, niveau, section,
    deb_min, fin_min, vol_prog}."""
    idx = {}
    for r in _lignes_dict("A3_Sessions"):
        sid = str(r.get("ID session", "")).strip()
        if not sid:
            continue
        idx[sid] = {
            "enseignant": str(r.get("Enseignant", "")).strip(),
            "matiere": str(r.get("Matiere", "")).strip(),
            "filiere": str(r.get("Filiere", "")).strip(),
            "niveau": str(r.get("Niveau", "")).strip(),
            "section": str(r.get("Section", "")).strip(),
            "deb_min": _parse_heure(r.get("Heure debut", "")),
            "fin_min": _parse_heure(r.get("Heure fin", "")),
            "vol_prog": _num_h(r.get("Vol. horaire prog.", 0)),
        }
    return idx


def _ens_label_index():
    """'nom prenom' (minuscule) -> [matricules] (detection des homonymes)."""
    m = {}
    for r in _lignes_dict("E1_Enseignants"):
        mat = str(r.get("Matricule ens.", "")).strip()
        lbl = ("%s %s" % (str(r.get("Nom", "")).strip(),
                          str(r.get("Prenom", "")).strip())).strip()
        if mat and lbl:
            m.setdefault(lbl.lower(), []).append(mat)
    return m


def _duree_h_a3(a3d):
    dm, fm = a3d.get("deb_min"), a3d.get("fin_min")
    if dm is not None and fm is not None and fm > dm:
        return round((fm - dm) / 60.0, 2)
    return _num_h(a3d.get("vol_prog", 0))


def _appels_distincts(mois_annee=None):
    """Triplets distincts (Date, Session/Matiere, Creneau) ayant au moins un
    appel dans A2 (une seance comptee une fois, quel que soit le nb d'eleves)."""
    vus = set()
    for r in _lignes_dict("A2_Presences"):
        d = str(r.get("Date", "")).strip()
        sv = str(r.get("Session / Matiere", "")).strip()
        cr = str(r.get("Creneau", "")).strip()
        if not (d and sv):
            continue
        if mois_annee and _mois_de_date(d) != str(mois_annee).strip():
            continue
        vus.add((d, sv, cr))
    return vus


def mois_appels_dispo():
    """Periodes 'MM/AAAA' distinctes presentes dans les appels A2 (selecteur)."""
    vus = set()
    for (d, _sv, _cr) in _appels_distincts():
        m = _mois_de_date(d)
        if m:
            vus.add(m)
    return sorted(vus, key=lambda x: (x.split("/")[1], x.split("/")[0]))


def _exceptions_index():
    """(Date, Session, Creneau) -> dict de l'exception E3."""
    idx = {}
    for r in _lignes_dict("E3_Seances_faites"):
        d = str(r.get("Date", "")).strip()
        sv = str(r.get("Session / Matiere", "")).strip()
        cr = str(r.get("Creneau", "")).strip()
        if not (d and sv):
            continue
        idx[(d, sv, cr)] = {
            "etat": str(r.get("Etat", "")).strip(),
            "assure_par": str(r.get("Assure par", "")).strip(),
            "matiere_reelle": str(r.get("Matiere reelle", "")).strip(),
            "vol_constate": str(r.get("Vol. constate h", "")).strip(),
            "motif": str(r.get("Motif", "")).strip(),
        }
    return idx


def enseignant_programme_session(session_val):
    """(matricule, label, ambigu) de l'enseignant programme d'une seance A3."""
    a3 = _a3_index().get(str(session_val).strip())
    lbl = a3["enseignant"] if a3 else ""
    cands = _ens_label_index().get(lbl.lower(), [])
    if len(cands) == 1:
        return cands[0], lbl, False
    return "", lbl, (len(cands) > 1)


def seance_faite_existante(date_fr, session_val, creneau):
    """Exception E3 deja enregistree pour ce triplet (prefill de l'ecran appel)."""
    return _exceptions_index().get(
        (str(date_fr).strip(), str(session_val).strip(), str(creneau).strip()), {})


def heures_constatees_detail(mois_annee):
    """Calcule, pour le mois 'MM/AAAA', les heures constatees par enseignant
    (matricule) avec le detail des seances et la liste des anomalies."""
    mois_annee = str(mois_annee).strip()
    a3 = _a3_index()
    labels = _ens_label_index()
    ens = _ens_index()
    exc = _exceptions_index()
    par_ens, anomalies = {}, []
    for (d, sv, cr) in sorted(_appels_distincts(mois_annee)):
        ex = exc.get((d, sv, cr), {})
        if ex.get("etat", "") == "Cours annule":
            continue
        a3d = a3.get(sv)
        mat = ex.get("assure_par", "")
        matiere = ex.get("matiere_reelle", "")
        duree = _num_h(ex.get("vol_constate")) if ex.get("vol_constate", "") else None
        remplacant = bool(ex.get("assure_par", ""))
        if a3d:
            if not matiere:
                matiere = a3d["matiere"]
            if duree is None:
                duree = _duree_h_a3(a3d)
            if not mat:
                lbl = a3d["enseignant"]
                cands = labels.get(lbl.lower(), [])
                if len(cands) == 1:
                    mat = cands[0]
                elif not cands:
                    anomalies.append("Seance %s du %s (%s) : enseignant programme "
                                     "'%s' introuvable dans E1 — non comptabilise."
                                     % (sv, d, cr, lbl or "?"))
                    continue
                else:
                    anomalies.append("Seance %s du %s (%s) : homonyme '%s' (%d "
                                     "matricules) — attribuer via une exception."
                                     % (sv, d, cr, lbl, len(cands)))
                    continue
        else:
            if not mat or duree is None:
                anomalies.append("Appel du %s (%s) sur '%s' : seance A3 introuvable "
                                 "— non comptabilise (preciser via une exception)."
                                 % (d, cr, sv))
                continue
            if not matiere:
                matiere = sv
        e = ens.get(mat, {})
        rec = par_ens.setdefault(mat, {"nom": e.get("nom", ""),
                                       "prenom": e.get("prenom", ""),
                                       "total": 0.0, "seances": []})
        rec["total"] += duree or 0.0
        rec["seances"].append({"date": d, "creneau": cr, "session": sv,
                               "matiere": matiere, "duree": duree,
                               "remplacant": remplacant})
    return {"mois": mois_annee, "par_ens": par_ens, "anomalies": anomalies}


def _h_aff(n):
    n = round(n, 2)
    return int(n) if n == int(n) else n


def _e2_constate_mois(mois_annee):
    """{matricule: 'Vol. horaire constate' (texte)} d'E2 pour le mois."""
    out = {}
    for r in _lignes_dict("E2_Releve_heures"):
        if str(r.get("Mois / Annee", "")).strip() == str(mois_annee).strip():
            out[str(r.get("Matricule ens.", "")).strip()] = \
                str(r.get("Vol. horaire constate", "")).strip()
    return out


def _e2_prog_mois(mois_annee):
    """{matricule: 'Vol. horaire prog.' (texte)} d'E2 pour le mois (= H prevues)."""
    out = {}
    for r in _lignes_dict("E2_Releve_heures"):
        if str(r.get("Mois / Annee", "")).strip() == str(mois_annee).strip():
            out[str(r.get("Matricule ens.", "")).strip()] = \
                str(r.get("Vol. horaire prog.", "")).strip()
    return out


def heures_constatees_apercu(mois_annee):
    """Tableau de l'ecran 'Heures du mois' : par enseignant, H prevues (E2) / faites
    (calcule A2xA3+E3) / enregistrees (E2), avec deux ecarts de controle (#8) :
    'prevu - fait' (seances prevues non assurees/non pointees) et 'calcule - enregistre'
    (feuilles non remplies ou retard de saisie). Statut de report inchange.
    L'union des matricules (calcules U E2 du mois) fait apparaitre aussi les enseignants
    qui ont un volume prevu mais aucune heure faite (controle 'feuilles non remplies')."""
    mois_annee = str(mois_annee).strip()
    det = heures_constatees_detail(mois_annee)
    e2c = _e2_constate_mois(mois_annee)
    e2p = _e2_prog_mois(mois_annee)
    ens = _ens_index()
    mats = set(det["par_ens"].keys()) | set(e2c.keys()) | set(e2p.keys())
    lignes = []
    tot_prevu = tot_calc = tot_actuel = 0.0
    for mat in mats:
        if not mat:
            continue
        rec = det["par_ens"].get(mat)
        reportable = rec is not None            # une heure calculee a reporter
        calc = round(rec["total"], 2) if rec else 0.0
        if rec:
            nom = ("%s %s" % (rec["nom"], rec["prenom"])).strip()
            nb_seances, seances = len(rec["seances"]), rec["seances"]
        else:
            e = ens.get(mat, {})
            nom = ("%s %s" % (e.get("nom", ""), e.get("prenom", ""))).strip()
            nb_seances, seances = 0, []
        actuel = e2c.get(mat, "")
        prevu = e2p.get(mat, "")
        if actuel == "":
            statut = "a_reporter"
        elif abs(_num_h(actuel) - calc) < 1e-6:
            statut = "identique"
        else:
            statut = "different"
        ecart_pf = _fmt_ecart(_num_h(prevu) - calc) if prevu != "" else ""
        ecart_ce = _fmt_ecart(calc - _num_h(actuel)) if actuel != "" else ""
        zero = ("", "0", "+0", "-0")
        alerte = (statut == "a_reporter") or (ecart_pf not in zero) or (ecart_ce not in zero)
        lignes.append({"matricule": mat, "nom": nom, "reportable": reportable,
                       "calc": _h_aff(calc), "actuel": actuel, "prevu": prevu,
                       "ecart_pf": ecart_pf, "ecart_ce": ecart_ce, "alerte": alerte,
                       "statut": statut, "nb_seances": nb_seances, "seances": seances})
        tot_calc += calc
        if prevu != "":
            tot_prevu += _num_h(prevu)
        if actuel != "":
            tot_actuel += _num_h(actuel)
    lignes.sort(key=lambda x: (x["nom"].lower(), x["matricule"]))
    totaux = {"prevu": _h_aff(tot_prevu), "calc": _h_aff(tot_calc),
              "actuel": _h_aff(tot_actuel),
              "ecart_pf": _fmt_ecart(tot_prevu - tot_calc),
              "ecart_ce": _fmt_ecart(tot_calc - tot_actuel)}
    return {"mois": mois_annee, "libelle": _periode_libelle(mois_annee),
            "lignes": lignes, "anomalies": det["anomalies"], "totaux": totaux,
            "nb": len(lignes),
            "nb_diff": len([l for l in lignes
                            if l["reportable"] and l["statut"] == "different"])}


def synthese_ecart_heures():
    """Synthese globale H prevues / enregistrees pour la tuile du tableau de bord (#8).
    Base = E2_Releve_heures (toutes periodes) : prevu = somme 'Vol. horaire prog.',
    enregistre = somme 'Vol. horaire constate', ecart = prevu - enregistre. nb_ens_ecart
    = enseignants dont le cumul prevu/enregistre differe (feuilles non remplies / retard)."""
    prevu = enr = 0.0
    par_ens = {}
    for r in _lignes_dict("E2_Releve_heures"):
        m = str(r.get("Matricule ens.", "")).strip()
        if not m:
            continue
        p = _num_h(r.get("Vol. horaire prog.", 0))
        c = _num_h(r.get("Vol. horaire constate", 0))
        prevu += p
        enr += c
        par_ens[m] = par_ens.get(m, 0.0) + (p - c)
    ecart = prevu - enr
    nb = len([1 for v in par_ens.values() if abs(v) > 1e-6])
    return {"prevu_aff": _h_aff(prevu), "enregistre_aff": _h_aff(enr),
            "ecart_aff": _fmt_ecart(ecart), "ecart_nonzero": abs(ecart) > 1e-6,
            "nb_ens_ecart": nb}


def reporter_heures_mois(mois_annee, matricules=None, forcer=False, motif=""):
    """Ecrit le Vol. horaire constate calcule dans E2 (cle Mois+Matricule, upsert).
    Garde-fou : une valeur E2 deja saisie et DIFFERENTE du calcul n'est PAS ecrasee
    sans forcer=True (correction manuelle : feuille disparue / heures contestees).
    V1.92 : tout ecrasement force d'une valeur divergente est une REGULARISATION ;
    on renvoie le detail (matricule, nom, ancien -> nouveau) pour journalisation."""
    mois_annee = str(mois_annee).strip()
    motif = str(motif or "").strip()
    det = heures_constatees_detail(mois_annee)
    e2 = _e2_constate_mois(mois_annee)
    bm = _brut("E2_Releve_heures", "Mois / Annee")
    bma = _brut("E2_Releve_heures", "Matricule ens.")
    bc = _brut("E2_Releve_heures", "Vol. horaire constate")
    cibles = set(matricules) if matricules else None
    lignes, divergences, regularisations = [], [], []
    for mat, rec in det["par_ens"].items():
        if cibles is not None and mat not in cibles:
            continue
        calc = round(rec["total"], 2)
        actuel = e2.get(mat, "")
        divergent = actuel != "" and abs(_num_h(actuel) - calc) > 1e-6
        if divergent and not forcer:
            divergences.append(mat)
            continue
        if divergent and forcer:                         # regularisation tracee
            regularisations.append({
                "matricule": mat,
                "nom": ("%s %s" % (rec["nom"], rec["prenom"])).strip(),
                "ancien": actuel, "nouveau": _h_aff(calc), "motif": motif})
        lignes.append({bm: mois_annee, bma: mat, bc: _h_aff(calc)})
    res = {"ajout": 0, "maj": 0}
    if lignes:
        res = _db.ecrire_lignes_lot("E2_Releve_heures", lignes, cles=[bm, bma])
    return {"reportes": res["ajout"] + res["maj"], "ajout": res["ajout"],
            "maj": res["maj"], "divergences": divergences,
            "regularisations": regularisations}


def enregistrer_seance_faite(seance_id, date_fr, creneau, etat, assure_par,
                             matiere_reelle, vol_constate, motif, saisi_par):
    """Upsert E3 (cle Date+Session+Creneau) UNIQUEMENT si exception reelle
    (annulation, remplacant, matiere/duree differente, motif). Sinon, si une
    exception existait, on la neutralise (Etat=Assuree, surcharges effacees).
    Renvoie True si une ligne E3 a ete ecrite/mise a jour."""
    sv = cle_session(_seance_par_id(seance_id)) or str(seance_id).strip()
    date_fr, creneau = str(date_fr).strip(), str(creneau).strip()
    etat = str(etat).strip() or "Assuree"
    assure_par = str(assure_par or "").strip()
    matiere_reelle = str(matiere_reelle or "").strip()
    vol_constate = str(vol_constate or "").strip().replace(",", ".")
    motif = str(motif or "").strip()
    prog_mat, _lbl, _amb = enseignant_programme_session(sv)
    est_exception = (etat == "Cours annule"
                     or (assure_par and assure_par != prog_mat)
                     or bool(matiere_reelle) or bool(vol_constate) or bool(motif))
    existe = (date_fr, sv, creneau) in _exceptions_index()
    if not est_exception and not existe:
        return False
    bmap = _brut_map("E3_Seances_faites")
    ligne = {
        bmap["Date"]: date_fr, bmap["Session / Matiere"]: sv,
        bmap["Creneau"]: creneau, bmap["Etat"]: etat,
        bmap["Assure par"]: assure_par, bmap["Matiere reelle"]: matiere_reelle,
        bmap["Vol. constate h"]: vol_constate, bmap["Motif"]: motif,
        bmap["Saisi par"]: str(saisi_par or "").strip(),
    }
    _db.ecrire_lignes_lot(
        "E3_Seances_faites", [ligne],
        cles=[bmap["Date"], bmap["Session / Matiere"], bmap["Creneau"]])
    return True


# --- Export Excel des 5 tables du tableau de bord ---------------------------
def export_tdb_xlsx():
    """Construit un vrai .xlsx (openpyxl) des 5 jeux du tableau de bord. Renvoie des bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    g = donnees_graphiques()
    k = kpis()
    wb = Workbook()
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill("solid", fgColor="FF1F4E79")
    titre_font = Font(name="Calibri", size=13, bold=True, color="FF1F4E79")

    def feuille(nom, entetes, lignes, titre):
        ws = wb.create_sheet(nom)
        ws.cell(1, 1, titre).font = titre_font
        for j, h in enumerate(entetes, start=1):
            c = ws.cell(2, j, h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        for i, lig in enumerate(lignes, start=3):
            for j, v in enumerate(lig, start=1):
                ws.cell(i, j, v)
        for j, h in enumerate(entetes, start=1):
            longueurs = [len(str(h))] + [len(str(l[j - 1])) for l in lignes if j - 1 < len(l)]
            ws.column_dimensions[chr(64 + j)].width = min(max(longueurs) + 3, 50)
        return ws

    # Synthese (KPI)
    ws0 = wb.active
    ws0.title = "Synthese"
    ws0.cell(1, 1, "Tableau de bord EMSP — synthese").font = titre_font
    ws0.cell(2, 1, "Genere le").font = Font(bold=True)
    ws0.cell(2, 2, fmt_date(_dt.date.today()))
    synth = [("Etudiants enregistres", k["etudiants"]), ("Dont actifs", k["actifs"]),
             ("Dont diplomes", k["diplomes"]), ("Enseignants references", k["enseignants"]),
             ("Recettes (KMF)", k["recettes"]), ("Depenses (KMF)", k["depenses"]),
             ("Solde global (KMF)", k["solde"]), ("Taux de presence (%)", k["taux_presence"]),
             ("Heures constatees (total)", k["heures"])]
    for i, (lib, v) in enumerate(synth, start=4):
        ws0.cell(i, 1, lib).font = Font(bold=True)
        ws0.cell(i, 2, v)
    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 18

    feuille("Effectif par filiere", ["Filiere", "Effectif"],
            list(zip(g["filieres"]["labels"], g["filieres"]["valeurs"])),
            "Effectif par filiere")
    feuille("Repartition par statut", ["Statut", "Nombre"],
            list(zip(g["statuts"]["labels"], g["statuts"]["valeurs"])),
            "Repartition des etudiants par statut")
    fin = g["finances"]
    feuille("Finances par categorie",
            ["Categorie", "Recettes (KMF)", "Depenses (KMF)"],
            [[fin["labels"][i], fin["series"][0]["valeurs"][i], fin["series"][1]["valeurs"][i]]
             for i in range(len(fin["labels"]))],
            "Recettes / Depenses par categorie (KMF)")
    feuille("Presence par creneau", ["Creneau", "Taux de presence (%)"],
            list(zip(g["presence"]["labels"], g["presence"]["valeurs"])),
            "Taux de presence par creneau")
    feuille("Heures par enseignant", ["Matricule", "Heures constatees"],
            list(zip(g["heures"]["labels"], g["heures"]["valeurs"])),
            "Heures constatees par enseignant")

    for ws in wb.worksheets:
        bandeau_xlsx(ws, ws.max_column)
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# SAISIE EN GRILLE — REGISTRE DE TRESORERIE (F1) — V1.20
# ---------------------------------------------------------------------------
# Reproduit le mode de travail du registre papier "Situation de compte" :
# on choisit UN compte/caisse, le solde d'ouverture = solde courant du compte,
# on saisit plusieurs lignes en grille (tabulation cellule a cellule), le solde
# se recalcule en direct (cote client), puis UN enregistrement par lot. Ecrit
# dans l'onglet F1 existant (aucune modification de structure du classeur).
# Solde = solde initial du compte + somme(recettes - depenses), comme F2.
# ===========================================================================

# Colonnes editables de la grille (libelles propres ; le compte est choisi en
# tete de grille, Sens est deduit du montant, Saisi par est auto).
_TRESO_COLS = ["Date operation", "Categorie", "Poste budgetaire",
               "Reference / N piece", "Mode paiement", "Libelle / description",
               "Tiers", "Source de financement / Bailleur",
               "Montant Recette (KMF)", "Montant Depense (KMF)"]


def solde_courant_compte(nom):
    """Solde courant d'un compte/caisse = solde initial (F2) + recettes - depenses (F1)."""
    nom = str(nom).strip()
    if not nom:
        return 0.0
    init = 0.0
    noms_f2 = _lignes_dict("F2_Comptes")
    for r in noms_f2:
        if str(r.get("Nom du compte / caisse", "")).strip() == nom:
            init = _num(r.get("Solde initial (KMF)", 0))
            break
    comptes = _db.colonne("F1_Mouvements", _brut("F1_Mouvements", "Compte / caisse"))
    rec = _db.colonne("F1_Mouvements", _brut("F1_Mouvements", "Montant Recette (KMF)"))
    dep = _db.colonne("F1_Mouvements", _brut("F1_Mouvements", "Montant Depense (KMF)"))
    net = init
    for i, c in enumerate(comptes):
        if str(c).strip() == nom:
            net += _num(rec[i] if i < len(rec) else 0) - _num(dep[i] if i < len(dep) else 0)
    return net


def comptes_treso():
    """Comptes/caisses selectionnables pour la grille + leur solde courant."""
    out = []
    vals = options_liste("Comptes_caisses") or []
    for v in vals:
        nom = v["value"] if isinstance(v, dict) else v
        out.append({"nom": nom, "solde": solde_courant_compte(nom)})
    return out


def treso_grille_colonnes():
    """Specs des colonnes editables de la grille (reprend types/options de F1)."""
    champs = {c["libelle"]: c for c in champs_saisie("F1_Mouvements")}
    cols = []
    for lib in _TRESO_COLS:
        c = champs.get(lib)
        if not c:
            continue
        cols.append({"brut": c["brut"], "libelle": lib, "type": c["type"],
                     "options": c["options"], "obligatoire": c["obligatoire"]})
    return cols


def _ligne_treso_vide(vals):
    """Une ligne de grille est 'vide' si aucune cellule n'est renseignee."""
    return not any(str(v).strip() for v in vals.values())


def enregistrer_treso_lot(compte, lignes, saisi_par):
    """Valide puis ecrit en lot les lignes de la grille dans F1 (atomique).
    `lignes` = liste de dicts {libelle_propre: valeur}. Renvoie (n, message, erreurs).
    Sens deduit du montant (recette OU depense). Si une ligne est invalide, RIEN
    n'est ecrit : on renvoie les erreurs (ligne par ligne) pour correction."""
    compte = str(compte).strip()
    if not compte:
        return 0, "Choisissez d'abord un compte / caisse.", []
    brut = lambda lib: _brut("F1_Mouvements", lib)
    a_ecrire, erreurs = [], []
    for idx, lig in enumerate(lignes, start=1):
        if _ligne_treso_vide(lig):
            continue
        rec = str(lig.get("Montant Recette (KMF)", "")).strip()
        dep = str(lig.get("Montant Depense (KMF)", "")).strip()
        # Sens deduit : recette si montant recette renseigne, sinon depense.
        if rec and not dep:
            sens = "Recette"
        elif dep and not rec:
            sens = "Depense"
        elif rec and dep:
            erreurs.append("Ligne %d : une seule colonne (Recette OU Depense)." % idx)
            continue
        else:
            erreurs.append("Ligne %d : indiquez un montant (Recette ou Depense)." % idx)
            continue
        valeurs = {brut(k): str(v).strip() for k, v in lig.items() if str(v).strip()}
        valeurs[brut("Sens")] = sens
        valeurs[brut("Compte / caisse")] = compte
        valeurs[brut("Saisi par")] = saisi_par
        ok, msg = valide_saisie("F1_Mouvements", valeurs)
        if not ok:
            erreurs.append("Ligne %d : %s" % (idx, msg))
            continue
        a_ecrire.append(valeurs)
    if erreurs:
        return 0, "Aucune ligne enregistree : corrigez les erreurs.", erreurs
    if not a_ecrire:
        return 0, "Aucune ligne a enregistrer.", []
    try:
        n = _db.ajouter_lignes("F1_Mouvements", a_ecrire)
    except OverflowError:
        return 0, "Capacite de l'onglet tresorerie atteinte.", []
    except Exception:
        return 0, "Echec de l'enregistrement.", []
    return n, "%d ligne(s) enregistree(s) dans la tresorerie." % n, []


# ===========================================================================
# V1.69 — DROITS D'INSCRIPTION PAR ETUDIANT (vue derivee de F1_Mouvements)
# ---------------------------------------------------------------------------
# Pas de double stockage : le "Paye" est agrege depuis F1 (recettes 706 du
# niveau courant, filtrees par matricule + annee academique, calcul en Python).
# Les tarifs sont lus dans P0 (colonnes appariees), editables par l'EMSP.
# ===========================================================================

def tarifs_inscription():
    """Dict {niveau: montant_int} depuis P0 (colonnes appariees
    Tarif_inscription_niveau / Tarif_inscription_KMF). Editable par l'EMSP."""
    par = {decoupe_provenance(b)[0]: v for b, v in _db.listes_parametres().items()}
    niveaux = par.get("Tarif_inscription_niveau", [])
    montants = par.get("Tarif_inscription_KMF", [])
    out = {}
    for i, niv in enumerate(niveaux):
        n = str(niv).strip()
        if not n:
            continue
        m = _num(montants[i]) if i < len(montants) else 0
        out[n] = int(round(m))
    return out


def _poste_inscription_code(niveau):
    """Code poste recette (706b/c/d) pour le niveau, ou '' si hors grille."""
    return config.POSTE_INSCRIPTION_PAR_NIVEAU.get(str(niveau).strip(), "")


def _poste_inscription_complet(code):
    """Libelle complet du poste ('706b — ...') depuis Postes_budgetaires (P0),
    ou le code seul si introuvable."""
    if not code:
        return ""
    for v in (options_liste("Postes_budgetaires") or []):
        s = str(v).strip()
        if s.replace(" ", "").startswith(code):
            return s
    return code


def droits_inscription(matricule):
    """Vue derivee des droits d'inscription d'un etudiant pour SON niveau et SON
    annee academique courants (depuis A1). Renvoie un dict ou None.
      tarif_defini False -> "Tarif non defini" (niveau hors grille P0).
      paye  = somme F1 (Recette, matricule, poste du niveau, annee academique).
      reste = tarif - paye. versements = detail date."""
    fiche = fiche_etudiant(matricule)
    if not fiche:
        return None
    mat = fiche["matricule"]
    niveau = str(fiche.get("Niveau", "")).strip()
    annee = str(fiche.get("Annee acad.", "")).strip()
    code = _poste_inscription_code(niveau)
    tarifs = tarifs_inscription()
    tarif_defini = bool(code) and (niveau in tarifs)
    tarif = tarifs.get(niveau, 0)

    paye, versements = 0, []
    if code:
        for r in _lignes_dict("F1_Mouvements"):
            if str(r.get("Sens", "")).strip().lower() != "recette":
                continue
            if str(r.get("Matricule", "")).strip().lower() != mat.lower():
                continue
            poste = str(r.get("Poste budgetaire", "")).strip().replace(" ", "")
            if not poste.startswith(code):
                continue
            an = str(r.get("Annee academique", "")).strip()
            if annee and an and an != annee:
                continue
            montant = int(round(_num(r.get("Montant Recette (KMF)", 0))))
            paye += montant
            versements.append({
                "date": str(r.get("Date operation", "")).strip(),
                "montant": montant, "montant_fmt": _fmt_kmf(montant),
                "annee": an, "reference": str(r.get("Reference / N piece", "")).strip()})
    reste = tarif - paye
    return {
        "matricule": mat, "niveau": niveau, "annee": annee,
        "poste_code": code, "poste_complet": _poste_inscription_complet(code),
        "tarif_defini": tarif_defini,
        "tarif": tarif, "tarif_fmt": _fmt_kmf(tarif),
        "paye": paye, "paye_fmt": _fmt_kmf(paye),
        "reste": reste, "reste_fmt": _fmt_kmf(reste),
        "solde": (reste <= 0),
        "versements": versements,
    }


def reste_du_par_filiere(filtres=None):
    """Reste du d'inscription agrege par filiere (V1.93). Par etudiant :
    reste = max(tarif(niveau) - paye(F1 inscription, son annee), 0) ; le trop-percu
    est plafonne a 0 (un etudiant a jour ou en avance compte 0, jamais negatif).
    Reagit aux filtres filiere/niveau/annee (jeu d'etudiants), contrairement aux
    finances brutes (F1 ne porte pas la filiere). Cle 'reste_du_filiere' du TDB."""
    f = filtres or {}
    tarifs = tarifs_inscription()
    et_rows = _lignes_filtrees("A1_Etudiants", f)
    idx = {}                                   # mat_lower -> (filiere, niveau, annee)
    for r in et_rows:
        mat = str(r.get("Matricule", "")).strip()
        if not mat:
            continue
        idx[mat.lower()] = (str(r.get("Filiere", "")).strip(),
                            str(r.get("Niveau", "")).strip(),
                            str(r.get("Annee acad.", "")).strip())
    code_de, paye = {}, {}                     # niveau->code ; mat_lower->paye inscription
    for r in _lignes_dict("F1_Mouvements"):
        if str(r.get("Sens", "")).strip().lower() != "recette":
            continue
        mat = str(r.get("Matricule", "")).strip().lower()
        if mat not in idx:
            continue
        _fil, niveau, annee = idx[mat]
        if niveau not in code_de:
            code_de[niveau] = _poste_inscription_code(niveau)
        code = code_de[niveau]
        if not code:
            continue
        poste = str(r.get("Poste budgetaire", "")).strip().replace(" ", "")
        if not poste.startswith(code):
            continue
        an = str(r.get("Annee academique", "")).strip()
        if annee and an and an != annee:
            continue
        paye[mat] = paye.get(mat, 0) + int(round(_num(r.get("Montant Recette (KMF)", 0))))
    reste_par, attendu_par, paye_par = {}, {}, {}
    for matl, (filiere, niveau, _an) in idx.items():
        if not filiere:
            continue
        tarif = int(tarifs.get(niveau, 0) or 0)
        p = paye.get(matl, 0)
        reste = tarif - p
        if reste < 0:
            reste = 0                          # trop-percu plafonne a 0
        attendu_par[filiere] = attendu_par.get(filiere, 0) + tarif
        paye_par[filiere] = paye_par.get(filiere, 0) + p
        reste_par[filiere] = reste_par.get(filiere, 0) + reste
    par = _db.listes_parametres()
    labels = [str(x).strip() for x in par.get("Filieres", []) if str(x).strip()]
    for k in reste_par:
        if k not in labels:
            labels.append(k)
    labels = [l for l in labels if l in reste_par]   # filieres effectivement presentes
    valeurs = [reste_par.get(l, 0) for l in labels]
    return {"labels": labels, "valeurs": valeurs,
            "attendu": [attendu_par.get(l, 0) for l in labels],
            "paye": [paye_par.get(l, 0) for l in labels],
            "total_reste": sum(valeurs)}


def enregistrer_encaissement(matricule, montant, annee, compte, mode_paiement,
                             reference, saisi_par):
    """Ecrit une recette d'inscription dans F1 (Sens=Recette, poste 706 du niveau,
    matricule, annee academique). Renvoie (ok, message). Compte/caisse et Mode de
    paiement obligatoires (rattachement caisse + solde). Categorie laissee vide :
    la comptabilite la completera dans /tresorerie."""
    fiche = fiche_etudiant(matricule)
    if not fiche:
        return False, "Matricule inconnu."
    niveau = str(fiche.get("Niveau", "")).strip()
    code = _poste_inscription_code(niveau)
    if not code or niveau not in tarifs_inscription():
        return False, ("Tarif non defini pour le niveau « %s » : renseignez-le dans "
                       "Parametres (P0) avant d'encaisser." % (niveau or "?"))
    m = int(round(_num(montant)))
    if m <= 0:
        return False, "Indiquez un montant valide."
    if not str(compte).strip():
        return False, "Choisissez un compte / caisse."
    if not str(mode_paiement).strip():
        return False, "Choisissez un mode de paiement."

    from datetime import date as _date
    auj = _date.today().strftime("%d/%m/%Y")
    b = lambda lib: _brut("F1_Mouvements", lib)
    valeurs = {
        b("Date operation"): auj,
        b("Sens"): "Recette",
        b("Poste budgetaire"): _poste_inscription_complet(code),
        b("Mode paiement"): str(mode_paiement).strip(),
        b("Compte / caisse"): str(compte).strip(),
        b("Reference / N piece"): str(reference).strip(),
        b("Libelle / description"): "Droit d'inscription %s — %s" % (niveau, fiche["nom_complet"]),
        b("Montant Recette (KMF)"): m,
        b("Tiers"): fiche["nom_complet"],
        b("Saisi par"): str(saisi_par).strip(),
        b("Matricule"): fiche["matricule"],
        b("Annee academique"): str(annee).strip() or str(fiche.get("Annee acad.", "")).strip(),
    }
    valeurs = {k: v for k, v in valeurs.items() if str(v).strip() != ""}
    try:
        n = _db.ajouter_lignes("F1_Mouvements", [valeurs])
    except OverflowError:
        return False, "Capacite de la tresorerie atteinte."
    except Exception:
        return False, "Echec de l'enregistrement."
    return (n == 1), ("Encaissement enregistre (%s KMF)." % _fmt_kmf(m)
                      if n == 1 else "Echec de l'enregistrement.")


# ===========================================================================
# EDITION "SITUATION DE COMPTE" — registre mensuel imprimable — V1.21
# ---------------------------------------------------------------------------
# Reproduit le document papier signe (Gestionnaire + Directeur) : un compte,
# une periode (mois), une ligne d'ouverture "report a nouveau" (solde au debut
# de periode), les lignes de la periode avec solde courant, puis "SOLDE AU ...".
# Convention du document : Solde = Debit - Credit. Mapping retenu :
#   Debit  = Montant Recette (entree, augmente le solde)
#   Credit = Montant Depense (sortie, diminue le solde)
# => Solde = recettes - depenses, coherent avec solde_courant_compte / F2.
# Lecture seule (lit F1/F2). Aucune modification du classeur.
# ===========================================================================
def _parse_date_fr(s):
    """Convertit une valeur en datetime.date. Accepte un objet date/datetime
    (cellule Excel typee) ou une chaine 'JJ/MM/AAAA' (format charte) ; tolere
    aussi 'JJ-MM-AAAA' et 'AAAA-MM-JJ'. Renvoie None si non interpretable."""
    if s is None or s == "":
        return None
    if isinstance(s, _dt.datetime):
        return s.date()
    if isinstance(s, _dt.date):
        return s
    txt = str(s).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return _dt.datetime.strptime(txt, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def mois_treso_dispo(compte=""):
    """Periodes MM/AAAA distinctes presentes dans F1 (option : pour un compte)."""
    compte = str(compte).strip()
    vus = set()
    for r in _lignes_dict("F1_Mouvements"):
        if compte and str(r.get("Compte / caisse", "")).strip() != compte:
            continue
        d = _parse_date_fr(r.get("Date operation", ""))
        if d:
            vus.add((d.year, d.month))
    return ["%02d/%04d" % (m, y) for (y, m) in sorted(vus, reverse=True)]


def _dernier_jour(annee, mois):
    if mois == 12:
        return _dt.date(annee, 12, 31)
    return _dt.date(annee, mois + 1, 1) - _dt.timedelta(days=1)


def situation_compte(compte, mois):
    """Registre 'Situation de compte' d'un compte pour une periode MM/AAAA.
    Renvoie en-tete, ligne d'ouverture (report a nouveau), lignes de la periode
    avec solde courant, et solde de cloture. Tout en lecture seule."""
    compte = str(compte).strip()
    mois = str(mois).strip()
    try:
        m, y = mois.split("/")
        m, y = int(m), int(y)
        debut = _dt.date(y, m, 1)
        fin = _dernier_jour(y, m)
    except Exception:
        return None
    # Solde initial du compte (F2)
    init = 0.0
    for r in _lignes_dict("F2_Comptes"):
        if str(r.get("Nom du compte / caisse", "")).strip() == compte:
            init = _num(r.get("Solde initial (KMF)", 0))
            break
    # Lignes du compte
    lignes_compte = [r for r in _lignes_dict("F1_Mouvements")
                     if str(r.get("Compte / caisse", "")).strip() == compte]
    # Report a nouveau = solde initial + mouvements AVANT le debut de periode
    ouverture = init
    periode = []
    for r in lignes_compte:
        d = _parse_date_fr(r.get("Date operation", ""))
        if d is None:
            continue
        net = _num(r.get("Montant Recette (KMF)", 0)) - _num(r.get("Montant Depense (KMF)", 0))
        if d < debut:
            ouverture += net
        elif d <= fin:
            periode.append((d, r))
    periode.sort(key=lambda t: (t[0], str(t[1].get("Reference / N piece", ""))))
    colonnes = ["N", "Chapitre", "Compte", "Date", "N piece", "Description", "Beneficiaire",
                "Debit (recette)", "Credit (depense)", "Solde"]
    solde = ouverture
    lignes = []
    for i, (d, r) in enumerate(periode, start=1):
        rec = _num(r.get("Montant Recette (KMF)", 0))
        dep = _num(r.get("Montant Depense (KMF)", 0))
        solde += rec - dep
        lignes.append([
            str(i),
            r.get("Poste budgetaire", ""),
            r.get("Compte", ""),
            r.get("Date operation", ""),
            r.get("Reference / N piece", ""),
            r.get("Libelle / description", ""),
            r.get("Tiers", ""),
            _kmf_aff(rec) if rec else "",
            _kmf_aff(dep) if dep else "",
            _kmf_aff(solde),
        ])
    return {
        "colonnes": colonnes,
        "ouverture": {"libelle": "Report a nouveau", "solde": _kmf_aff(ouverture)},
        "lignes": lignes,
        "cloture": {"libelle": "SOLDE AU " + fin.strftime("%d/%m/%Y"),
                    "solde": _kmf_aff(solde)},
        "contexte": {"compte": compte, "periode": mois, "nb": len(lignes),
                     "date_jour": fmt_date(_dt.date.today())},
    }


# ===========================================================================
# COUCHE REQUETES MULTICRITERES (LECTURE SEULE) — V1.22
# ---------------------------------------------------------------------------
# Donne a la Direction la puissance d'Excel sans ouvrir une copie du fichier :
#  - explorateur generique par table (filtres multicriteres, tri, colonnes),
#  - vues metier predefinies (questions recurrentes, croisements inter-tables),
#  - export Excel de CHAQUE selection (la soupape qui evite de "taper" ailleurs).
# Strictement en lecture : ne modifie jamais le classeur.
# ===========================================================================

# Tables interrogeables (onglets de donnees lisibles ; pas les onglets de config).
TABLES_INTERROGEABLES = ["A1_Etudiants", "A2_Presences", "A3_Sessions", "S1_Stages",
                         "E1_Enseignants", "E2_Releve_heures", "L1_Salles",
                         "L2_Reservations", "M1_Equipements", "F1_Mouvements", "F2_Comptes"]


def tables_interrogeables():
    out = []
    for t in TABLES_INTERROGEABLES:
        if t in _db.onglets():
            lib = config.TAB_INDEX.get(t, {}).get("libelle", t)
            out.append({"onglet": t, "libelle": lib})
    return out


def colonnes_table(onglet):
    return [m["libelle"] for m in entetes_meta(onglet)]


_OPERATEURS = ["contient", "egal", "debut", "sup", "inf", "nonvide", "vide"]


def _match_filtre(cellule, op, val):
    s = str(cellule or "").strip()
    v = str(val or "").strip()
    if op == "nonvide":
        return s != ""
    if op == "vide":
        return s == ""
    if op == "sup":
        return _num(s) > _num(v)
    if op == "inf":
        return _num(s) < _num(v)
    sl, vl = s.lower(), v.lower()
    if op == "egal":
        return sl == vl
    if op == "debut":
        return sl.startswith(vl)
    return vl in sl  # contient (defaut)


def explorer(onglet, filtres=None, colonnes_sel=None, tri_col="", tri_sens="asc"):
    """Filtre/trie/projette une table. filtres = [{col, op, val}]. colonnes_sel =
    sous-ensemble a afficher (None/[] = toutes). Renvoie colonnes, lignes, nb."""
    if onglet not in TABLES_INTERROGEABLES:
        return None
    cols_all = colonnes_table(onglet)
    rows = _lignes_dict(onglet)
    for f in (filtres or []):
        col, op = f.get("col", ""), f.get("op", "contient")
        if col not in cols_all or op not in _OPERATEURS:
            continue
        val = f.get("val", "")
        if op not in ("vide", "nonvide") and str(val).strip() == "":
            continue
        rows = [r for r in rows if _match_filtre(r.get(col, ""), op, val)]
    if tri_col in cols_all:
        vals = [str(r.get(tri_col, "")).strip() for r in rows]
        non_vides = [v for v in vals if v != ""]
        numerique = bool(non_vides) and all(
            re.match(r"^-?\d+([.,]\d+)?$", v.replace(" ", "")) for v in non_vides)
        cle = (lambda r: _num(r.get(tri_col, ""))) if numerique \
            else (lambda r: str(r.get(tri_col, "")).strip().lower())
        rows = sorted(rows, key=cle, reverse=(tri_sens == "desc"))
    cols_show = [c for c in cols_all if (not colonnes_sel or c in colonnes_sel)]
    lignes = [[r.get(c, "") for c in cols_show] for r in rows]
    return {"onglet": onglet, "libelle": config.TAB_INDEX.get(onglet, {}).get("libelle", onglet),
            "colonnes": cols_show, "cols_all": cols_all, "lignes": lignes, "nb": len(lignes)}


# --- Marquage FORMATION des exports Excel (V1.50) --------------------------
# En MODE_FORMATION, les exports portent une ligne d'en-tete rouge "FORMATION"
# en tete de chaque feuille et un suffixe "_FORMATION" dans le nom de fichier,
# en coherence avec le bandeau web et le filigrane PDF. En production : sans effet.
def bandeau_xlsx(ws, ncols=1):
    """Insere une ligne 'FORMATION' rouge en tete de la feuille si MODE_FORMATION.
    insert_rows(1) decale le contenu existant d'une ligne (aucune fusion en amont)."""
    if not config.MODE_FORMATION:
        return
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.insert_rows(1)
    c = ws.cell(1, 1, "FORMATION — données d'entraînement, sans valeur officielle")
    c.font = Font(name="Calibri", bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor="FFC0241F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    n = ncols or ws.max_column or 1
    if n > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)


def nom_export(nom):
    """Insere le suffixe _FORMATION avant l'extension si MODE_FORMATION.
    ex. 'export_A1_20260618.xlsx' -> 'export_A1_20260618_FORMATION.xlsx'."""
    if not config.MODE_FORMATION:
        return nom
    base, dot, ext = nom.rpartition(".")
    return "%s_FORMATION.%s" % (base, ext) if dot else nom + "_FORMATION"


def _xlsx_simple(titre, colonnes, lignes):
    """Un .xlsx d'une feuille (titre + en-tetes + lignes). Renvoie des bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Selection"
    ws.cell(1, 1, titre).font = Font(name="Calibri", size=13, bold=True, color="FF1F4E79")
    for j, h in enumerate(colonnes, start=1):
        c = ws.cell(2, j, h)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor="FF1F4E79")
        c.alignment = Alignment(horizontal="center")
    for i, lig in enumerate(lignes, start=3):
        for j, v in enumerate(lig, start=1):
            ws.cell(i, j, v)
    for j, h in enumerate(colonnes, start=1):
        longueurs = [len(str(h))] + [len(str(l[j - 1])) for l in lignes if j - 1 < len(l)]
        ws.column_dimensions[chr(64 + j)].width = min(max(longueurs) + 3, 55)
    bandeau_xlsx(ws, len(colonnes))
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- VUES METIER PREDEFINIES (croisements / questions recurrentes) ----------
def _mois_key(s):
    try:
        m, y = str(s).strip().split("/")
        return (int(y), int(m))
    except Exception:
        return None


def vue_equip_loc(args):
    """Q1 — Equipements et leur localisation (registre M1), filtrable."""
    fs = str(args.get("salle", "")).strip()
    fb = str(args.get("bailleur", "")).strip()
    fc = str(args.get("categorie", "")).strip()
    cols = ["Designation", "Categorie", "Salle / localisation",
            "Source de financement / Bailleur", "Etat", "N inventaire / serie", "Montant (KMF)"]
    lignes = []
    for r in _lignes_dict("M1_Equipements"):
        if not str(r.get("Designation", "")).strip():
            continue
        if fs and str(r.get("Salle / localisation", "")).strip() != fs:
            continue
        if fb and str(r.get("Source de financement / Bailleur", "")).strip() != fb:
            continue
        if fc and str(r.get("Categorie", "")).strip() != fc:
            continue
        lignes.append([r.get(c, "") for c in cols])
    lignes.sort(key=lambda x: (str(x[2]).lower(), str(x[0]).lower()))
    return {"titre": "Equipements et localisation", "colonnes": cols, "lignes": lignes,
            "nb": len(lignes), "sous_titre": "Inventaire M1 (filtres facultatifs)"}


def vue_equip_bailleur(args):
    """Q4 — Equipements finances par un bailleur (M1) + total montant."""
    fb = str(args.get("bailleur", "")).strip()
    cols = ["Source de financement / Bailleur", "Designation", "Categorie",
            "Salle / localisation", "Date d'acquisition", "Reference / N piece", "Montant (KMF)"]
    lignes, total = [], 0.0
    for r in _lignes_dict("M1_Equipements"):
        if not str(r.get("Designation", "")).strip():
            continue
        if fb and str(r.get("Source de financement / Bailleur", "")).strip() != fb:
            continue
        lignes.append([r.get(c, "") for c in cols])
        total += _num(r.get("Montant (KMF)", 0))
    lignes.sort(key=lambda x: (str(x[0]).lower(), str(x[1]).lower()))
    if lignes:
        lignes.append(["", "", "", "", "", "TOTAL", _kmf_aff(total)])
    st = ("Bailleur : " + fb) if fb else "Tous bailleurs"
    return {"titre": "Equipements par bailleur", "colonnes": cols, "lignes": lignes,
            "nb": max(len(lignes) - (1 if lignes else 0), 0), "sous_titre": st}


_L2_TAB = "L2_Reservations"


def _hhmm_min(s):
    """'H:MM' ou 'HH:MM' -> minutes depuis minuit, ou None si invalide."""
    s = str(s or "").strip()
    if ":" not in s:
        return None
    try:
        h, m = s.split(":", 1)
        return int(h) * 60 + int(m)
    except (ValueError, TypeError):
        return None


def planning_salles_par_date(date_fr, demo=False):
    """Occupation des salles pour une DATE (JJ/MM/AAAA), facon agenda.
    Fusionne les seances de cours (A3, projetees sur le jour de la semaine) et
    les reservations (L2) de cette date exacte. Ne renvoie QUE les salles
    effectivement occupees. Chaque salle : {salle, blocs:[...]} ou chaque bloc
    porte deb_min/fin_min/type/matiere/debut/fin/groupe (compatibles agenda)."""
    jour = _jour_de_date(date_fr)
    liste = salles(demo)
    sc = [x for x in seances(demo)
          if str(x.get("jour", "")).strip().lower() == (jour or "").lower()]
    out = []
    for s in liste:
        blocs = []
        for x in sc:                       # seances de cours (A3)
            if _seance_dans_salle(x, s):
                blocs.append(x)
        if not demo:                       # reservations (L2) de cette date
            for r in reservations_salle(s.get("nom", "")):
                if r["date"] != date_fr:
                    continue
                blocs.append({
                    "deb_min": _hhmm_min(r["h_debut"]),
                    "fin_min": _hhmm_min(r["h_fin"]),
                    "type": (r["type"] or "reservation"),
                    "matiere": (r["motif"] or r["matiere"] or r["type"] or "Réservation"),
                    "debut": r["h_debut"], "fin": r["h_fin"],
                    "groupe": " · ".join(t for t in [r["reserve_par"], r["statut"]] if t),
                })
        if blocs:
            blocs.sort(key=lambda b: (b.get("deb_min") if b.get("deb_min") is not None else 9999))
            out.append({"salle": s, "blocs": blocs})
    return out


def reservations_liste(date_fr=""):
    """Reservations L2 en dicts propres (toutes, ou d'une date), triees."""
    date_fr = str(date_fr or "").strip()
    out = []
    for r in _lignes_dict(_L2_TAB):
        d = str(r.get("Date", "")).strip()
        if date_fr and d != date_fr:
            continue
        out.append({
            "id": str(r.get("ID reservation", "")).strip(),
            "salle": str(r.get("Salle", "")).strip(),
            "date": d,
            "h_debut": str(r.get("Heure debut", "")).strip(),
            "h_fin": str(r.get("Heure fin", "")).strip(),
            "type": str(r.get("Type", "")).strip(),
            "reserve_par": str(r.get("Reserve par", "")).strip(),
            "motif": str(r.get("Motif / objet", "")).strip(),
            "statut": str(r.get("Statut", "")).strip(),
        })
    out.sort(key=lambda x: (x["date"], x["h_debut"], x["salle"]))
    return out


def reservations_salle(nom_salle):
    """Reservations (L2) d'une salle, en dicts propres, triees par date/heure."""
    cible = str(nom_salle or "").strip().lower()
    out = []
    for r in _lignes_dict(_L2_TAB):
        if str(r.get("Salle", "")).strip().lower() != cible:
            continue
        out.append({
            "id": str(r.get("ID reservation", "")).strip(),
            "date": str(r.get("Date", "")).strip(),
            "h_debut": str(r.get("Heure debut", "")).strip(),
            "h_fin": str(r.get("Heure fin", "")).strip(),
            "type": str(r.get("Type", "")).strip(),
            "reserve_par": str(r.get("Reserve par", "")).strip(),
            "motif": str(r.get("Motif / objet", "")).strip(),
            "statut": str(r.get("Statut", "")).strip(),
            "seance": str(r.get("Seance liee (ID session A3)", "")).strip(),
            "filiere": str(r.get("Filiere", "")).strip(),
            "niveau": str(r.get("Niveau", "")).strip(),
            "matiere": str(r.get("Matiere", "")).strip(),
            "matricule": str(r.get("Matricule ens.", "")).strip(),
            "enseignant": str(r.get("Enseignant", "")).strip(),
        })
    out.sort(key=lambda x: (x["date"], x["h_debut"]))
    return out


def seances_salle(nom_salle):
    """Seances de cours (A3) tenues dans la salle, pour pre-remplir une
    reservation : value = ID session, libelle = Matiere - Jour Heure (Enseignant)."""
    cible = str(nom_salle or "").strip().lower()
    out = []
    for r in _lignes_dict("A3_Sessions"):
        if str(r.get("Salle", "")).strip().lower() != cible:
            continue
        sid = str(r.get("ID session", "")).strip()
        if not sid:
            continue
        h = ("%s-%s" % (str(r.get("Heure debut", "")).strip(),
                        str(r.get("Heure fin", "")).strip())).strip("-")
        lib = "%s — %s %s" % (str(r.get("Matiere", "")).strip(),
                              str(r.get("Jour", "")).strip(), h)
        ens = str(r.get("Enseignant", "")).strip()
        if ens:
            lib += " (%s)" % ens
        out.append({"id": sid, "libelle": lib.strip(" —"),
                    "filiere": str(r.get("Filiere", "")).strip(),
                    "niveau": str(r.get("Niveau", "")).strip(),
                    "matiere": str(r.get("Matiere", "")).strip(),
                    "enseignant": ens})
    return out


def enseignants_choix():
    """[{matricule, nom}] depuis E1, pour la datalist du formulaire reservation."""
    out = []
    for mat, e in _ens_index().items():
        nom = ("%s %s" % (e.get("nom", ""), e.get("prenom", ""))).strip()
        out.append({"matricule": mat, "nom": nom})
    out.sort(key=lambda x: x["nom"].lower())
    return out


def _seance_a3(seance_id):
    sid = str(seance_id or "").strip().lower()
    if not sid:
        return None
    for r in _lignes_dict("A3_Sessions"):
        if str(r.get("ID session", "")).strip().lower() == sid:
            return r
    return None


def creer_reservation(acteur, d, login):
    """Cree une reservation de salle (L2). d = champs du formulaire (cles propres).
    - Si 'seance' (ID session A3) est fourni -> pre-remplit filiere/niveau/matiere/
      enseignant depuis A3 (sans ecraser une saisie explicite).
    - Si 'matricule' fourni et nom vide -> resout le nom depuis E1.
    Le matricule reste la cle (chainage prof -> heures -> compta). Renvoie (ok, msg)."""
    if not peut_ecrire(acteur, _L2_TAB):
        return False, "Droits insuffisants pour reserver une salle."
    salle = str(d.get("salle", "")).strip()
    date = str(d.get("date", "")).strip()
    if not salle or not date:
        return False, "Salle et date sont obligatoires."

    seance_id = str(d.get("seance", "")).strip()
    filiere = str(d.get("filiere", "")).strip()
    niveau = str(d.get("niveau", "")).strip()
    matiere = str(d.get("matiere", "")).strip()
    matricule = str(d.get("matricule", "")).strip()
    enseignant = str(d.get("enseignant", "")).strip()

    se = _seance_a3(seance_id) if seance_id else None
    if se is not None:
        filiere = filiere or str(se.get("Filiere", "")).strip()
        niveau = niveau or str(se.get("Niveau", "")).strip()
        matiere = matiere or str(se.get("Matiere", "")).strip()
        enseignant = enseignant or str(se.get("Enseignant", "")).strip()

    if matricule and not enseignant:
        e = _ens_index().get(matricule, {})
        enseignant = ("%s %s" % (e.get("nom", ""), e.get("prenom", ""))).strip()

    rid = _prochain_res()
    def B(c): return _brut(_L2_TAB, c)
    valeurs = {
        B("ID reservation"): rid,
        B("Salle"): salle,
        B("Date"): date,
        B("Heure debut"): str(d.get("heure_debut", "")).strip(),
        B("Heure fin"): str(d.get("heure_fin", "")).strip(),
        B("Type"): str(d.get("type", "")).strip(),
        B("Reserve par"): str(d.get("reserve_par", "")).strip(),
        B("Motif / objet"): str(d.get("motif", "")).strip(),
        B("Statut"): (str(d.get("statut", "")).strip() or "Confirmee"),
        B("Saisi par"): login,
        B("Seance liee (ID session A3)"): seance_id,
        B("Filiere"): filiere,
        B("Niveau"): niveau,
        B("Matiere"): matiere,
        B("Matricule ens."): matricule,
        B("Enseignant"): enseignant,
    }
    try:
        _db.ajouter_ligne(_L2_TAB, valeurs)
    except Exception as ex:
        return False, "Echec de l'enregistrement (%s)." % type(ex).__name__
    auth.journal(login, "Reservation salle creee", _L2_TAB, "%s / %s" % (rid, salle))
    return True, "Reservation %s enregistree." % rid


def vue_salle_occupation(args):
    """Q2 — Occupation d'une salle = cours (A3) + reservations (L2), avec qui/quand."""
    fs = str(args.get("salle", "")).strip()
    cols = ["Salle", "Origine", "Quand", "Horaire", "Detail"]
    lignes = []
    for r in _lignes_dict("A3_Sessions"):
        salle = str(r.get("Salle", "")).strip()
        if not salle or (fs and salle != fs):
            continue
        horaire = ("%s-%s" % (str(r.get("Heure debut", "")).strip(),
                              str(r.get("Heure fin", "")).strip())).strip("-")
        detail = ("%s — %s" % (r.get("Matiere", ""), r.get("Enseignant", ""))).strip(" —")
        lignes.append([salle, "Cours", r.get("Jour", ""), horaire, detail])
    for r in _lignes_dict("L2_Reservations"):
        salle = str(r.get("Salle", "")).strip()
        if not salle or (fs and salle != fs):
            continue
        horaire = ("%s-%s" % (str(r.get("Heure debut", "")).strip(),
                              str(r.get("Heure fin", "")).strip())).strip("-")
        det = "%s : %s" % (r.get("Type", ""), r.get("Motif / objet", ""))
        rp = str(r.get("Reserve par", "")).strip()
        if rp:
            det += " (%s)" % rp
        stt = str(r.get("Statut", "")).strip()
        if stt:
            det += " [%s]" % stt
        lignes.append([salle, "Reservation", r.get("Date", ""), horaire, det.strip(" :")])
    lignes.sort(key=lambda x: (str(x[0]).lower(), str(x[1]), str(x[2])))
    st = ("Salle : " + fs) if fs else "Toutes les salles (cours + reservations)"
    return {"titre": "Occupation des salles", "colonnes": cols, "lignes": lignes,
            "nb": len(lignes), "sous_titre": st}


def vue_ens_ecart(args):
    """Q5 — Ecart programme/constate par enseignant sur une periode de mois."""
    md, mf = _mois_key(args.get("mois_debut", "")), _mois_key(args.get("mois_fin", ""))
    ens = _ens_index()
    agg = {}
    for r in _lignes_dict("E2_Releve_heures"):
        k = _mois_key(r.get("Mois / Annee", ""))
        if md and (k is None or k < md):
            continue
        if mf and (k is None or k > mf):
            continue
        mat = str(r.get("Matricule ens.", "")).strip()
        if not mat:
            continue
        a = agg.setdefault(mat, [0.0, 0.0])
        a[0] += _num(r.get("Vol. horaire prog.", 0))
        a[1] += _num(r.get("Vol. horaire constate", 0))
    cols = ["Matricule", "Nom Prenom", "Vol. programme", "Vol. constate", "Ecart (h)"]
    lignes = []
    for mat, (prog, const) in agg.items():
        e = ens.get(mat, {})
        nom = ("%s %s" % (e.get("nom", ""), e.get("prenom", ""))).strip()
        lignes.append([mat, nom, _fmt_kmf(prog), _fmt_kmf(const), _fmt_kmf(prog - const)])
    lignes.sort(key=lambda x: _num(x[4]), reverse=True)
    bornes = []
    if args.get("mois_debut"):
        bornes.append("de " + str(args.get("mois_debut")))
    if args.get("mois_fin"):
        bornes.append("a " + str(args.get("mois_fin")))
    st = "Ecart = programme - constate" + ((" (" + " ".join(bornes) + ")") if bornes else " (toutes periodes)")
    return {"titre": "Heures enseignants — ecart programme/constate", "colonnes": cols,
            "lignes": lignes, "nb": len(lignes), "sous_titre": st}


VUES = {
    "equip_loc": {"libelle": "Equipements et leur localisation", "icone": "ti-map-pin",
                  "description": "Inventaire des equipements avec leur salle (Q1).",
                  "builder": vue_equip_loc},
    "salle_occupation": {"libelle": "Salles : qui occupe et quand", "icone": "ti-door",
                         "description": "Cours (planning) + reservations hors cours, par salle (Q2).",
                         "builder": vue_salle_occupation},
    "equip_bailleur": {"libelle": "Equipements par bailleur", "icone": "ti-businessplan",
                       "description": "Equipements finances par l'AFD ou un autre bailleur (Q4).",
                       "builder": vue_equip_bailleur},
    "ens_ecart": {"libelle": "Enseignants : ecart programme/constate", "icone": "ti-clock-exclamation",
                  "description": "Heures non assurees par enseignant, sur une periode (Q5).",
                  "builder": vue_ens_ecart},
}
VUES_ORDRE = ["equip_loc", "salle_occupation", "equip_bailleur", "ens_ecart"]


# --- Q3 : absences eleves (cours A2) + observations de stage (S1) -----------
# Nuances assumees : A2 n'a pas de champ commentaire ; l'absence en stage n'est
# pas modelisee comme telle => on remonte les OBSERVATIONS / plaintes de S1.
def _noms_etudiants():
    idx = {}
    for r in _lignes_dict("A1_Etudiants"):
        m = str(r.get("Matricule", "")).strip()
        if m:
            idx[m] = ("%s %s" % (r.get("Nom", ""), r.get("Prenom", ""))).strip()
    return idx


def vue_absences(args):
    """Q3 — Absences en cours (A2, Present != O) + observations/plaintes de stage (S1)."""
    fmat = str(args.get("matricule", "")).strip()
    forig = str(args.get("origine", "")).strip().lower()
    noms = _noms_etudiants()
    cols = ["Matricule", "Nom Prenom", "Origine", "Quand", "Detail"]
    lignes = []
    if forig != "stage":
        for r in _lignes_dict("A2_Presences"):
            p = str(r.get("Present (O/N)", "")).strip().upper()
            if not p or p == "O":          # absence = explicitement non present
                continue
            mat = str(r.get("Matricule", "")).strip()
            if fmat and mat != fmat:
                continue
            mat_lib = str(r.get("Session / Matiere", "")).strip()
            cre = str(r.get("Creneau", "")).strip()
            detail = mat_lib + ((" (%s)" % cre) if cre else "")
            lignes.append([mat, noms.get(mat, ""), "Absence cours", r.get("Date", ""), detail])
    if forig != "cours":
        for r in _lignes_dict("S1_Stages"):
            obs = str(r.get("Observation / plainte", "")).strip()
            if not obs:
                continue
            mat = str(r.get("Matricule", "")).strip()
            if fmat and mat != fmat:
                continue
            quand = ("%s-%s" % (str(r.get("Date debut", "")).strip(),
                                str(r.get("Date fin", "")).strip())).strip("-")
            lieu = str(r.get("Lieu de stage", "")).strip()
            detail = (lieu + " : " + obs) if lieu else obs
            lignes.append([mat, noms.get(mat, ""), "Observation stage", quand, detail])
    lignes.sort(key=lambda x: (str(x[0]), str(x[2]), str(x[3])))
    st = ("Etudiant " + fmat) if fmat else "Absences en cours (A2) + observations de stage (S1)"
    return {"titre": "Absences & observations", "colonnes": cols, "lignes": lignes,
            "nb": len(lignes), "sous_titre": st}


# ===========================================================================
# TABLEAU CROISE LEGER (PIVOT) — V1.23 — lecture seule
# ---------------------------------------------------------------------------
# Une dimension en ligne, une (optionnelle) en colonne, une mesure (Nombre,
# Somme, Moyenne). Totaux corrects (moyenne = somme/effectif, pas moyenne de
# moyennes). Renvoie colonnes/lignes prets pour le rendu et l'export.
# ===========================================================================
def pivot(onglet, lig, col="", mesure="count", mes_col=""):
    if onglet not in TABLES_INTERROGEABLES:
        return None
    cols_all = colonnes_table(onglet)
    if lig not in cols_all:
        return None
    rows = _lignes_dict(onglet)
    mlabel = {"count": "Nombre", "somme": "Somme (%s)" % mes_col,
              "moyenne": "Moyenne (%s)" % mes_col}.get(mesure, "Nombre")

    def part(r):  # contribution (somme, effectif) d'une ligne
        if mesure == "count":
            return (1.0, 1)
        return (_num(r.get(mes_col, 0)), 1)

    def fmt(somme, eff):
        if eff == 0:
            return ""
        v = (somme / eff) if mesure == "moyenne" else somme
        return str(int(round(v))) if mesure == "count" else _fmt_kmf(v)

    if col and col in cols_all:
        lig_vals, col_vals = [], []
        S, N = {}, {}
        for r in rows:
            lv = str(r.get(lig, "")).strip() or "(vide)"
            cv = str(r.get(col, "")).strip() or "(vide)"
            if lv not in lig_vals:
                lig_vals.append(lv)
            if cv not in col_vals:
                col_vals.append(cv)
            s, n = part(r)
            S[(lv, cv)] = S.get((lv, cv), 0.0) + s
            N[(lv, cv)] = N.get((lv, cv), 0) + n
        lig_vals.sort(); col_vals.sort()
        colonnes = [lig] + col_vals + ["Total"]
        lignes = []
        col_S = {cv: 0.0 for cv in col_vals}; col_N = {cv: 0 for cv in col_vals}
        grand_S, grand_N = 0.0, 0
        for lv in lig_vals:
            row = [lv]; rs, rn = 0.0, 0
            for cv in col_vals:
                s, n = S.get((lv, cv), 0.0), N.get((lv, cv), 0)
                row.append(fmt(s, n)); rs += s; rn += n
                col_S[cv] += s; col_N[cv] += n
            row.append(fmt(rs, rn)); grand_S += rs; grand_N += rn
            lignes.append(row)
        total = ["Total"] + [fmt(col_S[cv], col_N[cv]) for cv in col_vals] + [fmt(grand_S, grand_N)]
        lignes.append(total)
        st = "%s par %s x %s" % (mlabel, lig, col)
    else:
        S, N = {}, {}; lig_vals = []
        for r in rows:
            lv = str(r.get(lig, "")).strip() or "(vide)"
            if lv not in lig_vals:
                lig_vals.append(lv)
            s, n = part(r)
            S[lv] = S.get(lv, 0.0) + s; N[lv] = N.get(lv, 0) + n
        lig_vals.sort()
        colonnes = [lig, mlabel]
        lignes = [[lv, fmt(S[lv], N[lv])] for lv in lig_vals]
        tot_S = sum(S.values()); tot_N = sum(N.values())
        lignes.append(["Total", fmt(tot_S, tot_N)])
        st = "%s par %s" % (mlabel, lig)
    return {"titre": "Tableau croise", "colonnes": colonnes, "lignes": lignes,
            "nb": max(len(lignes) - 1, 0), "sous_titre": st}


# Enregistrer la vue Q3 dans le registre des vues metier.
VUES["absences"] = {"libelle": "Absences eleves & observations", "icone": "ti-user-x",
                    "description": "Absences en cours (A2) + observations/plaintes de stage (S1) (Q3).",
                    "builder": vue_absences}
VUES_ORDRE.append("absences")


# ---------------------------------------------------------------------------
# Referentiel des formations (maquettes) — onglet R1_Maquettes (V1.25)
# Lecture seule. Sert de socle aux matieres des seances (A3) et au volume
# horaire programme (suivi des heures), et a la structure UE pour les notes.
# ---------------------------------------------------------------------------
def _r1_col(nom):
    ent = _db.entetes("R1_Maquettes")
    return ent.index(nom) if nom in ent else None


def matieres_maquette(filiere=None, niveau=None, semestre=None):
    """Liste (sans doublon, ordre d'apparition) des matieres de la maquette,
    filtrable par filiere / niveau / semestre."""
    cf, cn, cs, cm = (_r1_col("Filière"), _r1_col("Niveau"),
                      _r1_col("Semestre"), _r1_col("Matière / Contenu"))
    out, seen = [], set()
    for r in _db.lignes("R1_Maquettes"):
        if filiere and str(r[cf]).strip() != str(filiere).strip():
            continue
        if niveau and str(r[cn]).strip() != str(niveau).strip():
            continue
        if semestre and str(r[cs]).strip() != str(semestre).strip():
            continue
        m = str(r[cm]).strip()
        if m and m not in seen:
            seen.add(m)
            out.append(m)
    return out


def heures_programmees(filiere=None, niveau=None, matiere=None):
    """Somme des heures programmees ('Total heures') de la maquette pour le
    filtre donne (sert au suivi des heures programme vs constate)."""
    cf, cn, cmat, ct = (_r1_col("Filière"), _r1_col("Niveau"),
                        _r1_col("Matière / Contenu"), _r1_col("Total heures"))
    tot = 0
    for r in _db.lignes("R1_Maquettes"):
        if filiere and str(r[cf]).strip() != str(filiere).strip():
            continue
        if niveau and str(r[cn]).strip() != str(niveau).strip():
            continue
        if matiere and str(r[cmat]).strip() != str(matiere).strip():
            continue
        tot += _num(r[ct])
    return tot


def documents_officiels_groupes():
    """Documents officiels (H1) groupes par CATEGORIE (colonne 'Type'), pour la
    consultation. Ordre des categories = celui de la liste editable 'Categories_doc'
    (P0), puis categories supplementaires, puis '(Sans categorie)'. Chaque document :
    titre, reference, date, responsable, chemin."""
    iT, iCat, iRef, iCh, iDate, iResp = _idx_map(_db, "H1_Biblio_docs",
        ["Titre du document", "Type", "Reference", "Chemin / lien local",
         "Date de mise a jour", "Responsable mise a jour"])
    # ordre de reference des categories
    ordre = []
    for brut, vals in _db.listes_parametres().items():
        if decoupe_provenance(brut)[0] == "Categories_doc":
            ordre = [str(v).strip() for v in vals if str(v).strip()]
            break
    par = {}
    for r in _db.lignes("H1_Biblio_docs"):
        titre = str(r[iT]).strip() if iT >= 0 else ""
        if not titre:
            continue
        cat = (str(r[iCat]).strip() if iCat >= 0 else "") or "(Sans catégorie)"
        par.setdefault(cat, []).append({
            "titre": titre,
            "reference": str(r[iRef]).strip() if iRef >= 0 else "",
            "chemin": str(r[iCh]).strip() if iCh >= 0 else "",
            "date": str(r[iDate]).strip() if iDate >= 0 else "",
            "responsable": str(r[iResp]).strip() if iResp >= 0 else "",
        })
    cats = [c for c in ordre if c in par] + \
           sorted(c for c in par if c not in ordre and c != "(Sans catégorie)") + \
           (["(Sans catégorie)"] if "(Sans catégorie)" in par else [])
    return [{"categorie": c, "documents": sorted(par[c], key=lambda d: d["titre"].lower())}
            for c in cats]


# --- A3 : matiere en suggestions (datalist) alimentees par la maquette (V1.26) -------
# Saisie LIBRE avec suggestions filtrees par Filiere/Niveau/Semestre. Pas de
# validation stricte (le Dictionnaire garde "Texte", classeur inchange). Le
# filtrage et le pre-remplissage du volume horaire programme se font cote client
# a partir des lignes compactes injectees ci-dessous. La maquette code le semestre
# en cursus (1..6) ; A3 saisit S1/S2 (annee) -> le client derive le semestre cursus
# de (Niveau, Semestre) : L1.S1=1, L1.S2=2, L2.S1=3, L2.S2=4, L3.S1=5, L3.S2=6.
def maquette_lignes_datalist():
    """Lignes compactes de la maquette pour le filtrage cote client :
    [{f, n, s, m, h}] (filiere, niveau, semestre cursus 1..6, matiere, total heures)."""
    cf, cn, cs, cm, ct = (_r1_col("Filière"), _r1_col("Niveau"),
                          _r1_col("Semestre"), _r1_col("Matière / Contenu"),
                          _r1_col("Total heures"))
    out = []
    for r in _db.lignes("R1_Maquettes"):
        m = str(r[cm]).strip()
        if not m:
            continue
        out.append({"f": str(r[cf]).strip(), "n": str(r[cn]).strip(),
                    "s": str(r[cs]).strip(), "m": m, "h": _num(r[ct])})
    return out


def maquette_datalist_cfg(onglet):
    """Resout la config datalist d'un onglet (libelles propres -> 'brut'/name HTML).
    Renvoie None si l'onglet n'est pas concerne."""
    spec = config.MAQUETTE_DATALIST.get(onglet)
    if not spec:
        return None
    lib2brut = {m["libelle"]: m["brut"] for m in entetes_meta(onglet)}
    cfg = {"id": "dl_maquette_" + onglet}
    for cle in ("matiere", "volume", "filiere", "niveau", "semestre"):
        cfg[cle] = lib2brut.get(spec[cle], spec[cle])
    return cfg


# --- Module Stages : controle de quota a la saisie + tableau de bord (V1.29) --------
# Aucune ecriture : on expose le referentiel des lieux (quota par seance) et les
# affectations existantes. Le calcul des places restantes et l'occupation se font
# cote client (hors-ligne), comme l'aide A3.
def _idx_lib(onglet, libelle_propre):
    for i, b in enumerate(_db.entetes(onglet)):
        if decoupe_provenance(b)[0] == libelle_propre:
            return i
    return -1


def _libelle_lieu(lieu, service):
    """Libelle composite identique a la liste S1 : 'Lieu — Service' (service facultatif)."""
    lieu = str(lieu).strip()
    service = str(service).strip()
    return (lieu + " — " + service) if (lieu and service) else lieu


def stages_referentiel_lieux():
    """Lieux de S2 : {lib (= valeur stockee dans S1), quota, niveau}."""
    il = _idx_lib("S2_Lieux_stage", "Lieu / structure")
    isv = _idx_lib("S2_Lieux_stage", "Service")
    iq = _idx_lib("S2_Lieux_stage", "Quota")
    inv = _idx_lib("S2_Lieux_stage", "Niveau concerne")
    out = []
    for r in _db.lignes("S2_Lieux_stage"):
        lib = _libelle_lieu(r[il] if il >= 0 else "", r[isv] if isv >= 0 else "")
        if not lib:
            continue
        out.append({"lib": lib,
                    "quota": _num_h(r[iq]) if iq >= 0 else 0,
                    "niveau": str(r[inv]).strip() if inv >= 0 else ""})
    return out


def stages_affectations():
    """Affectations existantes (S1) : {lieu, annee, seance, matricule}."""
    il = _idx_lib("S1_Stages", "Lieu de stage")
    ia = _idx_lib("S1_Stages", "Annee acad.")
    isn = _idx_lib("S1_Stages", "N seance (1-6)")
    im = _idx_lib("S1_Stages", "Matricule")
    out = []
    for r in _db.lignes("S1_Stages"):
        lieu = str(r[il]).strip() if il >= 0 else ""
        if not lieu:
            continue
        out.append({"lieu": lieu,
                    "annee": str(r[ia]).strip() if ia >= 0 else "",
                    "seance": str(r[isn]).strip() if isn >= 0 else "",
                    "matricule": str(r[im]).strip() if im >= 0 else ""})
    return out


def stages_cfg_saisie():
    """Noms 'bruts' (name HTML) des champs S1 utiles au controle de quota cote client."""
    e = _db.entetes("S1_Stages")
    def brut(lib):
        i = _idx_lib("S1_Stages", lib)
        return e[i] if i >= 0 else lib
    return {"lieu": brut("Lieu de stage"), "annee": brut("Annee acad."),
            "seance": brut("N seance (1-6)"), "matricule": brut("Matricule")}


# === Module Notes : moteur de calcul des moyennes + releve (V1.31) =================
# Regles decret 05-106 : moyenne matiere = 1/4 CC + 3/4 examen (ou note unique) ;
# moyenne UE = moyenne des matieres PONDEREE par 'Coef matiere' (defaut 1 =
# arithmetique, modele du releve officiel) ; moyenne semestre = moyenne des UE
# ponderee par le coefficient d'UE ; 2e session : la note de septembre remplace
# juin et ANNULE la 1re session (CC compris — decret art. 10, releve officiel).
# L'outil calcule et PROPOSE Admis/Ajourne ; la deliberation reste manuelle.
def _n(x):
    return str(x).strip().lower()


def _arrondi2(x):
    """Arrondi commercial (demi vers le haut) a 2 decimales. None reste None.
    Utilise pour l'AFFICHAGE et la comparaison de validation ; les calculs en
    cascade (matiere -> UE -> semestre) se font sur les valeurs EXACTES."""
    if x is None:
        return None
    from decimal import Decimal, ROUND_HALF_UP
    return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _moyenne_matiere(cc, examen):
    """1/4 CC + 3/4 examen si les deux notes ; sinon la note unique presente ; None si aucune."""
    scc, sex = str(cc).strip(), str(examen).strip()
    if scc != "" and sex != "":
        return 0.25 * _num_h(cc) + 0.75 * _num_h(examen)
    if sex != "":
        return _num_h(examen)        # note unique cote examen (ex. stage)
    if scc != "":
        return _num_h(cc)
    return None


def _mention(moy):
    if moy is None:
        return ""
    if moy >= 16:
        return "Tres bien"
    if moy >= 14:
        return "Bien"
    if moy >= 12:
        return "Assez bien"
    if moy >= 10:
        return "Passable"
    return "Insuffisant"


def _decision_passage(ects_requis, ects_acquis):
    """Decision annuelle de passage selon l'ecart de credits (V1.90).
    ecart = requis - acquis. 0 -> 'Admis' ; 1..SEUIL -> 'Admis conditionnel' ;
    > SEUIL -> 'Ajourne'. Seuil lu dans config.SEUIL_PASSAGE_CONDITIONNEL
    (parametrable). requis indeterminee (<=0, pas de bareme) -> ('—', None).
    Renvoie (libelle, ecart)."""
    try:
        req = float(ects_requis)
        acq = float(ects_acquis)
    except (TypeError, ValueError):
        return "—", None
    if req <= 0:
        return "—", None
    ecart = req - acq
    if ecart <= 0:
        return "Admis", 0.0
    seuil = float(seuil_passage())
    if ecart <= seuil:
        return "Admis conditionnel", ecart
    return "Ajourne", ecart


def decision_passage_officielle(moyenne, ects_requis, ects_acquis):
    """Decision officielle de passage (R7b, regle Dr Kamal 30/06/2026) :
      - moyenne >= 10                          -> 'Admis'
      - sinon, ecart de credits ECTS <= marge  -> 'Admis conditionnel'
      - sinon                                  -> 'Ajourne'
    La marge (en credits) est parametrable (seuil_passage / reglages.json).
    L'admission DEFINITIVE reste une action manuelle (validation par classe).
    Renvoie (libelle, ecart) ; moyenne None -> ('—', None)."""
    if moyenne is None:
        return "—", None
    try:
        m = float(moyenne)
    except (TypeError, ValueError):
        return "—", None
    if m >= 10:
        return "Admis", 0.0
    try:
        req = float(ects_requis); acq = float(ects_acquis)
    except (TypeError, ValueError):
        return "Ajourne", None
    if req <= 0:
        return "Ajourne", None
    ecart = req - acq
    if ecart <= float(seuil_passage()):
        return "Admis conditionnel", ecart
    return "Ajourne", ecart


def _idx_map(db, onglet, libelles):
    e = db.entetes(onglet)
    pos = {decoupe_provenance(b)[0]: i for i, b in enumerate(e)}
    return [pos.get(l, -1) for l in libelles]


def etudiant_a1(matricule):
    """(nom prenom, filiere, niveau) depuis A1 ; vides si introuvable."""
    im, inom, ipre, ifil, iniv = _idx_map(_db, "A1_Etudiants",
        ["Matricule", "Nom", "Prenom", "Filiere", "Niveau"])
    for r in _db.lignes("A1_Etudiants"):
        if im >= 0 and str(r[im]).strip() == str(matricule).strip():
            nom = ((str(r[inom]).strip() if inom >= 0 else "") + " " +
                   (str(r[ipre]).strip() if ipre >= 0 else "")).strip()
            return nom, (str(r[ifil]).strip() if ifil >= 0 else ""), (str(r[iniv]).strip() if iniv >= 0 else "")
    return "", "", ""


def recherche_etudiants():
    """Liste compacte de tous les etudiants pour la recherche/autocompletion
    (hors-ligne, cote client) : [{matricule, nom, prenom, filiere, niveau, label}].
    label = 'matricule — Nom Prenom (Filiere Niveau)'."""
    out = []
    for e in _lignes_dict("A1_Etudiants"):
        mat = str(e.get("Matricule", "")).strip()
        if not mat:
            continue
        nom = str(e.get("Nom", "")).strip()
        pre = str(e.get("Prenom", "")).strip()
        fil = str(e.get("Filiere", "")).strip()
        niv = str(e.get("Niveau", "")).strip()
        ctx = " ".join(x for x in [fil, niv] if x)
        label = "%s — %s %s%s" % (mat, nom, pre, (" (" + ctx + ")") if ctx else "")
        out.append({"matricule": mat, "nom": nom, "prenom": pre,
                    "filiere": fil, "niveau": niv, "label": label.strip()})
    out.sort(key=lambda x: (x["nom"].lower(), x["prenom"].lower(), x["matricule"]))
    return out


def fiche_etudiant(matricule):
    """Fiche complete d'un etudiant depuis A1 (tous les champs propres), ou None."""
    mat = str(matricule).strip()
    if not mat:
        return None
    for e in _lignes_dict("A1_Etudiants"):
        if str(e.get("Matricule", "")).strip().lower() == mat.lower():
            nom = str(e.get("Nom", "")).strip()
            pre = str(e.get("Prenom", "")).strip()
            d = {k: ("" if v is None else str(v).strip()) for k, v in e.items()}
            d["matricule"] = str(e.get("Matricule", "")).strip()
            d["nom_complet"] = (nom + " " + pre).strip()
            d["a_photo"] = _photo_existe(d["matricule"])
            return d
    return None


# --- Photos d'identite (fichier image, hors classeur) ---------------------
# JPEG / PNG uniquement, detectes par signature binaire (pas de dependance
# binaire type Pillow). Nom de fichier canonique : <matricule>.jpg.
_SIG_JPEG = b"\xFF\xD8\xFF"
_SIG_PNG = b"\x89PNG\r\n\x1a\n"


def _type_image(donnees):
    """'jpeg' / 'png' selon la signature binaire d'en-tete, sinon None."""
    if donnees[:3] == _SIG_JPEG:
        return "jpeg"
    if donnees[:8] == _SIG_PNG:
        return "png"
    return None


def chemin_photo(matricule):
    """Chemin canonique <matricule>.jpg (matricule assaini), ou None."""
    import os as _os
    safe = "".join(c for c in str(matricule).strip() if c.isalnum() or c in "-_")
    if not safe:
        return None
    return _os.path.join(getattr(config, "PHOTOS_DIR", ""), "%s.jpg" % safe)


def _photo_existe(matricule):
    import os as _os
    p = chemin_photo(matricule)
    return bool(p) and _os.path.exists(p)


def photo_servie(matricule):
    """(chemin, mimetype) si une photo existe, sinon (None, None).
    Le type est lu aux octets d'en-tete : un PNG depose sous .jpg est servi
    en image/png ; tout le reste en image/jpeg (defaut)."""
    import os as _os
    p = chemin_photo(matricule)
    if not p or not _os.path.exists(p):
        return None, None
    try:
        with open(p, "rb") as fh:
            sig = fh.read(8)
    except OSError:
        return None, None
    return p, ("image/png" if sig == _SIG_PNG else "image/jpeg")


def enregistrer_photo(matricule, donnees):
    """Valide (matricule connu, taille, format) puis ecrit <matricule>.jpg.
    Renvoie (ok, message). N'ecrit jamais dans le classeur."""
    import os as _os
    mat = str(matricule).strip()
    if not mat or fiche_etudiant(mat) is None:
        return False, "Matricule inconnu : aucune fiche correspondante."
    if not donnees:
        return False, "Aucun fichier recu."
    maxi = getattr(config, "PHOTO_MAX_OCTETS", 1024 * 1024)
    if len(donnees) > maxi:
        return False, "Image trop lourde (%d Mo maximum)." % max(1, maxi // (1024 * 1024))
    if _type_image(donnees) is None:
        return False, "Format non reconnu : seuls JPEG et PNG sont acceptes."
    p = chemin_photo(mat)
    if not p:
        return False, "Matricule invalide."
    try:
        _os.makedirs(_os.path.dirname(p), exist_ok=True)
        with open(p, "wb") as fh:
            fh.write(donnees)
    except OSError:
        return False, "Ecriture du fichier impossible."
    return True, "Photo enregistree."


def supprimer_photo(matricule):
    """Retire le fichier photo de l'etudiant. Renvoie (ok, message)."""
    import os as _os
    p = chemin_photo(matricule)
    if p and _os.path.exists(p):
        try:
            _os.remove(p)
            return True, "Photo retiree."
        except OSError:
            return False, "Suppression impossible."
    return False, "Aucune photo a retirer."


def stages_etudiant(matricule):
    """Stages (S1) d'un etudiant : [{annee, seance, session, lieu, debut, fin, note,
    retour, effective}]. 'effective' marque la ligne dont la note est retenue
    (Rattrapage si note saisie, sinon Normale)."""
    mat = str(matricule).strip()
    out = []
    for s in _lignes_dict("S1_Stages"):
        if str(s.get("Matricule", "")).strip().lower() != mat.lower():
            continue
        out.append({"annee": str(s.get("Annee acad.", "")).strip(),
                    "seance": str(s.get("N seance (1-6)", "")).strip(),
                    "session": _sess_norm(s.get("Session", "")),
                    "lieu": str(s.get("Lieu de stage", "")).strip(),
                    "debut": str(s.get("Date debut", "")).strip(),
                    "fin": str(s.get("Date fin", "")).strip(),
                    "note": str(s.get("Note stage /20", "")).strip(),
                    "retour": str(s.get("Fiche retour (O/N)", "")).strip()})
    # Note effective par (annee, seance) : Rattrapage notee prime, sinon Normale.
    par_cle = {}
    for o in out:
        par_cle.setdefault((o["annee"], o["seance"]), {})[o["session"]] = o["note"]
    for o in out:
        d = par_cle[(o["annee"], o["seance"])]
        sess_eff = "Rattrapage" if d.get("Rattrapage", "") != "" else "Normale"
        o["effective"] = (o["session"] == sess_eff)
    return out


def _bareme_ues(filiere, niveau, semestre):
    """UE du barème pour (filiere, niveau, semestre), ordonnees :
    [{num,intitule,coef,ects,confirme,matieres,coef_mat}].
    V1.99.50 : coef PAR MATIERE (colonne 'Coef matiere', defaut 1 si absente ou
    vide) -> la moyenne d'UE devient une moyenne PONDEREE des matieres ; avec
    tous les coefs a 1 elle reste la moyenne arithmetique (modele du releve
    officiel, verifie sur le releve L2 SI : UE11 = 11,78 ; S3 = 9,89)."""
    iF, iN, iS, iU, iI, iM, iC, iE, iK, iCM = _idx_map(_db_notes, "N1_Bareme_UE",
        ["Filiere", "Niveau", "Semestre", "N° UE", "Intitule UE", "Matiere",
         "Coef UE", "ECTS UE", "Coef confirme", "Coef matiere"])
    ues, parnum = [], {}
    for r in _db_notes.lignes("N1_Bareme_UE"):
        if _n(r[iF]) != _n(filiere) or _n(r[iN]) != _n(niveau) or str(r[iS]).strip() != str(semestre).strip():
            continue
        num = str(r[iU]).strip()
        if num not in parnum:
            conf = (iK >= 0 and str(r[iK]).strip().lower().startswith("o"))
            ue = {"num": num, "intitule": str(r[iI]).strip(),
                  "coef": _num_h(r[iC]), "ects": _num_h(r[iE]), "confirme": conf,
                  "matieres": [], "coef_mat": {}}
            parnum[num] = ue; ues.append(ue)
        mat = str(r[iM]).strip()
        if mat and mat not in parnum[num]["matieres"]:
            parnum[num]["matieres"].append(mat)
            cm = 1.0
            if iCM >= 0 and str(r[iCM]).strip() not in ("", "None"):
                cm = _num_h(r[iCM])
                if not cm or cm <= 0:
                    cm = 1.0
            parnum[num]["coef_mat"][mat] = cm
    return ues


# --- Controle de coherence du bareme ECTS (R1, V1.99.38) — NON BLOQUANT ----------
# Signale a l'ecran (ecran d'edition N1_Bareme_UE) les ecarts du bareme sans empecher
# la saisie ni le calcul. 5 regles : ECTS/semestre != 30, ECTS/niveau != 60,
# UE sans coef, coef anormal (<=0 ou >5), coef non confirme.
def coherence_bareme(filiere=None):
    """Bilan de coherence de N1_Bareme_UE. Renvoie {ok, nb, semestres, niveaux,
    ues, non_confirmes, nb_ue}. Dedoublonne par UE. Optionnellement filtre par filiere."""
    iF, iN, iS, iU, iI, iC, iE, iK = _idx_map(_db_notes, "N1_Bareme_UE",
        ["Filiere", "Niveau", "Semestre", "N° UE", "Intitule UE",
         "Coef UE", "ECTS UE", "Coef confirme"])
    cible = _n(filiere) if filiere else None
    vus = {}
    for r in _db_notes.lignes("N1_Bareme_UE"):
        fil = str(r[iF]).strip() if iF >= 0 else ""
        if not fil or (cible and _n(fil) != cible):
            continue
        niv = str(r[iN]).strip() if iN >= 0 else ""
        sem = str(r[iS]).strip() if iS >= 0 else ""
        ue = str(r[iU]).strip() if iU >= 0 else ""
        if not ue:
            continue
        cle = (fil, niv, sem, ue)
        if cle in vus:
            continue
        coef_s = str(r[iC]).strip() if iC >= 0 else ""
        ects_s = str(r[iE]).strip() if iE >= 0 else ""
        vus[cle] = {"intitule": str(r[iI]).strip() if iI >= 0 else "",
                    "coef": _num_h(coef_s) if coef_s else None,
                    "ects": _num_h(ects_s) if ects_s else None,
                    "confirme": (iK >= 0 and str(r[iK]).strip().lower().startswith("o"))}
    s_sem, s_niv = {}, {}
    ues_pb, non_conf = [], []
    for (fil, niv, sem, ue), d in sorted(vus.items()):
        if d["ects"] is not None:
            s_sem[(fil, niv, sem)] = s_sem.get((fil, niv, sem), 0.0) + d["ects"]
            s_niv[(fil, niv)] = s_niv.get((fil, niv), 0.0) + d["ects"]
        probs = []
        if d["coef"] is None:
            probs.append("coef manquant")
        elif not (0 < d["coef"] <= 5):
            probs.append("coef anormal (%g)" % d["coef"])
        ref = {"filiere": fil, "niveau": niv, "semestre": sem, "ue": ue,
               "intitule": d["intitule"]}
        if probs:
            ues_pb.append(dict(ref, probleme=" ; ".join(probs)))
        if not d["confirme"]:
            non_conf.append(ref)
    sem_pb = [{"filiere": k[0], "niveau": k[1], "semestre": k[2],
               "somme": _arrondi2(v), "ecart": _arrondi2(v - 30)}
              for k, v in sorted(s_sem.items()) if abs(v - 30) >= 0.5]
    niv_pb = [{"filiere": k[0], "niveau": k[1],
               "somme": _arrondi2(v), "ecart": _arrondi2(v - 60)}
              for k, v in sorted(s_niv.items()) if abs(v - 60) >= 0.5]
    nb = len(sem_pb) + len(niv_pb) + len(ues_pb) + len(non_conf)
    return {"ok": nb == 0, "nb": nb, "semestres": sem_pb, "niveaux": niv_pb,
            "ues": ues_pb, "non_confirmes": non_conf, "nb_ue": len(vus)}


def _cc_table(matricule, annee):
    """CC DERIVE des controles N4 : {(session, semestre, num_ue, matiere): cc} ou
    cc = moyenne PONDEREE des notes de controle (coef 1 par defaut ; en pratique un
    seul controle par matiere -> cc = sa note). Onglet absent (script de migration
    non encore joue) => {} : on retombe alors sur N2.CC (comportement V1.30)."""
    if "N4_Controles" not in _db_notes.onglets():
        return {}
    iMat, iAn, iSess, iSem, iU, iM, iNote, iCoef = _idx_map(_db_notes, "N4_Controles",
        ["Matricule", "Annee acad.", "Session", "Semestre", "N° UE", "Matiere", "Note /20", "Coef"])
    acc = {}
    for r in _db_notes.lignes("N4_Controles"):
        if iMat < 0 or str(r[iMat]).strip() != str(matricule).strip():
            continue
        if iAn >= 0 and _n(r[iAn]) != _n(annee):
            continue
        note = str(r[iNote]).strip() if iNote >= 0 else ""
        if note == "":
            continue
        sess = (str(r[iSess]).strip() if iSess >= 0 else "") or "1"
        sem = str(r[iSem]).strip() if iSem >= 0 else ""
        ue = str(r[iU]).strip() if iU >= 0 else ""
        mat = str(r[iM]).strip() if iM >= 0 else ""
        coef = _num_h(r[iCoef]) if (iCoef >= 0 and str(r[iCoef]).strip() != "") else 1.0
        if coef <= 0:
            coef = 1.0
        cle = (sess, sem, ue, mat)
        s, p = acc.get(cle, (0.0, 0.0))
        acc[cle] = (s + _num_h(note) * coef, p + coef)
    return {k: (s / p) for k, (s, p) in acc.items() if p > 0}


def _cc_session(cct, sessions, sess, sem, num_ue, mat):
    """CC de la session demandee : controle N4 de cette session prioritaire,
    sinon CC N2 de cette session. V1.99.50 — alignement sur le RELEVE OFFICIEL
    et le decret 05-106 art. 10 (« la note de la 1ere session est annulee ») :
    en session 2, la moyenne de la matiere repose sur les seules notes de la
    session 2 (CC de rattrapage s'il existe, sinon note unique d'examen) ; le
    CC de session 1 n'est PLUS conserve. Verifie sur le releve L2 SI : chaque
    'Moy. sess. 2' du document = examen session 2 seul (ex. 4,00 -> 10,00)."""
    ccd = cct.get((str(sess), sem, num_ue, mat))
    if ccd is not None:
        return ccd
    return sessions.get(str(sess), ("", ""))[0]


def _notes_effectives(matricule, annee, semestre, session_max=None):
    """{(num_ue, matiere): moyenne_matiere} ; la session 2 remplace la session 1.
    Le CC est DERIVE des controles N4 quand ils existent ; sinon repli sur N2.CC.
    session_max='1' (V1.99.50) : les lignes de session 2 sont ignorees (edition
    'Premiere session' du bulletin officiel)."""
    iMat, iAn, iSess, iSem, iU, iM, iCC, iEx = _idx_map(_db_notes, "N2_Notes",
        ["Matricule", "Annee acad.", "Session", "Semestre", "N° UE", "Matiere", "CC", "Examen"])
    cct = _cc_table(matricule, annee)
    sem = str(semestre).strip()
    parcle = {}
    for r in _db_notes.lignes("N2_Notes"):
        if str(r[iMat]).strip() != str(matricule).strip():
            continue
        if _n(r[iAn]) != _n(annee) or str(r[iSem]).strip() != sem:
            continue
        cle = (str(r[iU]).strip(), str(r[iM]).strip())
        sess = str(r[iSess]).strip() or "1"
        if str(session_max or "").strip() == "1" and sess != "1":
            continue                                     # edition Premiere session
        parcle.setdefault(cle, {})[sess] = (r[iCC], r[iEx])
    # Matieres avec controles N4 mais sans ligne N2 (examen pas encore saisi) :
    # on les fait apparaitre avec un examen vide => moyenne = CC seul.
    for (s, sm, ue, mat) in cct:
        if str(session_max or "").strip() == "1" and s != "1":
            continue
        if sm == sem:
            parcle.setdefault((ue, mat), {}).setdefault(s, ("", ""))
    out, sess_eff = {}, {}
    for cle, sessions in parcle.items():
        s_eff = "2" if "2" in sessions else "1"          # S2 remplace S1 (examen)
        ex = sessions.get(s_eff, ("", ""))[1]            # examen de la session effective
        cc = _cc_session(cct, sessions, s_eff, sem, cle[0], cle[1])  # notes de la session effective seule
        out[cle] = _moyenne_matiere(cc, ex)
        sess_eff[cle] = s_eff                            # V1.91 : session retenue (1/2)
    return out, sess_eff


# ===========================================================================
# SAISIE FACON BULLETIN (un etudiant -> sa grille UE/matieres, V1.65)
# ---------------------------------------------------------------------------
# Ecran de saisie d'un bulletin complet pour UN etudiant : grille du bareme
# (UE -> matieres), CC affiche (DERIVE des controles N4 ; lecture seule s'il
# existe, sinon saisissable), examen saisi, moyennes calculees en direct cote
# client. Ecrit N2_Notes en upsert (cle Matricule+Annee+Session+Semestre+UE+
# Matiere) UNIQUEMENT pour les matieres renseignees (CC ou Examen non vide).
# ---------------------------------------------------------------------------
def semestres_classe(filiere, niveau):
    """Semestres presents au bareme pour une classe (tries)."""
    iF, iN, iS = _idx_map(_db_notes, "N1_Bareme_UE", ["Filiere", "Niveau", "Semestre"])
    ss = set()
    for r in _db_notes.lignes("N1_Bareme_UE"):
        if _n(r[iF]) == _n(filiere) and _n(r[iN]) == _n(niveau):
            s = str(r[iS]).strip()
            if s:
                ss.add(s)
    return sorted(ss, key=lambda x: (len(x), x))


def _n2_par_session(matricule, annee, semestre):
    """{(num_ue, matiere): {session: (cc, examen)}} depuis N2 pour ce contexte."""
    iMat, iAn, iSess, iSem, iU, iM, iCC, iEx = _idx_map(_db_notes, "N2_Notes",
        ["Matricule", "Annee acad.", "Session", "Semestre", "N° UE", "Matiere", "CC", "Examen"])
    out = {}
    for r in _db_notes.lignes("N2_Notes"):
        if str(r[iMat]).strip() != str(matricule).strip():
            continue
        if _n(r[iAn]) != _n(annee) or str(r[iSem]).strip() != str(semestre).strip():
            continue
        cle = (str(r[iU]).strip(), str(r[iM]).strip())
        sess = str(r[iSess]).strip() or "1"
        out.setdefault(cle, {})[sess] = (r[iCC], r[iEx])
    return out


def bulletin_saisie(matricule, annee, semestre, session):
    """Modele de la grille de saisie facon bulletin. None si etudiant/bareme absent."""
    matricule = str(matricule).strip()
    nom, filiere, niveau = etudiant_a1(matricule)
    if not nom and not filiere:
        return None
    semestre = str(semestre).strip()
    session = str(session).strip() or "1"
    ues_bareme = _bareme_ues(filiere, niveau, semestre)
    if not ues_bareme:
        return {"ok": False, "nom": nom, "filiere": filiere, "niveau": niveau,
                "annee": annee, "semestre": semestre, "session": session,
                "ues": [], "provisoire": False}
    cct = _cc_table(matricule, annee)        # {(sess, sem, ue, mat): cc}
    n2 = _n2_par_session(matricule, annee, semestre)
    idx = 0
    ues = []
    for ue in ues_bareme:
        mats = []
        for m in ue["matieres"]:
            cle = (ue["num"], m)
            cc_der = cct.get((session, semestre, ue["num"], m))
            cc_n2, ex_n2 = n2.get(cle, {}).get(session, ("", ""))
            readonly = cc_der is not None
            cc_val = _arrondi2(cc_der) if readonly else (str(cc_n2).strip())
            # Reference session 1 (quand on saisit la session 2).
            ref = None
            if session == "2":
                cc1_der = cct.get(("1", semestre, ue["num"], m))
                cc1_n2, ex1 = n2.get(cle, {}).get("1", ("", ""))
                cc1 = cc1_der if cc1_der is not None else cc1_n2
                moy1 = _moyenne_matiere(cc1, ex1)
                ref = {"cc": _arrondi2(cc1) if cc1_der is not None else str(cc1).strip(),
                       "examen": str(ex1).strip(), "moyenne": _arrondi2(moy1)}
            mats.append({"idx": idx, "matiere": m, "cc": ("" if cc_val is None else cc_val),
                         "examen": str(ex_n2).strip(), "cc_readonly": readonly, "ref": ref,
                         "coefm": ue.get("coef_mat", {}).get(m, 1.0) or 1.0})
            idx += 1
        ues.append({"num": ue["num"], "intitule": ue["intitule"], "coef": ue["coef"],
                    "ects": ue["ects"], "confirme": ue["confirme"], "matieres": mats})
    provisoire = any(not ue["confirme"] for ue in ues_bareme)
    return {"ok": True, "matricule": matricule, "nom": nom, "filiere": filiere,
            "niveau": niveau, "annee": annee, "semestre": semestre, "session": session,
            "ues": ues, "provisoire": provisoire, "nb": idx}


def _note_valide(s):
    """(ok, texte_normalise) pour une note /20. '' = vide accepte (ok, '')."""
    s = str(s).strip().replace(",", ".")
    if s == "":
        return True, ""
    try:
        v = float(s)
    except ValueError:
        return False, None
    if v < 0 or v > 20:
        return False, None
    # normalisation d'affichage : on garde la saisie telle quelle (sans zero inutile)
    return True, ("%g" % v)


def enregistrer_bulletin(matricule, annee, semestre, session, items, saisi_par):
    """Upsert N2_Notes pour les matieres renseignees. items = [{ue, matiere, cc,
    examen, cc_readonly}]. cc d'une matiere a CC derive (readonly) n'est pas ecrit.
    Renvoie (ok, message)."""
    matricule = str(matricule).strip()
    annee = str(annee).strip()
    semestre = str(semestre).strip()
    session = str(session).strip() or "1"
    if not (matricule and semestre):
        return False, "Contexte incomplet (matricule / semestre)."
    bM, bA, bSe, bSm, bU, bMa, bCC, bEx = (
        _brut("N2_Notes", "Matricule"), _brut("N2_Notes", "Annee acad."),
        _brut("N2_Notes", "Session"), _brut("N2_Notes", "Semestre"),
        _brut("N2_Notes", "N° UE"), _brut("N2_Notes", "Matiere"),
        _brut("N2_Notes", "CC"), _brut("N2_Notes", "Examen"))
    lignes = []
    for it in items:
        cc_ro = bool(it.get("cc_readonly"))
        cc_in = "" if cc_ro else str(it.get("cc", "")).strip()
        ex_in = str(it.get("examen", "")).strip()
        # matiere renseignee = au moins une note saisie (CC manuel ou examen)
        if cc_in == "" and ex_in == "":
            continue
        okc, ccn = _note_valide(cc_in)
        oke, exn = _note_valide(ex_in)
        if not okc or not oke:
            return False, ("Note invalide pour « %s » (attendu : nombre entre 0 et 20)."
                           % it.get("matiere", "?"))
        ligne = {bM: matricule, bA: annee, bSe: session, bSm: semestre,
                 bU: str(it.get("ue", "")).strip(), bMa: str(it.get("matiere", "")).strip(),
                 bEx: exn}
        if not cc_ro:                 # CC manuel : on l'ecrit (vide compris -> efface)
            ligne[bCC] = ccn
        lignes.append(ligne)
    if not lignes:
        return False, "Aucune matiere renseignee : rien a enregistrer."
    res = _db_notes.ecrire_lignes_lot("N2_Notes", lignes, cles=[bM, bA, bSe, bSm, bU, bMa])
    return True, ("Bulletin enregistre : %d matiere(s) ; %d ajout(s), %d mise(s) a jour."
                  % (len(lignes), res["ajout"], res["maj"]))


def releve_semestre(matricule, annee, semestre, session_max=None):
    """Releve d'un semestre : UE/matieres avec moyennes, validation, mention, proposition.
    V1.99.50 : moyenne d'UE PONDEREE par le coef matiere (defaut 1 = arithmetique,
    modele du releve officiel). session_max='1' -> edition PREMIERE SESSION :
    les notes de session 2 sont ignorees (page session 1 du bulletin)."""
    nom, filiere, niveau = etudiant_a1(matricule)
    ues = _bareme_ues(filiere, niveau, semestre)
    notes, notes_sess = _notes_effectives(matricule, annee, semestre, session_max)
    nd = _matieres_nd(filiere, niveau, semestre, annee)   # V1.77 : (num_ue, matiere) non dispensees
    lignes, sc, sp, toutes_ok = [], 0.0, 0.0, True
    ects_acq, ects_tot = 0.0, 0.0
    a_nd = False        # au moins une matiere exclue (non dispensee)
    incomplet = False   # au moins une matiere dispensee mais pas encore notee
    for ue in ues:
        mats, s_m, p_m = [], 0.0, 0.0
        nb_prevues = len(ue["matieres"]); nb_nd = 0
        for m in ue["matieres"]:
            if (str(ue["num"]).strip(), str(m).strip()) in nd:   # affichage (b) : n'apparait pas
                a_nd = True; nb_nd += 1
                continue
            mv = notes.get((ue["num"], m))         # valeur EXACTE
            cm = ue.get("coef_mat", {}).get(m, 1.0) or 1.0
            mats.append({"matiere": m, "moyenne": _arrondi2(mv), "coefm": cm,
                         "session": notes_sess.get((ue["num"], m), "")})  # V1.91
            if mv is not None:
                s_m += mv * cm; p_m += cm          # cascade EXACTE, ponderee
            else:
                incomplet = True                   # dispensee mais pas encore notee
        if nb_prevues > 0 and nb_nd == nb_prevues:  # UE ENTIEREMENT non dispensee
            continue                                # -> retiree (prorata ECTS, option i)
        ects_tot += ue["ects"]
        moy_ue_exact = (s_m / p_m) if p_m > 0 else None
        moy_ue_aff = _arrondi2(moy_ue_exact)
        validee = moy_ue_aff is not None and moy_ue_aff >= 10
        if moy_ue_exact is not None:
            sc += ue["coef"]; sp += moy_ue_exact * ue["coef"]   # cascade EXACTE
            if not validee:
                toutes_ok = False
            if validee:
                ects_acq += ue["ects"]
        lignes.append({"num": ue["num"], "intitule": ue["intitule"], "coef": ue["coef"],
                       "ects": ue["ects"], "matieres": mats, "moyenne": moy_ue_aff, "validee": validee})
    moy_sem_exact = (sp / sc) if sc else None
    moy_sem = _arrondi2(moy_sem_exact)
    admis = moy_sem is not None and (toutes_ok or moy_sem >= 10)
    provisoire = bool(ues) and any(not ue["confirme"] for ue in ues)
    return {"matricule": matricule, "nom": nom, "filiere": filiere, "niveau": niveau,
            "annee": annee, "semestre": semestre, "ues": lignes, "moyenne": moy_sem,
            "moyenne_exacte": moy_sem_exact, "mention": _mention(moy_sem),
            "bareme_provisoire": provisoire,
            "a_non_dispensee": a_nd, "incomplet": incomplet,
            "proposition": ("—" if moy_sem is None else ("Admis" if admis else "Ajourne")),
            "ects_acquis": ects_acq, "ects_total": ects_tot,
            "signalements": signalements_etudiant(matricule, annee, semestre)}


def releve_annuel(matricule, annee, session_max=None):
    """Recapitulatif annuel : les deux semestres du niveau + moyenne annuelle.
    session_max='1' -> edition Premiere session (les notes de session 2 sont ignorees)."""
    nom, filiere, niveau = etudiant_a1(matricule)
    base = {"l1": ("1", "2"), "l2": ("3", "4"), "l3": ("5", "6")}.get(_n(niveau))
    if base:
        sems = list(base)
    else:
        # fallback : semestres presents dans les notes de l'etudiant pour l'annee
        iMat, iAn, iSem = _idx_map(_db_notes, "N2_Notes", ["Matricule", "Annee acad.", "Semestre"])
        sems = sorted({str(r[iSem]).strip() for r in _db_notes.lignes("N2_Notes")
                       if str(r[iMat]).strip() == str(matricule).strip() and _n(r[iAn]) == _n(annee)
                       and str(r[iSem]).strip()})
    rels = [releve_semestre(matricule, annee, s, session_max) for s in sems]
    moys = [r["moyenne_exacte"] for r in rels if r.get("moyenne_exacte") is not None]
    moy_an = _arrondi2(sum(moys) / len(moys)) if moys else None
    admis = moy_an is not None and moy_an >= 10 and all(
        r["proposition"] == "Admis" for r in rels if r["moyenne"] is not None)
    # V1.90 : credits annuels (somme des 2 semestres) + decision de passage.
    ects_acq = sum(r.get("ects_acquis", 0.0) or 0.0 for r in rels)
    ects_req = sum(r.get("ects_total", 0.0) or 0.0 for r in rels)
    decision, ecart = decision_passage_officielle(moy_an, ects_req, ects_acq)
    return {"matricule": matricule, "nom": nom, "filiere": filiere, "niveau": niveau,
            "annee": annee, "semestres": rels, "moyenne": moy_an, "mention": _mention(moy_an),
            "bareme_provisoire": any(r.get("bareme_provisoire") for r in rels),
            "a_non_dispensee": any(r.get("a_non_dispensee") for r in rels),
            "incomplet": any(r.get("incomplet") for r in rels),
            "proposition": ("—" if moy_an is None else ("Admis" if admis else "Ajourne")),
            "ects_acquis": ects_acq, "ects_requis": ects_req, "ects_ecart": ecart,
            "decision": decision,
            "signalements": signalements_etudiant(matricule, annee)}


def notes_assist():
    """Donnees pour l'assistance a la saisie des notes (N2) : barème compact, table
    matricule -> filiere/niveau (A1), et noms 'bruts' des champs N2. Filtrage et
    pre-remplissage cote client (offline), comme l'aide A3."""
    iF, iN, iS, iU, iM = _idx_map(_db_notes, "N1_Bareme_UE",
        ["Filiere", "Niveau", "Semestre", "N° UE", "Matiere"])
    bareme = []
    for r in _db_notes.lignes("N1_Bareme_UE"):
        bareme.append({"f": str(r[iF]).strip(), "n": str(r[iN]).strip(), "s": str(r[iS]).strip(),
                       "ue": str(r[iU]).strip(), "m": str(r[iM]).strip()})
    im, ifil, iniv = _idx_map(_db, "A1_Etudiants", ["Matricule", "Filiere", "Niveau"])
    etu = {}
    for r in _db.lignes("A1_Etudiants"):
        mat = str(r[im]).strip() if im >= 0 else ""
        if mat:
            etu[mat] = {"f": str(r[ifil]).strip() if ifil >= 0 else "",
                        "n": str(r[iniv]).strip() if iniv >= 0 else ""}
    e = _db_notes.entetes("N2_Notes")
    def brut(lib):
        for b in e:
            if decoupe_provenance(b)[0] == lib:
                return b
        return lib
    cfg = {"matricule": brut("Matricule"), "semestre": brut("Semestre"),
           "ue": brut("N° UE"), "matiere": brut("Matiere"),
           "dl_ue": "dl_ue_N2", "dl_mat": "dl_mat_N2"}
    return {"bareme": bareme, "etu": etu, "cfg": cfg}


def signalements_etudiant(matricule, annee, semestre=None):
    """Signalements (N3) d'un etudiant pour une annee. Si 'semestre' est fourni, on
    garde ceux de ce semestre + ceux sans semestre precis. INFORMATION de deliberation :
    n'affecte NI le calcul NI la proposition ; exclue du bulletin officiel."""
    iMat, iDate, iAn, iSem, iCtx, iFon, iNom, iMot = _idx_map(_db_notes, "N3_Signalements",
        ["Matricule", "Date", "Annee acad.", "Semestre", "Contexte",
         "Emis par - fonction", "Nom de l'emetteur", "Motif"])
    out = []
    for r in _db_notes.lignes("N3_Signalements"):
        if str(r[iMat]).strip() != str(matricule).strip() or _n(r[iAn]) != _n(annee):
            continue
        sem = str(r[iSem]).strip()
        if semestre not in (None, "") and sem not in ("", str(semestre).strip()):
            continue
        out.append({"date": str(r[iDate]).strip(), "semestre": sem,
                    "contexte": str(r[iCtx]).strip(), "fonction": str(r[iFon]).strip(),
                    "nom": str(r[iNom]).strip(), "motif": str(r[iMot]).strip()})
    return out


def etat_signalements(annee="", semestre="", contexte="", filiere="", niveau="", du="", au=""):
    """Compte rendu d'indiscipline GROUPE PAR ETUDIANT, filtrable (annee, semestre,
    contexte, filiere, niveau, plage de dates du/au). Sans notes : pur disciplinaire.
    Renvoie une liste triee par nom : [{matricule, nom, filiere, niveau, nb, signalements:[...]}]."""
    iMat, iDate, iAn, iSem, iCtx, iFon, iNom, iMot = _idx_map(_db_notes, "N3_Signalements",
        ["Matricule", "Date", "Annee acad.", "Semestre", "Contexte",
         "Emis par - fonction", "Nom de l'emetteur", "Motif"])
    bdu, bau = _parse_date_fr(du) if du else None, _parse_date_fr(au) if au else None
    parmat = {}
    for r in _db_notes.lignes("N3_Signalements"):
        mat = str(r[iMat]).strip()
        if not mat:
            continue
        if annee and _n(r[iAn]) != _n(annee):
            continue
        if semestre and str(r[iSem]).strip() != str(semestre).strip():
            continue
        if contexte and _n(r[iCtx]) != _n(contexte):
            continue
        d = _parse_date_fr(r[iDate])
        if bdu and (d is None or d < bdu):
            continue
        if bau and (d is None or d > bau):
            continue
        parmat.setdefault(mat, []).append({
            "date": str(r[iDate]).strip(), "semestre": str(r[iSem]).strip(),
            "contexte": str(r[iCtx]).strip(), "fonction": str(r[iFon]).strip(),
            "nom": str(r[iNom]).strip(), "motif": str(r[iMot]).strip(),
            "_tri": d or (0, 0, 0)})
    out = []
    for mat, sigs in parmat.items():
        nom, fil, niv = etudiant_a1(mat)
        if filiere and _n(fil) != _n(filiere):
            continue
        if niveau and _n(niv) != _n(niveau):
            continue
        sigs.sort(key=lambda x: x["_tri"])
        for x in sigs:
            x.pop("_tri", None)
        out.append({"matricule": mat, "nom": nom, "filiere": fil, "niveau": niv,
                    "nb": len(sigs), "signalements": sigs})
    out.sort(key=lambda e: (e["nom"] or e["matricule"]).lower())
    return out


# =========================================================================
# CLÔTURE / ARCHIVAGE / PASSATION (V1.44)
# Élèves : année SCOLAIRE (oct->juil), clôture au 31/07. Compta : année CIVILE.
# Déclenchement MANUEL par la Direction. Écrit dans les journaux permanents
# J1_Journal_eleves / J2_Journal_compta (onglets du classeur) ; archive les
# données dans des fichiers séparés sous config.ARCHIVES_DIR. Tous les calculs
# (totaux, soldes) sont faits EN PYTHON.
# =========================================================================
def _norm(s):
    """Normalise pour comparaison : sans accents, minuscules, sans espaces de bord."""
    s = unicodedata.normalize("NFD", str(s or "")).encode("ascii", "ignore").decode()
    return s.lower().strip()


def _brut_map(onglet):
    """Map {libelle_propre: en-tete_brut} pour écrire via ecrire_lignes_lot."""
    return {decoupe_provenance(b)[0]: b for b in _db.entetes(onglet)}


def _ecrire_journal(onglet, lignes_propre, cles_propre):
    """Écrit (upsert) des lignes dans un journal permanent, en traduisant les clés
    propres en en-têtes bruts exacts attendus par la couche d'accès."""
    bm = _brut_map(onglet)
    lignes = [{bm[k]: v for k, v in d.items() if k in bm} for d in lignes_propre]
    cles = [bm[k] for k in cles_propre if k in bm]
    return _db.ecrire_lignes_lot(onglet, lignes, cles=cles)


# --- Années scolaires --------------------------------------------------------
def annee_scolaire_label(d=None):
    """Libellé de l'année scolaire (oct->juil) contenant la date d : 'AAAA-AAAA'."""
    d = d or datetime.date.today()
    if d.month >= 8:                 # août..décembre -> année en cours -> +1
        a = d.year
    else:                            # janvier..juillet -> année-1 -> en cours
        a = d.year - 1
    return "%d-%d" % (a, a + 1)


def _annee_fin_label(label):
    """Année de fin (int) d'un libellé 'AAAA-AAAA' (ou None si non parsable)."""
    m = re.match(r"^\s*(\d{4})\s*-\s*(\d{4})\s*$", str(label or ""))
    return int(m.group(2)) if m else None


# --- Fichiers d'archive (séparés, charte EMSP) -------------------------------
def _ecrire_archive(chemin, banniere, entetes, lignes):
    """Crée/complète un fichier d'archive .xlsx (charte EMSP). Si le fichier existe,
    AJOUTE les lignes à la suite ; sinon le crée avec bannière + en-têtes. openpyxl
    est sans risque ici (fichier neuf, aucun dessin à préserver)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    os.makedirs(os.path.dirname(chemin), exist_ok=True)
    if os.path.exists(chemin):
        wb = openpyxl.load_workbook(chemin)
        ws = wb.active
        depart = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Archive"
        ws["A1"] = banniere
        ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
        for j, h in enumerate(entetes, 1):
            c = ws.cell(row=2, column=j, value=h)
            c.font = Font(name="Calibri", bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A3"
        depart = 3
    for i, lig in enumerate(lignes):
        for j, val in enumerate(lig, 1):
            c = ws.cell(row=depart + i, column=j, value=val)
            c.font = Font(name="Calibri")
    wb.save(chemin)
    wb.close()
    return len(lignes)


# --- ÉLÈVES : clôture annuelle (journalisation des sorties) -------------------
def _matricules_journalises():
    return {str(r.get("Matricule", "")).strip()
            for r in _lignes_dict("J1_Journal_eleves")
            if str(r.get("Matricule", "")).strip()}


def eleves_sortis_a_journaliser():
    """Élèves de A1 dont le statut marque une sortie (Diplômé/Abandonné/Radié) et
    qui ne sont pas encore au journal permanent. Pour la saisie diplôme/mention."""
    sortie = {_norm(s) for s in config.STATUTS_SORTIE}
    deja = _matricules_journalises()
    out = []
    for r in _lignes_dict("A1_Etudiants"):
        mat = str(r.get("Matricule", "")).strip()
        statut = str(r.get("Statut", "")).strip()
        if not mat or _norm(statut) not in sortie or mat in deja:
            continue
        out.append({
            "matricule": mat,
            "nom": str(r.get("Nom", "")).strip(),
            "prenom": str(r.get("Prenom", "")).strip(),
            "filiere": str(r.get("Filiere", "")).strip(),
            "niveau": str(r.get("Niveau", "")).strip(),
            "annee_entree": str(r.get("Annee acad.", "")).strip(),
            "statut": statut,
        })
    out.sort(key=lambda e: (e["nom"] or e["matricule"]).lower())
    return out


def cloturer_eleves(acteur, annee_label, saisies, login):
    """Journalise les élèves sortis. `saisies` = {matricule: {'diplome':..,'mention':..}}.
    Idempotent (n'ajoute pas un matricule déjà journalisé). Renvoie (n, message, lignes_pv)."""
    if not est_admin(acteur) and not peut_ecrire(acteur, "J1_Journal_eleves"):
        return 0, "Action reservee a la Direction.", []
    annee_label = str(annee_label or annee_scolaire_label()).strip()
    candidats = {e["matricule"]: e for e in eleves_sortis_a_journaliser()}
    aujourd = datetime.date.today().strftime("%d/%m/%Y")
    lignes, pv = [], []
    for mat, e in candidats.items():
        s = saisies.get(mat, {}) if saisies else {}
        periode = "%s -> %s" % (e["annee_entree"] or "?", annee_label)
        ligne = {
            "Matricule": mat, "Nom": e["nom"], "Prenom": e["prenom"],
            "Filiere": e["filiere"], "Niveau atteint": e["niveau"],
            "Periode (entree - sortie)": periode, "Statut final": e["statut"],
            "Diplome obtenu": str(s.get("diplome", "")).strip(),
            "Mention": str(s.get("mention", "")).strip(),
            "Ref. archive": "", "Cloture le": aujourd, "Cloture par": login,
        }
        lignes.append(ligne)
        pv.append(ligne)
    if not lignes:
        return 0, "Aucun eleve sorti a journaliser pour cette annee.", []
    _ecrire_journal("J1_Journal_eleves", lignes,
                    cles_propre=["Matricule"])
    return len(lignes), "%d eleve(s) journalise(s)." % len(lignes), pv


# --- ÉLÈVES : archivage différé (cohortes sorties depuis >= N ans) ------------
def eleves_archivables(today=None):
    """Lignes du journal permanent dont l'année de sortie est >= ANNEES_GARDE_ELEVES
    révolue, encore présentes dans A1 et pas encore archivées (Ref. archive vide)."""
    today = today or datetime.date.today()
    seuil = today.year - int(config.ANNEES_GARDE_ELEVES)
    en_a1 = {str(r.get("Matricule", "")).strip() for r in _lignes_dict("A1_Etudiants")}
    out = []
    for r in _lignes_dict("J1_Journal_eleves"):
        mat = str(r.get("Matricule", "")).strip()
        if not mat or mat not in en_a1:
            continue
        if str(r.get("Ref. archive", "")).strip():
            continue
        fin = _annee_fin_label(str(r.get("Periode (entree - sortie)", "")).split("->")[-1])
        if fin is not None and fin <= seuil:
            out.append({"matricule": mat, "label": "%d-%d" % (fin - 1, fin),
                        "nom": str(r.get("Nom", "")).strip()})
    return out


def archiver_eleves(acteur, login):
    """Déplace de A1 vers archives/EMSP_Archive_Eleves_AAAA-AAAA.xlsx les cohortes
    sorties depuis >= N ans, et renseigne Ref. archive dans le journal. Renvoie
    (n, message, recap)."""
    if not est_admin(acteur) and not peut_ecrire(acteur, "J1_Journal_eleves"):
        return 0, "Action reservee a la Direction.", []
    cibles = eleves_archivables()
    if not cibles:
        return 0, "Aucune cohorte a archiver (sorties depuis moins de %d ans)." \
               % config.ANNEES_GARDE_ELEVES, []
    # Regrouper par année scolaire de sortie -> un fichier par cohorte.
    par_label = {}
    for c in cibles:
        par_label.setdefault(c["label"], set()).add(c["matricule"])
    entetes_a1 = _db.entetes("A1_Etudiants")
    propres_a1 = [decoupe_provenance(b)[0] for b in entetes_a1]
    a1 = _lignes_dict("A1_Etudiants")
    a_retirer, recap = set(), []
    maj_journal = []
    for label, mats in par_label.items():
        chemin = os.path.join(config.ARCHIVES_DIR, "EMSP_Archive_Eleves_%s.xlsx" % label)
        lignes = [[r.get(p, "") for p in propres_a1] for r in a1
                  if str(r.get("Matricule", "")).strip() in mats]
        _ecrire_archive(chemin, "ARCHIVE DES ELEVES — annee scolaire %s" % label,
                        propres_a1, lignes)
        a_retirer |= mats
        recap.append({"label": label, "fichier": os.path.basename(chemin), "nb": len(lignes)})
        for m in mats:
            maj_journal.append({"Matricule": m, "Ref. archive": os.path.basename(chemin)})
    # Réécrire A1 sans les élèves archivés.
    restants = [{decoupe_provenance(b)[0]: r.get(decoupe_provenance(b)[0], "")
                 for b in entetes_a1}
                for r in a1 if str(r.get("Matricule", "")).strip() not in a_retirer]
    # remplacer_donnees attend des clés brutes
    bm = _brut_map("A1_Etudiants")
    restants_bruts = [{bm[k]: v for k, v in d.items() if k in bm} for d in restants]
    _db.remplacer_donnees("A1_Etudiants", restants_bruts)
    # Renseigner Ref. archive dans le journal.
    _ecrire_journal("J1_Journal_eleves", maj_journal, cles_propre=["Matricule"])
    return len(a_retirer), "%d eleve(s) archive(s)." % len(a_retirer), recap


# --- COMPTA : clôture d'exercice (année civile) ------------------------------
def _an_op(d):
    """Année (str 'AAAA') d'une date d'opération 'JJ/MM/AAAA', ou '' si illisible."""
    s = str(d or "").strip()
    if "/" in s:
        bout = s.split("/")[-1].strip()
        if re.match(r"^\d{4}$", bout):
            return bout
    return s[-4:] if re.match(r"^\d{4}$", s[-4:]) else ""


def _soldes_initiaux_f2():
    """{nom_compte: solde initial (float)} depuis F2."""
    out = {}
    for r in _lignes_dict("F2_Comptes"):
        nom = str(r.get("Nom du compte / caisse", "")).strip()
        if nom:
            out[nom] = _num(r.get("Solde initial (KMF)", 0))
    return out


def apercu_cloture_compta(annee):
    """Aperçu de la clôture compta d'un exercice (année civile). V1.94 : tout est
    STRICTEMENT borné à l'exercice. Totaux = mouvements de l'année ; solde de clôture
    par compte = solde ARRÊTÉ au 31/12 de l'année = solde initial F2 + net des
    mouvements dont l'année <= `annee` (les écritures de N+1 déjà saisies sont exclues)."""
    annee = str(annee).strip()
    mvts = _lignes_dict("F1_Mouvements")
    tot_rec = tot_dep = 0.0
    nb = 0
    net = {}                       # net cumulé par compte jusqu'au 31/12/annee
    for r in mvts:
        a = _an_op(r.get("Date operation", ""))
        rec = _num(r.get("Montant Recette (KMF)", 0))
        dep = _num(r.get("Montant Depense (KMF)", 0))
        est_report = "report a nouveau" in str(r.get("Libelle / description", "")).lower()
        if annee and a and a <= annee:                 # cumul borné (<= exercice)
            cpt = str(r.get("Compte / caisse", "")).strip()
            if cpt:
                net[cpt] = net.get(cpt, 0.0) + rec - dep   # inclut le report (= ouverture)
        if annee and a == annee and not est_report:    # totaux de l'exercice (hors report)
            tot_rec += rec
            tot_dep += dep
            nb += 1
    inits = _soldes_initiaux_f2()
    comptes = []
    solde_total = 0.0
    for c in comptes_treso():
        nom = c["nom"]
        s = inits.get(nom, 0.0) + net.get(nom, 0.0)     # solde arrêté au 31/12/annee
        comptes.append({"nom": nom, "solde": s, "solde_fmt": _fmt_kmf(s)})
        solde_total += s
    return {"annee": annee, "nb_mouvements": nb,
            "total_recettes": tot_rec, "total_recettes_fmt": _fmt_kmf(tot_rec),
            "total_depenses": tot_dep, "total_depenses_fmt": _fmt_kmf(tot_dep),
            "solde_cloture": solde_total, "solde_cloture_fmt": _fmt_kmf(solde_total),
            "comptes": comptes}


def cloturer_compta(acteur, annee, login):
    """Clôture l'exercice `annee` (civile). V1.94 — STRICTEMENT bornée à l'exercice :
    (1) archive les mouvements F1 dont l'année <= `annee` (les écritures de N+1 déjà
    saisies sont CONSERVÉES dans F1, jamais archivées ni perdues), (2) écrit le journal
    permanent J2 (totaux de l'année, solde arrêté au 31/12), (3) reporte à nouveau :
    une ligne 'Report a nouveau' par compte au 01/01/(N+1) = solde au 31/12/N, les
    écritures N+1 conservées, F2.Solde initial remis à 0. Renvoie (ok, msg, recap)."""
    if not est_admin(acteur) and not peut_ecrire(acteur, "J2_Journal_compta"):
        return False, "Action reservee a la Direction.", {}
    annee = str(annee).strip()
    if not re.match(r"^\d{4}$", annee):
        return False, "Annee civile invalide (AAAA).", {}
    ap = apercu_cloture_compta(annee)
    # 1) Archiver SEULEMENT les mouvements de l'exercice (année <= annee).
    entetes_f1 = _db.entetes("F1_Mouvements")
    propres_f1 = [decoupe_provenance(b)[0] for b in entetes_f1]
    mvts = _lignes_dict("F1_Mouvements")
    a_archiver = [r for r in mvts if (_an_op(r.get("Date operation", "")) or annee) <= annee]
    conserves = [r for r in mvts if (_an_op(r.get("Date operation", "")) or annee) > annee]
    chemin = os.path.join(config.ARCHIVES_DIR, "EMSP_Archive_Compta_%s.xlsx" % annee)
    lignes_arch = [[r.get(p, "") for p in propres_f1] for r in a_archiver
                   if any(str(r.get(p, "")).strip() for p in propres_f1)]
    nb_arch = _ecrire_archive(chemin, "ARCHIVE COMPTABLE — exercice %s" % annee,
                              propres_f1, lignes_arch) if lignes_arch else 0
    ref = os.path.basename(chemin)
    # 2) Journal permanent compta.
    _ecrire_journal("J2_Journal_compta", [{
        "Annee": annee,
        "Total recettes (KMF)": _fmt_kmf(ap["total_recettes"]),
        "Total depenses (KMF)": _fmt_kmf(ap["total_depenses"]),
        "Solde de cloture (KMF)": _fmt_kmf(ap["solde_cloture"]),
        "Ref. archive": ref if lignes_arch else "",
        "Cloture le": datetime.date.today().strftime("%d/%m/%Y"),
        "Cloture par": login,
    }], cles_propre=["Annee"])
    # 3) Report à nouveau (01/01/N+1 = solde au 31/12/N) + CONSERVATION des écritures N+1.
    bm = _brut_map("F1_Mouvements")
    report = []
    debut = "01/01/%d" % (int(annee) + 1)
    for c in ap["comptes"]:
        s = c["solde"]
        if abs(s) < 1e-9 and nb_arch == 0:
            continue
        ligne = {bm["Date operation"]: debut,
                 bm["Sens"]: "Recette" if s >= 0 else "Depense",
                 bm["Compte / caisse"]: c["nom"],
                 bm["Libelle / description"]: "Report a nouveau (exercice %s)" % annee,
                 bm["Saisi par"]: login}
        if s >= 0:
            ligne[bm["Montant Recette (KMF)"]] = _fmt_kmf(s)
        else:
            ligne[bm["Montant Depense (KMF)"]] = _fmt_kmf(-s)
        report.append(ligne)
    # Réinjecter les écritures conservées (année > N) en clés brutes.
    for r in conserves:
        ligne = {bm[k]: v for k, v in r.items() if k in bm and str(v).strip() != ""}
        if ligne:
            report.append(ligne)
    _db.remplacer_donnees("F1_Mouvements", report)
    # F2 : solde initial = 0 (l'ouverture est désormais la ligne de report en F1).
    bm2 = _brut_map("F2_Comptes")
    maj_f2 = [{bm2["Nom du compte / caisse"]: c["nom"], bm2["Solde initial (KMF)"]: 0}
              for c in ap["comptes"]]
    if maj_f2:
        _db.ecrire_lignes_lot("F2_Comptes", maj_f2,
                              cles=[bm2["Nom du compte / caisse"]])
    recap = {"annee": annee, "ref_archive": ref if lignes_arch else "(aucun mouvement)",
             "nb_archive": nb_arch, "nb_conserves": len(conserves),
             "total_recettes_fmt": ap["total_recettes_fmt"],
             "total_depenses_fmt": ap["total_depenses_fmt"],
             "solde_cloture_fmt": ap["solde_cloture_fmt"], "comptes": ap["comptes"]}
    return True, ("Exercice %s cloture (%d mouvement(s) archive(s)%s)."
                  % (annee, nb_arch,
                     "" if not conserves else ", %d ecriture(s) %s+ conservee(s)"
                     % (len(conserves), int(annee) + 1))), recap


def cloture_apercu():
    """Données pour l'écran Clôture & archivage : à journaliser, archivables, exercices
    compta disponibles, et le libellé d'année scolaire courant."""
    return {
        "annee_scolaire": annee_scolaire_label(),
        "annee_acad_courante": _annee_acad_defaut(),
        "annee_suivante": _annee_suivante_label(_annee_acad_defaut()) or "",
        "eleves_a_journaliser": eleves_sortis_a_journaliser(),
        "eleves_archivables": eleves_archivables(),
        "mentions": list(config.MENTIONS),
        "annee_civile": datetime.date.today().year,
    }


# --- PV de clôture / passation : Word (.docx sans dépendance) + page imprimable ---
import zipfile as _zipf
import html as _html


def _docx_esc(s):
    return _html.escape(str(s), quote=True)


def generer_docx(chemin, blocs):
    """Génère un .docx minimal et valide SANS dépendance externe (zipfile + XML brut),
    police Calibri, titres #1F4E79 (charte). `blocs` = liste de (type, texte) avec
    type in {titre, soustitre, gras, para, sign}. 100% hors-ligne (aucun moteur tiers)."""
    corps = []
    for typ, txt in blocs:
        t = _docx_esc(txt)
        if typ == "titre":
            corps.append('<w:p><w:pPr><w:spacing w:after="120"/><w:jc w:val="center"/></w:pPr>'
                         '<w:r><w:rPr><w:b/><w:color w:val="1F4E79"/><w:sz w:val="30"/></w:rPr>'
                         '<w:t xml:space="preserve">%s</w:t></w:r></w:p>' % t)
        elif typ == "soustitre":
            corps.append('<w:p><w:pPr><w:spacing w:after="160"/><w:jc w:val="center"/></w:pPr>'
                         '<w:r><w:rPr><w:color w:val="1F4E79"/><w:sz w:val="26"/></w:rPr>'
                         '<w:t xml:space="preserve">%s</w:t></w:r></w:p>' % t)
        elif typ == "gras":
            corps.append('<w:p><w:pPr><w:spacing w:before="160" w:after="60"/></w:pPr>'
                         '<w:r><w:rPr><w:b/><w:color w:val="1F4E79"/></w:rPr>'
                         '<w:t xml:space="preserve">%s</w:t></w:r></w:p>' % t)
        elif typ == "sign":
            corps.append('<w:p><w:pPr><w:spacing w:before="600"/></w:pPr>'
                         '<w:r><w:t xml:space="preserve">%s</w:t></w:r></w:p>' % t)
        else:
            corps.append('<w:p><w:pPr><w:spacing w:after="60"/></w:pPr>'
                         '<w:r><w:t xml:space="preserve">%s</w:t></w:r></w:p>' % t)
    document = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                '<w:body>%s<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
                '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440"/>'
                '</w:sectPr></w:body></w:document>' % "".join(corps))
    styles = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
              '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
              '<w:docDefaults><w:rPrDefault><w:rPr><w:rFonts w:ascii="Calibri" w:hAnsi="Calibri"/>'
              '<w:sz w:val="22"/></w:rPr></w:rPrDefault></w:docDefaults></w:styles>')
    ct = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
          '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
          '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
          '<Default Extension="xml" ContentType="application/xml"/>'
          '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
          '<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
          '</Types>')
    rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            '</Relationships>')
    drels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
             '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
             '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
             '</Relationships>')
    if os.path.dirname(chemin):
        os.makedirs(os.path.dirname(chemin), exist_ok=True)
    with _zipf.ZipFile(chemin, "w", _zipf.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", ct)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/document.xml", document)
        z.writestr("word/styles.xml", styles)
        z.writestr("word/_rels/document.xml.rels", drels)
    return chemin


def pv_blocs(pv):
    """Construit les blocs du PV (Word ET page imprimable) depuis le payload de clôture."""
    b = [("titre", "UNIVERSITE DES COMORES — ECOLE DE MEDECINE ET DE SANTE PUBLIQUE")]
    kind = pv.get("kind")
    if kind == "eleves":
        b.append(("soustitre", "Proces-verbal de cloture — Eleves (annee scolaire %s)" % pv.get("annee", "")))
        b.append(("gras", "Recapitulatif"))
        b.append(("para", "Nombre d'eleves journalises : %d" % pv.get("n", 0)))
        for e in pv.get("eleves", []):
            dip = (" — %s" % e["diplome"]) if e.get("diplome") else ""
            men = (" (mention %s)" % e["mention"]) if e.get("mention") else ""
            b.append(("para", "%s %s %s — %s — statut : %s%s%s"
                      % (e.get("matricule", ""), e.get("nom", ""), e.get("prenom", ""),
                         e.get("filiere", ""), e.get("statut", ""), dip, men)))
    elif kind == "archivage":
        b.append(("soustitre", "Proces-verbal d'archivage des eleves"))
        b.append(("gras", "Cohortes archivees"))
        b.append(("para", "Nombre total d'eleves archives : %d" % pv.get("n", 0)))
        for c in pv.get("cohortes", []):
            b.append(("para", "Annee %s : %d eleve(s) -> fichier %s"
                      % (c.get("label", ""), c.get("nb", 0), c.get("fichier", ""))))
    elif kind == "compta":
        b.append(("soustitre", "Proces-verbal de cloture — Exercice comptable %s" % pv.get("annee", "")))
        b.append(("gras", "Recapitulatif de l'exercice"))
        b.append(("para", "Total des recettes : %s KMF" % pv.get("total_recettes_fmt", "0")))
        b.append(("para", "Total des depenses : %s KMF" % pv.get("total_depenses_fmt", "0")))
        b.append(("para", "Solde de cloture (reporte a nouveau) : %s KMF" % pv.get("solde_cloture_fmt", "0")))
        b.append(("para", "Fichier d'archive des mouvements : %s" % pv.get("ref_archive", "")))
        b.append(("gras", "Soldes reportes par compte"))
        for c in pv.get("comptes", []):
            b.append(("para", "%s : %s KMF" % (c.get("nom", ""), c.get("solde_fmt", ""))))
    b.append(("para", "Fait a Moroni, le %s." % pv.get("date", "")))
    b.append(("para", "Cloture realisee par : %s" % pv.get("login", "")))
    b.append(("sign", "Le Directeur :                                              Le Gestionnaire :"))
    return b


# ===========================================================================
# EDITIONS COMPLEMENTAIRES (V1.46) — aucune modification du classeur.
# Rendu via imprimer.html (kinds : presence_semaine, table). Doc construit en
# Python (doc_adhoc) : pas besoin d'entree D1_Modeles_docs.
# ===========================================================================
_ENTETE_STD = ("UNION DES COMORES\n"
               "Unite - Solidarite - Developpement\n"
               "Ministere de l'Education Nationale, de l'Enseignement Superieur et de la Recherche\n"
               "Universite des Comores\n"
               "Ecole de Medecine et de Sante Publique")


def doc_adhoc(titre, corps="", mentions="", signataire="", nb_copies=1, orientation="portrait"):
    """Document d'impression construit en Python, meme rendu que les modeles D1 :
    entete standard + titre/corps/mentions/signataire/nb_copies (+ orientation)."""
    try:
        nbc = max(1, min(int(nb_copies), 5))
    except Exception:
        nbc = 1
    return {"cle": "", "libelle": titre, "entete": _ENTETE_STD, "titre": titre,
            "corps": corps, "mentions": mentions, "signataire": signataire,
            "nb_copies": nbc, "orientation": orientation, "jetons": {}, "tabulaire": True}


# --- Feuille de presence HEBDOMADAIRE (document A) : paysage, jours x creneaux ---
_PRESENCE_JOURS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
_PRESENCE_CRENEAUX = ["10h", "12h", "15h", "17h"]


def feuille_presence_semaine(filiere="", niveau="", section="", annee=""):
    """Feuille de presence de la semaine (modele EMSP) : etudiants filtres
    PRE-REMPLIS (N, matricule, nom, prenom, date et lieu de naissance) + grille
    jours x creneaux (10h/12h/15h/17h) a cocher. Format paysage. Si aucun
    etudiant ne correspond, on produit des lignes vierges (feuille a remplir)."""
    rows = _lignes_dict("A1_Etudiants")

    def garde(r):
        return ((not filiere or str(r.get("Filiere", "")).strip() == filiere)
                and (not niveau or str(r.get("Niveau", "")).strip() == niveau)
                and (not section or str(r.get("Section", "")).strip() == section)
                and (not annee or str(r.get("Annee acad.", "")).strip() == annee))

    sel = [r for r in rows if garde(r)]
    lignes = []
    for i, r in enumerate(sel, start=1):
        lignes.append([str(i), r.get("Matricule", ""), r.get("Nom", ""),
                       r.get("Prenom", ""), r.get("Date naissance", ""),
                       r.get("Lieu naissance", "")])
    if not lignes:                       # feuille vierge a remplir a la main
        lignes = [[str(i), "", "", "", "", ""] for i in range(1, 26)]
    return {
        "fixes": ["N", "Matricule", "Nom", "Prenom", "Date de naissance", "Lieu de naissance"],
        "jours": _PRESENCE_JOURS, "creneaux": _PRESENCE_CRENEAUX, "lignes": lignes,
        "contexte": {"departement": filiere or "..............................",
                     "niveau": niveau or "..............................",
                     "section": section or "..................",
                     "annee": annee or "..................",
                     "effectif": len([x for x in lignes if x[1]]),
                     "date_jour": fmt_date(_dt.date.today())},
    }


# --- Feuille de presence VIERGE contextualisee (R3, V1.99.37) : seance ponctuelle
#     (date + horaire libre), roster de la classe pre-rempli, emargement vide ------
def feuille_presence_vierge(filiere, niveau, section="", date_fr="", debut="",
                            fin="", matiere="", enseignant="", n_defaut=25):
    """Feuille de presence d'une seance ponctuelle, PRE-REMPLIE avec les etudiants
    de la classe (Filiere + Niveau [+ Section]). Colonne emargement vide. Si aucun
    etudiant ne correspond -> lignes numerotees vierges (repli). Mentionne les
    horaires (debut - fin). Filiere et Niveau sont requis (controle en amont)."""
    rows = _lignes_dict("A1_Etudiants")

    def garde(r):
        return ((str(r.get("Filiere", "")).strip() == filiere)
                and (str(r.get("Niveau", "")).strip() == niveau)
                and (not section or str(r.get("Section", "")).strip() == section))

    sel = [r for r in rows if garde(r)]
    lignes = []
    for i, r in enumerate(sel, start=1):
        nom = str(r.get("Nom", "")).strip()
        prenom = str(r.get("Prenom", "")).strip()
        lignes.append([str(i), str(r.get("Matricule", "")).strip(),
                       (nom + " " + prenom).strip()])
    vierge = not lignes
    if vierge:                       # repli : grille a remplir a la main
        try:
            nb = min(max(int(n_defaut), 1), 200)
        except (TypeError, ValueError):
            nb = 25
        lignes = [[str(i), "", ""] for i in range(1, nb + 1)]

    horaire = (debut + " – " + fin) if (debut and fin) else (debut or fin or "—")
    classe = " ".join(x for x in [filiere, niveau, section] if x)
    return {
        "contexte": {
            "classe": classe,
            "filiere": filiere, "niveau": niveau, "section": section or "—",
            "date": date_fr or fmt_date(_dt.date.today()),
            "horaire": horaire,
            "matiere": matiere or "—",
            "enseignant": enseignant or "—",
            "effectif": len(sel),
        },
        "lignes": lignes,
        "vierge": vierge,
    }


# --- Plan d'action (G1) : edition tableau --------------------------------------
def plan_action_liste():
    """Plan d'action — suivi des ecarts et axes de progres (G1) en tableau,
    en ordre logique : axe, constat, objectif, action, responsable, priorite,
    temporalite, echeance, indicateur, etat d'avancement, observations, type."""
    rows = _lignes_dict("G1_Plan_action")
    colonnes = [
        ("N", "N"),
        ("Axe / thème", "Axe / thème"),
        ("Domaine / module", "Domaine / module"),
        ("Constat / écart", "Ecart constate"),
        ("Objectif (résultat attendu)", "Objectif (résultat attendu)"),
        ("Action corrective", "Action corrective"),
        ("Responsable", "Responsable"),
        ("Priorité", "Priorité"),
        ("Temporalité", "Temporalité"),
        ("Échéance", "Echeance"),
        ("Indicateur de réussite et preuves", "Indicateur de réussite et preuves"),
        ("État d'avancement", "Statut"),
        ("Observations", "Observations"),
        ("Type d'écart", "Type d'écart"),
    ]
    cols = [c[0] for c in colonnes]
    keys = [c[1] for c in colonnes]
    lignes = []
    for i, r in enumerate(rows, start=1):
        prem = str(r.get("N", "")).strip() or str(i)
        lignes.append([prem] + [r.get(k, "") for k in keys[1:]])
    return {"colonnes": cols, "lignes": lignes,
            "contexte": {"legende": "Suivi des ecarts et axes de progres — %d action(s) · Edite le %s"
                         % (len(lignes), fmt_date(_dt.date.today()))}}


def plan_action_kpis():
    """Tableau de bord du plan d'action (lecture G1) : volumes par etat
    d'avancement, par priorite, par axe ; actions en retard ; % achevees."""
    rows = _lignes_dict("G1_Plan_action")
    rows = [r for r in rows if any(str(r.get(k, "")).strip() for k in
            ("Axe / thème", "Domaine / module", "Ecart constate", "Action corrective"))]
    total = len(rows)
    par_etat, par_prio, par_axe = {}, {}, {}
    acheve_mots = ("achev", "atteint", "termin", "clotur", "réalis", "realis", "fait")
    nb_acheve = nb_retard = 0
    auj = _dt.date.today()
    for r in rows:
        etat = str(r.get("Statut", "")).strip() or "(non renseigné)"
        par_etat[etat] = par_etat.get(etat, 0) + 1
        est_acheve = any(m in etat.lower() for m in acheve_mots)
        if est_acheve:
            nb_acheve += 1
        prio = str(r.get("Priorité", "")).strip() or "(non renseignée)"
        par_prio[prio] = par_prio.get(prio, 0) + 1
        axe = str(r.get("Axe / thème", "")).strip() or str(r.get("Domaine / module", "")).strip() or "(non renseigné)"
        par_axe[axe] = par_axe.get(axe, 0) + 1
        ech = _parse_date_fr(r.get("Echeance", ""))
        if ech is not None and ech < auj and not est_acheve:
            nb_retard += 1
    def _tri(d):
        return sorted(d.items(), key=lambda kv: (-kv[1], kv[0]))
    pct = round(100.0 * nb_acheve / total) if total else 0
    return {"total": total, "acheve": nb_acheve, "pct_acheve": pct, "retard": nb_retard,
            "par_etat": _tri(par_etat), "par_prio": _tri(par_prio), "par_axe": _tri(par_axe)}


# --- Etats comptables complementaires ------------------------------------------
def journal_treso(compte="", mois=""):
    """Journal de tresorerie : toutes les operations (recettes ET depenses) d'une
    periode MM/AAAA, pour un compte donne ou TOUS les comptes, avec totaux."""
    compte = str(compte).strip()
    mois = str(mois).strip()
    debut = fin = None
    try:
        m, y = mois.split("/")
        debut = _dt.date(int(y), int(m), 1)
        fin = _dernier_jour(int(y), int(m))
    except Exception:
        pass
    sel = []
    for r in _lignes_dict("F1_Mouvements"):
        if compte and str(r.get("Compte / caisse", "")).strip() != compte:
            continue
        d = _parse_date_fr(r.get("Date operation", ""))
        if debut and (d is None or d < debut or d > fin):
            continue
        sel.append((d or _dt.date.min, r))
    sel.sort(key=lambda t: (t[0], str(t[1].get("Reference / N piece", ""))))
    cols = ["Date", "N piece", "Compte / caisse", "Description", "Beneficiaire",
            "Recette", "Depense"]
    lignes, tr, td = [], 0.0, 0.0
    for _d, r in sel:
        rec = _num(r.get("Montant Recette (KMF)", 0))
        dep = _num(r.get("Montant Depense (KMF)", 0))
        tr += rec
        td += dep
        lignes.append([r.get("Date operation", ""), r.get("Reference / N piece", ""),
                       r.get("Compte / caisse", ""), r.get("Libelle / description", ""),
                       r.get("Tiers", ""), _kmf_aff(rec) if rec else "",
                       _kmf_aff(dep) if dep else ""])
    total = ["", "", "", "", "TOTAL", _kmf_aff(tr), _kmf_aff(td)]
    return {"colonnes": cols, "lignes": lignes, "total": total,
            "contexte": {"legende": "Compte : %s · Periode : %s · %d operation(s) · Edite le %s"
                         % (compte or "Tous les comptes",
                            _periode_libelle(mois) if mois else "Toutes periodes",
                            len(lignes), fmt_date(_dt.date.today()))}}


def balance_comptes():
    """Situation globale (balance) : un compte/caisse par ligne — solde initial,
    total recettes, total depenses, solde courant — + ligne TOTAL."""
    init_map = {}
    for r in _lignes_dict("F2_Comptes"):
        init_map[str(r.get("Nom du compte / caisse", "")).strip()] = _num(r.get("Solde initial (KMF)", 0))
    agg = {}
    for r in _lignes_dict("F1_Mouvements"):
        c = str(r.get("Compte / caisse", "")).strip()
        if not c:
            continue
        a = agg.setdefault(c, [0.0, 0.0])
        a[0] += _num(r.get("Montant Recette (KMF)", 0))
        a[1] += _num(r.get("Montant Depense (KMF)", 0))
    noms = [c["nom"] for c in comptes_treso()] or sorted(set(list(init_map) + list(agg)))
    cols = ["Compte / caisse", "Solde initial", "Total recettes", "Total depenses", "Solde courant"]
    lignes, ti, tr, td, ts = [], 0.0, 0.0, 0.0, 0.0
    for nom in noms:
        ini = init_map.get(nom, 0.0)
        rec, dep = agg.get(nom, [0.0, 0.0])
        sc = ini + rec - dep
        ti += ini; tr += rec; td += dep; ts += sc
        lignes.append([nom, _kmf_aff(ini), _kmf_aff(rec), _kmf_aff(dep), _kmf_aff(sc)])
    total = ["TOTAL", _kmf_aff(ti), _kmf_aff(tr), _kmf_aff(td), _kmf_aff(ts)]
    return {"colonnes": cols, "lignes": lignes, "total": total,
            "contexte": {"legende": "Tous les comptes et caisses · %d compte(s) · Edite le %s"
                         % (len(lignes), fmt_date(_dt.date.today()))}}


def etat_par_poste(mois="", compte=""):
    """Etat des recettes et depenses par POSTE BUDGETAIRE (regroupement comptable),
    pour une periode MM/AAAA (option) et un compte (option). Avec ligne TOTAL."""
    mois = str(mois).strip()
    compte = str(compte).strip()
    debut = fin = None
    try:
        m, y = mois.split("/")
        debut = _dt.date(int(y), int(m), 1)
        fin = _dernier_jour(int(y), int(m))
    except Exception:
        pass
    agg = {}
    ordre = []
    for r in _lignes_dict("F1_Mouvements"):
        if compte and str(r.get("Compte / caisse", "")).strip() != compte:
            continue
        if debut:
            d = _parse_date_fr(r.get("Date operation", ""))
            if d is None or d < debut or d > fin:
                continue
        poste = str(r.get("Poste budgetaire", "")).strip() or "(non precise)"
        if poste not in agg:
            agg[poste] = [0.0, 0.0]
            ordre.append(poste)
        agg[poste][0] += _num(r.get("Montant Recette (KMF)", 0))
        agg[poste][1] += _num(r.get("Montant Depense (KMF)", 0))
    cols = ["Poste budgetaire", "Recettes", "Depenses", "Solde (R - D)"]
    lignes, tr, td = [], 0.0, 0.0
    for poste in sorted(ordre):
        rec, dep = agg[poste]
        tr += rec
        td += dep
        lignes.append([poste, _kmf_aff(rec), _kmf_aff(dep), _kmf_aff(rec - dep)])
    total = ["TOTAL", _kmf_aff(tr), _kmf_aff(td), _kmf_aff(tr - td)]
    return {"colonnes": cols, "lignes": lignes, "total": total,
            "contexte": {"legende": "Compte : %s · Periode : %s · %d poste(s) · Edite le %s"
                         % (compte or "Tous les comptes",
                            _periode_libelle(mois) if mois else "Toutes periodes",
                            len(lignes), fmt_date(_dt.date.today()))}}


# ===========================================================================
# V1.70 — BUDGET PAR POSTE (prevu / realise / ecart) + TAUX DE CHANGE
# ---------------------------------------------------------------------------
# F3_Budget_poste porte le PREVU (par poste, par exercice = annee civile). Le
# REALISE est agrege depuis F1_Mouvements par poste, filtre sur l'annee civile.
# P2_Taux porte les taux de reference (EUR = parite fixe 491,967). Calculs en
# Python (openpyxl ne recalcule pas les formules). Lecture seule pour ces vues.
# ===========================================================================
def taux_change(devise):
    """Taux de change en KMF d'une devise depuis P2_Taux (ex. taux_change('EUR')
    = 491.967). Renvoie un float, ou None si la devise/onglet est absent."""
    dev = str(devise or "").strip().upper()
    if not dev or "P2_Taux" not in _db.onglets():
        return None
    for r in _lignes_dict("P2_Taux"):
        if str(r.get("Devise", "")).strip().upper() == dev:
            t = r.get("Taux en KMF", "")
            return _num(t) if str(t).strip() != "" else None
    return None


def budget_par_poste(exercice=""):
    """Montant budgete (prevu) agrege par poste pour un exercice (annee civile).
    Renvoie {poste: montant}. Exercice vide = toutes lignes confondues."""
    exercice = str(exercice).strip()
    out = {}
    if "F3_Budget_poste" not in _db.onglets():
        return out
    for r in _lignes_dict("F3_Budget_poste"):
        poste = str(r.get("Poste budgetaire", "")).strip()
        if not poste:
            continue
        ex = str(r.get("Exercice", "")).strip()
        if exercice and ex and ex != exercice:
            continue
        out[poste] = out.get(poste, 0.0) + _num(r.get("Montant budgete (KMF)", 0))
    return out


def _realise_par_poste(exercice=""):
    """Realise (recettes + depenses) agrege par poste depuis F1, filtre annee
    civile (meme logique de filtre que apercu_cloture_compta). {poste: montant}."""
    exercice = str(exercice).strip()
    out = {}
    for r in _lignes_dict("F1_Mouvements"):
        if exercice:
            d = str(r.get("Date operation", "")).strip()
            if ("/%s" % exercice) not in d and not d.endswith(exercice):
                continue
        poste = str(r.get("Poste budgetaire", "")).strip()
        if not poste:
            continue
        out[poste] = out.get(poste, 0.0) + _num(r.get("Montant Recette (KMF)", 0)) \
            + _num(r.get("Montant Depense (KMF)", 0))
    return out


def _sens_par_poste(exercice=""):
    """Sens (Recette/Depense) declare au budget pour chaque poste (1er rencontre)."""
    exercice = str(exercice).strip()
    out = {}
    if "F3_Budget_poste" not in _db.onglets():
        return out
    for r in _lignes_dict("F3_Budget_poste"):
        p = str(r.get("Poste budgetaire", "")).strip()
        if not p or p in out:
            continue
        ex = str(r.get("Exercice", "")).strip()
        if exercice and ex and ex != exercice:
            continue
        s = str(r.get("Sens", "")).strip()
        if s:
            out[p] = s
    return out


def etat_poste_budget(exercice):
    """Comparatif PREVU / REALISE / ECART + taux de realisation par poste, pour
    un exercice (annee civile). Prevu = F3 ; Realise = F1 (annee civile). Renvoie
    colonnes / lignes / total / contexte, prets pour l'edition (etat-poste)."""
    exercice = str(exercice).strip()
    prevu = budget_par_poste(exercice)
    realise = _realise_par_poste(exercice)
    sens = _sens_par_poste(exercice)
    postes = sorted(set(prevu) | set(realise))
    cols = ["Poste budgetaire", "Sens", "Prevu (KMF)", "Realise (KMF)",
            "Ecart (KMF)", "Taux real. %"]
    lignes, tp, tr = [], 0.0, 0.0
    for p in postes:
        pv = prevu.get(p, 0.0)
        rl = realise.get(p, 0.0)
        tp += pv
        tr += rl
        taux = ("%.0f" % (100.0 * rl / pv)) if pv else "—"
        lignes.append([p, sens.get(p, ""), _kmf_aff(pv), _kmf_aff(rl),
                       _kmf_aff(rl - pv), taux])
    total = ["TOTAL", "", _kmf_aff(tp), _kmf_aff(tr), _kmf_aff(tr - tp),
             ("%.0f" % (100.0 * tr / tp)) if tp else "—"]
    return {"colonnes": cols, "lignes": lignes, "total": total,
            "contexte": {"exercice": exercice,
                         "legende": "Budget par poste — exercice %s · %d poste(s) · Edite le %s"
                         % (exercice or "tous", len(lignes), fmt_date(_dt.date.today()))}}


# ===========================================================================
# C-5 (V1.99.13) — Synthese budgetaire : prevu (F5) / realise (F1) / ecart
# Prevu = F5 (budget previsionnel detaille). Realise = F1 (depenses), borne sur la
# PLAGE DE LA SESSION academique. Trois mailles : poste-code, bailleur, formation.
# La maille formation est en PREVU SEUL (F1 ne porte pas la formation -> non
# ventilable, coherent option B). Tout en Python (jamais de cellule formule).
# ===========================================================================
def _bornes_session(label):
    """Plage (debut, fin) d'une session academique 'AAAA-AAAA'. Annee demarrant a
    config.MOIS_DEBUT_ANNEE_ACAD (defaut octobre) : '2025-2026' -> 01/10/2025 ..
    30/09/2026. (None, None) si label illisible."""
    m = re.match(r"^\s*(\d{4})\s*-\s*(\d{4})\s*$", str(label or ""))
    if not m:
        return None, None
    a1 = int(m.group(1))
    mois = int(getattr(config, "MOIS_DEBUT_ANNEE_ACAD", 10))
    debut = _dt.date(a1, mois, 1)
    fin = _dt.date(a1 + 1, mois, 1) - _dt.timedelta(days=1)
    return debut, fin


def _f5_prevu_agrege(cle, session="", formation="", bailleur=""):
    """Prevu (F5) agrege par 'cle' ('poste'|'bailleur'|'formation') -> {valeur: KMF}.
    Filtres optionnels session / formation / bailleur."""
    champ = {"poste": "Poste budgetaire",
             "bailleur": "Source de financement / Bailleur",
             "formation": "Formation"}[cle]
    sC = str(session).strip().lower()
    fC = str(formation).strip().lower()
    bC = str(bailleur).strip().lower()
    out = {}
    for r in _lignes_dict(_F5_TAB):
        if sC and str(r.get("Session", "")).strip().lower() != sC:
            continue
        if fC and str(r.get("Formation", "")).strip().lower() != fC:
            continue
        if bC and str(r.get("Source de financement / Bailleur", "")).strip().lower() != bC:
            continue
        v = str(r.get(champ, "")).strip() or "(non renseigné)"
        m = _montant_ligne_kmf(r.get("Qte1"), r.get("Qte2"), r.get("Cout unitaire (KMF)"))
        out[v] = out.get(v, 0.0) + m
    return out


def _f1_realise_agrege(cle, debut=None, fin=None, bailleur=""):
    """Realise DEPENSE (F1) agrege par 'cle' ('poste'|'bailleur'), borne sur la plage
    [debut, fin] (Date operation). Seules les depenses comptent (prevu F5 = depenses).
    -> {valeur: KMF}."""
    champ = {"poste": "Poste budgetaire",
             "bailleur": "Source de financement / Bailleur"}[cle]
    bC = str(bailleur).strip().lower()
    out = {}
    for r in _lignes_dict("F1_Mouvements"):
        if debut or fin:
            d = _parse_date_fr(r.get("Date operation", ""))
            if d is None:
                continue
            if debut and d < debut:
                continue
            if fin and d > fin:
                continue
        if bC and str(r.get("Source de financement / Bailleur", "")).strip().lower() != bC:
            continue
        dep = _num(r.get("Montant Depense (KMF)", 0))
        if dep <= 0:
            continue
        v = str(r.get(champ, "")).strip() or "(non renseigné)"
        out[v] = out.get(v, 0.0) + dep
    return out


def synthese_budgetaire(session="", formation="", bailleur=""):
    """Synthese prevu (F5) / realise (F1) / ecart. Ecart = Prevu - Realise (positif =
    sous-consomme, negatif = depassement). Trois mailles ; formation = prevu seul.
    Realise borne sur la plage de la session (point 3 verrouille)."""
    debut, fin = _bornes_session(session)
    intitules = {str(r.get("Code", "")).strip(): str(r.get("Intitule", "")).strip()
                 for r in _lignes_dict("P3_Nomenclature")}

    def _table_ecart(prevu, realise):
        lignes, tp, tr = [], 0.0, 0.0
        for k in sorted(set(prevu) | set(realise)):
            pv = prevu.get(k, 0.0)
            rl = realise.get(k, 0.0)
            tp += pv
            tr += rl
            lignes.append({"cle": k, "prevu": pv, "realise": rl, "ecart": pv - rl,
                           "pct": (round(100.0 * rl / pv) if pv else None)})
        return {"lignes": lignes, "total_prevu": tp, "total_realise": tr,
                "total_ecart": tp - tr, "pct": (round(100.0 * tr / tp) if tp else None)}

    par_poste = _table_ecart(_f5_prevu_agrege("poste", session, formation, bailleur),
                             _f1_realise_agrege("poste", debut, fin, bailleur))
    for l in par_poste["lignes"]:
        l["intitule"] = intitules.get(l["cle"], "")

    par_bailleur = _table_ecart(_f5_prevu_agrege("bailleur", session, formation, bailleur),
                                _f1_realise_agrege("bailleur", debut, fin, bailleur))

    f_prev = _f5_prevu_agrege("formation", session, formation, bailleur)
    par_formation = {
        "lignes": sorted(({"cle": k, "prevu": v} for k, v in f_prev.items()),
                         key=lambda x: -x["prevu"]),
        "total_prevu": sum(f_prev.values()),
    }

    return {"session": session, "formation": formation, "bailleur": bailleur,
            "debut": fmt_date(debut) if debut else "", "fin": fmt_date(fin) if fin else "",
            "par_poste": par_poste, "par_bailleur": par_bailleur,
            "par_formation": par_formation}


def kpis_budget(session=""):
    """Indicateurs BUDGET du tableau de bord direction (#20, brique C).
    Bornes sur UNE session (defaut = annee academique courante) ; non filtrables
    par filiere/niveau. Source : synthese_budgetaire (maille poste). Ecart = prevu
    - realise (negatif = depassement). Renvoie les 5 KPI + les donnees du graphe
    Prevu/Realise par poste."""
    sess = session or _annee_acad_defaut()
    pp = synthese_budgetaire(sess)["par_poste"]
    lignes = pp["lignes"]
    nb_dep = len([l for l in lignes if l["ecart"] < 0])
    postes = [{"libelle": (l.get("intitule") or l["cle"]),
               "prevu": l["prevu"], "realise": l["realise"]}
              for l in lignes if (l["prevu"] or l["realise"])]
    return {
        "session": sess,
        "bud_prevu": pp["total_prevu"],
        "bud_realise": pp["total_realise"],
        "bud_taux": pp["pct"] if pp["pct"] is not None else 0,
        "bud_ecart": pp["total_ecart"],
        "bud_depassement": nb_dep,
        "postes": postes,
    }


def sessions_budget():
    """Sessions distinctes presentes dans F5_Budget_Prev (pour les filtres)."""
    vus = []
    for r in _lignes_dict(_F5_TAB):
        s = str(r.get("Session", "")).strip()
        if s and s not in vus:
            vus.append(s)
    return sorted(vus)


# === Bulletin officiel (mise en page identique au RELEVE_NOTES.pdf) ============
# Reutilise releve_semestre/releve_annuel pour les MOYENNES (calcul decret 05-106
# inchange) et lit N2 pour exposer C.Continu / Examen / session 2 par matiere.
def _fmt_note(v):
    """Note d'affichage facon bulletin : '' si vide, sinon nombre a virgule FR."""
    s = str(v).strip()
    if s == "":
        return ""
    n = _arrondi2(_num_h(v))
    if n is None:
        return ""
    if abs(n - round(n)) < 1e-9:
        return str(int(round(n)))
    return ("%.2f" % n).rstrip("0").rstrip(".").replace(".", ",")


def _notes_brutes(matricule, annee, semestre):
    """{(num_ue, matiere): {cc1, ex1, moy1, ex2, moy2}} pour l'affichage du bulletin."""
    iMat, iAn, iSess, iSem, iU, iM, iCC, iEx = _idx_map(_db_notes, "N2_Notes",
        ["Matricule", "Annee acad.", "Session", "Semestre", "N° UE", "Matiere", "CC", "Examen"])
    cct = _cc_table(matricule, annee)
    sem = str(semestre).strip()
    parcle = {}
    for r in _db_notes.lignes("N2_Notes"):
        if str(r[iMat]).strip() != str(matricule).strip():
            continue
        if _n(r[iAn]) != _n(annee) or str(r[iSem]).strip() != sem:
            continue
        cle = (str(r[iU]).strip(), str(r[iM]).strip())
        sess = str(r[iSess]).strip() or "1"
        parcle.setdefault(cle, {})[sess] = (r[iCC], r[iEx])
    for (s, sm, ue, mat) in cct:
        if sm == sem:
            parcle.setdefault((ue, mat), {}).setdefault(s, ("", ""))
    out = {}
    for cle, ss in parcle.items():
        cc1, ex1 = ss.get("1", ("", ""))
        cc1d = cct.get(("1", sem, cle[0], cle[1]))
        if cc1d is not None:
            cc1 = cc1d
        d = {"cc1": _fmt_note(cc1), "ex1": _fmt_note(ex1),
             "moy1": _fmt_note(_moyenne_matiere(cc1, ex1)), "ex2": "", "moy2": ""}
        if "2" in ss:
            ex2 = ss["2"][1]
            cc2 = _cc_session(cct, ss, "2", sem, cle[0], cle[1])  # notes session 2 seules
            d["ex2"] = _fmt_note(ex2)
            d["moy2"] = _fmt_note(_moyenne_matiere(cc2, ex2))
        out[cle] = d
    return out


_ORDINAL_SEM = {"1": "Premier", "2": "Deuxieme", "3": "Troisieme",
                "4": "Quatrieme", "5": "Cinquieme", "6": "Sixieme"}


def bulletin_officiel(matricule, annee, session=""):
    """Releve de notes ANNUEL au format officiel : un bloc par semestre
    (matieres avec C.Continu / Examen / Moyenne / session 2), moyenne de semestre,
    decision et mention, puis moyenne annuelle. Calculs = moteur decret 05-106.
    V1.99.50 — edition par session (modele du releve officiel scanne) :
      session='1' -> page(s) PREMIERE SESSION : notes de session 1 uniquement,
                     pas de colonnes session 2 ;
      session='2' ou '' -> DEUXIEME SESSION si des notes de rattrapage existent
                     (colonnes Exam. sess. 2 / Moy. sess. 2), sinon 1re session."""
    session = str(session or "").strip()
    session_max = "1" if session == "1" else None
    an = releve_annuel(matricule, annee, session_max)
    sems = []
    for rel in an["semestres"]:
        brut = _notes_brutes(matricule, annee, rel["semestre"])
        a_s2 = False
        ues = []
        for ue in rel["ues"]:
            mats = []
            for m in ue["matieres"]:
                b = brut.get((ue["num"], m["matiere"]), {})
                if session_max == "1":
                    b = dict(b); b["ex2"] = ""; b["moy2"] = ""   # page session 1
                if b.get("moy2"):
                    a_s2 = True
                mats.append({"matiere": m["matiere"], "cc": b.get("cc1", ""),
                             "ex": b.get("ex1", ""), "moy": b.get("moy1", _fmt_note(m["moyenne"])),
                             "ex2": b.get("ex2", ""), "moy2": b.get("moy2", "")})
            ues.append({"num": ue["num"], "intitule": ue["intitule"], "coef": ue["coef"],
                        "ects": ue["ects"], "moyenne": _fmt_note(ue["moyenne"]),
                        "validee": ue["validee"], "matieres": mats})
        sems.append({"semestre": rel["semestre"],
                     "ordinal": _ORDINAL_SEM.get(str(rel["semestre"]), str(rel["semestre"])),
                     "ues": ues, "a_session2": a_s2, "moyenne": _fmt_note(rel["moyenne"]),
                     "session_libelle": ("Deuxieme session" if a_s2 else "Premiere session"),
                     "mention": rel["mention"], "proposition": rel["proposition"],
                     "ects_acquis": rel["ects_acquis"], "ects_total": rel["ects_total"]})
    return {"matricule": matricule, "nom": an["nom"], "filiere": an["filiere"],
            "niveau": an["niveau"], "annee": annee, "semestres": sems,
            "session_edition": session,
            "moyenne": _fmt_note(an["moyenne"]), "mention": an["mention"],
            "a_non_dispensee": an.get("a_non_dispensee", False),
            "incomplet": an.get("incomplet", False),
            "ects_acquis": an.get("ects_acquis"), "ects_requis": an.get("ects_requis"),
            "ects_ecart": an.get("ects_ecart"), "decision": an.get("decision"),
            "proposition": an["proposition"]}


# ===========================================================================
# V1.71 — FILTRAGE MULTICRITERE + IMPRESSION DE SELECTION (coeur logique)
# Criteres standard : filiere / niveau / annee academique / periode (du-au).
# Un onglet ne reagit a un critere que s'il figure dans la map config
# correspondante. A2_Presences et S1_Stages : filiere/niveau resolus via le
# Matricule (enrichissement depuis A1). Aucune ecriture classeur.
# ===========================================================================
def _col_propre(s):
    """Libelle de colonne sans marqueur (*)/(**) — aligne sur _lignes_dict."""
    return decoupe_provenance(str(s))[0]


def _matricule_index():
    """dict Matricule -> {filiere, niveau} construit depuis A1_Etudiants."""
    idx = {}
    for r in _lignes_dict("A1_Etudiants"):
        mat = str(r.get("Matricule", "")).strip()
        if mat:
            idx[mat] = {"filiere": str(r.get("Filiere", "")).strip(),
                        "niveau": str(r.get("Niveau", "")).strip()}
    return idx


def valeurs_filtres():
    """Valeurs distinctes des criteres, lues depuis P0 (commun a tout l'app).
    Generalise liste_filtres() : filiere / niveau / annee academique."""
    par = _db.listes_parametres()
    def liste(cle_propre):
        for k, vals in par.items():
            if _col_propre(k) == cle_propre:
                return sorted({str(v).strip() for v in vals if str(v).strip()})
        return []
    return {"filiere": liste("Filieres"),
            "niveau": liste("Niveaux"),
            "annee": liste("Annees_acad")}


def onglet_supporte(onglet):
    """Quels criteres l'onglet sait honorer (pour masquer les controles IHM)."""
    enr = onglet in config.ONGLETS_ENRICHIS_PAR_MATRICULE
    return {
        "filiere": onglet in config.COLONNE_FILIERE_PAR_ONGLET or enr,
        "niveau": onglet in config.COLONNE_NIVEAU_PAR_ONGLET or enr,
        "annee": onglet in config.COLONNE_ANNEE_PAR_ONGLET,
        "periode": onglet in config.COLONNE_DATE_PAR_ONGLET,
    }


def _filtres_actifs(f):
    return bool(f.get("filiere") or f.get("niveau") or f.get("annee")
                or f.get("du") or f.get("au"))


def _val_filiere_niveau(onglet, get, matidx):
    """Renvoie (filiere, niveau) d'une ligne via colonne propre ou Matricule."""
    cf = config.COLONNE_FILIERE_PAR_ONGLET.get(onglet)
    cn = config.COLONNE_NIVEAU_PAR_ONGLET.get(onglet)
    fil = get(cf) if cf else None
    niv = get(cn) if cn else None
    if (fil is None or niv is None) and matidx is not None:
        mat = str(get("Matricule") or "").strip()
        info = matidx.get(mat)
        if info:
            if fil is None:
                fil = info.get("filiere")
            if niv is None:
                niv = info.get("niveau")
    return fil, niv


def _ligne_passe(onglet, get, f, matidx):
    """get(libelle_propre)->valeur. True si la ligne satisfait les filtres.
    Un critere non supporte par l'onglet n'exclut jamais (onglet inerte)."""
    if f.get("filiere") or f.get("niveau"):
        fil, niv = _val_filiere_niveau(onglet, get, matidx)
        if f.get("filiere") and fil is not None and str(fil).strip() != f["filiere"]:
            return False
        if f.get("niveau") and niv is not None and str(niv).strip() != f["niveau"]:
            return False
    if f.get("annee"):
        ca = config.COLONNE_ANNEE_PAR_ONGLET.get(onglet)
        if ca is not None and str(get(ca) or "").strip() != f["annee"]:
            return False
    if f.get("du") or f.get("au"):
        cd = config.COLONNE_DATE_PAR_ONGLET.get(onglet)
        if cd is not None:
            d = _parse_date_fr(get(cd))
            if d is not None:
                du = _parse_date_fr(f.get("du"))
                au = _parse_date_fr(f.get("au"))
                if du and d < du:
                    return False
                if au and d > au:
                    return False
    return True


def _lignes_filtrees(onglet, filtres):
    """Lignes (dicts, cles propres) d'un onglet apres application des filtres."""
    f = filtres or {}
    rows = _lignes_dict(onglet)
    if not _filtres_actifs(f):
        return rows
    matidx = _matricule_index() if onglet in config.ONGLETS_ENRICHIS_PAR_MATRICULE else None
    return [r for r in rows if _ligne_passe(onglet, lambda c: r.get(c), f, matidx)]


def _bandeau_contexte(onglet, f):
    """Contexte d'affichage du bandeau de filtres (libelles + supports)."""
    return {
        "filiere": f.get("filiere") or "Toutes",
        "niveau": f.get("niveau") or "Tous",
        "annee": f.get("annee") or "Toutes",
        "du": f.get("du") or "", "au": f.get("au") or "",
        "actif": _filtres_actifs(f),
        "supporte": onglet_supporte(onglet),
        "date_jour": fmt_date(_dt.date.today()),
    }


def filtrer_table(onglet, filtres=None):
    """table(onglet) restreinte aux lignes satisfaisant les filtres + bandeau.
    Conserve le contrat de table() (titre/entetes/lignes/nb)."""
    f = filtres or {}
    data = table(onglet)
    data["bandeau"] = _bandeau_contexte(onglet, f)
    if not _filtres_actifs(f):
        return data
    libelles = [m["libelle"] for m in data["entetes"]]
    pos = {_col_propre(lib): i for i, lib in enumerate(libelles)}
    matidx = _matricule_index() if onglet in config.ONGLETS_ENRICHIS_PAR_MATRICULE else None
    def getter(lig):
        return lambda c: (lig[pos[_col_propre(c)]]
                          if c and _col_propre(c) in pos
                          and pos[_col_propre(c)] < len(lig) else None)
    gardees = [lig for lig in data["lignes"]
               if _ligne_passe(onglet, getter(lig), f, matidx)]
    data["lignes"] = gardees
    data["nb"] = len(gardees)
    return data


def export_selection(onglet, filtres=None):
    """Classeur .xlsx (bytes) de la selection courante (reutilise _xlsx_simple).
    La vue PDF paysage est rendue cote app via filtrer_table + template."""
    data = filtrer_table(onglet, filtres)
    colonnes = [m["libelle"] for m in data["entetes"]]
    titre = "%s — selection" % data.get("titre", onglet)
    return _xlsx_simple(titre, colonnes, data["lignes"])


# === Fiche enseignant (V1.73) — symetrique de la fiche etudiant ===========
# Lecture seule (le CRUD enseignant reste dans le module E1_Enseignants).
# Le lien seances/heures se fait par Matricule ens. (E2) et par le NOM porte
# dans le champ texte 'Enseignant' de A3 / 'Assure par' de E3 (matricule pris
# en charge aussi, des qu'il sera propose comme valeur de ce champ).

def _fnum(v):
    """Nombre tolerant ('12', '12,5', '12.5', '' -> 0.0)."""
    try:
        s = str(v).replace(",", ".").replace(" ", "").strip()
        return float(s) if s != "" else 0.0
    except Exception:
        return 0.0


def _fmt_h(x):
    """Heures sans decimale inutile (charte)."""
    try:
        x = float(x)
    except Exception:
        return "0"
    if x == int(x):
        return str(int(x))
    return ("%.2f" % x).rstrip("0").rstrip(".")


def _noms_possibles(nom, prenom, matricule):
    """Formes a comparer au champ texte 'Enseignant' / 'Assure par'
    (insensible casse/espaces) : 'Nom Prenom', 'Prenom Nom', 'Nom', matricule."""
    nom = (nom or "").strip()
    prenom = (prenom or "").strip()
    formes = set()
    if nom or prenom:
        formes.add((nom + " " + prenom).strip().lower())
        formes.add((prenom + " " + nom).strip().lower())
        if nom:
            formes.add(nom.lower())
    if matricule:
        formes.add(str(matricule).strip().lower())
    return {re.sub(r"\s+", " ", f) for f in formes if f.strip()}


def recherche_enseignants():
    """Liste compacte des enseignants pour recherche/autocompletion :
    [{matricule, nom, prenom, statut, departement, label}].
    label = 'matricule — Nom Prenom (Statut · Departement)'."""
    out = []
    for e in _lignes_dict("E1_Enseignants"):
        mat = str(e.get("Matricule ens.", "")).strip()
        nom = str(e.get("Nom", "")).strip()
        pre = str(e.get("Prenom", "")).strip()
        if not (mat or nom or pre):
            continue
        sta = str(e.get("Statut", "")).strip()
        dep = str(e.get("Departement", "")).strip()
        ctx = " · ".join(x for x in [sta, dep] if x)
        label = "%s — %s %s%s" % (mat, nom, pre, (" (" + ctx + ")") if ctx else "")
        out.append({"matricule": mat, "nom": nom, "prenom": pre,
                    "statut": sta, "departement": dep,
                    "label": label.strip().strip("—").strip()})
    out.sort(key=lambda x: (x["nom"].lower(), x["prenom"].lower(), x["matricule"]))
    return out


def fiche_enseignant(matricule):
    """Fiche complete d'un enseignant depuis E1 (tous champs propres), ou None."""
    mat = str(matricule).strip()
    if not mat:
        return None
    for e in _lignes_dict("E1_Enseignants"):
        if str(e.get("Matricule ens.", "")).strip().lower() == mat.lower():
            nom = str(e.get("Nom", "")).strip()
            pre = str(e.get("Prenom", "")).strip()
            d = {k: ("" if v is None else str(v).strip()) for k, v in e.items()}
            d["matricule"] = str(e.get("Matricule ens.", "")).strip()
            d["nom_complet"] = (nom + " " + pre).strip()
            d["a_photo"] = _photo_existe(d["matricule"])
            return d
    return None


def heures_enseignant(matricule):
    """Releve d'heures (E2) pour un matricule : lignes + totaux."""
    mat = str(matricule).strip().lower()
    lignes = []
    tot_p = tot_c = tot_a = 0.0
    for r in _lignes_dict("E2_Releve_heures"):
        if str(r.get("Matricule ens.", "")).strip().lower() != mat:
            continue
        p = _fnum(r.get("Vol. horaire prog.", ""))
        c = _fnum(r.get("Vol. horaire constate", ""))
        a = _fnum(r.get("Total heures a payer", ""))
        tot_p += p; tot_c += c; tot_a += a
        lignes.append({"mois": str(r.get("Mois / Annee", "")).strip(),
                       "prog": _fmt_h(p), "constate": _fmt_h(c), "a_payer": _fmt_h(a)})
    return {"lignes": lignes, "tot_prog": _fmt_h(tot_p),
            "tot_constate": _fmt_h(tot_c), "tot_a_payer": _fmt_h(tot_a)}


def seances_enseignant(nom, prenom, matricule):
    """Seances planifiees (A3) ou le champ Enseignant correspond a cet
    enseignant (Nom+Prenom, ou matricule a terme)."""
    formes = _noms_possibles(nom, prenom, matricule)
    out = []
    for s in _lignes_dict("A3_Sessions"):
        ens = re.sub(r"\s+", " ", str(s.get("Enseignant", "")).strip().lower())
        if ens and ens in formes:
            out.append({
                "annee": str(s.get("Annee acad.", "")).strip(),
                "filiere": str(s.get("Filiere", "")).strip(),
                "niveau": str(s.get("Niveau", "")).strip(),
                "matiere": str(s.get("Matiere", "")).strip(),
                "jour": str(s.get("Jour", "")).strip(),
                "debut": str(s.get("Heure debut", "")).strip(),
                "fin": str(s.get("Heure fin", "")).strip(),
                "salle": str(s.get("Salle", "")).strip(),
                "type": str(s.get("Type", "")).strip(),
                "vol": _fmt_h(_fnum(s.get("Vol. horaire prog.", ""))),
            })
    return out


def seances_faites_enseignant(nom, prenom, matricule):
    """Seances reellement assurees / exceptions (E3) ou 'Assure par'
    correspond a cet enseignant."""
    formes = _noms_possibles(nom, prenom, matricule)
    out = []
    for s in _lignes_dict("E3_Seances_faites"):
        ap = re.sub(r"\s+", " ", str(s.get("Assure par", "")).strip().lower())
        if ap and ap in formes:
            out.append({
                "date": str(s.get("Date", "")).strip(),
                "session": str(s.get("Session / Matiere", "")).strip(),
                "creneau": str(s.get("Creneau", "")).strip(),
                "etat": str(s.get("Etat", "")).strip(),
                "matiere": str(s.get("Matiere reelle", "")).strip(),
                "vol": _fmt_h(_fnum(s.get("Vol. constate h", ""))),
                "motif": str(s.get("Motif", "")).strip(),
            })
    return out


def enregistrer_photo_enseignant(matricule, donnees):
    """Comme enregistrer_photo mais valide le matricule contre E1 (enseignant).
    Reutilise les memes controles taille/format et le meme dossier photos."""
    import os as _os
    mat = str(matricule).strip()
    if not mat or fiche_enseignant(mat) is None:
        return False, "Matricule inconnu : aucune fiche enseignant correspondante."
    if not donnees:
        return False, "Aucun fichier recu."
    maxi = getattr(config, "PHOTO_MAX_OCTETS", 1024 * 1024)
    if len(donnees) > maxi:
        return False, "Image trop lourde (%d Mo maximum)." % max(1, maxi // (1024 * 1024))
    if _type_image(donnees) is None:
        return False, "Format non reconnu : seuls JPEG et PNG sont acceptes."
    p = chemin_photo(mat)
    if not p:
        return False, "Matricule invalide."
    try:
        _os.makedirs(_os.path.dirname(p), exist_ok=True)
        with open(p, "wb") as fh:
            fh.write(donnees)
    except Exception as exc:
        return False, "Echec d'ecriture de la photo : %s" % exc
    return True, "Photo enregistree."


# =====================================================================
# V1.75 — Editions logistiques & financieres (inventaire / besoins /
# etat par source de financement). Lecture seule, aucun ecrit classeur.
# Rendu generique kind="table" (colonnes / lignes / total / contexte).
# =====================================================================

_M1_AXES = {
    "salle":    ("Salle / localisation", "Salle / localisation"),
    "bailleur": ("Source de financement / Bailleur", "Source de financement / bailleur"),
    "etat":     ("Etat", "Etat"),
}


def inventaire_equipements(axe="salle"):
    """Inventaire M1 regroupe par axe (salle | bailleur | etat), avec
    sous-totaux Quantite + Montant par groupe et total general.
    Colonnes : ID, Designation, Categorie, Salle, Etat, Bailleur, Quantite, Montant (KMF)."""
    axe = (axe or "salle").strip().lower()
    if axe not in _M1_AXES:
        axe = "salle"
    cle_groupe, libelle_axe = _M1_AXES[axe]
    cols = ["ID", "Designation", "Categorie", "Salle", "Etat", "Bailleur", "Quantite", "Montant (KMF)"]

    groupes = {}
    for r in _lignes_dict("M1_Equipements"):
        idv = str(r.get("ID equipement", "")).strip()
        des = str(r.get("Designation", "")).strip()
        if not idv and not des:
            continue
        g = str(r.get(cle_groupe, "")).strip() or "(non renseigne)"
        qte = _num(r.get("Quantite", 0)) or 0
        mtt = _num(r.get("Montant (KMF)", 0)) or 0
        groupes.setdefault(g, []).append({
            "id": idv, "des": des,
            "cat": str(r.get("Categorie", "")).strip(),
            "salle": str(r.get("Salle / localisation", "")).strip(),
            "etat": str(r.get("Etat", "")).strip(),
            "bailleur": str(r.get("Source de financement / Bailleur", "")).strip(),
            "qte": qte, "mtt": mtt,
        })

    lignes = []
    tot_q, tot_m, nb = 0.0, 0.0, 0
    for g in sorted(groupes):
        items = groupes[g]
        lignes.append(["— %s : %s —" % (libelle_axe.upper(), g), "", "", "", "", "", "", ""])
        sq, sm = 0.0, 0.0
        for it in sorted(items, key=lambda x: (x["des"].lower(), x["id"])):
            lignes.append([it["id"], it["des"], it["cat"], it["salle"], it["etat"],
                           it["bailleur"], _kmf_aff(it["qte"]) if it["qte"] else "",
                           _kmf_aff(it["mtt"]) if it["mtt"] else ""])
            sq += it["qte"]; sm += it["mtt"]; nb += 1
        lignes.append(["", "", "", "", "", "Sous-total %s" % g, _kmf_aff(sq), _kmf_aff(sm)])
        tot_q += sq; tot_m += sm
    total = ["TOTAL GENERAL", "", "", "", "", "", _kmf_aff(tot_q), _kmf_aff(tot_m)]
    leg = "Inventaire des equipements par %s · %d equipement(s) · %d groupe(s) · Edite le %s" % (
        libelle_axe, nb, len(groupes), fmt_date(_dt.date.today()))
    return {"colonnes": cols, "lignes": lignes, "total": total,
            "contexte": {"legende": leg}}


def salles_equip_dispo():
    vals = {str(r.get("Salle / localisation", "")).strip()
            for r in _lignes_dict("M1_Equipements")}
    return sorted(v for v in vals if v)


def bailleurs_equip_dispo():
    vals = {str(r.get("Source de financement / Bailleur", "")).strip()
            for r in _lignes_dict("M1_Equipements")}
    return sorted(v for v in vals if v)


def etats_equip_dispo():
    vals = {str(r.get("Etat", "")).strip() for r in _lignes_dict("M1_Equipements")}
    return sorted(v for v in vals if v)


def bon_de_besoin(statut="", priorite="", salle=""):
    """Etat des expressions de besoin (L3), filtrable Statut / Priorite / Salle.
    Colonnes : ID, Date, Type, Equipement, Libelle, Quantite, Salle, Priorite,
    Statut, Cout estime (KMF), Demandeur. Total = somme des couts estimes."""
    statut = (statut or "").strip().lower()
    priorite = (priorite or "").strip().lower()
    salle = (salle or "").strip().lower()
    cols = ["ID", "Date", "Type", "Equipement", "Libelle", "Quantite", "Salle",
            "Priorite", "Statut", "Cout estime (KMF)", "Demandeur"]
    lignes, tot, nb = [], 0.0, 0
    for r in _lignes_dict("L3_Besoins"):
        idv = str(r.get("ID besoin", "")).strip()
        lib = str(r.get("Libelle du besoin", "")).strip()
        if not idv and not lib:
            continue
        st = str(r.get("Statut", "")).strip()
        pr = str(r.get("Priorite", "")).strip()
        sa = str(r.get("Localisation / salle", "")).strip()
        if statut and st.lower() != statut:
            continue
        if priorite and pr.lower() != priorite:
            continue
        if salle and sa.lower() != salle:
            continue
        cout = _num(r.get("Cout estime (KMF)", 0)) or 0
        tot += cout; nb += 1
        lignes.append([idv, str(r.get("Date d'expression", "")).strip(),
                       str(r.get("Type de besoin", "")).strip(),
                       str(r.get("Equipement concerne", "")).strip(),
                       lib,
                       _kmf_aff(r.get("Quantite", "")) if str(r.get("Quantite", "")).strip() else "",
                       sa, pr, st,
                       _kmf_aff(cout) if cout else "",
                       str(r.get("Demandeur", "")).strip()])
    total = ["", "", "", "", "", "", "", "", "TOTAL", _kmf_aff(tot), ""]
    crit = " · ".join(filter(None, [
        ("Statut : %s" % statut.capitalize()) if statut else "",
        ("Priorite : %s" % priorite.capitalize()) if priorite else "",
        ("Salle : %s" % salle) if salle else ""])) or "Tous les besoins"
    leg = "Expression de besoin (L3) · %s · %d ligne(s) · Edite le %s" % (
        crit, nb, fmt_date(_dt.date.today()))
    return {"colonnes": cols, "lignes": lignes, "total": total,
            "contexte": {"legende": leg}}


def besoin_statuts_dispo():
    vals = {str(r.get("Statut", "")).strip() for r in _lignes_dict("L3_Besoins")}
    return sorted(v for v in vals if v)


def besoin_priorites_dispo():
    vals = {str(r.get("Priorite", "")).strip() for r in _lignes_dict("L3_Besoins")}
    return sorted(v for v in vals if v)


def _annee_de_date(txt):
    """Extrait l'annee civile d'une date 'JJ/MM/AAAA' (ou autre) -> str ou ''."""
    s = str(txt or "").strip()
    m = re.search(r"(\d{4})", s)
    return m.group(1) if m else ""


def annees_civiles_f1():
    vals = {_annee_de_date(r.get("Date operation", "")) for r in _lignes_dict("F1_Mouvements")}
    return sorted((v for v in vals if v), reverse=True)


def etat_par_bailleur(annee=""):
    """Etat des recettes / depenses par source de financement (F1), regroupe par
    bailleur. Filtre annee civile optionnel (sur Date operation).
    Colonnes : Bailleur, Total recettes, Total depenses, Solde (KMF)."""
    annee = str(annee or "").strip()
    cols = ["Source de financement / Bailleur", "Total recettes", "Total depenses", "Solde (KMF)"]
    agg = {}
    for r in _lignes_dict("F1_Mouvements"):
        if annee and _annee_de_date(r.get("Date operation", "")) != annee:
            continue
        b = str(r.get("Source de financement / Bailleur", "")).strip() or "(non renseigne)"
        a = agg.setdefault(b, [0.0, 0.0])
        a[0] += _num(r.get("Montant Recette (KMF)", 0)) or 0
        a[1] += _num(r.get("Montant Depense (KMF)", 0)) or 0
    lignes, tr, td, ts = [], 0.0, 0.0, 0.0
    for b in sorted(agg):
        rec, dep = agg[b]
        sol = rec - dep
        tr += rec; td += dep; ts += sol
        lignes.append([b, _kmf_aff(rec), _kmf_aff(dep), _kmf_aff(sol)])
    total = ["TOTAL", _kmf_aff(tr), _kmf_aff(td), _kmf_aff(ts)]
    leg = "Recettes / depenses par source de financement%s · %d bailleur(s) · Edite le %s" % (
        (" · annee civile %s" % annee) if annee else "", len(lignes), fmt_date(_dt.date.today()))
    return {"colonnes": cols, "lignes": lignes, "total": total,
            "contexte": {"legende": leg}}


# =====================================================================
# V1.99 — Fiche bailleur (F4_Bailleurs). Referentiel des financeurs au-dessus
# de F1/F3 : contrat, budget, perimetre, documents. L'indice de tracabilite
# (budget F4, depense F1, equipements M1, % localises) est livre en Touche 4
# (V1.99.2) : voir tracabilite_bailleur(). Le Bloc 2b d'origine (registre
# d'achats F1 + jointure M1<->F1 sur la reference piece) est ABANDONNE faute
# de source achats ; le financement se tague directement sur M1/F1/F3.
# =====================================================================

def liste_bailleurs():
    """Liste compacte des bailleurs (F4) pour recherche/affichage :
    [{id, nom, type, statut, label}]. label = 'ID — Nom (Type · Statut)'."""
    out = []
    for b in _lignes_dict("F4_Bailleurs"):
        idb = str(b.get("ID bailleur", "")).strip()
        nom = str(b.get("Nom / raison sociale", "")).strip()
        if not (idb or nom):
            continue
        typ = str(b.get("Type", "")).strip()
        sta = str(b.get("Statut", "")).strip()
        label = idb + (" — " + nom if nom else "")
        meta = " · ".join(x for x in (typ, sta) if x)
        if meta:
            label += " (" + meta + ")"
        out.append({"id": idb, "nom": nom, "type": typ, "statut": sta, "label": label})
    out.sort(key=lambda x: (x["nom"] or x["id"]).lower())
    return out


def fiche_bailleur(id_bailleur):
    """Fiche complete d'un bailleur (F4) par ID, ou None. Tous champs propres
    + cles d'affichage id / nom."""
    idb = str(id_bailleur or "").strip()
    if not idb:
        return None
    for b in _lignes_dict("F4_Bailleurs"):
        if str(b.get("ID bailleur", "")).strip().lower() == idb.lower():
            d = {k: ("" if v is None else str(v).strip()) for k, v in b.items()}
            d["id"] = str(b.get("ID bailleur", "")).strip()
            d["nom"] = str(b.get("Nom / raison sociale", "")).strip()
            return d
    return None


# ---------------------------------------------------------------------
# Documents lies aux bailleurs (3b). Le dossier fait foi (data.py) ; ici
# validation (role module Financier, format, taille, type P0) + mise en
# forme pour l'affichage. Pas de suppression : un document obsolete est
# MARQUE (statut), jamais efface. L'audit journal.csv est ecrit par la
# route (app.py), pour respecter l'ordre des couches.
# ---------------------------------------------------------------------
def _taille_aff(octets):
    try:
        o = int(octets)
    except (TypeError, ValueError):
        return ""
    if o < 1024:
        return "%d o" % o
    if o < 1024 * 1024:
        return "%d Ko" % round(o / 1024)
    return "%.1f Mo".replace(".", ",") % (o / (1024 * 1024))


def types_document_bailleur():
    """Liste P0 des natures de piece (selecteur d'ajout). Extensible (Dictionnaire)."""
    return list(_db.listes_parametres().get(config.P0_TYPES_DOC_BAILLEUR, []))


def documents_bailleur(id_bailleur):
    """Documents d'un bailleur, plus recents en premier :
    [{nom_stocke, nom_original, type, date_ajout, taille, taille_aff,
      ajoute_par, statut, obsolete}]."""
    out = []
    for d in _db.lire_index_documents_bailleur(id_bailleur):
        st = str(d.get("statut", "")).strip()
        out.append({
            "nom_stocke": str(d.get("nom_stocke", "")),
            "nom_original": str(d.get("nom_original", "")),
            "type": str(d.get("type", "")),
            "date_ajout": str(d.get("date_ajout", "")),
            "taille": d.get("taille", 0),
            "taille_aff": _taille_aff(d.get("taille", 0)),
            "ajoute_par": str(d.get("ajoute_par", "")),
            "statut": st,
            "obsolete": bool(st),
        })
    out.reverse()  # l'index ajoute en fin de liste -> plus recents d'abord
    return out


def ajouter_document_bailleur(acteur, id_bailleur, donnees_octets, nom_original, type_doc):
    """Ajoute une piece a un bailleur. Reserve au module Financier.
    Renvoie (ok, message, meta|None)."""
    if not est_admin(acteur) and not peut_ecrire(acteur, "F4_Bailleurs"):
        return False, "Action reservee au module Financier.", None
    if fiche_bailleur(id_bailleur) is None:
        return False, "Bailleur introuvable.", None
    nom_original = str(nom_original or "").strip()
    if not nom_original:
        return False, "Aucun fichier fourni.", None
    if not donnees_octets:
        return False, "Fichier vide.", None
    ext = os.path.splitext(nom_original)[1].lower()
    if ext not in config.DOC_BAILLEUR_EXT:
        return False, "Format non autorise (PDF, JPG, PNG, DOCX).", None
    if len(donnees_octets) > config.DOC_BAILLEUR_MAX_OCTETS:
        return False, "Fichier trop volumineux (10 Mo maximum).", None
    type_doc = str(type_doc or "").strip()
    types = types_document_bailleur()
    if types and type_doc and type_doc not in types:
        return False, "Type de document non reconnu.", None
    meta = _db.ajouter_document_bailleur(
        id_bailleur, donnees_octets, nom_original, type_doc,
        str(acteur.get("login", "")) if isinstance(acteur, dict) else "")
    try:  # compteur F4 (3c) : cosmetique, ne doit jamais bloquer l'ajout
        _db.definir_nb_documents_bailleur(
            id_bailleur, len(_db.lire_index_documents_bailleur(id_bailleur)))
    except Exception:
        pass
    return True, "Document ajoute.", meta


def marquer_statut_document_bailleur(acteur, id_bailleur, nom_stocke, statut):
    """Marque un document (ex. obsolete) sans le supprimer. Renvoie (ok, message)."""
    if not est_admin(acteur) and not peut_ecrire(acteur, "F4_Bailleurs"):
        return False, "Action reservee au module Financier."
    statut = str(statut or "").strip()
    if not _db.definir_statut_document_bailleur(id_bailleur, nom_stocke, statut):
        return False, "Document introuvable."
    return True, ("Document marque." if statut else "Marquage retire.")


def chemin_document_bailleur(id_bailleur, nom_stocke):
    """Chemin disque d'un document INDEXE pour telechargement, ou None.
    L'appartenance a l'index fait autorite : seuls les vrais documents sont servis
    (le fichier index.json et les fichiers internes ne sont jamais telechargeables)."""
    connus = {d["nom_stocke"] for d in documents_bailleur(id_bailleur)}
    if nom_stocke not in connus:
        return None
    return _db.chemin_document_bailleur(id_bailleur, nom_stocke)


# ---------------------------------------------------------------------
# Indice de tracabilite par bailleur (Touche 4). Honnete : vide tant que
# rien n'est tague. En-tete = budget alloue (F4), depense (F1), reste,
# taux de consommation, nb equipements (M1), % localises. Detail par poste
# (facultatif) = depense F1 / budgete F3 par poste, affiche seulement s'il
# y a des donnees. Le champ « Source de financement / Bailleur » de M1/F1/F3
# reference la liste P0 Sources_financement ; F4.ID = cette valeur (jointure).
# ---------------------------------------------------------------------
def _m1_localise(salle):
    """Vrai si l'equipement a une salle REELLE (exclut « (a ventiler) » et
    « (non localise) » et le vide)."""
    s = str(salle or "").strip().lower()
    return bool(s) and "ventiler" not in s and "non localis" not in s


def tracabilite_bailleur(id_bailleur):
    """Indice de tracabilite d'un bailleur (None si fiche absente)."""
    idb = str(id_bailleur or "").strip()
    if fiche_bailleur(idb) is None:
        return None
    cle = idb.lower()
    fiche = fiche_bailleur(idb)
    budget = _num(fiche.get("Budget alloue (KMF)", 0)) or 0

    depense, dep_poste = 0.0, {}
    for r in _lignes_dict("F1_Mouvements"):
        if str(r.get("Source de financement / Bailleur", "")).strip().lower() != cle:
            continue
        m = _num(r.get("Montant Depense (KMF)", 0)) or 0
        depense += m
        if m:
            p = str(r.get("Poste budgetaire", "")).strip() or "(sans poste)"
            dep_poste[p] = dep_poste.get(p, 0) + m

    bud_poste = {}
    for r in _lignes_dict("F3_Budget_poste"):
        if str(r.get("Source de financement / Bailleur", "")).strip().lower() != cle:
            continue
        m = _num(r.get("Montant budgete (KMF)", 0)) or 0
        if m:
            p = str(r.get("Poste budgetaire", "")).strip() or "(sans poste)"
            bud_poste[p] = bud_poste.get(p, 0) + m

    nb_equip, nb_loc = 0, 0
    for r in _lignes_dict("M1_Equipements"):
        if str(r.get("Source de financement / Bailleur", "")).strip().lower() != cle:
            continue
        nb_equip += 1
        if _m1_localise(r.get("Salle / localisation", "")):
            nb_loc += 1

    reste = (budget - depense) if budget else None
    taux = round(depense / budget * 100) if budget else None
    pct_loc = round(nb_loc / nb_equip * 100) if nb_equip else None
    postes = sorted(set(dep_poste) | set(bud_poste))
    detail = [{"poste": p, "budgete": bud_poste.get(p, 0),
               "budgete_aff": _kmf_aff(bud_poste.get(p, 0)) if bud_poste.get(p) else "",
               "depense": dep_poste.get(p, 0),
               "depense_aff": _kmf_aff(dep_poste.get(p, 0)) if dep_poste.get(p) else ""}
              for p in postes]
    return {
        "id": idb,
        "budget": budget, "budget_aff": _kmf_aff(budget) if budget else "",
        "depense": depense, "depense_aff": _kmf_aff(depense),
        "reste": reste, "reste_aff": (_kmf_aff(reste) if reste is not None else ""),
        "taux_conso": taux,
        "nb_equip": nb_equip, "nb_localises": nb_loc, "pct_localises": pct_loc,
        "detail_postes": detail, "a_detail": bool(detail),
        "a_budget": bool(budget),
    }


# =====================================================================
# V1.76 — Grille de saisie des notes par CLASSE et par MATIERE.
# Liste d'eleves facon feuille de presence : plusieurs controles
# (coef par colonne) -> CC pondere -> moyenne = 1/4 CC + 3/4 examen
# (decret 05-106, art. 8). Controles ecrits dans N4_Controles ; examen
# dans N2_Notes (CC reste DERIVE de N4, non ecrit dans N2).
# =====================================================================

NB_CONTROLES_DEFAUT = 3


def classe_matieres(filiere, niveau, semestre):
    """[{num_ue, intitule, coef, matiere}] pour le selecteur matiere d'une classe."""
    out = []
    for ue in _bareme_ues(filiere, niveau, semestre):
        for m in ue["matieres"]:
            out.append({"num_ue": ue["num"], "intitule": ue["intitule"],
                        "coef": ue["coef"], "matiere": m})
    return out


def _controles_classe(annee, session, semestre, num_ue, matiere):
    """({matricule: {n_ctrl(int): note_str}}, {n_ctrl: coef}) depuis N4 pour le contexte."""
    iMat, iAn, iSess, iSem, iU, iM, iNum, iNote, iCoef = _idx_map(
        _db_notes, "N4_Controles",
        ["Matricule", "Annee acad.", "Session", "Semestre", "N° UE", "Matiere",
         "N° de controle", "Note /20", "Coef"])
    parmat, coefs = {}, {}
    for r in _db_notes.lignes("N4_Controles"):
        if _n(r[iAn]) != _n(annee):
            continue
        if (str(r[iSess]).strip() or "1") != str(session).strip():
            continue
        if str(r[iSem]).strip() != str(semestre).strip():
            continue
        if str(r[iU]).strip() != str(num_ue).strip():
            continue
        if _n(r[iM]) != _n(matiere):
            continue
        try:
            nc = int(float(str(r[iNum]).strip() or "0"))
        except ValueError:
            nc = 0
        if nc <= 0:
            continue
        note = str(r[iNote]).strip()
        if note == "":
            continue          # controle sans note : pas de colonne fantome au rechargement
        mat = str(r[iMat]).strip()
        parmat.setdefault(mat, {})[nc] = note
        c = _num_h(r[iCoef]) if str(r[iCoef]).strip() != "" else 1.0
        coefs[nc] = c if c > 0 else 1.0
    return parmat, coefs


def _examens_classe(annee, session, semestre, num_ue, matiere):
    """{matricule: examen_str} depuis N2 pour le contexte (session demandee)."""
    iMat, iAn, iSess, iSem, iU, iM, iCC, iEx = _idx_map(
        _db_notes, "N2_Notes",
        ["Matricule", "Annee acad.", "Session", "Semestre", "N° UE", "Matiere", "CC", "Examen"])
    out = {}
    for r in _db_notes.lignes("N2_Notes"):
        if _n(r[iAn]) != _n(annee):
            continue
        if (str(r[iSess]).strip() or "1") != str(session).strip():
            continue
        if str(r[iSem]).strip() != str(semestre).strip():
            continue
        if str(r[iU]).strip() != str(num_ue).strip():
            continue
        if _n(r[iM]) != _n(matiere):
            continue
        out[str(r[iMat]).strip()] = str(r[iEx]).strip()
    return out


def _cc_pondere(notes_par_num, coefs):
    sn = sp = 0.0
    for nc, note in notes_par_num.items():
        if str(note).strip() == "":
            continue
        c = coefs.get(nc, 1.0) or 1.0
        sn += _num_h(note) * c
        sp += c
    return (sn / sp) if sp > 0 else None


def notes_grille(filiere, niveau, section, annee, semestre, num_ue, matiere, session="1"):
    """Modele de la grille classe x matiere : liste d'eleves avec leurs controles,
    examen, CC (pondere) et moyenne (1/4-3/4). nums = numeros de controle affiches."""
    session = str(session).strip() or "1"
    roster = liste_etudiants(filiere, niveau, section, annee)["lignes"]
    ctrl, coefs = _controles_classe(annee, session, semestre, num_ue, matiere)
    nums = sorted(coefs) if coefs else list(range(1, NB_CONTROLES_DEFAUT + 1))
    if not nums:
        nums = list(range(1, NB_CONTROLES_DEFAUT + 1))
    coefs_list = [coefs.get(n, 1.0) for n in nums]
    exams = _examens_classe(annee, session, semestre, num_ue, matiere)
    intitule = ""
    ects_ue = ""
    for u in _bareme_ues(filiere, niveau, semestre):
        if str(u["num"]).strip() == str(num_ue).strip():
            intitule = u["intitule"]
            ects_ue = u.get("ects", "")
            break
    if isinstance(ects_ue, float) and ects_ue == int(ects_ue):
        ects_ue = int(ects_ue)
    eleves = []
    for l in roster:
        mat = str(l[1]).strip()
        nom = ("%s %s" % (l[3], l[4])).strip()
        cnotes = ctrl.get(mat, {})
        notes = [cnotes.get(n, "") for n in nums]
        cc = _cc_pondere({n: cnotes.get(n, "") for n in nums},
                         {n: coefs.get(n, 1.0) for n in nums})
        ex = exams.get(mat, "")
        moy = _moyenne_matiere(cc if cc is not None else "", ex)
        eleves.append({"matricule": mat, "nom": nom, "controles": notes,
                       "examen": ex, "cc": _arrondi2(cc), "moyenne": _arrondi2(moy)})
    return {"ok": True, "filiere": filiere, "niveau": niveau, "section": section,
            "annee": annee, "semestre": semestre, "session": session,
            "num_ue": num_ue, "intitule_ue": intitule, "ects_ue": ects_ue, "matiere": matiere,
            "nums": nums, "coefs": coefs_list, "nb": len(eleves), "eleves": eleves,
            "non_dispensee": _nd_etat_exact(filiere, niveau, section, annee, session,
                                            semestre, num_ue, matiere)[0],
            "motif_nd": _nd_etat_exact(filiere, niveau, section, annee, session,
                                       semestre, num_ue, matiere)[1]}


def enregistrer_notes_grille(filiere, niveau, section, annee, semestre, num_ue,
                             matiere, session, coefs, eleves, saisi_par,
                             non_dispensee=False, motif=""):
    """Ecrit N4 (controles, un par eleve x controle) + N2 (examen). CC derive de N4,
    non ecrit dans N2. eleves = [{matricule, controles:[...], examen}].
    V1.77 : si non_dispensee, trace la matiere dans N5 (exclue du calcul) et n'ecrit
    aucune note ; sinon, retire toute trace N5 prealable (reversible) puis ecrit."""
    session = str(session).strip() or "1"
    annee = str(annee).strip()
    semestre = str(semestre).strip()
    num_ue = str(num_ue).strip()
    matiere = str(matiere).strip()
    if not (semestre and num_ue and matiere):
        return False, "Contexte incomplet (semestre / UE / matiere)."
    nd_on = str(non_dispensee).strip().lower() in ("1", "true", "oui", "on", "vrai")
    enregistrer_matiere_nd(filiere, niveau, section, annee, session, semestre,
                           num_ue, matiere, nd_on, motif, saisi_par)
    if nd_on:
        return True, ("Matiere marquee non dispensee : exclue du calcul et tracee (N5). "
                      "Les notes eventuelles ne sont pas enregistrees.")
    n = len(coefs)
    if n == 0:
        return False, "Aucun controle defini."
    kM = _brut("N4_Controles", "Matricule"); kA = _brut("N4_Controles", "Annee acad.")
    kSe = _brut("N4_Controles", "Session"); kSm = _brut("N4_Controles", "Semestre")
    kU = _brut("N4_Controles", "N° UE"); kMa = _brut("N4_Controles", "Matiere")
    kNc = _brut("N4_Controles", "N° de controle"); kNo = _brut("N4_Controles", "Note /20")
    kCo = _brut("N4_Controles", "Coef"); kSp = _brut("N4_Controles", "Saisi par")
    existant_ctrl, existant_coefs = _controles_classe(annee, session, semestre, num_ue, matiere)
    lignes4 = []
    for el in eleves:
        m = str(el.get("matricule", "")).strip()
        if not m:
            continue
        notes = el.get("controles", [])
        for i in range(n):
            note = str(notes[i]).strip() if i < len(notes) else ""
            nc = i + 1
            deja = nc in existant_ctrl.get(m, {})
            if note == "" and not deja:
                continue
            ok, noten = _note_valide(note)
            if not ok:
                return False, "Note de controle invalide pour %s (attendu 0 a 20)." % m
            coef = coefs[i]
            coef = ("%g" % coef) if str(coef).strip() != "" else "1"
            lignes4.append({kM: m, kA: annee, kSe: session, kSm: semestre, kU: num_ue,
                            kMa: matiere, kNc: str(nc), kNo: noten, kCo: coef, kSp: saisi_par})
    # Nettoyage des controles supprimes (n° > n) : on vide leur note (upsert) pour
    # qu'ils ne pesent plus dans le CC.
    for m, ncs in existant_ctrl.items():
        for nc in ncs:
            if nc > n:
                co = ("%g" % existant_coefs.get(nc, 1.0))
                lignes4.append({kM: m, kA: annee, kSe: session, kSm: semestre, kU: num_ue,
                                kMa: matiere, kNc: str(nc), kNo: "", kCo: co, kSp: saisi_par})
    bM = _brut("N2_Notes", "Matricule"); bA = _brut("N2_Notes", "Annee acad.")
    bSe = _brut("N2_Notes", "Session"); bSm = _brut("N2_Notes", "Semestre")
    bU = _brut("N2_Notes", "N° UE"); bMa = _brut("N2_Notes", "Matiere")
    bEx = _brut("N2_Notes", "Examen")
    exist_ex = _examens_classe(annee, session, semestre, num_ue, matiere)
    lignes2 = []
    for el in eleves:
        m = str(el.get("matricule", "")).strip()
        if not m:
            continue
        ex = str(el.get("examen", "")).strip()
        if ex == "" and m not in exist_ex:
            continue
        ok, exn = _note_valide(ex)
        if not ok:
            return False, "Note d'examen invalide pour %s (attendu 0 a 20)." % m
        lignes2.append({bM: m, bA: annee, bSe: session, bSm: semestre, bU: num_ue,
                        bMa: matiere, bEx: exn})
    r4 = (_db_notes.ecrire_lignes_lot("N4_Controles", lignes4,
          cles=[kM, kA, kSe, kSm, kU, kMa, kNc]) if lignes4 else {"ajout": 0, "maj": 0})
    r2 = (_db_notes.ecrire_lignes_lot("N2_Notes", lignes2,
          cles=[bM, bA, bSe, bSm, bU, bMa]) if lignes2 else {"ajout": 0, "maj": 0})
    nb = len({str(el.get("matricule", "")).strip() for el in eleves
              if str(el.get("matricule", "")).strip()})
    return True, ("Notes enregistrees : %d eleve(s) ; controles %d ajout / %d maj ; "
                  "examens %d ajout / %d maj." % (nb, r4["ajout"], r4["maj"],
                                                  r2["ajout"], r2["maj"]))


def feuille_notes_edition(filiere, niveau, section, annee, semestre, num_ue, matiere, session="1"):
    """Edition imprimable (kind=table) de la grille : colonnes dynamiques C1..Cn + CC + Examen + Moyenne."""
    g = notes_grille(filiere, niveau, section, annee, semestre, num_ue, matiere, session)
    cols = ["N°", "Matricule", "Nom et prenom"]
    for i, nc in enumerate(g["nums"]):
        cols.append("C%d (coef %g)" % (nc, g["coefs"][i]))
    cols += ["CC /20", "Examen /20", "Moyenne"]
    lignes = []
    for idx, e in enumerate(g["eleves"], start=1):
        row = [str(idx), e["matricule"], e["nom"]]
        row += [("%s" % c) if str(c).strip() != "" else "" for c in e["controles"]]
        row += [_kmf_aff(e["cc"]) if e["cc"] is not None else "",
                e["examen"], _kmf_aff(e["moyenne"]) if e["moyenne"] is not None else ""]
        lignes.append(row)
    leg = ("%s · %s · Section %s · Semestre %s · Session %s · UE %s %s · Matiere : %s · "
           "%d eleve(s) · Moyenne = 1/4 CC + 3/4 examen (decret 05-106, art. 8) · Edite le %s" % (
               filiere, niveau, section, semestre, session, num_ue, g["intitule_ue"],
               matiere, g["nb"], fmt_date(_dt.date.today())))
    return {"colonnes": cols, "lignes": lignes, "total": None, "contexte": {"legende": leg}}


# =====================================================================
# V1.77 — MATIERE NON DISPENSEE (statut explicite, niveau classe)
# Onglet N5_Matieres_ND. Affichage (b) : la matiere ND n'apparait pas au
# releve ; mention de tracabilite (texte dans les gabarits d'impression) ;
# moyenne recalculee sur les seules matieres faites. ECTS = option (i) :
# l'UE garde ses ECTS tant qu'une matiere est faite ; seule une UE
# ENTIEREMENT non dispensee perd ses ECTS (prorata = retrait). Distinction :
# ND (exclue + tracee) vs pas encore notee (releve marque incomplet, NON
# exclue). Statut au niveau CLASSE (pas par etudiant).
# =====================================================================
ONGLET_ND = "N5_Matieres_ND"


def _matieres_nd(filiere, niveau, semestre, annee=None):
    """Set des (num_ue, matiere) non dispensees pour cette classe/semestre.
    Match (filiere, niveau, semestre [, annee]) ; section/session ignorees (ND =
    fait de semestre). Onglet absent => set vide (retro-compatible)."""
    if ONGLET_ND not in _db_notes.onglets():
        return set()
    iF, iN, iSm, iAn, iU, iM = _idx_map(_db_notes, ONGLET_ND,
        ["Filiere", "Niveau", "Semestre", "Annee acad.", "N° UE", "Matiere"])
    out = set()
    for r in _db_notes.lignes(ONGLET_ND):
        if iF >= 0 and _n(r[iF]) != _n(filiere):
            continue
        if iN >= 0 and _n(r[iN]) != _n(niveau):
            continue
        if iSm >= 0 and str(r[iSm]).strip() != str(semestre).strip():
            continue
        if annee not in (None, "") and iAn >= 0 and _n(r[iAn]) != _n(annee):
            continue
        out.add((str(r[iU]).strip() if iU >= 0 else "",
                 str(r[iM]).strip() if iM >= 0 else ""))
    return out


def _nd_meme_contexte(d, filiere, niveau, section, annee, session, semestre, num_ue, matiere):
    """d = {entete_brut: valeur}. True si la ligne N5 vise EXACTEMENT ce contexte."""
    def g(lib):
        return str(d.get(_brut(ONGLET_ND, lib), "") or "").strip()
    return (_n(g("Filiere")) == _n(filiere) and _n(g("Niveau")) == _n(niveau)
            and _n(g("Section")) == _n(section) and _n(g("Annee acad.")) == _n(annee)
            and g("Session") == (str(session).strip() or "1")
            and g("Semestre") == str(semestre).strip()
            and g("N° UE") == str(num_ue).strip()
            and _n(g("Matiere")) == _n(matiere))


def _nd_etat_exact(filiere, niveau, section, annee, session, semestre, num_ue, matiere):
    """(bool, motif) pour CE contexte exact (avec section/session) : etat de l'interrupteur."""
    if ONGLET_ND not in _db_notes.onglets():
        return False, ""
    ent = _db_notes.entetes(ONGLET_ND)
    bMo = _brut(ONGLET_ND, "Motif")
    for r in _db_notes.lignes(ONGLET_ND):
        d = {ent[i]: ("" if r[i] is None else r[i]) for i in range(min(len(ent), len(r)))}
        if _nd_meme_contexte(d, filiere, niveau, section, annee, session, semestre, num_ue, matiere):
            return True, str(d.get(bMo, "") or "").strip()
    return False, ""


def enregistrer_matiere_nd(filiere, niveau, section, annee, session, semestre,
                           num_ue, matiere, nd_on, motif, saisi_par):
    """Upsert (nd_on=True) / retrait (nd_on=False) d'une ligne N5 pour le contexte
    exact. Reecrit l'onglet (petit volume, sans colonne calcul) : idempotent et
    reversible. Sans effet si l'onglet N5 est absent (retro-compatible)."""
    if ONGLET_ND not in _db_notes.onglets():
        return
    ent = _db_notes.entetes(ONGLET_ND)
    restantes = []
    for r in _db_notes.lignes(ONGLET_ND):
        d = {ent[i]: ("" if r[i] is None else r[i]) for i in range(min(len(ent), len(r)))}
        if all(str(v or "").strip() == "" for v in d.values()):
            continue
        if _nd_meme_contexte(d, filiere, niveau, section, annee, session, semestre, num_ue, matiere):
            continue
        restantes.append(d)
    if nd_on:
        restantes.append({
            _brut(ONGLET_ND, "Filiere"): filiere, _brut(ONGLET_ND, "Niveau"): niveau,
            _brut(ONGLET_ND, "Section"): section, _brut(ONGLET_ND, "Annee acad."): annee,
            _brut(ONGLET_ND, "Session"): str(session).strip() or "1",
            _brut(ONGLET_ND, "Semestre"): semestre, _brut(ONGLET_ND, "N° UE"): num_ue,
            _brut(ONGLET_ND, "Matiere"): matiere, _brut(ONGLET_ND, "Motif"): motif or "",
            _brut(ONGLET_ND, "Saisi par"): saisi_par or ""})
    _db_notes.remplacer_donnees(ONGLET_ND, restantes)


# ===========================================================================
# V1.80 — Verrou anti-suppression des listes P0 + sauvegarde horodatee
# ===========================================================================

def _norm_liste(x):
    """Normalise un libelle de colonne/liste : retire un suffixe (*) / (**) / (...)
    final, minuscule, espaces compactes."""
    x = re.sub(r"\s*\([^)]*\)\s*$", "", str(x or "")).strip().lower()
    return re.sub(r"\s+", " ", x)


def valeur_liste_utilisee(colonne, valeur):
    """Verrou P0 : nombre d'enregistrements employant encore `valeur` pour la
    liste `colonne`. Renvoie (total, apercu) ou apercu = liste de 'Onglet: n'.
    Liste hors config.P0_CONSOMMATEURS -> (0, []) (suppression libre)."""
    cible = _norm_liste(colonne)
    consommateurs = getattr(config, "P0_CONSOMMATEURS", {}).get(cible)
    if not consommateurs:
        return 0, []
    v = str(valeur or "").strip().lower()
    if not v:
        return 0, []
    total, par_onglet = 0, {}
    for onglet, base in consommateurs:
        n = 0
        for r in _lignes_dict(onglet):
            for k, val in r.items():
                if _norm_liste(k) == base and str(val or "").strip().lower() == v:
                    n += 1
                    break
        if n:
            par_onglet[onglet] = par_onglet.get(onglet, 0) + n
            total += n
    apercu = ["%s: %d" % (o, n) for o, n in par_onglet.items()]
    return total, apercu


def sauvegarder_classeurs():
    """Copie horodatee des classeurs de donnees (EMSP_V1 + EMSP_Notes) dans
    donnees/sauvegardes/EMSP_sauvegarde_AAAA-MM-JJ_HHMMSS/. Copie fichier brute
    (shutil) : ne passe JAMAIS par openpyxl, donc aucun dessin n'est altere.
    Renvoie (ok, message, nom_dossier, fichiers_copies)."""
    import os as _os, shutil as _sh, datetime as _dt
    horod = _dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    nom = "EMSP_sauvegarde_" + horod
    dest = _os.path.join(config.SAUVEGARDES_DIR, nom)
    try:
        _os.makedirs(dest, exist_ok=True)
        copies = []
        for src in [config.WORKBOOK, getattr(config, "WORKBOOK_NOTES", None)]:
            if src and _os.path.exists(src):
                _sh.copy2(src, _os.path.join(dest, _os.path.basename(src)))
                copies.append(_os.path.basename(src))
        if not copies:
            return False, "Aucun classeur a sauvegarder (introuvable sur disque).", nom, []
        return True, "Sauvegarde creee : %d classeur(s)." % len(copies), nom, copies
    except Exception as exc:
        return False, "Echec de la sauvegarde : %s" % exc, nom, []


def liste_sauvegardes():
    """Sauvegardes existantes, plus recentes d'abord : nom, horodatage lisible,
    nb de classeurs, taille (Ko)."""
    import os as _os
    d = getattr(config, "SAUVEGARDES_DIR", "")
    out = []
    if d and _os.path.isdir(d):
        for nom in sorted(_os.listdir(d), reverse=True):
            p = _os.path.join(d, nom)
            if not _os.path.isdir(p):
                continue
            fichiers = [f for f in _os.listdir(p) if f.lower().endswith((".xlsx", ".xls"))]
            taille = 0
            for f in fichiers:
                try:
                    taille += _os.path.getsize(_os.path.join(p, f))
                except OSError:
                    pass
            aff = nom
            base = nom.replace("EMSP_sauvegarde_", "")
            if "_" in base:
                dpart, hpart = base.split("_", 1)
                ymd = dpart.split("-")
                if len(ymd) == 3 and len(hpart) >= 6:
                    aff = "%s/%s/%s à %s:%s:%s" % (ymd[2], ymd[1], ymd[0],
                                                   hpart[:2], hpart[2:4], hpart[4:6])
            out.append({"nom": nom, "affiche": aff, "n": len(fichiers),
                        "ko": int(round(taille / 1024.0))})
    return out


# ===========================================================================
# V1.81 — Module stages : synthese multicritere + tableau de bord (lecture)
# ===========================================================================

def _stg_get(row, base):
    """Lit une valeur de `row` (dict _lignes_dict) par libelle normalise
    (resilient aux suffixes (*) / (**))."""
    b = _norm_liste(base)
    for k, v in row.items():
        if _norm_liste(k) == b:
            return ("" if v is None else str(v)).strip()
    return ""


def stages_synthese(lieu="", seance="", niveau="", filiere="", nom="", annee=""):
    """Affectations de stage (S1) enrichies par l'etudiant (A1) et le lieu (S2),
    filtrees, plus un tableau de bord. Lecture/agregation pure (aucune ecriture).
    Renvoie {affectations, tdb, options, filtres}."""
    lieu = (lieu or "").strip(); seance = (seance or "").strip()
    niveau = (niveau or "").strip(); filiere = (filiere or "").strip()
    nom = (nom or "").strip(); annee = (annee or "").strip()

    # Index etudiants (matricule -> infos)
    etu = {}
    for e in _lignes_dict("A1_Etudiants"):
        m = _stg_get(e, "matricule")
        if not m:
            continue
        etu[m.lower()] = {"nom": _stg_get(e, "nom"), "prenom": _stg_get(e, "prenom"),
                          "filiere": _stg_get(e, "filiere"), "niveau": _stg_get(e, "niveau"),
                          "section": _stg_get(e, "section")}

    # Referentiel des lieux (S2) : libelle = "Lieu — Service" (comme stocke dans S1)
    lieux = {}
    for r in _lignes_dict("S2_Lieux_stage"):
        nl = _stg_get(r, "lieu / structure")
        if not nl:
            continue
        serv = _stg_get(r, "service")
        lib = nl + (" — " + serv if serv else "")
        lieux[lib.lower()] = {"lib": lib, "service": serv, "commune": _stg_get(r, "commune"),
                              "niveau": _stg_get(r, "niveau concerne"),
                              "quota": int(_fnum(_stg_get(r, "quota")) or 0),
                              "periode": _stg_get(r, "periode de disponibilite")}

    # Affectations (S1) enrichies
    aff = []
    for r in _lignes_dict("S1_Stages"):
        m = _stg_get(r, "matricule"); ls = _stg_get(r, "lieu de stage")
        if not (m or ls):
            continue
        info = etu.get(m.lower(), {}); linfo = lieux.get(ls.lower(), {})
        aff.append({
            "matricule": m, "nom": info.get("nom", ""), "prenom": info.get("prenom", ""),
            "filiere": info.get("filiere", ""),
            "niveau": info.get("niveau", "") or linfo.get("niveau", ""),
            "lieu": ls, "commune": linfo.get("commune", ""),
            "annee": _stg_get(r, "annee acad."), "seance": _stg_get(r, "n seance (1-6)"),
            "debut": _stg_get(r, "date debut"), "fin": _stg_get(r, "date fin"),
            "session": _sess_norm(_stg_get(r, "session")),
            "retour": _stg_get(r, "fiche retour (o/n)"), "note": _stg_get(r, "note stage /20"),
            "heures_abs": _stg_get(r, "heures d'absence"),
        })

    def _ok(x):
        if lieu and lieu.lower() not in x["lieu"].lower():
            return False
        if seance and seance != str(x["seance"]).strip():
            return False
        if niveau and niveau.lower() != x["niveau"].lower():
            return False
        if filiere and filiere.lower() != x["filiere"].lower():
            return False
        if annee and annee != str(x["annee"]).strip():
            return False
        if nom:
            cible = ("%s %s %s" % (x["nom"], x["prenom"], x["matricule"])).lower()
            if nom.lower() not in cible:
                return False
        return True

    aff_f = [x for x in aff if _ok(x)]

    # Tableau de bord (sur le jeu filtre)
    mats = set(x["matricule"].lower() for x in aff_f if x["matricule"])
    occ = {}
    for x in aff_f:
        occ[x["lieu"].lower()] = occ.get(x["lieu"].lower(), 0) + 1
    lieux_detail, quota_tot, places_dispo = [], 0, 0
    for k, l in lieux.items():
        if niveau and l["niveau"] and l["niveau"].lower() != niveau.lower():
            continue
        o = occ.get(k, 0); q = l["quota"]
        quota_tot += q; places_dispo += max(q - o, 0)
        lieux_detail.append({"lib": l["lib"], "niveau": l["niveau"], "quota": q, "occupe": o,
                             "dispo": max(q - o, 0), "periode": l["periode"],
                             "taux": int(round(100.0 * o / q)) if q else 0})
    lieux_detail.sort(key=lambda d: d["lib"].lower())

    promo = {}
    for x in aff_f:
        promo.setdefault(x["niveau"] or "—", set()).add(x["matricule"].lower())
    repartition = [{"niveau": n, "n": len(s)} for n, s in sorted(promo.items())]

    tdb = {"etudiants": len(mats), "lieux": len(lieux_detail),
           "places_dispo": places_dispo, "quota_total": quota_tot,
           "occupe_total": sum(d["occupe"] for d in lieux_detail),
           "taux_global": (int(round(100.0 * sum(d["occupe"] for d in lieux_detail) / quota_tot))
                           if quota_tot else 0),
           "lieux_detail": lieux_detail, "repartition": repartition}

    options = {
        "lieux": sorted({l["lib"] for l in lieux.values()}),
        "seances": sorted({x["seance"] for x in aff if x["seance"]}),
        "niveaux": sorted({x["niveau"] for x in aff if x["niveau"]}
                          | {l["niveau"] for l in lieux.values() if l["niveau"]}),
        "filieres": sorted({x["filiere"] for x in aff if x["filiere"]}),
        "annees": sorted({x["annee"] for x in aff if x["annee"]}),
    }
    aff_f.sort(key=lambda x: (str(x["seance"]), x["nom"].lower(), x["prenom"].lower()))
    return {"affectations": aff_f, "tdb": tdb, "options": options,
            "filtres": {"lieu": lieu, "seance": seance, "niveau": niveau,
                        "filiere": filiere, "nom": nom, "annee": annee}}


# ===========================================================================
# V1.85 — Resultats par classe (synthese 1 ligne/eleve) + edition par lot
# ===========================================================================

def resultats_classe(filiere, niveau, section, annee, periode):
    """Synthese des resultats d'une classe : 1 ligne par eleve (moyenne, mention,
    decision, ECTS). periode = 'annuel' -> recapitulatif annuel ; sinon numero de
    semestre (cf. semestres_classe). Lecture/agregation pure. Renvoie {lignes, stats}."""
    periode = str(periode or "").strip()
    annuel = periode.lower() in ("annuel", "annee", "année", "an")
    roster = roster_classe(filiere, niveau, section)
    lignes, moys = [], []
    n_admis = n_ajourne = n_incomplet = n_conditionnel = 0
    for i, e in enumerate(roster, 1):
        mat = e["matricule"]
        r = (releve_annuel(mat, annee) if annuel
             else releve_semestre(mat, annee, periode))
        moy = r.get("moyenne")
        # Decision officielle : annuel -> regle combinee (moyenne + ECTS) ;
        # semestre -> proposition (Admis/Ajourne).
        dec = r.get("decision") if annuel else r.get("proposition", "\u2014")
        dec = dec or "\u2014"
        if dec == "Admis":
            n_admis += 1
        elif dec == "Admis conditionnel":
            n_conditionnel += 1
        elif dec == "Ajourne":
            n_ajourne += 1
        if r.get("incomplet"):
            n_incomplet += 1
        if moy is not None:
            moys.append(moy)
        flags = []
        if r.get("incomplet"):
            flags.append("incomplet")
        if r.get("bareme_provisoire"):
            flags.append("barème prov.")
        if r.get("a_non_dispensee"):
            flags.append("matière ND")
        ects = "" if annuel else ("%g / %g" % (r.get("ects_acquis", 0) or 0,
                                               r.get("ects_total", 0) or 0))
        lignes.append({"n": i, "matricule": mat, "nom": e.get("nom", ""),
                       "prenom": e.get("prenom", ""),
                       "moyenne": ("%.2f" % moy) if moy is not None else "",
                       "mention": r.get("mention", "") or "", "decision": dec,
                       "ects": ects, "note": ", ".join(flags)})
    moy_classe = ("%.2f" % (sum(moys) / len(moys))) if moys else ""
    stats = {"effectif": len(roster), "admis": n_admis, "ajourne": n_ajourne,
             "conditionnel": n_conditionnel,
             "incomplet": n_incomplet, "moyenne_classe": moy_classe, "annuel": annuel}
    return {"lignes": lignes, "stats": stats}


# ===========================================================================
# R7a — Attestation de passage (PDF) + workflow de liberation — V1.99.43
# Deliberation (conseil des professeurs) par classe+annee : metadonnee de
# workflow dans instance/deliberations.json (hors classeur). Apres la
# deliberation, un delai de contestation (config.DELAI_CONTESTATION_JOURS) doit
# s'ecouler avant que l'attestation devienne imprimable/distribuable (CR 11/06).
# Decision affichee = Admis/Ajourne (deja calculee par releve_annuel). Le
# "passage conditionnel" (regle ECTS a confirmer par le Directeur) = R7b, non
# traite ici.
# ===========================================================================

_DELIBERATIONS_FILE = config.DELIBERATIONS_FILE


def _delib_cle(annee, filiere, niveau):
    return "%s|%s|%s" % (str(annee).strip(), str(filiere).strip(), str(niveau).strip())


def _lire_deliberations():
    try:
        with open(_DELIBERATIONS_FILE, encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}


def _parse_date_fr(s):
    """JJ/MM/AAAA -> datetime.date ; None si invalide."""
    s = str(s or "").strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _lat(s):
    """Rend une chaine sure pour les polices coeur (latin-1) de fpdf2 :
    remplace les caracteres typographiques hors latin-1."""
    s = str(s or "")
    repl = {"\u2014": "-", "\u2013": "-", "\u2019": "'", "\u2018": "'",
            "\u201c": '"', "\u201d": '"', "\u2026": "...", "\u00a0": " "}
    for k, v in repl.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")


def deliberation_classe(annee, filiere, niveau):
    """Etat de la deliberation d'une classe+annee : date saisie, date de
    liberation (date + delai de contestation), liberable (aujourd'hui >= lib.)."""
    e = _lire_deliberations().get(_delib_cle(annee, filiere, niveau)) or {}
    d = _parse_date_fr(e.get("date_deliberation", ""))
    delai = int(getattr(config, "DELAI_CONTESTATION_JOURS", 7))
    out = {"date": e.get("date_deliberation", ""), "saisi_par": e.get("saisi_par", ""),
           "saisi_le": e.get("saisi_le", ""), "delai_jours": delai,
           "date_liberation": "", "liberable": False, "jours_restants": None}
    if d:
        lib = d + _dt.timedelta(days=delai)
        out["date_liberation"] = lib.strftime("%d/%m/%Y")
        today = _dt.date.today()
        out["liberable"] = today >= lib
        out["jours_restants"] = max((lib - today).days, 0)
    return out


def enregistrer_deliberation(acteur, annee, filiere, niveau, date_str, login):
    """Enregistre/maj la date de deliberation d'une classe. Reserve au droit
    d'ecriture des notes (scolarite/direction). Renvoie (ok, msg)."""
    if not peut_ecrire(acteur, "N2_Notes"):
        return False, "Action non autorisee."
    if not (str(annee).strip() and str(filiere).strip() and str(niveau).strip()):
        return False, "Classe incomplete (filiere, niveau, annee)."
    d = _parse_date_fr(date_str)
    if d is None:
        return False, "Date de deliberation invalide (attendu JJ/MM/AAAA)."
    if d > _dt.date.today():
        return False, "La date de deliberation ne peut pas etre dans le futur."
    try:
        os.makedirs(config.INSTANCE_DIR, exist_ok=True)
        data = _lire_deliberations()
        data[_delib_cle(annee, filiere, niveau)] = {
            "date_deliberation": d.strftime("%d/%m/%Y"),
            "saisi_par": login, "saisi_le": _dt.date.today().strftime("%d/%m/%Y")}
        with open(_DELIBERATIONS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as ex:
        return False, "Echec d'enregistrement : %s" % ex
    auth.journal(login, "Deliberation enregistree",
                 "%s %s %s" % (filiere, niveau, annee), d.strftime("%d/%m/%Y"))
    return True, "Deliberation du %s enregistree." % d.strftime("%d/%m/%Y")


# --- Validation DEFINITIVE des bulletins (R7b — action manuelle) -----------
# Liste a cocher par classe : l'admission definitive est validee a la main
# (sceau des bulletins). Etat hors classeur dans instance/validations_bulletins.json,
# cle 'annee|filiere|niveau' -> liste de matricules valides. C'est cette
# validation (et non le delai de contestation) qui libere l'attestation.

_VALIDATIONS_FILE = os.path.join(config.INSTANCE_DIR, "validations_bulletins.json")


def _lire_validations():
    try:
        with open(_VALIDATIONS_FILE, encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}


def bulletins_valides(annee, filiere, niveau):
    """Ensemble (minuscules) des matricules dont le bulletin est valide
    definitivement pour cette classe+annee."""
    lst = _lire_validations().get(_delib_cle(annee, filiere, niveau)) or []
    return set(str(m).strip().lower() for m in lst)


def est_bulletin_valide(matricule, annee, filiere, niveau):
    return str(matricule).strip().lower() in bulletins_valides(annee, filiere, niveau)


def enregistrer_validations(acteur, annee, filiere, niveau, matricules, login):
    """Remplace la liste des bulletins valides d'une classe par 'matricules'
    (les eleves coches). Reserve au droit d'ecriture des notes. (ok, msg)."""
    if not peut_ecrire(acteur, "N2_Notes"):
        return False, "Action non autorisee."
    if not (str(annee).strip() and str(filiere).strip() and str(niveau).strip()):
        return False, "Classe incomplete (filiere, niveau, annee)."
    mats = sorted({str(m).strip() for m in (matricules or []) if str(m).strip()})
    try:
        os.makedirs(config.INSTANCE_DIR, exist_ok=True)
        data = _lire_validations()
        data[_delib_cle(annee, filiere, niveau)] = mats
        with open(_VALIDATIONS_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as ex:
        return False, "Echec d'enregistrement : %s" % ex
    auth.journal(login, "Bulletins valides (definitif)",
                 "%s %s %s" % (filiere, niveau, annee), "%d eleve(s)" % len(mats))
    return True, "Validation definitive enregistree (%d eleve(s))." % len(mats)


def attestation_data(matricule, annee):
    """Donnees de l'attestation : identite + decision annuelle (Admis/Ajourne)
    depuis releve_annuel. Renvoie un dict ou None."""
    r = releve_annuel(matricule, annee)
    if not r:
        return None
    prenom = ""
    for e in _lignes_dict("A1_Etudiants"):
        if _stg_get(e, "matricule").lower() == str(matricule).strip().lower():
            prenom = _stg_get(e, "prenom"); break
    return {"matricule": r.get("matricule", matricule), "nom": r.get("nom", ""),
            "prenom": prenom, "filiere": r.get("filiere", ""), "niveau": r.get("niveau", ""),
            "annee": r.get("annee", annee),
            "moyenne": ("%.2f" % r["moyenne"]) if r.get("moyenne") is not None else "",
            "mention": r.get("mention", "") or "",
            "decision": r.get("decision", r.get("proposition", "\u2014")),
            "passe": (r.get("decision") in ("Admis", "Admis conditionnel"))}


def _att_pdf_bytes(d, delib):
    """Construit le PDF de l'attestation (fpdf2, police coeur Helvetica + _lat).
    Le modele Word officiel n'etant pas fourni, mise en page chartee generique."""
    try:
        from fpdf import FPDF
    except Exception:
        raise RuntimeError("Bibliotheque PDF (fpdf2) absente : a vendoriser dans le "
                           "kit (wheel hors-ligne) avant generation d'attestation.")
    BLEU = (31, 78, 121); GRIS = (90, 90, 90)
    pdf = FPDF(format="A4", unit="mm")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    pdf.set_margins(20, 18, 20)
    logo = os.path.join(config.BASE_DIR, "static", "img", "logo_udc.jpg")
    if os.path.exists(logo):
        try:
            pdf.image(logo, x=20, y=14, w=22)
        except Exception:
            pass
    pdf.set_xy(45, 16); pdf.set_text_color(*BLEU)
    for i, ligne in enumerate(getattr(config, "ATTESTATION_ENTETE", [])):
        pdf.set_x(45)
        pdf.set_font("Helvetica", "B" if i < 2 else "", 13 if i == 0 else 11)
        pdf.cell(0, 6, _lat(ligne), ln=1)
    pdf.ln(12)
    pdf.set_text_color(*BLEU); pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "ATTESTATION DE PASSAGE", ln=1, align="C")
    pdf.ln(6)
    pdf.set_text_color(0, 0, 0); pdf.set_font("Helvetica", "", 11)
    nomc = ("%s %s" % (d["prenom"], d["nom"])).strip() or d["nom"]
    corps = ("Le Directeur de l'Ecole de Medecine et de Sante Publique atteste que "
             "l'etudiant(e) designe(e) ci-dessous a fait l'objet d'une deliberation "
             "pour l'annee academique %s :" % d["annee"])
    pdf.multi_cell(0, 7, _lat(corps)); pdf.ln(2)
    for lib, val in [("Nom et prenom", nomc), ("Matricule", d["matricule"]),
                     ("Filiere", d["filiere"]), ("Niveau", d["niveau"]),
                     ("Annee academique", d["annee"]),
                     ("Moyenne annuelle", (d["moyenne"] + " / 20") if d["moyenne"] else "\u2014"),
                     ("Mention", d["mention"] or "\u2014"),
                     ("Decision", d["decision"])]:
        pdf.set_font("Helvetica", "B", 11); pdf.cell(55, 8, _lat(lib + " :"))
        pdf.set_font("Helvetica", "", 11); pdf.cell(0, 8, _lat(str(val)), ln=1)
    pdf.ln(4)
    pdf.set_text_color(*GRIS); pdf.set_font("Helvetica", "I", 9)
    pdf.multi_cell(0, 5, _lat("Delibere le %s par le conseil des professeurs. "
                              "Delai de contestation expire le %s."
                              % (delib.get("date", "\u2014"),
                                 delib.get("date_liberation", "\u2014"))))
    pdf.ln(8)
    pdf.set_text_color(0, 0, 0); pdf.set_font("Helvetica", "", 11)
    today = _dt.date.today().strftime("%d/%m/%Y")
    pdf.cell(0, 7, _lat("Fait a %s, le %s"
                        % (getattr(config, "ATTESTATION_LIEU", "Moroni"), today)),
             ln=1, align="R")
    pdf.ln(12)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, _lat(getattr(config, "ATTESTATION_SIGNATAIRE", "Le Directeur")), ln=1, align="R")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 6, _lat(getattr(config, "ATTESTATION_SIGNATAIRE_NOM", "")), ln=1, align="R")
    return bytes(pdf.output())


def _att_chemin(annee, matricule):
    dossier = os.path.join(config.DOCS_ETUDIANTS_DIR, str(annee).strip())
    nom = "Attestation_passage_%s.pdf" % re.sub(r"[^A-Za-z0-9_-]+", "_", str(matricule).strip())
    return dossier, nom


def generer_attestation(matricule, annee, filiere, niveau, login, sauver=True):
    """Genere l'attestation PDF d'un eleve si son bulletin est VALIDE
    DEFINITIVEMENT (action manuelle) et qu'il passe (Admis ou Admis conditionnel).
    Renvoie (ok, msg, octets, nom_fichier). Copie dans documents/etudiants/<annee>/."""
    if not est_bulletin_valide(matricule, annee, filiere, niveau):
        return (False, "Bulletin non valide definitivement : cochez l'eleve dans la "
                "liste de validation avant de generer l'attestation.", None, None)
    d = attestation_data(matricule, annee)
    if not d:
        return False, "Eleve introuvable ou releve indisponible.", None, None
    if not d["passe"]:
        return (False, "Pas d'attestation de passage (decision : %s)."
                % (d["decision"] or "non admis"), None, None)
    delib = deliberation_classe(annee, filiere, niveau)
    octets = _att_pdf_bytes(d, delib)
    dossier, nom = _att_chemin(annee, matricule)
    if sauver:
        try:
            os.makedirs(dossier, exist_ok=True)
            with open(os.path.join(dossier, nom), "wb") as f:
                f.write(octets)
        except Exception:
            pass
    auth.journal(login, "Attestation de passage generee",
                 "%s %s %s" % (filiere, niveau, annee), matricule)
    return True, "Attestation generee.", octets, nom


def generer_attestations_lot(filiere, niveau, section, annee, login):
    """Genere les attestations de tous les eleves VALIDES qui passent (Admis ou
    Admis conditionnel) d'une classe, copie sur disque, et renvoie un ZIP.
    (ok, msg, octets_zip, n)."""
    valides = bulletins_valides(annee, filiere, niveau)
    if not valides:
        return (False, "Aucun bulletin valide definitivement pour cette classe : "
                "cochez les eleves a attester puis enregistrez la validation.", None, 0)
    delib = deliberation_classe(annee, filiere, niveau)
    roster = roster_classe(filiere, niveau, section)
    buf = _io.BytesIO(); n = 0
    with _zipf.ZipFile(buf, "w", _zipf.ZIP_DEFLATED) as z:
        for e in roster:
            if e["matricule"].strip().lower() not in valides:
                continue
            d = attestation_data(e["matricule"], annee)
            if not d or not d["passe"]:
                continue
            octets = _att_pdf_bytes(d, delib)
            dossier, nom = _att_chemin(annee, e["matricule"])
            z.writestr(nom, octets)
            try:
                os.makedirs(dossier, exist_ok=True)
                with open(os.path.join(dossier, nom), "wb") as f:
                    f.write(octets)
            except Exception:
                pass
            n += 1
    if n == 0:
        return False, "Aucun eleve valide qui passe a attester dans cette classe.", None, 0
    auth.journal(login, "Attestations (lot) generees",
                 "%s %s %s" % (filiere, niveau, annee), "%d eleve(s)" % n)
    return True, "%d attestation(s) generee(s)." % n, buf.getvalue(), n


# ===========================================================================
# R11 Brique 1 — Referentiel des volumes horaires (lecture R1_Maquettes) — V1.99.46
# Source : onglet R1_Maquettes (Filiere, Niveau, Semestre, N UE, Intitule UE,
# Matiere/Contenu, Enseignant, CM, TD, TP, Total heures, Vol. horaire UE).
# Donne les "heures dues" par UE/matiere d'une classe : reference pour le
# compteur de la grille (Brique 2). Override/saisie = via le CRUD generique R1.
# Lecture/agregation pure.
# ===========================================================================

def _r1key(s):
    """Normalise un libelle R1 : minuscules, sans accents, alphanumerique
    (gere 'Filiere', 'Matiere / Contenu', 'N° UE'...)."""
    t = unicodedata.normalize("NFKD", str(s or ""))
    t = "".join(c for c in t if not unicodedata.combining(c)).lower()
    return re.sub(r"[^a-z0-9]+", " ", t).strip()


def _r1_get(row, base):
    b = _r1key(base)
    for k, v in row.items():
        if _r1key(k) == b:
            return ("" if v is None else str(v)).strip()
    return ""


def volumes_options():
    """Filieres / niveaux / semestres distincts presents dans R1_Maquettes."""
    fil, niv, sem = set(), set(), set()
    for r in _lignes_dict("R1_Maquettes"):
        f = _r1_get(r, "filiere"); n = _r1_get(r, "niveau"); s = _r1_get(r, "semestre")
        if f:
            fil.add(f)
        if n:
            niv.add(n)
        if s:
            sem.add(s)
    return {"filieres": sorted(fil), "niveaux": sorted(niv), "semestres": sorted(sem)}


def volumes_classe(filiere, niveau, semestre=""):
    """Volumes horaires d'une classe (filiere, niveau, semestre) depuis
    R1_Maquettes, groupes par UE. Renvoie {ues, total_classe, options, filtres}.
    'heures dues' par UE = somme des Total heures de ses contenus."""
    filiere = (filiere or "").strip(); niveau = (niveau or "").strip()
    semestre = (semestre or "").strip()
    par_ue = {}
    ordre = []
    for r in _lignes_dict("R1_Maquettes"):
        if _r1_get(r, "filiere").lower() != filiere.lower():
            continue
        if _r1_get(r, "niveau").lower() != niveau.lower():
            continue
        if semestre and _r1_get(r, "semestre") != semestre:
            continue
        nue = _r1_get(r, "n ue")
        intit = _r1_get(r, "intitule ue module")
        cle = (nue, intit)
        cm = _fnum(_r1_get(r, "cm")) or 0
        td = _fnum(_r1_get(r, "td")) or 0
        tp = _fnum(_r1_get(r, "tp")) or 0
        tot = _fnum(_r1_get(r, "total heures")) or (cm + td + tp)
        mat = {"matiere": _r1_get(r, "matiere contenu"),
               "enseignant": _r1_get(r, "enseignant"),
               "cm": cm, "td": td, "tp": tp, "total": tot}
        if cle not in par_ue:
            par_ue[cle] = {"n_ue": nue, "intitule": intit, "matieres": [], "total": 0.0}
            ordre.append(cle)
        par_ue[cle]["matieres"].append(mat)
        par_ue[cle]["total"] += tot
    ues = [par_ue[c] for c in ordre]
    total_classe = sum(u["total"] for u in ues)
    return {"ues": ues, "total_classe": total_classe,
            "options": volumes_options(),
            "filtres": {"filiere": filiere, "niveau": niveau, "semestre": semestre}}


# ===========================================================================
# R11 Brique 2 — Editeur de grille hebdomadaire (A3_Sessions) + compteur — V1.99.47
# Gabarit hebdo recurrent (V1a) : on pose Matiere + Salle + Enseignant (ou n°)
# dans (Jour x creneau) par classe. Compteur : heures placees (Σ Vol. horaire
# prog. des seances A3) vs heures dues (R1 via volumes_classe). Detection de
# conflits salle/enseignant sur un meme creneau+jour.
# ===========================================================================

_JOURS_GRILLE = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]


def _a3_get(row, base):
    b = _r1key(base)
    for k, v in row.items():
        if _r1key(k) == b:
            return ("" if v is None else str(v)).strip()
    return ""


def _seances_classe(filiere, niveau, section, annee, semestre):
    """Seances A3 d'une classe (lecture), filtrees. Liste de dicts."""
    out = []
    for r in _lignes_dict("A3_Sessions"):
        if _a3_get(r, "filiere").lower() != filiere.lower():
            continue
        if _a3_get(r, "niveau").lower() != niveau.lower():
            continue
        if section and _a3_get(r, "section").lower() != section.lower():
            continue
        if annee and _a3_get(r, "annee acad") != annee:
            continue
        if semestre and _a3_get(r, "semestre") != semestre:
            continue
        out.append({
            "id": _a3_get(r, "id session"), "jour": _a3_get(r, "jour"),
            "debut": _a3_get(r, "heure debut"), "fin": _a3_get(r, "heure fin"),
            "matiere": _a3_get(r, "matiere"), "salle": _a3_get(r, "salle"),
            "enseignant": _a3_get(r, "enseignant"),
            "vol": _fnum(_a3_get(r, "vol horaire prog")) or 0,
        })
    return out


def grille_classe(filiere, niveau, section, annee, semestre):
    """Grille hebdo d'une classe : seances rangees par (creneau x jour), creneaux
    distincts, compteur heures placees/dues/reste par matiere, conflits, options.
    Lecture/agregation pure."""
    filiere = (filiere or "").strip(); niveau = (niveau or "").strip()
    section = (section or "").strip(); annee = (annee or "").strip()
    semestre = (semestre or "").strip()
    seances = _seances_classe(filiere, niveau, section, annee, semestre)

    creneaux = sorted({(s["debut"], s["fin"]) for s in seances if s["debut"]},
                      key=lambda x: x[0])
    # grille[(debut,fin)][jour] = [seances]
    grille = {}
    conflits = []
    for s in seances:
        cle = (s["debut"], s["fin"])
        grille.setdefault(cle, {}).setdefault(s["jour"], []).append(s)
    # Conflits : meme jour+creneau, meme salle ou meme enseignant sur 2 seances.
    for cle, parj in grille.items():
        for jour, lst in parj.items():
            if len(lst) < 2:
                continue
            for champ, lib in (("salle", "salle"), ("enseignant", "enseignant")):
                vus = {}
                for s in lst:
                    v = (s[champ] or "").strip().lower()
                    if not v:
                        continue
                    if v in vus:
                        conflits.append("%s %s-%s : %s « %s » en double"
                                        % (jour, cle[0], cle[1], lib, s[champ]))
                    vus[v] = True

    # Compteur : dues (R1) par libelle (UE intitule + matiere/contenu), place (A3) par matiere.
    vols = volumes_classe(filiere, niveau, semestre)
    dues = {}
    for u in vols["ues"]:
        if u["intitule"]:
            dues[_r1key(u["intitule"])] = dues.get(_r1key(u["intitule"]), 0) + u["total"]
        for m in u["matieres"]:
            if m["matiere"]:
                dues[_r1key(m["matiere"])] = m["total"]
    place = {}
    for s in seances:
        if s["matiere"]:
            place[s["matiere"]] = place.get(s["matiere"], 0) + s["vol"]
    compteur = []
    for mat in sorted(place):
        d = dues.get(_r1key(mat))
        compteur.append({"matiere": mat, "place": place[mat], "dues": d,
                         "reste": (d - place[mat]) if d is not None else None})
    total_place = sum(place.values())
    total_dues = vols["total_classe"]

    options = {
        "matieres": sorted({m["matiere"] for u in vols["ues"] for m in u["matieres"] if m["matiere"]}
                           | {u["intitule"] for u in vols["ues"] if u["intitule"]}),
        "salles": options_liste("Salles (L1)") or [],
        "enseignants": options_liste("Enseignants (E1)") or [],
        "jours": _JOURS_GRILLE,
        "filieres": vols["options"]["filieres"], "niveaux": vols["options"]["niveaux"],
        "semestres": vols["options"]["semestres"],
    }
    return {"creneaux": creneaux, "jours": _JOURS_GRILLE, "grille": grille,
            "seances": seances, "compteur": compteur, "conflits": conflits,
            "total_place": total_place, "total_dues": total_dues,
            "reste_total": (total_dues - total_place),
            "options": options,
            "filtres": {"filiere": filiere, "niveau": niveau, "section": section,
                        "annee": annee, "semestre": semestre}}


def creer_seance_grille(acteur, annee, semestre, filiere, niveau, section, matiere,
                        enseignant, salle, jour, debut, fin, vol_prog, login):
    """Ajoute une seance A3 (gabarit hebdo) pour la grille. Reserve a l'ecriture
    A3_Sessions. Renvoie (ok, msg)."""
    if not peut_ecrire(acteur, "A3_Sessions"):
        return False, "Action non autorisee."
    if not (filiere and niveau and matiere and jour):
        return False, "Filiere, niveau, matiere et jour sont obligatoires."
    if jour not in _JOURS_GRILLE:
        return False, "Jour invalide."
    if not (_RE_HEURE.match(str(debut or "")) and _RE_HEURE.match(str(fin or ""))):
        return False, "Heures invalides (format HH:MM)."
    if str(fin) <= str(debut):
        return False, "L'heure de fin doit suivre l'heure de debut."
    sid = _prochain_id_session()
    valeurs = {
        _brut("A3_Sessions", "ID session"): sid,
        _brut("A3_Sessions", "Annee acad."): str(annee or _annee_acad_defaut()).strip(),
        _brut("A3_Sessions", "Semestre"): str(semestre or "").strip(),
        _brut("A3_Sessions", "Filiere"): str(filiere).strip(),
        _brut("A3_Sessions", "Niveau"): str(niveau).strip(),
        _brut("A3_Sessions", "Section"): str(section or "").strip(),
        _brut("A3_Sessions", "Matiere"): str(matiere).strip(),
        _brut("A3_Sessions", "Enseignant"): str(enseignant or "").strip(),
        _brut("A3_Sessions", "Salle"): str(salle or "").strip(),
        _brut("A3_Sessions", "Jour"): str(jour).strip(),
        _brut("A3_Sessions", "Heure debut"): str(debut).strip(),
        _brut("A3_Sessions", "Heure fin"): str(fin).strip(),
        _brut("A3_Sessions", "Type"): "Cours",
        _brut("A3_Sessions", "Vol. horaire prog."): (_fnum(vol_prog) or 0),
    }
    _db.ajouter_ligne("A3_Sessions", valeurs)
    auth.journal(login, "Seance planifiee (grille)",
                 "%s %s %s %s-%s" % (filiere, niveau, jour, debut, fin), matiere)
    return True, "Seance ajoutee a la grille (%s)." % sid


def supprimer_seance_grille(acteur, sid, login):
    """Supprime une seance A3 par ID. Reserve a l'ecriture A3_Sessions. (ok, msg)."""
    if not peut_ecrire(acteur, "A3_Sessions"):
        return False, "Action non autorisee."
    sid = str(sid or "").strip()
    if not sid:
        return False, "Identifiant de seance manquant."
    n = _db.supprimer_ligne_par_cle("A3_Sessions", _brut("A3_Sessions", "ID session"), sid)
    if not n:
        return False, "Seance introuvable (%s)." % sid
    auth.journal(login, "Seance supprimee (grille)", sid, "")
    return True, "Seance %s supprimee." % sid


# --- R11 Brique 3 — Impression de la grille hebdomadaire (PDF fpdf2) --------
# Mise en page facon plnfiction : creneaux (lignes) x jours Lundi->Samedi
# (colonnes), une classe = une page A4 paysage ; pied de page = compteur
# (place/du/reste par matiere) + conflits. Lecture pure via grille_classe.

def _grid_nlines(pdf, w, text, lh):
    """Nombre de lignes qu'occupe 'text' dans une cellule de largeur w
    (defensif selon la version de fpdf2)."""
    text = _lat(text)
    try:
        return max(len(pdf.multi_cell(w, lh, text, dry_run=True, output="LINES")), 1)
    except TypeError:
        pass
    try:
        return max(len(pdf.multi_cell(w, lh, text, split_only=True)), 1)
    except TypeError:
        return max(text.count("\n") + 1, 1)


def _grid_entete(pdf, BLEU, titre, sous_titre):
    """En-tete chartee (logo + identite UdC/EMSP + titre + classe)."""
    pdf.set_margins(12, 12, 12)
    logo = os.path.join(config.BASE_DIR, "static", config.LOGO) \
        if getattr(config, "LOGO", "") else ""
    if logo and os.path.exists(logo):
        try:
            pdf.image(logo, x=12, y=9, w=16)
        except Exception:
            pass
    pdf.set_xy(31, 10); pdf.set_text_color(*BLEU)
    for i, ligne in enumerate(getattr(config, "ATTESTATION_ENTETE", [])):
        pdf.set_x(31)
        pdf.set_font("Helvetica", "B" if i < 2 else "", 11 if i == 0 else 9)
        pdf.cell(0, 4.6, _lat(ligne), ln=1)
    pdf.ln(2)
    pdf.set_text_color(*BLEU); pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 7, _lat(titre), ln=1, align="C")
    pdf.set_text_color(0, 0, 0); pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, _lat(sous_titre), ln=1, align="C")
    pdf.ln(2)


def grille_pdf_bytes(filiere, niveau, section, annee, semestre):
    """PDF (A4 paysage) de la grille hebdomadaire d'une classe : creneaux x
    jours Lundi->Samedi, cases (matiere + n + salle + enseignant + volume),
    pied compteur + conflits. Renvoie les octets PDF."""
    try:
        from fpdf import FPDF
    except Exception:
        raise RuntimeError("Bibliotheque PDF (fpdf2) absente : a vendoriser dans le "
                           "kit (wheel hors-ligne) avant generation de la grille.")
    g = grille_classe(filiere, niveau, section, annee, semestre)
    jours = g["jours"]
    BLEU = (31, 78, 121); GRIS = (90, 90, 90); CLAIR = (231, 238, 246)
    ROUGE = (155, 44, 44)

    classe = ("%s %s" % (filiere, niveau)).strip()
    if section:
        classe += " / %s" % section
    bouts = []
    if semestre:
        bouts.append("Semestre %s" % semestre)
    if annee:
        bouts.append(annee)
    sous = classe + ((" — " + " · ".join(bouts)) if bouts else "")

    pdf = FPDF(format="A4", unit="mm", orientation="L")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    _grid_entete(pdf, BLEU, "GRILLE HEBDOMADAIRE", sous)

    # Geometrie du tableau
    x0 = 12.0
    largeur = 297.0 - 2 * 12.0          # 273 mm utiles en paysage
    col_h = 24.0                        # colonne "creneau"
    col_j = (largeur - col_h) / len(jours)
    bas = 200.0                         # bord bas utile (A4 paysage = 210 mm)
    lh = 3.4                            # interligne des cases

    def entete_colonnes(y):
        pdf.set_xy(x0, y); pdf.set_fill_color(*BLEU); pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(col_h, 7, "Creneau", border=1, align="C", fill=True)
        for j in jours:
            pdf.cell(col_j, 7, _lat(j), border=1, align="C", fill=True)
        pdf.ln(7)
        pdf.set_text_color(0, 0, 0)
        return y + 7

    y = entete_colonnes(pdf.get_y())

    def texte_case(seances):
        blocs = []
        for s in seances:
            t = _lat(s.get("matiere") or "(sans matiere)")
            l2 = []
            if s.get("id"):
                l2.append("n %s" % s["id"])
            if s.get("salle"):
                l2.append(str(s["salle"]))
            ligne2 = " · ".join(l2)
            ens = _lat(s.get("enseignant") or "")
            vol = s.get("vol") or 0
            bloc = t
            if ligne2:
                bloc += "\n" + _lat(ligne2)
            if ens:
                bloc += "\n" + ens
            if vol:
                bloc += "\n%g h" % vol
            blocs.append(bloc)
        return "\n— — —\n".join(blocs) if blocs else ""

    if not g["creneaux"]:
        pdf.set_font("Helvetica", "I", 10); pdf.set_text_color(*GRIS)
        pdf.set_xy(x0, y + 4)
        pdf.cell(largeur, 8, _lat("Aucune seance placee pour cette classe."),
                 border=1, align="C")
        y += 12
    else:
        for (deb, fin) in g["creneaux"]:
            parj = g["grille"].get((deb, fin), {})
            textes = [texte_case(parj.get(j, [])) for j in jours]
            # hauteur de la ligne = max des cases (en lignes), bornee
            nl = max([1] + [_grid_nlines(pdf, col_j - 2, t, lh) for t in textes if t])
            row_h = max(10.0, nl * lh + 3.0)
            if y + row_h > bas:
                pdf.add_page()
                _grid_entete(pdf, BLEU, "GRILLE HEBDOMADAIRE (suite)", sous)
                y = entete_colonnes(pdf.get_y())
            # cellule creneau
            pdf.set_fill_color(*CLAIR); pdf.set_draw_color(150, 150, 150)
            pdf.rect(x0, y, col_h, row_h, style="DF")
            pdf.set_xy(x0, y + row_h / 2 - 3)
            pdf.set_font("Helvetica", "B", 8); pdf.set_text_color(*BLEU)
            pdf.multi_cell(col_h, 3, _lat("%s\n%s" % (deb, fin)), align="C")
            # cases jours
            pdf.set_text_color(0, 0, 0)
            for k, t in enumerate(textes):
                cx = x0 + col_h + k * col_j
                pdf.rect(cx, y, col_j, row_h)
                if t:
                    pdf.set_xy(cx + 1, y + 1.4)
                    pdf.set_font("Helvetica", "", 7)
                    pdf.multi_cell(col_j - 2, lh, _lat(t), align="L")
            y += row_h

    # ---- Pied : compteur (place/du/reste) + conflits ----
    def saut_si_besoin(h):
        nonlocal y
        if y + h > bas:
            pdf.add_page()
            _grid_entete(pdf, BLEU, "GRILLE HEBDOMADAIRE (suite)", sous)
            y = pdf.get_y()

    y += 4
    saut_si_besoin(40)
    pdf.set_xy(x0, y); pdf.set_text_color(*BLEU); pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, _lat("Couverture horaire par matiere"), ln=1); y += 6
    pdf.set_text_color(255, 255, 255); pdf.set_fill_color(*BLEU)
    pdf.set_font("Helvetica", "B", 8)
    cw = [largeur * 0.46, largeur * 0.18, largeur * 0.18, largeur * 0.18]
    pdf.set_xy(x0, y)
    for lib, w in zip(("Matiere", "Heures placees", "Heures dues", "Reste"), cw):
        pdf.cell(w, 6, _lat(lib), border=1, align="C", fill=True)
    pdf.ln(6); y += 6
    pdf.set_text_color(0, 0, 0); pdf.set_font("Helvetica", "", 8)
    for c in g["compteur"]:
        saut_si_besoin(6)
        d = c["dues"]; r = c["reste"]
        pdf.set_xy(x0, y)
        pdf.cell(cw[0], 5.5, _lat(c["matiere"]), border=1)
        pdf.cell(cw[1], 5.5, "%g h" % c["place"], border=1, align="C")
        pdf.cell(cw[2], 5.5, _lat(("%g h" % d) if d is not None else "—"), border=1, align="C")
        if r is None:
            pdf.set_text_color(*GRIS); txt = "—"
        elif r < 0:
            pdf.set_text_color(*ROUGE); txt = "%g h (depassement)" % r
        else:
            pdf.set_text_color(0, 0, 0); txt = "%g h" % r
        pdf.cell(cw[3], 5.5, _lat(txt), border=1, align="C")
        pdf.set_text_color(0, 0, 0)
        pdf.ln(5.5); y += 5.5
    # total classe
    saut_si_besoin(6)
    pdf.set_xy(x0, y); pdf.set_font("Helvetica", "B", 8); pdf.set_fill_color(*CLAIR)
    pdf.cell(cw[0], 5.5, _lat("Total classe"), border=1, fill=True)
    pdf.cell(cw[1], 5.5, "%g h" % g["total_place"], border=1, align="C", fill=True)
    pdf.cell(cw[2], 5.5, "%g h" % g["total_dues"], border=1, align="C", fill=True)
    rt = g["reste_total"]
    pdf.set_text_color(*(ROUGE if rt < 0 else (0, 0, 0)))
    pdf.cell(cw[3], 5.5, "%g h" % rt, border=1, align="C", fill=True)
    pdf.set_text_color(0, 0, 0); pdf.ln(5.5); y += 5.5

    if g["conflits"]:
        y += 4; saut_si_besoin(10 + 4 * len(g["conflits"]))
        pdf.set_xy(x0, y); pdf.set_text_color(*ROUGE); pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 6, _lat("Conflits detectes (%d)" % len(g["conflits"])), ln=1); y += 6
        pdf.set_font("Helvetica", "", 8); pdf.set_text_color(0, 0, 0)
        for c in g["conflits"]:
            pdf.set_x(x0); pdf.multi_cell(largeur, 4.4, _lat("- " + c)); y = pdf.get_y()

    # bas de page : edition
    pdf.set_y(202); pdf.set_text_color(*GRIS); pdf.set_font("Helvetica", "I", 7)
    pdf.cell(0, 4, _lat("Edite le %s — EMSP, Universite des Comores"
                        % _dt.date.today().strftime("%d/%m/%Y")), align="C")
    return bytes(pdf.output())


def grille_pdf_nom(filiere, niveau, section, annee, semestre):
    """Nom de fichier propose pour le PDF de grille d'une classe."""
    bout = "_".join([str(x).strip() for x in (filiere, niveau, section, semestre, annee)
                     if str(x).strip()])
    bout = re.sub(r"[^A-Za-z0-9_-]+", "_", bout).strip("_") or "classe"
    return "Grille_hebdo_%s.pdf" % bout


# ===========================================================================
# V1.87 — Affectation automatique des stages : MOTEUR DE PROPOSITION
# (lecture seule ; regles validees par le Dr Kamal, reunion du 26/06/2026)
# ===========================================================================

def _sess_norm(s):
    """Normalise la Session d'un stage : '' -> 'Normale' ; tout ce qui commence
    par 'rattr' -> 'Rattrapage' ; sinon 'Normale'. Insensible casse/accents."""
    t = _norm_liste(s or "")
    return "Rattrapage" if t.startswith("rattr") else "Normale"


def donnees_affectation_stages(filiere, niveau, annee, seance, session="Normale"):
    """Donnees de l'ecran d'affectation MANUELLE des stages (cochage par groupe),
    pour une SESSION donnee (Normale = stage initial ; Rattrapage). Lecture seule.
      - roster : eleves de A1 avec leur affectation pour cette seance ET session ;
                 en mode Rattrapage, restreint aux eleves ayant un stage Normal
                 (rattrapables) pour cette seance ;
      - lieux  : lieux du niveau (S2) avec quota / occupe / restant POUR cette session ;
      - sections : affectations (cette session) groupees par lieu ;
      - stats  : effectif, affectes, non affectes, capacite.
    Le lieu est identifie par le libelle composite "Lieu / structure — Service"."""
    annee = str(annee or "").strip()
    seance = str(seance or "").strip()
    session = _sess_norm(session)
    roster = roster_classe(filiere, niveau, "")   # sections de cours ignorees

    # Lignes S1 de (annee, seance), indexees par session normalisee.
    par_sess = {"Normale": {}, "Rattrapage": {}}
    for r in _lignes_dict("S1_Stages"):
        if _stg_get(r, "annee acad.") != annee or _stg_get(r, "n seance (1-6)") != seance:
            continue
        m = _stg_get(r, "matricule")
        if not m:
            continue
        sess = _sess_norm(_stg_get(r, "session"))
        par_sess[sess][m.lower()] = {"lieu": _stg_get(r, "lieu de stage"),
                                     "debut": _stg_get(r, "date debut"),
                                     "fin": _stg_get(r, "date fin"),
                                     "note": _stg_get(r, "note stage /20")}
    affect = par_sess[session]
    normaux = par_sess["Normale"]   # base des rattrapables

    # Roster (restreint aux rattrapables si mode Rattrapage).
    rost = []
    for e in roster:
        ml = e["matricule"].lower()
        if session == "Rattrapage" and ml not in normaux:
            continue
        a = affect.get(ml)
        rost.append({"matricule": e["matricule"], "nom": e["nom"], "prenom": e["prenom"],
                     "affecte": bool(a), "lieu_actuel": a["lieu"] if a else "",
                     "normal_lieu": normaux.get(ml, {}).get("lieu", "")})

    # Occupation par lieu POUR cette session.
    occ_par_lieu = {}
    for a in affect.values():
        k = (a["lieu"] or "").strip().lower()
        if k:
            occ_par_lieu[k] = occ_par_lieu.get(k, 0) + 1
    lieux = []
    for r in _lignes_dict("S2_Lieux_stage"):
        nl = _stg_get(r, "lieu / structure")
        if not nl:
            continue
        niv = _stg_get(r, "niveau concerne")
        if niv and niveau and niv.strip().lower() != str(niveau).strip().lower():
            continue
        serv = _stg_get(r, "service")
        lib = nl + (" \u2014 " + serv if serv else "")
        quota = int(_fnum(_stg_get(r, "quota")) or 0)
        occ = occ_par_lieu.get(lib.strip().lower(), 0)
        lieux.append({"lib": lib, "quota": quota, "occupe": occ,
                      "restant": max(quota - occ, 0), "commune": _stg_get(r, "commune"),
                      "periode": _stg_get(r, "periode de disponibilite")})
    lieux.sort(key=lambda l: l["lib"].lower())

    # Sections de fait (cette session) groupees par lieu.
    nom_par_mat = {e["matricule"].lower(): e for e in roster}
    par_lieu = {}
    for mat_l, a in affect.items():
        lib = a["lieu"] or "(sans lieu)"
        e = nom_par_mat.get(mat_l, {"matricule": mat_l, "nom": "", "prenom": ""})
        par_lieu.setdefault(lib, []).append(
            {"matricule": e.get("matricule", mat_l), "nom": e.get("nom", ""),
             "prenom": e.get("prenom", ""), "debut": a["debut"], "fin": a["fin"]})
    sections = []
    for lib in sorted(par_lieu):
        membres = sorted(par_lieu[lib], key=lambda x: (x["nom"].lower(), x["prenom"].lower()))
        sections.append({"lieu": lib, "eleves": membres})

    stats = {"effectif": len(rost),
             "affectes": sum(1 for e in rost if e["affecte"]),
             "non_affectes": sum(1 for e in rost if not e["affecte"]),
             "capacite": sum(l["quota"] for l in lieux), "nb_lieux": len(lieux)}
    return {"roster": rost, "lieux": lieux, "sections": sections, "stats": stats,
            "filiere": filiere, "niveau": niveau, "annee": annee, "seance": seance,
            "session": session}


def affecter_groupe_stages(matricules, annee, seance, lieu, date_debut, date_fin,
                           depassement=False, session="Normale"):
    """Affecte un GROUPE d'eleves coches a un lieu/service pour une seance et une
    SESSION (Normale / Rattrapage). Ecrit 1 ligne/eleve dans S1_Stages (upsert
    Matricule + Annee + Seance + Session) : la ligne Rattrapage COEXISTE avec la
    Normale (elle ne l'ecrase pas). Note, fiche retour et observation existantes
    sont conservees. Controle de quota par lieu POUR cette session : refus si le
    groupe depasse les places restantes, sauf depassement=True. Renvoie (ok, msg, n)."""
    annee = str(annee or "").strip()
    seance = str(seance or "").strip()
    lieu = str(lieu or "").strip()
    session = _sess_norm(session)
    date_debut = str(date_debut or "").strip()
    date_fin = str(date_fin or "").strip()
    mats = [str(m).strip() for m in (matricules or []) if str(m).strip()]
    if not mats:
        return False, "Aucun eleve coche.", 0
    if not lieu:
        return False, "Choisissez un lieu de stage.", 0
    if not (date_debut and date_fin):
        return False, "Renseignez les dates de debut et de fin.", 0

    # Quota du lieu choisi (S2).
    quota = None
    for r in _lignes_dict("S2_Lieux_stage"):
        nl = _stg_get(r, "lieu / structure")
        if not nl:
            continue
        serv = _stg_get(r, "service")
        lib = nl + (" \u2014 " + serv if serv else "")
        if lib.strip().lower() == lieu.strip().lower():
            quota = int(_fnum(_stg_get(r, "quota")) or 0)
            break

    # Occupation du lieu pour (annee, seance, SESSION), hors eleves (re)affectes.
    mats_low = set(m.lower() for m in mats)
    occ = 0
    for r in _lignes_dict("S1_Stages"):
        if _stg_get(r, "annee acad.") != annee or _stg_get(r, "n seance (1-6)") != seance:
            continue
        if _sess_norm(_stg_get(r, "session")) != session:
            continue
        if _stg_get(r, "lieu de stage").strip().lower() != lieu.strip().lower():
            continue
        if _stg_get(r, "matricule").lower() in mats_low:
            continue
        occ += 1

    depasse = (quota is not None and len(mats) > (quota - occ))
    if depasse and not depassement:
        restant = max((quota or 0) - occ, 0)
        return (False, "Capacite insuffisante : %d place(s) restante(s) sur ce lieu pour "
                       "%d eleve(s) coche(s). Cochez « autoriser le depassement » pour forcer."
                       % (restant, len(mats)), 0)

    lignes = [{"Matricule (*)": m, "Annee acad. (*)": annee, "N seance (1-6) (*)": seance,
               "Lieu de stage (*)": lieu, "Date debut (*)": date_debut,
               "Date fin (*)": date_fin, "Session": session} for m in mats]
    res = _db.ecrire_lignes_lot(
        "S1_Stages", lignes,
        cles=["Matricule (*)", "Annee acad. (*)", "N seance (1-6) (*)", "Session"])
    n = res.get("ajout", 0) + res.get("maj", 0)
    sur = " (depassement de quota signale)" if depasse else ""
    return True, "Groupe affecte (%s) : %d eleve(s) sur « %s »%s." % (session, n, lieu, sur), n


def _stages_note_effective(stages):
    """A partir d'une liste de stages bruts [{matricule, annee, seance, session,
    note, ...}], renvoie un dict {(matricule, annee, seance): note_effective} :
    la note du Rattrapage prime SI elle est saisie, sinon celle de la Normale."""
    par_cle = {}
    for s in stages:
        cle = (str(s.get("matricule", "")).strip().lower(),
               str(s.get("annee", "")).strip(), str(s.get("seance", "")).strip())
        sess = _sess_norm(s.get("session", ""))
        note = str(s.get("note", "")).strip()
        par_cle.setdefault(cle, {})[sess] = note
    eff = {}
    for cle, d in par_cle.items():
        ratt = d.get("Rattrapage", "")
        eff[cle] = ratt if ratt != "" else d.get("Normale", "")
    return eff


# ===========================================================================
# BLOC 3 — ETATS DE PAIEMENT DES VACATIONS (heures -> montant -> compta)
# V1.99.3. Hypotheses (signalees, ajustables) :
#   - 1 ligne par ENSEIGNANT (source E2_Releve_heures, "pour la paie") ;
#     Matiere/Niveau restent optionnels (non remplis automatiquement en V1).
#   - L'etat couvre une LISTE DE MOIS choisis (Mois / Annee de E2) ;
#     Semestre + Annee academique sont les etiquettes de l'etat.
#   - Heures autorisees a payer = heures effectuees par defaut (plafond
#     ajustable avant l'arrete).
#   - Mode Forfait mensuel (moniteurs) : repris depuis E1 (Mode + Cout
#     mensuel), nb de mois = nb de mois inclus par defaut (ajustable).
#   - Passage en compta : 1 depense F1 par enseignant, reference commune
#     "PAIE-<id>", garde-fou anti-double-passage par le Statut de l'etat.
# ===========================================================================

_ETAT_TAB = "E4_Etats_paiement"


def _be4(lib):
    """En-tete brut d'une colonne E4 (sans marqueur -> libelle tel quel)."""
    return _brut(_ETAT_TAB, lib)


def taux_horaire_effectif(ens):
    """Taux horaire applicable (KMF/h) : override E1 sinon defaut global."""
    v = _fnum((ens or {}).get("Taux horaire (KMF/h)", ""))
    return v if v > 0 else float(config.TAUX_HORAIRE_DEFAUT)


def _mode_remu(ens):
    """Mode de remuneration d'un enseignant : 'Horaire' (defaut) ou 'Forfait mensuel'."""
    m = str((ens or {}).get("Mode de remuneration", "")).strip()
    return m if m in config.MODES_REMUNERATION else "Horaire"


def id_etat_paiement(annee_acad, semestre):
    """Identifiant stable d'un etat : 'PAIE-<annee>-<semestre>'."""
    a = str(annee_acad or "").strip().replace(" ", "")
    s = str(semestre or "").strip().upper()
    return "PAIE-%s-%s" % (a, s)


# --- Montant en toutes lettres (francais, entiers KMF, hors-ligne) ----------
_UNITES = ["zero", "un", "deux", "trois", "quatre", "cinq", "six", "sept",
           "huit", "neuf", "dix", "onze", "douze", "treize", "quatorze",
           "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
_DIZAINES = {20: "vingt", 30: "trente", 40: "quarante", 50: "cinquante",
             60: "soixante", 80: "quatre-vingt"}


def _lettres_sous_cent(n):
    if n < 20:
        return _UNITES[n]
    if n < 70 or (80 <= n < 100):
        d = (n // 10) * 10
        u = n % 10
        base = _DIZAINES[80 if d in (80, 90) else d]
        if d == 80 and u == 0:
            return base + "s"
        if u == 0:
            return base
        if u == 1 and d in (20, 30, 40, 50, 60):
            return base + "-et-un"
        return base + "-" + _UNITES[u]
    # 70-79 et 90-99 : soixante-dix / quatre-vingt-dix
    d = 60 if n < 80 else 80
    base = _DIZAINES[d]
    reste = n - d
    if reste == 11 and d == 60:
        return base + "-et-onze"
    return base + "-" + _UNITES[reste]


def _lettres_sous_mille(n):
    if n == 0:
        return ""
    c = n // 100
    r = n % 100
    if c == 0:
        return _lettres_sous_cent(r)
    cent = "cent" if c == 1 else (_UNITES[c] + " cent")
    if r == 0:
        return cent + ("s" if c > 1 else "")
    return cent + " " + _lettres_sous_cent(r)


def montant_en_lettres(n):
    """Entier positif -> francais ('cent trois mille cinq cents'). KMF entiers."""
    n = int(round(_fnum(n)))
    if n == 0:
        return "zero"
    if n < 0:
        return "moins " + montant_en_lettres(-n)
    tranches = []  # (valeur, libelle d'echelle, pluriel possible)
    echelles = [(10 ** 9, "milliard"), (10 ** 6, "million"),
                (10 ** 3, "mille"), (1, "")]
    out = []
    for val, nom in echelles:
        q = n // val
        n = n % val
        if q == 0:
            continue
        if nom == "mille":
            mot = "" if q == 1 else (_lettres_sous_mille(q) + " ")
            out.append((mot + "mille").strip())
        elif nom == "":
            out.append(_lettres_sous_mille(q))
        else:
            mot = _lettres_sous_mille(q)
            pl = "s" if q > 1 else ""
            out.append(mot + " " + nom + pl)
    return " ".join(x for x in out if x).strip()


def mois_disponibles_paie():
    """Liste triee des 'Mois / Annee' presents dans E2 (pour le selecteur d'etat)."""
    vus = []
    for r in _lignes_dict("E2_Releve_heures"):
        m = str(r.get("Mois / Annee", "")).strip()
        if m and m not in vus:
            vus.append(m)
    def _cle(m):
        mm = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{4})\s*$", m)
        return (int(mm.group(2)), int(mm.group(1))) if mm else (9999, 99)
    return sorted(vus, key=_cle)


def _heures_par_enseignant(mois_inclus=None):
    """Agrege E2 par matricule sur les mois retenus (None = tous).
    -> {matricule: {'prog': float, 'constate': float}}."""
    inc = set(m.strip() for m in mois_inclus) if mois_inclus else None
    agg = {}
    for r in _lignes_dict("E2_Releve_heures"):
        mat = str(r.get("Matricule ens.", "")).strip()
        if not mat:
            continue
        mois = str(r.get("Mois / Annee", "")).strip()
        if inc is not None and mois not in inc:
            continue
        a = agg.setdefault(mat, {"prog": 0.0, "constate": 0.0})
        a["prog"] += _fnum(r.get("Vol. horaire prog.", ""))
        a["constate"] += _fnum(r.get("Vol. horaire constate", ""))
    return agg


def apercu_etat_paiement(annee_acad, semestre, mois_inclus=None,
                         autorisations=None, mois_forfait=None):
    """Calcule (LECTURE SEULE) les lignes d'un etat de paiement.
    autorisations = {matricule: heures} pour plafonner (defaut = effectuees).
    mois_forfait  = {matricule: nb_mois} pour les forfaits (defaut = nb mois inclus).
    -> {'id', 'annee', 'semestre', 'mois', 'lignes':[...], 'total', 'total_lettres',
        'statut'} ; statut = etat existant s'il y en a un, sinon 'Brouillon'."""
    autorisations = autorisations or {}
    mois_forfait = mois_forfait or {}
    eid = id_etat_paiement(annee_acad, semestre)
    # index E1 par matricule (dict propre complet)
    e1 = {}
    for x in _lignes_dict("E1_Enseignants"):
        mat = str(x.get("Matricule ens.", "")).strip()
        if mat:
            e1[mat] = x

    heures = _heures_par_enseignant(mois_inclus)
    nb_mois_defaut = len(mois_inclus) if mois_inclus else len(mois_disponibles_paie())
    lignes = []

    # 1) HORAIRE : tous les enseignants ayant des heures sur la periode
    for mat, h in heures.items():
        ens = e1.get(mat, {})
        if _mode_remu(ens) == "Forfait mensuel":
            continue  # traite en 2)
        taux = taux_horaire_effectif(ens)
        effect = h["constate"]
        auto = _fnum(autorisations.get(mat, effect))
        if auto < 0:
            auto = 0.0
        if auto > effect and mat not in autorisations:
            auto = effect
        montant = int(round(auto * taux))
        lignes.append(_ligne_etat(eid, ens, mat, "Horaire",
                                  h["prog"], effect, auto, taux, "", "", montant,
                                  annee_acad, semestre))

    # 2) FORFAIT mensuel : enseignants E1 en mode forfait avec un cout mensuel
    for mat, ens in e1.items():
        if _mode_remu(ens) != "Forfait mensuel":
            continue
        cout = _fnum(ens.get("Cout mensuel (KMF)", ""))
        if cout <= 0:
            continue
        nbm = mois_forfait.get(mat, nb_mois_defaut)
        try:
            nbm = int(nbm)
        except (TypeError, ValueError):
            nbm = nb_mois_defaut
        montant = int(round(nbm * cout))
        lignes.append(_ligne_etat(eid, ens, mat, "Forfait mensuel",
                                  "", "", "", "", nbm, cout, montant,
                                  annee_acad, semestre))

    lignes.sort(key=lambda d: (d["nom"].lower(), d["prenom"].lower(), d["matricule"]))
    total = sum(d["montant"] for d in lignes)
    # statut courant si l'etat existe deja
    statut = "Brouillon"
    for r in _lignes_dict(_ETAT_TAB):
        if str(r.get("ID etat", "")).strip() == eid:
            statut = str(r.get("Statut etat", "")).strip() or "Brouillon"
            break
    return {"id": eid, "annee": str(annee_acad).strip(),
            "semestre": str(semestre).strip(), "mois": list(mois_inclus or []),
            "lignes": lignes, "total": total,
            "total_lettres": montant_en_lettres(total) + " francs comoriens (KMF)",
            "statut": statut}


def _ligne_etat(eid, ens, mat, mode, prog, effect, auto, taux,
                nb_mois, cout, montant, annee, semestre):
    nom = str((ens or {}).get("Nom", "")).strip()
    pre = str((ens or {}).get("Prenom", "")).strip()
    return {
        "id": eid, "matricule": mat, "nom": nom, "prenom": pre,
        "nom_complet": (nom + " " + pre).strip(),
        "statut_prof": str((ens or {}).get("Statut", "")).strip(),
        "mode": mode,
        "matiere": str((ens or {}).get("Matieres enseignees", "")).strip(),
        "niveau": "",
        "h_prevues": prog, "h_effectuees": effect, "h_autorisees": auto,
        "taux": taux, "nb_mois": nb_mois, "cout_mensuel": cout,
        "montant": montant,
    }


def constituer_etat_paiement(annee_acad, semestre, mois_inclus, saisi_par,
                             autorisations=None, mois_forfait=None):
    """Ecrit/MET A JOUR un etat (Statut 'Brouillon') dans E4 (upsert ID+Matricule).
    Refuse si l'etat est deja Arrete ou Passe en compta. -> (ok, message, id)."""
    ap = apercu_etat_paiement(annee_acad, semestre, mois_inclus,
                              autorisations, mois_forfait)
    eid = ap["id"]
    if ap["statut"] in ("Arrete", "Passe en compta"):
        return False, ("Etat « %s » deja %s : il ne peut plus etre recalcule."
                       % (eid, ap["statut"].lower())), eid
    if not ap["lignes"]:
        return False, "Aucune ligne a payer pour cette periode.", eid
    lignes_brutes = []
    for d in ap["lignes"]:
        lignes_brutes.append({
            _be4("ID etat"): eid,
            _be4("Semestre"): ap["semestre"],
            _be4("Annee academique"): ap["annee"],
            _be4("Statut etat"): "Brouillon",
            _be4("Date arrete"): "",
            _be4("Arrete par"): "",
            _be4("Ref ecriture F1"): "",
            _be4("Matricule ens."): d["matricule"],
            _be4("Nom-Prenom"): d["nom_complet"],
            _be4("Statut prof."): d["statut_prof"],
            _be4("Mode remuneration"): d["mode"],
            _be4("Matiere enseignee"): d["matiere"],
            _be4("Niveau"): d["niveau"],
            _be4("Heures prevues"): d["h_prevues"],
            _be4("Heures effectuees"): d["h_effectuees"],
            _be4("Heures autorisees a payer"): d["h_autorisees"],
            _be4("Taux horaire (KMF/h)"): (d["taux"] if d["mode"] == "Horaire" else ""),
            _be4("Mois (forfait)"): d["nb_mois"],
            _be4("Cout mensuel (KMF)"): d["cout_mensuel"],
            _be4("Montant (KMF)"): d["montant"],
        })
    cles = [_be4("ID etat"), _be4("Matricule ens.")]
    res = _db.ecrire_lignes_lot(_ETAT_TAB, lignes_brutes, cles=cles)
    return True, ("Etat « %s » constitue : %d ligne(s), total %s KMF."
                  % (eid, len(lignes_brutes), _fmt_kmf(ap["total"]))), eid


def etats_paiement():
    """Liste des etats presents dans E4, groupes par ID -> en-tete + total."""
    grp = {}
    for r in _lignes_dict(_ETAT_TAB):
        eid = str(r.get("ID etat", "")).strip()
        if not eid:
            continue
        g = grp.setdefault(eid, {"id": eid,
                                 "semestre": str(r.get("Semestre", "")).strip(),
                                 "annee": str(r.get("Annee academique", "")).strip(),
                                 "statut": str(r.get("Statut etat", "")).strip() or "Brouillon",
                                 "date_arrete": str(r.get("Date arrete", "")).strip(),
                                 "arrete_par": str(r.get("Arrete par", "")).strip(),
                                 "nb": 0, "total": 0})
        g["nb"] += 1
        g["total"] += int(round(_fnum(r.get("Montant (KMF)", ""))))
    return sorted(grp.values(), key=lambda x: x["id"], reverse=True)


def detail_etat_paiement(eid):
    """Detail d'un etat : en-tete + lignes + total + total en lettres."""
    eid = str(eid).strip()
    lignes, statut, sem, annee, darr, parr, ref = [], "", "", "", "", "", ""
    for r in _lignes_dict(_ETAT_TAB):
        if str(r.get("ID etat", "")).strip() != eid:
            continue
        statut = str(r.get("Statut etat", "")).strip() or "Brouillon"
        sem = str(r.get("Semestre", "")).strip()
        annee = str(r.get("Annee academique", "")).strip()
        darr = str(r.get("Date arrete", "")).strip()
        parr = str(r.get("Arrete par", "")).strip()
        ref = str(r.get("Ref ecriture F1", "")).strip()
        lignes.append({
            "matricule": str(r.get("Matricule ens.", "")).strip(),
            "nom_complet": str(r.get("Nom-Prenom", "")).strip(),
            "statut_prof": str(r.get("Statut prof.", "")).strip(),
            "mode": str(r.get("Mode remuneration", "")).strip(),
            "matiere": str(r.get("Matiere enseignee", "")).strip(),
            "niveau": str(r.get("Niveau", "")).strip(),
            "h_prevues": _fmt_h(_fnum(r.get("Heures prevues", ""))),
            "h_effectuees": _fmt_h(_fnum(r.get("Heures effectuees", ""))),
            "h_autorisees": _fmt_h(_fnum(r.get("Heures autorisees a payer", ""))),
            "taux": _fmt_kmf(_fnum(r.get("Taux horaire (KMF/h)", ""))),
            "nb_mois": str(r.get("Mois (forfait)", "")).strip(),
            "cout_mensuel": _fmt_kmf(_fnum(r.get("Cout mensuel (KMF)", ""))),
            "montant": int(round(_fnum(r.get("Montant (KMF)", "")))),
        })
    if not lignes:
        return None
    lignes.sort(key=lambda d: d["nom_complet"].lower())
    total = sum(d["montant"] for d in lignes)
    return {"id": eid, "semestre": sem, "annee": annee, "statut": statut,
            "date_arrete": darr, "arrete_par": parr, "ref_f1": ref,
            "lignes": lignes, "total": total,
            "total_lettres": montant_en_lettres(total) + " francs comoriens (KMF)"}


def _maj_statut_etat(eid, valeurs_propres):
    """Met a jour des colonnes d'en-tete sur TOUTES les lignes d'un etat (upsert
    sur cle ID+Matricule -> les montants/heures sont preserves)."""
    lignes = []
    for r in _lignes_dict(_ETAT_TAB):
        if str(r.get("ID etat", "")).strip() != str(eid).strip():
            continue
        d = {_be4("ID etat"): eid,
             _be4("Matricule ens."): str(r.get("Matricule ens.", "")).strip()}
        for k, v in valeurs_propres.items():
            d[_be4(k)] = v
        lignes.append(d)
    if not lignes:
        return 0
    cles = [_be4("ID etat"), _be4("Matricule ens.")]
    _db.ecrire_lignes_lot(_ETAT_TAB, lignes, cles=cles)
    return len(lignes)


def arreter_etat_paiement(eid, par):
    """Fige l'etat : Statut Brouillon -> Arrete, pose Date arrete + Arrete par.
    Refuse si deja Arrete/Passe. -> (ok, message)."""
    d = detail_etat_paiement(eid)
    if d is None:
        return False, "Etat introuvable."
    if d["statut"] == "Passe en compta":
        return False, "Etat deja passe en compta : non modifiable."
    if d["statut"] == "Arrete":
        return False, "Etat deja arrete."
    auj = datetime.date.today().strftime("%d/%m/%Y")
    n = _maj_statut_etat(eid, {"Statut etat": "Arrete",
                               "Date arrete": auj, "Arrete par": str(par).strip()})
    return True, ("Etat arrete (%d ligne(s)) le %s. Total %s KMF — %s."
                  % (n, auj, _fmt_kmf(d["total"]), d["total_lettres"]))


def passer_etat_en_compta(eid, par, compte, mode_paiement, poste=None, date_op=None):
    """Garde-fou : exige Statut=Arrete. Ecrit 1 depense F1 par enseignant, pose la
    reference commune, passe l'etat a 'Passe en compta', journalise. -> (ok, msg)."""
    d = detail_etat_paiement(eid)
    if d is None:
        return False, "Etat introuvable."
    if d["statut"] == "Passe en compta":
        return False, "Etat deja passe en compta (anti-double-passage)."
    if d["statut"] != "Arrete":
        return False, "Arretez l'etat avant de le passer en compta."
    if not str(compte).strip():
        return False, "Choisissez un compte / caisse."
    if not str(mode_paiement).strip():
        return False, "Choisissez un mode de paiement."

    date_op = (date_op or datetime.date.today().strftime("%d/%m/%Y")).strip()
    poste = (poste if poste is not None else config.POSTE_DEPENSE_VACATIONS) or ""
    ref = "PAIE-%s" % eid.replace("PAIE-", "")
    b = lambda lib: _brut("F1_Mouvements", lib)
    depenses = []
    for ln in d["lignes"]:
        if ln["montant"] <= 0:
            continue
        depenses.append({
            b("Date operation"): date_op,
            b("Sens"): "Depense",
            b("Poste budgetaire"): poste,
            b("Mode paiement"): str(mode_paiement).strip(),
            b("Compte / caisse"): str(compte).strip(),
            b("Reference / N piece"): ref,
            b("Libelle / description"): ("Vacation %s %s — %s"
                                         % (d["semestre"], d["annee"], ln["nom_complet"])),
            b("Montant Depense (KMF)"): ln["montant"],
            b("Tiers"): ln["nom_complet"],
            b("Saisi par"): str(par).strip(),
            b("Matricule"): ln["matricule"],
            b("Annee academique"): d["annee"],
        })
    if not depenses:
        return False, "Aucun montant a passer en compta."
    try:
        n = _db.ajouter_lignes("F1_Mouvements", depenses)
    except OverflowError:
        return False, "Capacite de la tresorerie atteinte."
    except Exception:
        return False, "Echec de l'ecriture comptable."

    _maj_statut_etat(eid, {"Statut etat": "Passe en compta", "Ref ecriture F1": ref})
    # NB : l'audit (journal.csv) est ecrit par la couche app/route via auth,
    # apres succes (separation des couches) — pas ici.
    return True, ("Passe en compta : %d depense(s) F1 (ref %s), total %s KMF."
                  % (n, ref, _fmt_kmf(d["total"])))


# ===========================================================================
# BIBLIOTHEQUE DOCUMENTAIRE (V1.99.22) — magasin de fichiers hors-ligne.
# Inspire de la bibliotheque GMAO : un dossier racine fixe (config.BIBLIOTHEQUE_DIR),
# rangement libre en sous-dossiers. Aucune donnee dans le classeur : tout est
# fichier sur le poste. Anti-traversee de repertoire systematique.
# ===========================================================================
def a_droit_ecriture(role_row):
    """Vrai si le role possede UN droit d'ecriture (sur au moins un module) ou
    est administrateur. Sert de garde au DEPOT dans la bibliotheque (la
    suppression, elle, reste reservee a l'administration : est_admin)."""
    return bool(est_admin(role_row)
                or _expanse_groupes(role_row.get("ecriture", "")))


def _biblio_base():
    base = os.path.abspath(config.BIBLIOTHEQUE_DIR)
    if not os.path.isdir(base):
        os.makedirs(base, exist_ok=True)
        try:
            with open(os.path.join(base, "_LISEZ-MOI.txt"), "w",
                      encoding="utf-8") as fh:
                fh.write("Bibliotheque documentaire de l'EMSP.\r\n"
                         "Deposez ici les documents importants : strategiques, "
                         "officiels, supports de cours...\r\n"
                         "Rangement libre en sous-dossiers.\r\n")
        except OSError:
            pass
    return base


def _biblio_abspath(sous):
    """Chemin absolu d'un sous-dossier, borne a la racine (anti-traversee).
    Retourne None si la cible sort de la racine."""
    base = _biblio_base()
    sous = (sous or "").strip().strip("/").replace("\\", "/")
    cible = os.path.abspath(os.path.join(base, sous.replace("/", os.sep)))
    if cible != base and not cible.startswith(base + os.sep):
        return None
    return cible


def bibliotheque_existe(relpath):
    base = os.path.abspath(config.BIBLIOTHEQUE_DIR)
    if not relpath:
        return False
    cible = os.path.abspath(os.path.join(base, str(relpath).replace("/", os.sep)))
    return (cible == base or cible.startswith(base + os.sep)) and os.path.isfile(cible)


def bibliotheque_lister(sous=""):
    """Contenu d'un sous-dossier : {sous, parent, dossiers:[{nom,rel}],
    fichiers:[{nom,rel,ext}]}. Lecture pure ; arborescence libre sous racine fixe."""
    base = _biblio_base()
    abs_sous = _biblio_abspath(sous)
    sous = (sous or "").strip().strip("/").replace("\\", "/")
    if abs_sous is None or not os.path.isdir(abs_sous):
        sous, abs_sous = "", base
    dossiers, fichiers = [], []
    for nom in sorted(os.listdir(abs_sous), key=lambda s: s.lower()):
        if nom.startswith("."):
            continue
        chemin = os.path.join(abs_sous, nom)
        rel = (sous + "/" + nom) if sous else nom
        if os.path.isdir(chemin):
            dossiers.append({"nom": nom, "rel": rel})
        elif nom != "_LISEZ-MOI.txt":
            ext = os.path.splitext(nom)[1].lstrip(".").upper()
            fichiers.append({"nom": nom, "rel": rel, "ext": ext})
    parent = "/".join(sous.split("/")[:-1]) if sous else None
    return {"sous": sous, "parent": parent, "dossiers": dossiers, "fichiers": fichiers}


def bibliotheque_enregistrer(sous, filename, fileobj):
    """Depose un fichier uploade dans un sous-dossier (nom assaini)."""
    dest = _biblio_abspath(sous)
    if dest is None:
        return None
    os.makedirs(dest, exist_ok=True)
    safe = re.sub(r"[^A-Za-z0-9._()\- ]", "_", os.path.basename(filename or "")).strip()
    safe = safe or "document"
    fileobj.save(os.path.join(dest, safe))
    sous = (sous or "").strip().strip("/")
    return (sous + "/" + safe) if sous else safe


def bibliotheque_creer_dossier(sous, nom):
    """Cree un sous-dossier (nom assaini) sous `sous`."""
    parent = _biblio_abspath(sous)
    if parent is None:
        return False
    nom = re.sub(r"[^A-Za-z0-9._()\- ]", "_", (nom or "").strip()).strip()
    if not nom or nom.startswith("."):
        return False
    os.makedirs(os.path.join(parent, nom), exist_ok=True)
    return True


def bibliotheque_supprimer_fichier(relpath):
    """Supprime un FICHIER (anti-traversee ; les dossiers ne sont pas touches)."""
    if not bibliotheque_existe(relpath):
        return False
    base = os.path.abspath(config.BIBLIOTHEQUE_DIR)
    cible = os.path.abspath(os.path.join(base, str(relpath).replace("/", os.sep)))
    os.remove(cible)
    return True
