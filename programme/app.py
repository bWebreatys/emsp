# -*- coding: utf-8 -*-
"""APPLICATION FLASK — couche de presentation / routage.
Lance un serveur local hors-ligne. Aucune ressource externe (pas de CDN).
"""
from flask import (Flask, render_template, jsonify, abort, request,
                   redirect, url_for, session, flash, send_file, Response)
import io
import os
import config
import metier
import auth
from data import AccesDonnees

app = Flask(__name__)
# Session locale hors-ligne : memorise l'utilisateur AUTHENTIFIE (login + mot de
# passe). Les empreintes sont dans instance/comptes.json (hors depot) ; les droits
# dans P1_Roles (definis par l'admin). secret_key local (pas de secret distant).
app.secret_key = "emsp-local-session"
auth.ensure_superadmin()   # bootstrap : garantit un superadmin connectable


def _matricule_saisi(v):
    """Extrait le matricule d'une valeur qui peut etre le libelle complet
    'matricule — Nom Prenom (...)' choisi dans une liste deroulante (filet de
    securite cote serveur si le nettoyage JS n'a pas eu lieu)."""
    v = (v or "").strip()
    if "\u2014" in v:                      # tiret cadratin du libelle
        v = v.split("\u2014", 1)[0].strip()
    return v


def _role_courant():
    """Role de l'utilisateur AUTHENTIFIE (session). Les droits viennent de P1_Roles
    (par login) ; un superutilisateur garde l'acces total meme sans ligne P1_Roles ;
    un login sans ligne ni statut superuser n'a aucun droit (pas de repli permissif)."""
    login = session.get("user", "")
    if not login:
        return {"login": "", "role": "(non connecte)", "lecture": "", "ecriture": "",
                "financier": "N", "admin": "N", "superuser": False}
    for r in metier.roles():
        if r["login"] == login:
            return r
    if metier.est_superuser(login):
        return {"login": login, "role": "Super-administrateur", "lecture": "Tous",
                "ecriture": "Tous", "financier": "O", "admin": "O", "superuser": True}
    return {"login": login, "role": "(aucun droit defini)", "lecture": "", "ecriture": "",
            "financier": "N", "admin": "N", "superuser": False}


# Endpoints accessibles SANS etre connecte.
_PUBLIC = {"login", "static"}

# --- MULTI-POSTE (reseau cable interne) — V1.54 ---------------------------
# Modele sur : le poste PRINCIPAL detient les donnees et ECRIT ; les autres
# postes, qui se connectent par le reseau au serveur du poste principal,
# sont automatiquement en LECTURE SEULE (consultation), avec voyant rouge.
# Aucun risque de corruption : seul le poste local (127.0.0.1) ecrit.
# En mode mono-poste (defaut, ecoute sur 127.0.0.1), remote_addr est toujours
# local -> ce garde-fou est inerte et le comportement habituel est inchange.
# Endpoints POST autorises depuis un poste secondaire (sinon il ne pourrait pas
# se connecter) : authentification uniquement.
_POST_AUTORISES_SECONDAIRE = {"login", "logout", "mot_de_passe"}


def _est_poste_local():
    """True si la requete provient du poste qui heberge le serveur (et les donnees)."""
    ra = request.remote_addr or ""
    return ra in ("127.0.0.1", "::1", "localhost") or ra.startswith("127.")


@app.before_request
def _exige_connexion():
    ep = request.endpoint or ""
    if ep in _PUBLIC:
        return
    if not session.get("user"):
        return redirect(url_for("login", suite=request.path))
    # Mot de passe a changer : on bloque tout sauf le changement et la deconnexion.
    if auth.doit_changer(session["user"]) and ep not in ("mot_de_passe", "logout"):
        return redirect(url_for("mot_de_passe"))


@app.before_request
def _lecture_seule_si_poste_secondaire():
    """Bloque TOUTE ecriture (POST) provenant d'un poste secondaire (reseau) :
    consultation seulement. L'authentification reste permise."""
    if request.method == "POST" and not _est_poste_local():
        if (request.endpoint or "") not in _POST_AUTORISES_SECONDAIRE:
            flash("Poste secondaire en lecture seule : la saisie se fait sur le poste principal.", "err")
            return redirect(request.referrer or url_for("accueil"))


@app.context_processor
def injecte_contexte():
    login_courant = _role_courant().get("login", "")
    _role = _role_courant()
    _gov = set(config.ONGLETS_DIRECTION) | set(config.ONGLETS_FINANCIERS)
    for _v in config.MODULES_ONGLETS.values():
        _gov.update(_v)
    _modules_caches = set() if metier.est_admin(_role) else {c for c in _gov if not metier.peut_lire(_role, c)}
    return {
        "cfg": config,
        "GUIDE": config.GUIDE_STRUCTURE,
        "MODULES_CACHES": _modules_caches,
        "PAGES_REF": config.PAGES_REF,
        "TAB_INDEX": config.TAB_INDEX,
        "role_courant": _role_courant(),
        "couleur_user": metier.couleur_login(login_courant),
        "formation": config.MODE_FORMATION,
        "poste_secondaire": not _est_poste_local(),
        "est_admin_courant": metier.est_admin(_role_courant()),
        "alertes_capacite": metier.alertes_capacite(),
        "compte_expire": bool(login_courant) and auth.est_expire(login_courant),
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        login_ = request.form.get("login", "").strip()
        mdp = request.form.get("mdp", "")
        if auth.verifier(login_, mdp):
            session["user"] = login_
            auth.journal(login_, "Connexion", "", "")
            suite = request.form.get("suite") or url_for("accueil")
            return redirect(suite)
        auth.journal(login_ or "(inconnu)", "Connexion refusee", "", "")
        flash("Identifiant ou mot de passe incorrect.", "err")
    if session.get("user"):
        return redirect(url_for("accueil"))
    return render_template("login.html", suite=request.args.get("suite", ""),
                           titre_page="Connexion")


@app.route("/logout")
def logout():
    u = session.get("user", "")
    if u:
        auth.journal(u, "Deconnexion", "", "")
    session.clear()
    return redirect(url_for("login"))


@app.route("/mot-de-passe", methods=["GET", "POST"])
def mot_de_passe():
    u = session.get("user", "")
    force = auth.doit_changer(u)
    # Self-service supprime (V1.43) : l'ecran n'est accessible QUE pour le changement
    # FORCE au 1er login / apres reinitialisation. Hors de ce cas, la gestion du mot
    # de passe passe par le responsable informatique (reinitialisation).
    if not force:
        flash("Le changement de mot de passe se fait aupres du responsable informatique "
              "(reinitialisation).", "warn")
        return redirect(url_for("accueil"))
    if request.method == "POST":
        nouveau = request.form.get("nouveau", "")
        if nouveau != request.form.get("confirmer", ""):
            flash("La confirmation ne correspond pas.", "err")
        else:
            ok, msg = auth.changer_mdp(u, request.form.get("ancien", ""), nouveau)
            flash(msg, "ok" if ok else "err")
            if ok:
                auth.journal(u, "Changement mot de passe", "", "")
                return redirect(url_for("accueil"))
    return render_template("mot_de_passe.html", force=force, titre_page="Mot de passe")


@app.route("/")
def accueil():
    compteurs = {}
    db = AccesDonnees()
    onglets_existants = set(db.onglets())
    for sec in config.GUIDE_STRUCTURE:
        for cle, _, _ in sec["modules"]:
            if cle in onglets_existants and cle not in config.READONLY_TABS:
                try:
                    compteurs[cle] = db.nb_lignes(cle)
                except Exception:
                    compteurs[cle] = None
    return render_template("accueil.html", compteurs=compteurs,
                           kpis=metier.kpis(), titre_page="Accueil")


def _lire_filtres():
    """Lit les criteres de selection depuis l'URL (bandeau de filtres, V1.71).
    Periode = plage de dates du/au (conforme a metier._filtres_actifs)."""
    a = request.args
    return {
        "filiere": (a.get("filiere", "") or "").strip(),
        "niveau": (a.get("niveau", "") or "").strip(),
        "annee": (a.get("annee", "") or "").strip(),
        "du": (a.get("du", "") or "").strip(),
        "au": (a.get("au", "") or "").strip(),
    }


def _bandeau_dashboard(f):
    """Bandeau de contexte du tableau de bord : il agrege plusieurs onglets, donc
    supporte les 4 criteres. Construit dans la couche presentation, sans toucher
    metier. Memes cles que metier._bandeau_contexte (pour le partiel commun)."""
    return {
        "filiere": f.get("filiere") or "Toutes",
        "niveau": f.get("niveau") or "Tous",
        "annee": f.get("annee") or "Toutes",
        "du": f.get("du") or "", "au": f.get("au") or "",
        "actif": bool(f.get("filiere") or f.get("niveau") or f.get("annee")
                      or f.get("du") or f.get("au")),
        "supporte": {"filiere": True, "niveau": True, "annee": True, "periode": True},
        "date_jour": metier.fmt_date(metier._dt.date.today()),
    }


def _peut_choisir_indicateurs():
    """Droit de modifier la selection (globale) des indicateurs du TDB direction.
    Reserve a la Direction (meme perimetre que _exige_direction)."""
    r = _role_courant()
    return bool(metier.est_admin(r)
                or metier.peut_ecrire(r, "J1_Journal_eleves")
                or metier.peut_ecrire(r, "J2_Journal_compta"))


@app.route("/tableau-de-bord")
def tableau_bord():
    f = _lire_filtres()
    sel = metier.tdb_selection()
    charts_sel = [c for c in config.DASHBOARD_CHARTS if c["id"] in sel["charts"]]
    return render_template("tableau_bord.html",
                           charts=charts_sel,
                           tdb_kpis=config.TDB_KPIS,
                           tdb_charts_all=config.DASHBOARD_CHARTS,
                           tdb_kpis_budget=config.TDB_KPIS_BUDGET,
                           tdb_charts_budget=config.TDB_CHARTS_BUDGET,
                           kpis_budget=metier.kpis_budget(),
                           sel_kpis=sel["kpis"],
                           sel_charts=sel["charts"],
                           peut_choisir=_peut_choisir_indicateurs(),
                           kpis=metier.kpis(f),
                           reste_du_filiere=metier.reste_du_par_filiere(f),
                           bandeau=_bandeau_dashboard(f),
                           valeurs=metier.valeurs_filtres(),
                           ecart_heures=metier.synthese_ecart_heures(),
                           info=config.TAB_INDEX["TDB_Direction"],
                           titre_page="Tableau de bord")


@app.route("/tableau-de-bord/indicateurs", methods=["POST"])
def tableau_bord_indicateurs():
    _exige_direction()
    if request.form.get("action") == "reset":
        metier.tdb_selection_reset()
        flash("Selection des indicateurs reinitialisee (tous affiches).", "ok")
    else:
        metier.tdb_selection_set(request.form.getlist("kpi"),
                                 request.form.getlist("chart"))
        flash("Selection des indicateurs enregistree.", "ok")
    return redirect(url_for("tableau_bord"))


@app.route("/api/dashboard")
def api_dashboard():
    f = _lire_filtres()
    return jsonify({"kpis": metier.kpis(f),
                    "graphiques": metier.donnees_graphiques(f)})


@app.route("/plan-action/tableau-de-bord")
def tableau_bord_plan_action():
    # V1.99.33 (#19) : ecran unifie. La synthese est desormais en tete de l'ecran
    # de saisie du plan d'action ; cette route reste un alias retro-compatible.
    return redirect(url_for("module", onglet="G1_Plan_action"))


@app.route("/module/<onglet>")
def module(onglet):
    if onglet not in config.TAB_INDEX:
        abort(404)
    if onglet in config.SPECIAL_ROUTES:
        return redirect(url_for(config.SPECIAL_ROUTES[onglet]))
    info = config.TAB_INDEX[onglet]
    f = _lire_filtres()
    data = metier.filtrer_table(onglet, f)
    dico = metier.dictionnaire_par_onglet().get(onglet, [])
    peut = metier.peut_ecrire(_role_courant(), onglet)
    extra = {}
    if onglet == "P0_Parametres":
        extra["parametres"] = metier.parametres_editables()
    if onglet in config.ONGLETS_SAISIE_ACTIVE:
        extra["champs"] = metier.champs_saisie(onglet)
    # Suggestions maquette (datalist) : config des champs + lignes compactes pour
    # le filtrage et le pre-remplissage cote client (offline, aucun reseau).
    mq_cfg = metier.maquette_datalist_cfg(onglet)
    if mq_cfg:
        extra["maquette_cfg"] = mq_cfg
        extra["maquette_lignes"] = metier.maquette_lignes_datalist()
    # Module Stages : controle de quota a la saisie + tableau de bord d'occupation
    # (calcul cote client a partir du referentiel S2 et des affectations S1).
    if onglet == "S1_Stages":
        extra["stages_cfg"] = metier.stages_cfg_saisie()
        extra["stages_lieux"] = metier.stages_referentiel_lieux()
        extra["stages_affectations"] = metier.stages_affectations()
    if onglet == "N2_Notes":
        extra["notes_assist"] = metier.notes_assist()
    if onglet == "N1_Bareme_UE":
        extra["coherence_bareme"] = metier.coherence_bareme()
        extra["seuil_passage"] = metier.seuil_passage()
    if onglet == "H1_Biblio_docs":
        extra["docs_groupes"] = metier.documents_officiels_groupes()
    if onglet == "M1_Equipements":
        extra["materiels_panne"] = metier.materiels_en_panne()
    if onglet == "L3_Besoins":
        eq = request.args.get("equip", "").strip()
        ty = request.args.get("type", "").strip()
        if (eq or ty) and "champs" in extra:
            for ch in extra["champs"]:
                if eq and ch["libelle"] == "Equipement concerne":
                    ch["defaut"] = eq
                if ty and ch["libelle"] == "Type de besoin":
                    ch["defaut"] = ty
    if onglet == "G1_Plan_action":
        extra["pa_kpis"] = metier.plan_action_kpis()
        # Pre-remplissage depuis la synthese budgetaire (C-6, deep-link) : un ecart
        # budgetaire -> une action corrective pre-cadree. Matching tolerant (accents).
        _g1 = {"domaine / module": request.args.get("domaine", "").strip(),
               "type d'ecart": request.args.get("type_ecart", "").strip(),
               "axe / theme": request.args.get("axe", "").strip(),
               "ecart constate": request.args.get("ecart", "").strip()}
        if any(_g1.values()) and "champs" in extra:
            import unicodedata as _ud

            def _norm(s):
                s = "".join(x for x in _ud.normalize("NFD", str(s))
                            if _ud.category(x) != "Mn")
                return s.lower().strip()

            for ch in extra["champs"]:
                v = _g1.get(_norm(ch["libelle"]))
                if v:
                    ch["defaut"] = v
    return render_template("module.html", onglet=onglet, info=info,
                           data=data, dico=dico, peut_ecrire=peut,
                           cap=metier.capacite_onglet(onglet),
                           bandeau=data.get("bandeau"),
                           valeurs=metier.valeurs_filtres(),
                           titre_page=info["libelle"], **extra)


@app.route("/module/<onglet>/imprimer")
def module_imprimer(onglet):
    """Impression navigateur de la selection courante d'un module (paysage).
    L'utilisateur sort un PDF via Ctrl+P -> Enregistrer en PDF (zero dependance)."""
    if onglet not in config.TAB_INDEX:
        abort(404)
    info = config.TAB_INDEX[onglet]
    f = _lire_filtres()
    data = metier.filtrer_table(onglet, f)
    return render_template("impression_selection.html",
                           mode="module", onglet=onglet, info=info,
                           data=data, bandeau=data.get("bandeau"),
                           titre_page="Impression - %s" % info["libelle"])


@app.route("/tableau-de-bord/imprimer")
def tableau_bord_imprimer():
    """Impression navigateur du tableau de bord filtre (paysage). V1.99.15 :
    n'imprime que les KPI COCHES (selection globale, #20 brique B). Les graphes
    coches restent imprimables depuis l'ecran via @media print ; aucun canvas
    Chart.js n'est rendu dans la page d'impression (arbitrage 3)."""
    f = _lire_filtres()
    k = metier.kpis(f)
    sel = metier.tdb_selection()["kpis"]
    _money = {"recettes", "depenses", "solde", "reste_du"}
    _fin = {"recettes", "depenses", "solde"}
    lignes = []
    for kpi in config.TDB_KPIS:
        kid = kpi["id"]
        if kid not in sel:
            continue
        val = metier._fmt_kmf(k.get(kid, 0)) if kid in _money else k.get(kid, "")
        grise = (kid in _fin and k.get("finances_non_filtrable")) \
            or (kid == "heures" and k.get("heures_non_filtrable"))
        if grise:
            val = "%s (global)" % val
        lignes.append({"libelle": kpi["libelle"], "valeur": val, "grise": bool(grise)})
    # KPI budget coches (bornes sur la session courante, toujours "global")
    kb = metier.kpis_budget()
    _bud_money = {"bud_prevu", "bud_realise", "bud_ecart"}
    for kpi in config.TDB_KPIS_BUDGET:
        bid = kpi["id"]
        if bid not in sel:
            continue
        if bid in _bud_money:
            v = metier._fmt_kmf(kb.get(bid, 0))
        elif bid == "bud_taux":
            v = "%s %%" % kb.get(bid, 0)
        else:
            v = kb.get(bid, "")
        lignes.append({"libelle": kpi["libelle"], "valeur": v, "grise": False})
    # R4 (V1.99.40) : detail du reste du par filiere a l'impression, uniquement si
    # le KPI 'reste_du' fait partie de la selection cochee.
    rdf = metier.reste_du_par_filiere(f) if "reste_du" in sel else None
    return render_template("impression_selection.html",
                           mode="dashboard",
                           lignes=lignes,
                           kpis=k,
                           reste_du_filiere=rdf,
                           bandeau=_bandeau_dashboard(f),
                           info=config.TAB_INDEX["TDB_Direction"],
                           titre_page="Impression - Tableau de bord")


@app.route("/module/<onglet>/ajouter", methods=["POST"])
def module_ajouter(onglet):
    if onglet not in config.ONGLETS_SAISIE_ACTIVE:
        abort(404)
    _garde_ecriture(onglet)
    champs = metier.champs_saisie(onglet)
    valeurs = {c["brut"]: request.form.get(c["brut"], "").strip() for c in champs}
    valeurs = {k: v for k, v in valeurs.items() if v != ""}
    # Champs auto-remplis cote serveur : "Saisi par" = login du role courant.
    login = _role_courant().get("login", "")
    for libelle in config.CHAMPS_AUTO_LOGIN.get(onglet, []):
        valeurs[metier._brut(onglet, libelle)] = login
    ok, msg = metier.valide_saisie(onglet, valeurs)
    if not ok:
        flash(msg, "err")
        return redirect(url_for("module", onglet=onglet))
    try:
        chemin = config.WORKBOOK_NOTES if onglet in config.ONGLETS_NOTES else config.WORKBOOK
        AccesDonnees(chemin).ajouter_ligne(onglet, valeurs)
        flash("Ligne ajoutee.", "ok")
        auth.journal(login, "Ajout ligne", onglet, "")
        # Mode formation : plafond INDICATIF par onglet -> on n'empeche jamais
        # l'ajout, on signale simplement le depassement (donnees jetables).
        if config.MODE_FORMATION:
            try:
                n = AccesDonnees(chemin).nb_lignes(onglet)
                if n > config.FORMATION_MAX:
                    flash("Formation : l'onglet %s depasse %d lignes (%d). Donnees d'entrainement, "
                          "pensez a reinitialiser l'environnement de formation." % (onglet, config.FORMATION_MAX, n), "warn")
            except Exception:
                pass
    except OverflowError:
        flash("Capacite de l'onglet atteinte.", "err")
    except Exception:
        flash("Echec de l'enregistrement.", "err")
    return redirect(url_for("module", onglet=onglet))


@app.route("/module/<onglet>/modifier", methods=["POST"])
def module_modifier(onglet):
    # L'edition en place est restreinte (cf. ONGLETS_EDITION_LIGNE) : le journal
    # financier reste en ajout seul. L'ajout, lui, demeure ouvert a tout SAISIE_ACTIVE.
    if onglet not in config.ONGLETS_EDITION_LIGNE:
        abort(404)
    _garde_ecriture(onglet)
    try:
        index = int(request.form.get("_index", "-1"))
    except ValueError:
        index = -1
    if index < 0:
        flash("Ligne a modifier non precisee.", "err")
        return redirect(url_for("module", onglet=onglet))
    champs = metier.champs_saisie(onglet)
    # Modification : on prend TOUTES les valeurs soumises (vide = effacement).
    valeurs = {c["brut"]: request.form.get(c["brut"], "").strip() for c in champs}
    login = _role_courant().get("login", "")
    for libelle in config.CHAMPS_AUTO_LOGIN.get(onglet, []):
        valeurs[metier._brut(onglet, libelle)] = login
    ok, msg = metier.valide_saisie(onglet, valeurs)
    if not ok:
        flash(msg, "err")
        return redirect(url_for("module", onglet=onglet))
    try:
        chemin = config.WORKBOOK_NOTES if onglet in config.ONGLETS_NOTES else config.WORKBOOK
        AccesDonnees(chemin).modifier_ligne(onglet, index, valeurs)
        flash("Fiche modifiee.", "ok")
        auth.journal(login, "Modif ligne", onglet, "#%d" % index)
    except IndexError:
        flash("Ligne introuvable (a peut-etre ete deplacee).", "err")
    except Exception:
        flash("Echec de la modification.", "err")
    return redirect(url_for("module", onglet=onglet))


# --- Edition P0_Parametres (premier ecran ecrivant) -----------------------
def _garde_ecriture(onglet):
    """Verifie le droit d'ecriture du role courant ; abort(403) sinon."""
    if not metier.peut_ecrire(_role_courant(), onglet):
        abort(403)


@app.route("/parametres/ajouter", methods=["POST"])
def parametres_ajouter():
    _garde_ecriture("P0_Parametres")
    colonne = request.form.get("colonne", "")
    valeur = request.form.get("valeur", "").strip()
    if not valeur:
        flash("Saisissez une valeur avant d'ajouter.", "err")
    else:
        try:
            ok = AccesDonnees().ajouter_valeur_liste(colonne, valeur)
            flash("Valeur ajoutee a la liste." if ok
                  else "Cette valeur existe deja dans la liste.",
                  "ok" if ok else "err")
            if ok:
                auth.journal(session.get("user", ""), "Ajout valeur liste", colonne, valeur)
        except OverflowError:
            flash("Capacite de la liste atteinte.", "err")
        except KeyError:
            flash("Liste inconnue.", "err")
    return redirect(url_for("module", onglet="P0_Parametres"))


@app.route("/parametres/supprimer", methods=["POST"])
def parametres_supprimer():
    _garde_ecriture("P0_Parametres")
    colonne = request.form.get("colonne", "")
    valeur = request.form.get("valeur", "")
    # Verrou V1.80 : interdire le retrait d'une valeur encore employee.
    nb, ou = metier.valeur_liste_utilisee(colonne, valeur)
    if nb:
        flash("Retrait impossible : « %s » est encore utilise dans %d enregistrement(s) "
              "(%s). Reaffectez-les avant de retirer cette valeur."
              % (valeur, nb, ", ".join(ou)), "err")
        return redirect(url_for("module", onglet="P0_Parametres"))
    try:
        ok = AccesDonnees().supprimer_valeur_liste(colonne, valeur)
        flash("Valeur retiree." if ok else "Valeur introuvable.",
              "ok" if ok else "err")
        if ok:
            auth.journal(session.get("user", ""), "Retrait valeur liste", colonne, valeur)
    except KeyError:
        flash("Liste inconnue.", "err")
    return redirect(url_for("module", onglet="P0_Parametres"))


@app.route("/presences")
def presences_saisie():
    """Ecran unique des presences : le professeur choisit SA CLASSE et LA DATE
    (jour courant par defaut + calendrier), puis IMPRIME la feuille ou SAISIT
    (comptage de ses heures et de celles des eleves). La saisie part des seances
    prevues ce jour-la ; repli hors planning par creneau + duree (1h/2h)."""
    _garde_ecriture("A2_Presences")
    info = config.TAB_INDEX["A2_Presences"]
    filiere = request.args.get("filiere", "").strip()
    niveau = request.args.get("niveau", "").strip()
    section = request.args.get("section", "").strip()
    date_fr, date_iso = metier.normaliser_date_saisie(request.args.get("date", "").strip())
    seance_id = request.args.get("seance", "").strip()
    mode = request.args.get("mode", "").strip()          # 'libre' = hors planning
    creneau = request.args.get("creneau", "").strip()
    duree = request.args.get("duree", "").strip()
    matiere = request.args.get("matiere", "").strip()
    enseignant = request.args.get("enseignant", "").strip()
    vue = request.args.get("vue", "classe").strip()       # 'classe' ou 'prof'
    ens_prof = request.args.get("ens_prof", "").strip()   # matricule (vue par professeur)

    classes = metier.classes_en_service()
    jour = metier._jour_de_date(date_fr)

    # Seances prevues pour cette classe, ce jour-la (source : planning A3).
    prevues = []
    if filiere and niveau:
        for s in metier.seances_saisie():
            if s["filiere"] == filiere and s["niveau"] == niveau \
                    and (not section or s["section"] == section) \
                    and s["jour"] and jour and s["jour"].lower() == jour.lower():
                sess = metier.cle_session(metier._seance_par_id(s["id"])) if s["id"] else ""
                plbl = metier.enseignant_programme_session(sess)[1] if sess else ""
                dh = ""
                try:
                    if s["debut"] and s["fin"]:
                        dh = str(int(s["fin"][:2]) - int(s["debut"][:2]))
                except Exception:  # noqa: BLE001
                    dh = ""
                prevues.append({"id": s["id"], "matiere": s["matiere"],
                                "creneau": s["creneau_defaut"], "enseignant": plbl,
                                "salle": s["salle"], "debut": s["debut"],
                                "fin": s["fin"], "duree": dh})

    # Vue PAR PROFESSEUR : ses seances ce jour-la (toutes classes confondues).
    prof_seances = []
    if vue == "prof" and ens_prof and jour:
        for s in metier.seances_saisie():
            if not (s["jour"] and s["jour"].lower() == jour.lower()):
                continue
            sess = metier.cle_session(metier._seance_par_id(s["id"])) if s["id"] else ""
            pmat, plbl, _amb = metier.enseignant_programme_session(sess) if sess else ("", "", False)
            if pmat != ens_prof:
                continue
            dh = ""
            try:
                if s["debut"] and s["fin"]:
                    dh = str(int(s["fin"][:2]) - int(s["debut"][:2]))
            except Exception:  # noqa: BLE001
                dh = ""
            prof_seances.append({"id": s["id"], "matiere": s["matiere"],
                                 "creneau": s["creneau_defaut"], "enseignant": plbl,
                                 "filiere": s["filiere"], "niveau": s["niveau"],
                                 "section": s["section"], "salle": s["salle"],
                                 "debut": s["debut"], "fin": s["fin"], "duree": dh})

    # Contexte de pointage : soit une seance prevue, soit une seance hors planning.
    ctx, roster, deja = None, [], {}
    if seance_id:
        courante = metier._seance_par_id(seance_id)
        if courante:
            cren = creneau or courante.get("creneau_defaut", "")
            sess = metier.cle_session(courante)
            roster = metier.etudiants_de_seance(seance_id)
            deja = metier.presences_existantes(date_fr, sess, cren) if date_fr else {}
            plbl = metier.enseignant_programme_session(sess)[1]
            ctx = {"type": "prevue", "seance": seance_id, "creneau": cren,
                   "matiere": courante.get("matiere", ""), "enseignant": plbl,
                   "filiere": courante.get("filiere", filiere),
                   "niveau": courante.get("niveau", niveau),
                   "section": courante.get("section", section)}
    elif mode == "libre" and filiere and niveau and creneau and duree and matiere:
        debut, fin = metier.creneau_duree_heures(creneau, duree)
        if debut:
            roster = metier.roster_classe(filiere, niveau, section)
            deja = metier.presences_existantes_libre(date_fr, filiere, niveau,
                                                     matiere, enseignant, debut, fin)
            ctx = {"type": "libre", "creneau": creneau, "duree": duree,
                   "debut": debut, "fin": fin, "matiere": matiere,
                   "enseignant": enseignant, "filiere": filiere,
                   "niveau": niveau, "section": section}

    return render_template(
        "presences.html", info=info, titre_page="Presences",
        classes=classes, filiere=filiere, niveau=niveau, section=section,
        date_fr=date_fr, date_iso=date_iso, jour=jour,
        creneaux=config.CRENEAUX, prevues=prevues, ctx=ctx,
        vue=vue, ens_prof=ens_prof, prof_seances=prof_seances,
        roster=roster, deja=deja,
        matieres_sug=(metier.matieres_maquette(filiere, niveau)
                      if (filiere and niveau) else []),
        ens_options=(metier.options_liste("Enseignants matricule (E1)") or []),
        cap=metier.capacite_onglet("A2_Presences"))


@app.route("/presences/feuilles-prof")
def presences_feuilles_prof():
    """Imprime en une fois TOUTES les feuilles d'un professeur pour une date :
    une feuille (avec les eleves) par seance, saut de page entre chaque."""
    if not metier.peut_lire(_role_courant(), "A2_Presences"):
        abort(403)
    ens_prof = request.args.get("ens_prof", "").strip()
    date_fr, _ = metier.normaliser_date_saisie(request.args.get("date", "").strip())
    jour = metier._jour_de_date(date_fr)
    feuilles = []
    if ens_prof and jour:
        for s in metier.seances_saisie():
            if not (s["jour"] and s["jour"].lower() == jour.lower()):
                continue
            sess = metier.cle_session(metier._seance_par_id(s["id"])) if s["id"] else ""
            pmat, plbl, _amb = metier.enseignant_programme_session(sess) if sess else ("", "", False)
            if pmat != ens_prof:
                continue
            cren = s.get("creneau_defaut", "")
            roster = metier.etudiants_de_seance(s["id"])
            deja = metier.presences_existantes(date_fr, sess, cren) if date_fr else {}
            feuilles.append({
                "entete": {"filiere": s["filiere"], "niveau": s["niveau"],
                           "section": s["section"], "annee": "",
                           "matiere": s["matiere"], "enseignant": plbl,
                           "salle": s["salle"], "jour": s["jour"],
                           "debut": s["debut"], "fin": s["fin"], "duree": "",
                           "date": date_fr, "creneau": cren},
                "roster": roster, "deja": deja, "saisi": bool(deja)})
    return render_template("feuilles_prof.html", feuilles=feuilles,
                           enseignant=(feuilles[0]["entete"]["enseignant"] if feuilles else ens_prof),
                           date_fr=date_fr, jour=jour, titre_page="Feuilles du professeur",
                           date_jour=metier.fmt_date(metier._dt.date.today()))


@app.route("/presences/feuille")
def presences_feuille():
    """Feuille de presence imprimable AVEC les eleves (a signer). Deux entrees :
    - seance planifiee : ?seance=..&date=..&creneau=..
    - seance libre     : ?filiere=..&niveau=..&section=..&date=..&creneau=..&duree=..
                         (ou &debut=..&fin=..)&matiere=..&enseignant=..
    Si des presences existent deja pour ce contexte, la colonne affiche
    Present / Absent (releve saisi) ; sinon une case vide (a signer)."""
    if not metier.peut_lire(_role_courant(), "A2_Presences"):
        abort(403)
    date_fr, _ = metier.normaliser_date_saisie(request.args.get("date", "").strip())
    seance_id = request.args.get("seance", "").strip()
    if seance_id:
        courante = metier._seance_par_id(seance_id)
        if not courante:
            flash("Choisissez une seance pour imprimer la feuille.", "err")
            return redirect(url_for("presences_saisie"))
        creneau = request.args.get("creneau", "").strip() or courante.get("creneau_defaut", "")
        session_val = metier.cle_session(courante)
        roster = metier.etudiants_de_seance(seance_id)
        _pm, prog_lbl, _pa = metier.enseignant_programme_session(session_val)
        a3d = metier._a3_index().get(session_val) or {}
        annee = str(a3d.get("Annee acad.", "")).strip()
        duree = metier._h_aff(metier._duree_h_a3(a3d)) if a3d else ""
        deja = metier.presences_existantes(date_fr, session_val, creneau) if date_fr else {}
        entete = {"filiere": courante.get("filiere"), "niveau": courante.get("niveau"),
                  "section": courante.get("section"), "annee": annee,
                  "matiere": courante.get("matiere"), "enseignant": prog_lbl,
                  "salle": courante.get("salle"), "jour": courante.get("jour"),
                  "debut": courante.get("debut"), "fin": courante.get("fin"),
                  "duree": duree, "date": date_fr, "creneau": creneau}
    else:
        filiere = request.args.get("filiere", "").strip()
        niveau = request.args.get("niveau", "").strip()
        section = request.args.get("section", "").strip()
        if not (filiere and niveau):
            flash("Choisissez une classe (filiere et niveau) pour imprimer la feuille.", "err")
            return redirect(url_for("presences_saisie"))
        creneau = request.args.get("creneau", "").strip()
        duree = request.args.get("duree", "").strip()
        debut = request.args.get("debut", "").strip()
        fin = request.args.get("fin", "").strip()
        if creneau and duree and not (debut and fin):
            debut, fin = metier.creneau_duree_heures(creneau, duree)
        matiere = request.args.get("matiere", "").strip()
        enseignant = request.args.get("enseignant", "").strip()
        roster = metier.roster_classe(filiere, niveau, section)
        deja = (metier.presences_existantes_libre(date_fr, filiere, niveau, matiere,
                                                  enseignant, debut, fin)
                if (date_fr and debut and fin and matiere) else {})
        entete = {"filiere": filiere, "niveau": niveau, "section": section, "annee": "",
                  "matiere": matiere, "enseignant": enseignant, "salle": "",
                  "jour": metier._jour_de_date(date_fr),
                  "debut": debut, "fin": fin, "duree": "", "date": date_fr, "creneau": creneau}
    return render_template("feuille_seance.html", entete=entete, roster=roster,
                           deja=deja, saisi=bool(deja),
                           titre_page="Feuille de presence",
                           date_jour=metier.fmt_date(metier._dt.date.today()))


@app.route("/presences/enregistrer", methods=["POST"])
def presences_enregistrer():
    _garde_ecriture("A2_Presences")
    seance_id = request.form.get("seance", "").strip()
    date_fr = request.form.get("date", "").strip()
    creneau = request.form.get("creneau", "").strip()
    presents = request.form.getlist("present")  # matricules coches
    ok, msg = metier.enregistrer_presences_lot(
        seance_id, date_fr, creneau, presents,
        saisi_par=_role_courant().get("login", ""))
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Saisie presences", seance_id,
                     "%s %s" % (date_fr, creneau))
        # En-tete de seance realisee : exception (remplacant / annulation / duree).
        # Ecrit dans E3 UNIQUEMENT si c'est un ecart au programme (sinon rien).
        try:
            ecrit = metier.enregistrer_seance_faite(
                seance_id, date_fr, creneau,
                etat=request.form.get("etat", "").strip(),
                assure_par=request.form.get("assure_par", "").strip(),
                matiere_reelle=request.form.get("matiere_reelle", "").strip(),
                vol_constate=request.form.get("vol_constate", "").strip(),
                motif=request.form.get("motif", "").strip(),
                saisi_par=_role_courant().get("login", ""))
            if ecrit:
                flash("Exception de seance enregistree "
                      "(remplacant / annulation / duree).", "ok")
                auth.journal(session.get("user", ""),
                             "Seance realisee (exception)", seance_id,
                             "%s %s" % (date_fr, creneau))
        except Exception as e:  # noqa: BLE001
            flash("Presences enregistrees, mais l'exception de seance n'a pas pu "
                  "etre enregistree : %s" % e, "err")
    return redirect(url_for("presences_saisie", seance=seance_id,
                            date=date_fr, creneau=creneau))



@app.route("/presences/libre")
def presences_libre():
    """Presences en liste - seance AD HOC (option B). Aucune ligne A3 requise :
    on choisit classe + date + plage horaire + matiere + enseignant, le roster
    vient de A1, l'ecriture va dans A2 (upsert). Creation A3 optionnelle."""
    _garde_ecriture("A2_Presences")
    role = _role_courant()
    f = request.args.get("filiere", "").strip()
    n = request.args.get("niveau", "").strip()
    sec = request.args.get("section", "").strip()
    date_fr = request.args.get("date", "").strip()
    if not date_fr:                       # date du jour par defaut (modifiable)
        date_fr = metier.fmt_date(metier.datetime.date.today())
    debut = request.args.get("debut", "").strip()
    fin = request.args.get("fin", "").strip()
    matiere = request.args.get("matiere", "").strip()
    enseignant = request.args.get("enseignant", "").strip()
    salle = request.args.get("salle", "").strip()
    classes = metier.classes_en_service()
    roster, deja, matieres = [], {}, []
    if f and n:
        roster = metier.roster_classe(f, n, sec)
        matieres = metier.matieres_maquette(f, n)
        if date_fr and debut and fin and matiere:
            deja = metier.presences_existantes_libre(date_fr, f, n, matiere,
                                                     enseignant, debut, fin)
    return render_template(
        "presences_libre.html", info=config.TAB_INDEX["A2_Presences"],
        titre_page="Presences - seance libre", classes=classes,
        filiere=f, niveau=n, section=sec, date_fr=date_fr, debut=debut, fin=fin,
        matiere=matiere, enseignant=enseignant, salle=salle,
        roster=roster, deja=deja, matieres=matieres,
        ens_options=(metier.options_liste("Enseignants matricule (E1)") or []),
        peut_creer_a3=metier.peut_ecrire(role, "A3_Sessions"),
        cap=metier.capacite_onglet("A2_Presences"))


@app.route("/presences/libre/enregistrer", methods=["POST"])
def presences_libre_enregistrer():
    _garde_ecriture("A2_Presences")
    role = _role_courant()
    f = request.form.get("filiere", "").strip()
    n = request.form.get("niveau", "").strip()
    sec = request.form.get("section", "").strip()
    date_fr = request.form.get("date", "").strip()
    debut = request.form.get("debut", "").strip()
    fin = request.form.get("fin", "").strip()
    creneau = request.form.get("creneau", "").strip()
    duree = request.form.get("duree", "").strip()
    if creneau and duree and not (debut and fin):
        debut, fin = metier.creneau_duree_heures(creneau, duree)
    matiere = request.form.get("matiere", "").strip()
    enseignant = request.form.get("enseignant", "").strip()
    salle = request.form.get("salle", "").strip()
    recurrente = bool(request.form.get("recurrente"))
    presents = request.form.getlist("present")
    ok, msg = metier.enregistrer_presences_libre(
        f, n, sec, date_fr, debut, fin, matiere, enseignant, salle, presents,
        recurrente, saisi_par=role.get("login", ""),
        peut_creer_a3=metier.peut_ecrire(role, "A3_Sessions"))
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Saisie presences (seance libre)",
                     "%s %s" % (f, n), "%s %s-%s" % (date_fr, debut, fin))
    if request.form.get("retour", "").strip() == "presences":
        return redirect(url_for("presences_saisie", filiere=f, niveau=n, section=sec,
                                date=date_fr, mode="libre", creneau=creneau,
                                duree=duree, matiere=matiere, enseignant=enseignant))
    return redirect(url_for("presences_libre", filiere=f, niveau=n, section=sec,
                            date=date_fr, debut=debut, fin=fin, matiere=matiere,
                            enseignant=enseignant, salle=salle))



@app.route("/bulletin")
def bulletin():
    """Saisie facon bulletin : un etudiant -> sa grille UE/matieres (CC derive N4
    + examen), moyennes en direct, ecriture N2 (upsert) des matieres renseignees."""
    _garde_ecriture("N2_Notes")
    matricule = _matricule_saisi(request.args.get("matricule", ""))
    annee = request.args.get("annee", "").strip()
    semestre = request.args.get("semestre", "").strip()
    session = request.args.get("session", "1").strip() or "1"
    fiche = metier.fiche_etudiant(matricule) if matricule else None
    if fiche and not annee:
        annee = fiche.get("Annee acad.", "") or metier._annee_acad_defaut()
    semestres = []
    grille = None
    if fiche:
        semestres = metier.semestres_classe(fiche.get("Filiere", ""), fiche.get("Niveau", ""))
        if not semestre and semestres:
            semestre = semestres[0]
        if semestre:
            grille = metier.bulletin_saisie(matricule, annee, semestre, session)
    return render_template(
        "bulletin_saisie.html", info=config.TAB_INDEX["BUL_Saisie"],
        titre_page="Saisie facon bulletin", matricule=matricule, annee=annee,
        semestre=semestre, session=session, fiche=fiche, semestres=semestres,
        grille=grille, etudiants=metier.recherche_etudiants(),
        introuvable=(bool(matricule) and fiche is None))


@app.route("/bulletin/enregistrer", methods=["POST"])
def bulletin_enregistrer():
    _garde_ecriture("N2_Notes")
    import json as _json
    matricule = request.form.get("matricule", "").strip()
    annee = request.form.get("annee", "").strip()
    semestre = request.form.get("semestre", "").strip()
    sess = request.form.get("session", "1").strip() or "1"
    try:
        meta = _json.loads(request.form.get("meta", "[]"))
    except ValueError:
        meta = []
    items = []
    for it in meta:
        i = it.get("i")
        items.append({"ue": it.get("ue", ""), "matiere": it.get("mat", ""),
                      "cc_readonly": bool(it.get("ro")),
                      "cc": request.form.get("cc_%s" % i, ""),
                      "examen": request.form.get("examen_%s" % i, "")})
    ok, msg = metier.enregistrer_bulletin(matricule, annee, semestre, sess, items,
                                          saisi_par=_role_courant().get("login", ""))
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Saisie bulletin",
                     "%s S%s sess%s" % (matricule, semestre, sess),
                     "%s" % annee)
    return redirect(url_for("bulletin", matricule=matricule, annee=annee,
                            semestre=semestre, session=sess))


@app.route("/notes-classe")
def saisie_notes_classe():
    """V1.76 : grille de saisie des notes par classe et matiere (liste d'eleves,
    multi-controles -> CC pondere -> moyenne 1/4-3/4)."""
    _garde_ecriture("N2_Notes")
    f = request.args
    filiere = f.get("filiere", "").strip()
    niveau = f.get("niveau", "").strip()
    section = f.get("section", "").strip()
    annee = f.get("annee", "").strip() or metier._annee_acad_defaut()
    semestre = f.get("semestre", "").strip()
    session = f.get("session", "1").strip() or "1"
    matiere_cle = f.get("matiere", "").strip()       # "num_ue|||matiere"
    semestres = metier.semestres_classe(filiere, niveau) if (filiere and niveau) else []
    if not semestre and semestres:
        semestre = semestres[0]
    matieres = (metier.classe_matieres(filiere, niveau, semestre)
                if (filiere and niveau and semestre) else [])
    grille = None
    num_ue = mat = ""
    if matiere_cle and "|||" in matiere_cle:
        num_ue, mat = matiere_cle.split("|||", 1)
        if filiere and niveau and semestre:
            grille = metier.notes_grille(filiere, niveau, section, annee,
                                         semestre, num_ue, mat, session)
    return render_template(
        "saisie_notes_classe.html", info=config.TAB_INDEX["NOT_Grille"],
        titre_page="Saisie des notes par classe", filiere=filiere, niveau=niveau,
        section=section, annee=annee, semestre=semestre, session=session,
        matiere_cle=matiere_cle, num_ue=num_ue, matiere=mat,
        filieres=metier.options_liste("Filieres") or [],
        niveaux=metier.options_liste("Niveaux") or [],
        sections=metier.options_liste("Sections") or [],
        annees=metier.options_liste("Annees_acad") or [],
        semestres=semestres, matieres=matieres, grille=grille)


@app.route("/notes-classe/enregistrer", methods=["POST"])
def saisie_notes_classe_enregistrer():
    _garde_ecriture("N2_Notes")
    import json as _json
    f = request.form
    filiere = f.get("filiere", "").strip(); niveau = f.get("niveau", "").strip()
    section = f.get("section", "").strip(); annee = f.get("annee", "").strip()
    semestre = f.get("semestre", "").strip(); sess = f.get("session", "1").strip() or "1"
    num_ue = f.get("num_ue", "").strip(); matiere = f.get("matiere", "").strip()
    try:
        payload = _json.loads(f.get("payload", "{}"))
    except ValueError:
        payload = {}
    coefs = payload.get("coefs", [])
    eleves = payload.get("eleves", [])
    non_dispensee = f.get("non_dispensee", "").strip()   # V1.77 : interrupteur de la grille
    motif = f.get("motif", "").strip()
    ok, msg = metier.enregistrer_notes_grille(
        filiere, niveau, section, annee, semestre, num_ue, matiere, sess,
        coefs, eleves, saisi_par=_role_courant().get("login", ""),
        non_dispensee=non_dispensee, motif=motif)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Saisie notes (grille)",
                     "%s %s %s S%s sess%s — UE %s / %s" % (filiere, niveau, section,
                      semestre, sess, num_ue, matiere), annee)
    return redirect(url_for("saisie_notes_classe", filiere=filiere, niveau=niveau,
                            section=section, annee=annee, semestre=semestre,
                            session=sess, matiere="%s|||%s" % (num_ue, matiere)))


@app.route("/impressions/feuille-notes")
def impr_feuille_notes():
    """V1.76 : edition imprimable de la grille (C1..Cn + CC + Examen + Moyenne)."""
    f = request.args
    filiere = f.get("filiere", "").strip(); niveau = f.get("niveau", "").strip()
    section = f.get("section", "").strip(); annee = f.get("annee", "").strip()
    semestre = f.get("semestre", "").strip(); session = f.get("session", "1").strip() or "1"
    num_ue = f.get("num_ue", "").strip(); matiere = f.get("matiere", "").strip()
    if not (filiere and niveau and semestre and matiere):
        flash("Selection incomplete pour la feuille de notes.", "err")
        return redirect(url_for("saisie_notes_classe"))
    data = metier.feuille_notes_edition(filiere, niveau, section, annee, semestre,
                                        num_ue, matiere, session)
    doc = metier.doc_adhoc("Feuille de notes — %s" % matiere,
                           mentions="Notes sur 20. CC = moyenne ponderee des controles. "
                                    "Moyenne = 1/4 CC + 3/4 examen (decret 05-106, art. 8).",
                           signataire="L'enseignant | Le Chef de departement",
                           orientation="paysage")
    return _page_impression(doc, "table", table=data)


@app.route("/heures-constatees")
def heures_constatees():
    """Ecran 'Heures constatees du mois' : calcule (depuis A2 x A3 + exceptions E3)
    vs valeur E2 actuelle, avec report vers E2 (base de la paie)."""
    role = _role_courant()
    if not metier.peut_lire(role, "E2_Releve_heures"):
        abort(403)
    mois = request.args.get("mois", "").strip()
    apercu = metier.heures_constatees_apercu(mois) if mois else None
    return render_template("heures_constatees.html",
                           info=config.TAB_INDEX["HRS_Heures"],
                           titre_page="Heures constatees du mois",
                           mois=mois, apercu=apercu,
                           mois_dispo=metier.mois_appels_dispo(),
                           peut_reporter=metier.peut_ecrire(role, "E2_Releve_heures"))


@app.route("/heures-constatees/reporter", methods=["POST"])
def heures_constatees_reporter():
    _garde_ecriture("E2_Releve_heures")
    mois = request.form.get("mois", "").strip()
    forcer = request.form.get("forcer", "") == "1"
    motif = request.form.get("motif", "").strip()
    matricules = request.form.getlist("mat")  # vide => tous
    if forcer and not motif:
        flash("Report force : le motif de regularisation est obligatoire "
              "(il documente la correction pour la comptabilite).", "err")
        return redirect(url_for("heures_constatees", mois=mois))
    res = metier.reporter_heures_mois(mois, matricules or None, forcer=forcer, motif=motif)
    regs = res.get("regularisations", [])
    if res["divergences"] and not forcer:
        flash("%d report(s) effectue(s). %d enseignant(s) ont deja dans E2 une "
              "valeur differente du calcul (correction manuelle) : non ecrases. "
              "Pour les remplacer, cochez-les, saisissez un motif et confirmez le "
              "report force." % (res["reportes"], len(res["divergences"])), "warn")
    elif regs:
        flash("Report effectue : %d ajout(s), %d mise(s) a jour, dont %d "
              "regularisation(s) tracee(s) au journal." % (res["ajout"], res["maj"], len(regs)), "ok")
    else:
        flash("Report effectue : %d ajout(s), %d mise(s) a jour dans E2."
              % (res["ajout"], res["maj"]), "ok")
    auth.journal(session.get("user", ""), "Report heures -> E2", mois,
                 "reportes=%d forcer=%s" % (res["reportes"], forcer))
    for rg in regs:                                   # V1.92 : une ligne par regularisation
        auth.journal(session.get("user", ""), "Regularisation heures",
                     "%s %s (%s)" % (rg["matricule"], rg["nom"], mois),
                     "%s -> %s h · motif: %s" % (rg["ancien"], rg["nouveau"], rg["motif"]))
    return redirect(url_for("heures_constatees", mois=mois))


@app.route("/reference/<onglet>")
def reference(onglet):
    if onglet == "Dictionnaire":
        return render_template("dictionnaire.html",
                               groupes=metier.dictionnaire_par_onglet(),
                               titre_page="Dictionnaire des donnees")
    if onglet in ("Guide", "Legende"):
        lignes = AccesDonnees().lignes_libres(onglet)
        return render_template("texte.html", onglet=onglet, lignes=lignes,
                               titre_page=config.TAB_INDEX[onglet]["libelle"])
    abort(404)


@app.route("/calendrier")
def calendrier():
    vue = request.args.get("vue", "semaine")
    if vue not in ("mois", "semaine", "jour"):
        vue = "semaine"
    date_iso = request.args.get("date", "")
    demo = request.args.get("demo") == "1"
    ctx = {"vue": vue, "demo": demo, "date_param": date_iso,
           "info": config.TAB_INDEX["CAL_Calendrier"], "titre_page": "Calendrier"}
    if vue == "mois":
        ctx["mois"] = metier.calendrier_mois(date_iso, demo)
    elif vue == "jour":
        ctx["jour"] = metier.calendrier_jour(date_iso, demo)
    else:
        ctx["semaine"] = metier.calendrier_semaine(date_iso, demo)
    return render_template("calendrier.html", **ctx)


@app.route("/salles")
def salles():
    demo = request.args.get("demo") == "1"
    date_fr, date_iso = metier.normaliser_date_saisie(request.args.get("date", "").strip())
    jour = metier._jour_de_date(date_fr)
    occupation = metier.planning_salles_par_date(date_fr, demo)
    return render_template("salles.html", occupation=occupation,
                           date_fr=date_fr, date_iso=date_iso, jour=jour, demo=demo,
                           heure_min=config.CAL_HEURE_MIN, heure_max=config.CAL_HEURE_MAX,
                           info=config.TAB_INDEX["VUE_Salles"], titre_page="Salles — planning")


@app.route("/reservations")
def reservations():
    """Ecran de reservation des salles : choisir la date (jour du calendrier),
    reserver par CRENEAU + DUREE (pas d'heures libres), lister les reservations."""
    _garde_ecriture("L2_Reservations")
    date_fr, date_iso = metier.normaliser_date_saisie(request.args.get("date", "").strip())
    liste = metier.reservations_liste(date_fr)
    salles_opt = [s.get("nom", "") for s in metier.salles() if s.get("nom")]
    return render_template("reservations.html", info=config.TAB_INDEX["L2_Reservations"],
                           titre_page="Réservations de salles",
                           date_fr=date_fr, date_iso=date_iso,
                           creneaux=config.CRENEAUX, salles_opt=salles_opt,
                           types_res=config.TYPES_RESERVATION,
                           statuts_res=config.STATUTS_RESERVATION,
                           reservations=liste,
                           cap=metier.capacite_onglet("L2_Reservations"))


@app.route("/reservations/creer", methods=["POST"])
def reservations_creer():
    _garde_ecriture("L2_Reservations")
    creneau = request.form.get("creneau", "").strip()
    duree = request.form.get("duree", "").strip()
    hd, hf = metier.creneau_duree_heures(creneau, duree)
    date_fr = request.form.get("date", "").strip()
    login = _role_courant().get("login", "")
    d = {"salle": request.form.get("salle", "").strip(),
         "date": date_fr, "heure_debut": hd, "heure_fin": hf,
         "type": request.form.get("type", "").strip(),
         "reserve_par": request.form.get("reserve_par", "").strip(),
         "motif": request.form.get("motif", "").strip(),
         "statut": request.form.get("statut", "").strip()}
    ok, msg = metier.creer_reservation(_role_courant(), d, login)
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("reservations", date=date_fr))


@app.route("/salles/<salle_id>/inventaire/imprimer")
def salle_inventaire_imprimer(salle_id):
    """R5 (V1.99.41) : edition imprimable de l'inventaire materiel d'UNE salle."""
    occ = metier.salle_occupation(salle_id)
    if occ is None:
        abort(404)
    nom = occ["salle"].get("nom", "")
    data = metier.inventaire_salle(nom)
    doc = metier.doc_adhoc("Inventaire du materiel — salle %s" % nom,
                           mentions="Montants en KMF. Source : module Equipements (M1).",
                           signataire="Le Responsable logistique | Le Directeur de l'EMSP",
                           orientation="paysage")
    return _page_impression(doc, "table", table=data)


@app.route("/salles/<salle_id>")
def salle_detail(salle_id):
    demo = request.args.get("demo") == "1"
    occ = metier.salle_occupation(salle_id, demo)
    if occ is None:
        abort(404)
    # Edition de la fiche : reservee aux roles ayant le droit d'ecriture sur
    # L1_Salles (meme droit que l'editeur de ligne generique). En demo ou pour
    # une salle derivee du materiel (sans ID), pas d'edition.
    peut_modifier = (not demo
                     and not occ["salle"].get("depuis_materiel")
                     and occ.get("index") is not None
                     and metier.peut_ecrire(_role_courant(), "L1_Salles"))
    types = list(config.TYPES_SALLE)
    t_actuel = str(occ["salle"].get("type", "")).strip()
    if t_actuel and t_actuel not in types:
        types.append(t_actuel)   # ne jamais perdre une valeur deja saisie
    peut_reserver = (not demo
                     and metier.peut_ecrire(_role_courant(), "L2_Reservations"))
    return render_template("salle_detail.html", occ=occ, demo=demo,
                           peut_modifier=peut_modifier, peut_reserver=peut_reserver,
                           types_salle=types,
                           info=config.TAB_INDEX["VUE_Salles"],
                           titre_page="Salle — " + occ["salle"]["nom"])


@app.route("/salles/<salle_id>/modifier", methods=["POST"])
def salle_modifier(salle_id):
    # Meme chaine que l'editeur de ligne generique (source unique) : champs_saisie
    # -> valide_saisie -> modifier_ligne sur L1_Salles. L'ID salle (cle) n'est PAS
    # soumis : il est preserve. Le materiel n'est pas touche ici (ecran M1 dedie).
    _garde_ecriture("L1_Salles")
    occ_cur = metier.salle_occupation(salle_id)
    if occ_cur is None or occ_cur.get("index") is None:
        flash("Salle introuvable.", "err")
        return redirect(url_for("salles"))
    index = occ_cur["index"]
    champs = metier.champs_saisie("L1_Salles")
    valeurs = {c["brut"]: request.form.get(c["brut"], "").strip()
               for c in champs if c["brut"] != "ID salle"}
    # L'ID (cle) n'est pas editable : on reinjecte la valeur courante pour la
    # preserver ET satisfaire la validation (champ obligatoire).
    valeurs["ID salle"] = occ_cur["salle"]["id"]
    ok, msg = metier.valide_saisie("L1_Salles", valeurs)
    if not ok:
        flash(msg, "err")
        return redirect(url_for("salle_detail", salle_id=salle_id))
    login = _role_courant().get("login", "")
    try:
        AccesDonnees().modifier_ligne("L1_Salles", index, valeurs)
        flash("Fiche salle modifiee.", "ok")
        auth.journal(login, "Modif salle", "L1_Salles", salle_id)
    except IndexError:
        flash("Ligne introuvable (a peut-etre ete deplacee).", "err")
    except Exception:
        flash("Echec de la modification.", "err")
    return redirect(url_for("salle_detail", salle_id=salle_id))


@app.route("/salles/<salle_id>/reserver", methods=["POST"])
def salle_reserver(salle_id):
    # Activation de l'action « reserver » sur la fiche salle. Ecrit une ligne L2
    # (chaine metier.creer_reservation : pre-remplissage A3 + resolution E1).
    _garde_ecriture("L2_Reservations")
    occ = metier.salle_occupation(salle_id)
    if occ is None:
        flash("Salle introuvable.", "err")
        return redirect(url_for("salles"))
    login = _role_courant().get("login", "")
    d = {
        "salle": occ["salle"]["nom"],
        "date": request.form.get("date", "").strip(),
        "heure_debut": request.form.get("heure_debut", "").strip(),
        "heure_fin": request.form.get("heure_fin", "").strip(),
        "type": request.form.get("type", "").strip(),
        "reserve_par": request.form.get("reserve_par", "").strip(),
        "motif": request.form.get("motif", "").strip(),
        "statut": request.form.get("statut", "").strip(),
        "seance": request.form.get("seance", "").strip(),
        "filiere": request.form.get("filiere", "").strip(),
        "niveau": request.form.get("niveau", "").strip(),
        "matiere": request.form.get("matiere", "").strip(),
        "matricule": request.form.get("matricule", "").strip(),
        "enseignant": request.form.get("enseignant", "").strip(),
    }
    ok, msg = metier.creer_reservation(_role_courant(), d, login)
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("salle_detail", salle_id=salle_id))


@app.route("/autorisations")
def autorisations():
    role = _role_courant()
    admin = metier.est_admin(role)
    users = metier.utilisateurs_admin() if admin else []
    # Mot de passe genere a afficher UNE seule fois (apres creation / reinitialisation).
    mdp_genere = session.pop("mdp_genere", None) if admin else None
    return render_template("autorisations.html",
                           matrice=metier.matrice_autorisations(),
                           est_admin=admin,
                           utilisateurs=users,
                           rubriques=metier.rubriques(),
                           couleurs=config.COULEURS_UTILISATEUR,
                           mdp_genere=mdp_genere,
                           info=config.TAB_INDEX["MAT_Autorisations"],
                           titre_page="Comptes & droits d'acces")


def _exige_admin():
    if not metier.est_admin(_role_courant()):
        abort(403)


@app.route("/autorisations/utilisateur", methods=["POST"])
def autorisations_utilisateur():
    _exige_admin()
    f = request.form
    login_ = f.get("login", "").strip()
    rubrique = f.get("rubrique", "").strip()
    couleur = f.get("couleur", "").strip()
    nouveau_compte = bool(login_) and not auth.existe(login_)
    # Droits par module (cases cochees) + bascules. "Tous (Direction)" est un
    # interrupteur distinct qui prime (cf. _compose_groupes).
    lecture = [g for g in f.getlist("lecture") if g != config.GROUPE_TOUS]
    ecriture = [g for g in f.getlist("ecriture") if g != config.GROUPE_TOUS]
    lecture_tous = f.get("lecture_tous") == "on"
    ecriture_tous = f.get("ecriture_tous") == "on"
    financier = f.get("financier") == "on"
    admin = f.get("admin") == "on"
    ok, msg = metier.enregistrer_utilisateur(
        _role_courant(), login=login_, role=f.get("role", ""),
        lecture_groupes=lecture, ecriture_groupes=ecriture,
        lecture_tous=lecture_tous, ecriture_tous=ecriture_tous,
        financier=financier, admin=admin)
    if not ok:
        flash(msg, "err")
        return redirect(url_for("autorisations"))
    # Attributs operationnels (rubrique, couleur) — stockes hors depot.
    auth.definir_attributs(login_, rubrique=rubrique or None, couleur=couleur or None)
    if nouveau_compte:
        # Mot de passe ALEATOIRE + validite annee scolaire ; affiche une seule fois.
        ok2, mdp = auth.reinitialiser(login_)
        auth.initialiser_validite(login_)
        session["mdp_genere"] = {"login": login_, "mdp": mdp, "motif": "creation"}
        auth.journal(session.get("user", ""), "Creation compte", login_,
                     "rubrique=%s" % (rubrique or "-"))
        flash(msg + " Mot de passe initial genere (a remettre a l'utilisateur, a changer au 1er login).", "ok")
    else:
        auth.journal(session.get("user", ""), "MAJ compte", login_,
                     "rubrique=%s" % (rubrique or "-"))
        flash(msg, "ok")
    return redirect(url_for("autorisations"))


@app.route("/autorisations/matrice", methods=["POST"])
def autorisations_matrice():
    """Enregistre la MATRICE des droits editable : pour chaque compte (hors
    superutilisateur), recompose lecture/ecriture par module + financier + admin
    et appelle le MEME writer que l'editeur de compte (enregistrer_utilisateur),
    donc les memes garde-fous (superutilisateur protege, dernier admin preserve)."""
    _exige_admin()
    f = request.form
    acteur = _role_courant()
    logins = [l for l in f.getlist("logins") if l.strip()]
    ok_n, erreurs = 0, []
    for login_ in logins:
        if metier.est_superuser(login_):
            continue
        lecture = [g for g in f.getlist("lec_%s" % login_) if g != config.GROUPE_TOUS]
        ecriture = [g for g in f.getlist("ecr_%s" % login_) if g != config.GROUPE_TOUS]
        ok, msg = metier.enregistrer_utilisateur(
            acteur, login=login_, role=f.get("role_%s" % login_, ""),
            lecture_groupes=lecture, ecriture_groupes=ecriture,
            lecture_tous=("lectous_%s" % login_) in f,
            ecriture_tous=("ecrtous_%s" % login_) in f,
            financier=("fin_%s" % login_) in f,
            admin=("adm_%s" % login_) in f)
        if ok:
            ok_n += 1
        else:
            erreurs.append("%s : %s" % (login_, msg))
    if erreurs:
        flash("Matrice enregistree pour %d compte(s). Non applique : %s"
              % (ok_n, " ; ".join(erreurs)), "err")
    else:
        flash("Matrice des droits enregistree (%d compte(s))." % ok_n, "ok")
    auth.journal(session.get("user", ""), "MAJ matrice des droits", "P1_Roles",
                 "%d compte(s)" % ok_n)
    return redirect(url_for("autorisations"))


@app.route("/autorisations/reinitialiser", methods=["POST"])
def autorisations_reinitialiser():
    _exige_admin()
    login_ = request.form.get("login", "").strip()
    if not login_:
        flash("Compte introuvable.", "err")
        return redirect(url_for("autorisations"))
    ok, mdp = auth.reinitialiser(login_)
    if ok:
        session["mdp_genere"] = {"login": login_, "mdp": mdp, "motif": "reinitialisation"}
        auth.journal(session.get("user", ""), "Reinitialisation mot de passe", login_, "")
        flash("Nouveau mot de passe genere pour %s (a changer au prochain login)." % login_, "ok")
    else:
        flash("Reinitialisation impossible.", "err")
    return redirect(url_for("autorisations"))


@app.route("/autorisations/renouveler", methods=["POST"])
def autorisations_renouveler():
    _exige_admin()
    login_ = request.form.get("login", "").strip()
    nouvelle = auth.renouveler(login_)
    if nouvelle:
        auth.journal(session.get("user", ""), "Renouvellement validite", login_, nouvelle)
        flash("Validite de %s renouvelee jusqu'au %s." % (login_, nouvelle), "ok")
    else:
        flash("Compte introuvable (aucun mot de passe defini).", "err")
    return redirect(url_for("autorisations"))


@app.route("/autorisations/supprimer", methods=["POST"])
def autorisations_supprimer():
    _exige_admin()
    login_ = request.form.get("login", "")
    ok, msg = metier.supprimer_utilisateur(_role_courant(), login_)
    if ok:
        auth.supprimer(login_)
        auth.journal(session.get("user", ""), "Suppression compte", login_, "")
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("autorisations"))


# --- Cloture / archivage / passation (V1.44) — reserve a la Direction -------
def _exige_direction():
    r = _role_courant()
    if not (metier.peut_ecrire(r, "J1_Journal_eleves")
            or metier.peut_ecrire(r, "J2_Journal_compta")):
        abort(403)


@app.route("/cloture")
def cloture():
    _exige_direction()
    pv = session.get("pv")
    return render_template("cloture.html", apercu=metier.cloture_apercu(),
                           pv=pv, info=config.TAB_INDEX["CLO_Cloture"],
                           titre_page="Cloture & archivage")


@app.route("/parametres/nomenclature")
def nomenclature():
    if not metier.peut_ecrire(_role_courant(), "P3_Nomenclature"):
        abort(403)
    filtres = {k: request.args.get(k, "").strip()
               for k in ("sens", "source", "actif", "q")}
    return render_template("nomenclature.html",
                           data=metier.donnees_nomenclature(filtres),
                           filtres=filtres,
                           info=config.TAB_INDEX["NOM_Curation"],
                           titre_page="Nomenclature — curation")


@app.route("/parametres/nomenclature/actif", methods=["POST"])
def nomenclature_actif():
    if not metier.peut_ecrire(_role_courant(), "P3_Nomenclature"):
        abort(403)
    codes = request.form.getlist("codes")
    actif = request.form.get("actif", "")
    n = metier.basculer_actif_codes(codes, actif)
    if n:
        flash("%d code(s) %s." % (n, "active(s)" if actif == "Oui" else "desactive(s)"), "ok")
        auth.journal(session.get("user", ""), "Nomenclature Actif=%s" % actif, "",
                     "%d code(s)" % n)
    else:
        flash("Aucun code selectionne.", "err")
    filtres = {k: request.form.get("f_" + k, "").strip()
               for k in ("sens", "source", "actif", "q")}
    return redirect(url_for("nomenclature",
                            **{k: v for k, v in filtres.items() if v}))


# --- C-4 (V1.99.12) : budget previsionnel par formation --------------------
def _exige_budget(ecriture=False):
    """Garde d'acces a l'ecran budget previsionnel. Lecture = financier (ou admin) ;
    ecriture = droit d'ecriture sur F5_Budget_Prev (ou admin). abort(403) sinon."""
    role = _role_courant()
    if metier.est_admin(role):
        return role
    if ecriture:
        if not metier.peut_ecrire(role, "F5_Budget_Prev"):
            abort(403)
    else:
        if not (metier.peut_lire(role, "F5_Budget_Prev")
                or metier.peut_ecrire(role, "F5_Budget_Prev")):
            abort(403)
    return role


_F5_CHAMPS = ["Formation", "Niveau", "Rubrique", "Designation", "Unite1", "Qte1",
              "Unite2", "Qte2", "Cout unitaire (KMF)", "Poste budgetaire",
              "Source de financement / Bailleur", "Session"]


def _f5_form(f):
    return {c: f.get(c, "").strip() for c in _F5_CHAMPS}


@app.route("/budget/previsionnel")
def budget_previsionnel():
    role = _exige_budget()
    formation = request.args.get("formation", "").strip()
    session_ = request.args.get("session", "").strip()
    data = metier.donnees_budget_prev(formation, session_)
    return render_template(
        "budget_previsionnel.html",
        data=data, formation=formation, session_sel=session_,
        formations_opt=metier.options_liste("Filieres") or [],
        niveaux_opt=config.LISTES_INLINE.get("Niveaux budget", []),
        postes_opt=metier.options_liste("Codes depense actifs (P3)") or [],
        bailleurs_opt=metier.options_liste("Sources_financement") or [],
        session_defaut=metier._annee_acad_defaut(),
        peut_ecrire=(metier.est_admin(role)
                     or metier.peut_ecrire(role, "F5_Budget_Prev")),
        info=config.TAB_INDEX["BUD_Previsionnel"],
        titre_page="Budget previsionnel par formation")


@app.route("/budget/previsionnel/ajouter", methods=["POST"])
def budget_previsionnel_ajouter():
    role = _exige_budget(ecriture=True)
    ok, msg = metier.ajouter_ligne_budget_prev(role, _f5_form(request.form),
                                               session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("budget_previsionnel",
                            formation=request.form.get("formation", "").strip(),
                            session=request.form.get("Session", "").strip()))


@app.route("/budget/previsionnel/modifier", methods=["POST"])
def budget_previsionnel_modifier():
    role = _exige_budget(ecriture=True)
    ok, msg = metier.modifier_ligne_budget_prev(role, request.form.get("index", ""),
                                                _f5_form(request.form),
                                                session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("budget_previsionnel",
                            formation=request.form.get("formation", "").strip(),
                            session=request.form.get("Session", "").strip()))


@app.route("/budget/previsionnel/supprimer", methods=["POST"])
def budget_previsionnel_supprimer():
    role = _exige_budget(ecriture=True)
    ok, msg = metier.supprimer_ligne_budget_prev(role, request.form.get("index", ""),
                                                 session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("budget_previsionnel",
                            formation=request.form.get("formation_sel", "").strip(),
                            session=request.form.get("session_sel", "").strip()))


@app.route("/budget/previsionnel/reglages", methods=["POST"])
def budget_previsionnel_reglages():
    role = _exige_budget(ecriture=True)
    msgs = []
    if request.form.get("taux_eur", "").strip():
        ok, m = metier.definir_taux_eur(role, request.form.get("taux_eur"),
                                        session.get("user", ""))
        msgs.append(m)
    if request.form.get("frais_admin_pct", "").strip():
        ok, m = metier.definir_frais_admin_pct(role, request.form.get("frais_admin_pct"),
                                               session.get("user", ""))
        msgs.append(m)
    flash(" ".join(msgs) if msgs else "Aucun reglage modifie.", "ok")
    return redirect(url_for("budget_previsionnel",
                            formation=request.form.get("formation_sel", "").strip(),
                            session=request.form.get("session_sel", "").strip()))


@app.route("/bareme/seuil", methods=["POST"])
def bareme_seuil():
    ok, m = metier.definir_seuil_passage(_role_courant(),
                                         request.form.get("seuil", ""),
                                         session.get("user", ""))
    flash(m, "ok" if ok else "err")
    return redirect(url_for("module", onglet="N1_Bareme_UE"))


@app.route("/budget/previsionnel/imprimer")
def budget_previsionnel_imprimer():
    _exige_budget()
    formation = request.args.get("formation", "").strip()
    session_ = request.args.get("session", "").strip()
    return render_template(
        "budget_previsionnel_print.html",
        data=metier.donnees_budget_prev(formation, session_),
        formation=formation, session_sel=session_,
        date_jour=metier.fmt_date(metier._dt.date.today()),
        titre_page="Budget previsionnel par formation")


# --- C-5 (V1.99.13) : synthese budgetaire prevu / realise / ecart ----------
def _synthese_ctx():
    """Contexte commun aux deux vues synthese (filtres + data)."""
    session_ = request.args.get("session", "").strip() or metier._annee_acad_defaut()
    formation = request.args.get("formation", "").strip()
    bailleur = request.args.get("bailleur", "").strip()
    return {
        "data": metier.synthese_budgetaire(session_, formation, bailleur),
        "session_sel": session_, "formation": formation, "bailleur": bailleur,
        "sessions_opt": metier.sessions_budget(),
        "formations_opt": metier.options_liste("Filieres") or [],
        "bailleurs_opt": metier.options_liste("Sources_financement") or [],
        "peut_g1": (metier.est_admin(_role_courant())
                    or metier.peut_ecrire(_role_courant(), "G1_Plan_action")),
    }


@app.route("/budget/synthese")
def budget_synthese():
    _exige_budget()
    ctx = _synthese_ctx()
    return render_template("budget_synthese.html",
                           info=config.TAB_INDEX["SYN_Budget"],
                           titre_page="Synthese budgetaire", **ctx)


@app.route("/budget/synthese/imprimer")
def budget_synthese_imprimer():
    _exige_budget()
    ctx = _synthese_ctx()
    return render_template("budget_synthese_print.html",
                           date_jour=metier.fmt_date(metier._dt.date.today()),
                           titre_page="Synthese budgetaire", **ctx)


@app.route("/cloture/eleves", methods=["POST"])
def cloture_eleves():
    _exige_direction()
    f = request.form
    annee = f.get("annee_scolaire", "").strip()
    # Saisies diplome/mention par matricule (champs nommes diplome_<mat> / mention_<mat>).
    saisies = {}
    for k in f:
        if k.startswith("diplome_"):
            saisies.setdefault(k[8:], {})["diplome"] = f.get(k, "").strip()
        elif k.startswith("mention_"):
            saisies.setdefault(k[8:], {})["mention"] = f.get(k, "").strip()
    n, msg, pv_eleves = metier.cloturer_eleves(_role_courant(), annee, saisies,
                                               session.get("user", ""))
    flash(msg, "ok" if n else "err")
    if n:
        auth.journal(session.get("user", ""), "Cloture eleves", annee, "%d journalise(s)" % n)
        session["pv"] = {"kind": "eleves", "annee": annee, "n": n,
                         "eleves": pv_eleves, "login": session.get("user", ""),
                         "date": metier.datetime.date.today().strftime("%d/%m/%Y")}
    return redirect(url_for("cloture"))


@app.route("/cloture/eleves/archiver", methods=["POST"])
def cloture_archiver():
    _exige_direction()
    n, msg, recap = metier.archiver_eleves(_role_courant(), session.get("user", ""))
    flash(msg, "ok" if n else "err")
    if n:
        auth.journal(session.get("user", ""), "Archivage eleves", "", "%d archive(s)" % n)
        session["pv"] = {"kind": "archivage", "n": n, "cohortes": recap,
                         "login": session.get("user", ""),
                         "date": metier.datetime.date.today().strftime("%d/%m/%Y")}
    return redirect(url_for("cloture"))


@app.route("/cloture/compta", methods=["POST"])
def cloture_compta():
    _exige_direction()
    annee = request.form.get("annee_civile", "").strip()
    ok, msg, recap = metier.cloturer_compta(_role_courant(), annee, session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Cloture compta", annee, recap.get("ref_archive", ""))
        recap.update({"kind": "compta", "login": session.get("user", ""),
                      "date": metier.datetime.date.today().strftime("%d/%m/%Y")})
        session["pv"] = recap
    return redirect(url_for("cloture"))


@app.route("/cloture/annee-suivante", methods=["POST"])
def cloture_annee_suivante():
    _exige_direction()
    ok, msg = metier.passer_annee_suivante(_role_courant(), session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("cloture"))


@app.route("/cloture/annee-courante", methods=["POST"])
def cloture_annee_courante():
    _exige_direction()
    label = request.form.get("annee_courante", "").strip()
    ok, msg = metier.definir_annee_courante(_role_courant(), label, session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("cloture"))


@app.route("/cloture/pv")
def cloture_pv():
    _exige_direction()
    pv = session.get("pv")
    if not pv:
        flash("Aucun proces-verbal recent a afficher.", "err")
        return redirect(url_for("cloture"))
    return render_template("pv_cloture.html", blocs=metier.pv_blocs(pv),
                           titre_page="Proces-verbal de cloture")


@app.route("/cloture/pv.docx")
def cloture_pv_docx():
    _exige_direction()
    pv = session.get("pv")
    if not pv:
        flash("Aucun proces-verbal recent a telecharger.", "err")
        return redirect(url_for("cloture"))
    import tempfile
    nom = "PV_Cloture_%s.docx" % (pv.get("annee") or pv.get("kind") or "EMSP")
    chemin = metier.os.path.join(tempfile.gettempdir(), nom)
    metier.generer_docx(chemin, metier.pv_blocs(pv))
    return send_file(chemin, as_attachment=True, download_name=nom)


@app.route("/journal")
def journal_audit():
    _exige_admin()
    f_login = request.args.get("login", "").strip()
    f_action = request.args.get("action", "").strip()
    return render_template("journal.html",
                           entrees=auth.lire_journal(400, f_login, f_action),
                           f_login=f_login, f_action=f_action,
                           titre_page="Journal d'audit")


@app.route("/import")
def import_csv():
    if not metier.peut_ecrire(_role_courant(), "IMPORT_zone"):
        abort(403)
    return render_template("import.html",
                           data=metier.import_resume(),
                           info=config.TAB_INDEX["IMPORT_zone"],
                           titre_page="Import CSV national")


def _garde_import():
    if not metier.peut_ecrire(_role_courant(), "IMPORT_zone"):
        abort(403)


@app.route("/import/importer", methods=["POST"])
def import_importer():
    _garde_import()
    try:
        n, msg = metier.importer_csv(request.form.get("csv", ""))
        flash(msg, "ok" if n else "err")
        if n:
            auth.journal(session.get("user", ""), "Import CSV", "IMPORT_zone", "%d ligne(s)" % n)
    except OverflowError:
        flash("Trop de lignes pour la zone d'import.", "err")
    except Exception:
        flash("Echec de l'import.", "err")
    return redirect(url_for("import_csv"))


@app.route("/import/vider", methods=["POST"])
def import_vider():
    _garde_import()
    flash(metier.vider_zone_import(), "ok")
    return redirect(url_for("import_csv"))


@app.route("/import/annuler", methods=["POST"])
def import_annuler():
    _garde_import()
    ok, msg = metier.annuler_import()
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("import_csv"))


# ===========================================================================
# MODULE IMPRESSIONS & EDITIONS — V1.18
# ===========================================================================
@app.route("/modeles")
def modeles_docs():
    """Parametrages -> Modeles de documents (edition des parties fixes)."""
    if not metier.peut_ecrire(_role_courant(), config.MODELE_TAB):
        abort(403)
    return render_template("modeles.html",
                           modeles=metier.modeles_docs(),
                           info=config.TAB_INDEX["ED_Modeles"],
                           titre_page="Modeles de documents")


@app.route("/modeles/enregistrer", methods=["POST"])
def modeles_enregistrer():
    if not metier.peut_ecrire(_role_courant(), config.MODELE_TAB):
        abort(403)
    cle = request.form.get("cle", "").strip()
    valeurs = {col: request.form.get(col, "")
               for col in ["En-tete", "Titre", "Corps", "Mentions / pied",
                           "Libelle signataire", "Nb copies"]}
    ok, msg = metier.enregistrer_modele(cle, valeurs)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "MAJ modele document", cle, "")
    return redirect(url_for("modeles_docs"))


# --- Bibliotheque documentaire (V1.99.22) ---------------------------------
def _garde_biblio_depot():
    if not metier.a_droit_ecriture(_role_courant()):
        abort(403)


def _garde_biblio_admin():
    if not metier.est_admin(_role_courant()):
        abort(403)


@app.route("/bibliotheque")
def bibliotheque():
    sous = request.args.get("d", "")
    contenu = metier.bibliotheque_lister(sous)
    return render_template("bibliotheque.html", c=contenu,
                           peut_deposer=metier.a_droit_ecriture(_role_courant()),
                           peut_supprimer=metier.est_admin(_role_courant()),
                           info=config.TAB_INDEX["BIBLIO_Docs"],
                           titre_page="Bibliotheque documentaire")


@app.route("/bibliotheque/fichier/<path:relpath>")
def bibliotheque_fichier(relpath):
    if not metier.bibliotheque_existe(relpath):
        abort(404)
    base = os.path.abspath(config.BIBLIOTHEQUE_DIR)
    chemin = os.path.abspath(os.path.join(base, relpath.replace("/", os.sep)))
    return send_file(chemin, as_attachment=False,
                     download_name=os.path.basename(chemin))


@app.route("/bibliotheque/deposer", methods=["POST"])
def bibliotheque_deposer():
    _garde_biblio_depot()
    sous = request.form.get("sous", "")
    f = request.files.get("fichier")
    if not f or not f.filename:
        flash("Choisissez un fichier a deposer.", "err")
        return redirect(url_for("bibliotheque", d=sous))
    rel = metier.bibliotheque_enregistrer(sous, f.filename, f)
    if rel:
        flash("Document depose.", "ok")
        auth.journal(session.get("user", ""), "Depot bibliotheque", "", rel)
    else:
        flash("Depot impossible (dossier invalide).", "err")
    return redirect(url_for("bibliotheque", d=sous))


@app.route("/bibliotheque/dossier", methods=["POST"])
def bibliotheque_dossier():
    _garde_biblio_depot()
    sous = request.form.get("sous", "")
    nom = request.form.get("nom", "")
    if metier.bibliotheque_creer_dossier(sous, nom):
        flash("Dossier cree.", "ok")
        auth.journal(session.get("user", ""), "Creation dossier bibliotheque", "", nom)
    else:
        flash("Nom de dossier invalide.", "err")
    return redirect(url_for("bibliotheque", d=sous))


@app.route("/bibliotheque/supprimer", methods=["POST"])
def bibliotheque_supprimer():
    _garde_biblio_admin()
    rel = request.form.get("rel", "")
    sous = "/".join(rel.split("/")[:-1])
    if metier.bibliotheque_supprimer_fichier(rel):
        flash("Document supprime.", "ok")
        auth.journal(session.get("user", ""), "Suppression bibliotheque", "", rel)
    else:
        flash("Suppression impossible.", "err")
    return redirect(url_for("bibliotheque", d=sous))


@app.route("/impressions")
def impressions():
    """Hub des impressions : lance chaque document avec ses criteres."""
    return render_template("impressions.html",
                           modeles=metier.modeles_docs(),
                           filtres=metier.liste_filtres(),
                           etudiants=metier.etudiants_dispo(),
                           recettes=metier.recettes_dispo(),
                           mois=metier.mois_dispo(),
                           comptes_treso=metier.comptes_treso(),
                           mois_treso=metier.mois_treso_dispo(),
                           presence_lignes=config.PRESENCE_LIGNES_DEFAUT,
                           peut_financier=(_role_courant().get("financier") == "O"),
                           annee_civile=metier._dt.date.today().year,
                           salles_equip=metier.salles_equip_dispo(),
                           bailleurs_equip=metier.bailleurs_equip_dispo(),
                           etats_equip=metier.etats_equip_dispo(),
                           besoin_statuts=metier.besoin_statuts_dispo(),
                           besoin_priorites=metier.besoin_priorites_dispo(),
                           annees_f1=metier.annees_civiles_f1(),
                           info=config.TAB_INDEX["ED_Impressions"],
                           titre_page="Impressions & editions")


def _page_impression(doc, kind, **ctx):
    return render_template("imprimer.html", doc=doc, kind=kind,
                           titre_page="Impression — " + doc["titre"], **ctx)


@app.route("/impressions/liste")
def impr_liste():
    f = request.args
    data = metier.liste_etudiants(f.get("filiere", ""), f.get("niveau", ""),
                                  f.get("section", ""), f.get("annee", ""))
    doc = metier.rendre_modele("LISTE_ETUD", {})
    return _page_impression(doc, "liste", table=data)


@app.route("/impressions/presence-vierge")
def impr_presence():
    filiere = request.args.get("filiere", "").strip()
    niveau = request.args.get("niveau", "").strip()
    section = request.args.get("section", "").strip()
    date_fr = request.args.get("date", "").strip()
    debut = request.args.get("debut", "").strip()
    fin = request.args.get("fin", "").strip()
    matiere = request.args.get("matiere", "").strip()
    enseignant = request.args.get("enseignant", "").strip()
    # Selection obligatoire (R3) : au moins la classe = Filiere + Niveau.
    if not (filiere and niveau):
        flash("Choisissez au moins une classe (filiere et niveau) pour la "
              "feuille de presence.", "err")
        return redirect(url_for("impressions"))
    try:
        n = int(request.args.get("lignes", config.PRESENCE_LIGNES_DEFAUT))
    except ValueError:
        n = config.PRESENCE_LIGNES_DEFAUT
    n = min(max(n, 1), 200)
    pv = metier.feuille_presence_vierge(filiere, niveau, section, date_fr,
                                        debut, fin, matiere, enseignant, n_defaut=n)
    doc = metier.rendre_modele("PRESENCE_VIERGE", {})
    return _page_impression(doc, "presence", pv=pv)


@app.route("/impressions/releve-individuel")
def impr_releve_ind():
    mois = request.args.get("mois", "").strip()
    mat = request.args.get("matricule", "").strip()
    jetons = metier.releve_individuel(mois, mat) if (mois and mat) else None
    if jetons is None:
        flash("Aucun releve pour cet enseignant sur cette periode.", "err")
        return redirect(url_for("impressions"))
    doc = metier.rendre_modele("RELEVE_IND", jetons)
    return _page_impression(doc, "prose")


@app.route("/impressions/releve-recap")
def impr_releve_recap():
    mois = request.args.get("mois", "").strip()
    if not mois:
        flash("Choisissez une periode (Mois / Annee).", "err")
        return redirect(url_for("impressions"))
    data = metier.releve_recap(mois)
    doc = metier.rendre_modele("RELEVE_RECAP", {})
    return _page_impression(doc, "recap", table=data)


@app.route("/impressions/recu")
def impr_recu():
    if _role_courant().get("financier") != "O":
        abort(403)
    idx = request.args.get("i", "").strip()
    jetons = metier.recu_jetons(idx) if idx != "" else None
    if jetons is None:
        flash("Recette introuvable (selectionnez une recette de la tresorerie).", "err")
        return redirect(url_for("impressions"))
    doc = metier.rendre_modele("RECU", jetons)
    return _page_impression(doc, "prose")


@app.route("/impressions/attestation")
def impr_attestation():
    mat = request.args.get("matricule", "").strip()
    jetons = metier.attestation_jetons(mat) if mat else None
    if jetons is None:
        flash("Etudiant introuvable.", "err")
        return redirect(url_for("impressions"))
    doc = metier.rendre_modele("ATTESTATION", jetons)
    return _page_impression(doc, "prose")


@app.route("/impressions/situation-compte")
def impr_situation_compte():
    if _role_courant().get("financier") != "O":
        abort(403)
    compte = request.args.get("compte", "").strip()
    mois = request.args.get("mois", "").strip()
    data = metier.situation_compte(compte, mois) if (compte and mois) else None
    if data is None:
        flash("Choisissez un compte et une periode (MM/AAAA).", "err")
        return redirect(url_for("impressions"))
    ctx = data["contexte"]
    doc = metier.rendre_modele("SITUATION_COMPTE", {
        "compte": ctx["compte"], "periode": ctx["periode"], "date_jour": ctx["date_jour"]})
    return _page_impression(doc, "situation", table=data)


@app.route("/impressions/export-tdb")
def impr_export_tdb():
    from flask import send_file
    data = metier.export_tdb_xlsx()
    buf = __import__("io").BytesIO(data)
    nom = "tableau_de_bord_EMSP_%s.xlsx" % metier._dt.date.today().strftime("%Y%m%d")
    return send_file(buf, as_attachment=True, download_name=metier.nom_export(nom),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/impressions/presence-semaine")
def impr_presence_semaine():
    f = request.args
    data = metier.feuille_presence_semaine(f.get("filiere", ""), f.get("niveau", ""),
                                           f.get("section", ""), f.get("annee", ""))
    doc = metier.doc_adhoc("Feuille de presence de la semaine",
                           mentions="En cas de rectification, les services concernes seront avertis.",
                           signataire="Signature de l'enseignant / du delegue", orientation="paysage")
    return _page_impression(doc, "presence_semaine", table=data)


@app.route("/impressions/appreciation-stage")
def impr_appreciation_stage():
    return render_template("appreciation_stage.html",
                           titre_page="Fiche d'appreciation de stage")


@app.route("/impressions/plan-action")
def impr_plan_action():
    data = metier.plan_action_liste()
    doc = metier.doc_adhoc("Plan d'action — suivi des ecarts et axes de progres",
                           signataire="La Direction", orientation="paysage")
    return _page_impression(doc, "table", table=data)


@app.route("/impressions/journal-treso")
def impr_journal_treso():
    if _role_courant().get("financier") != "O":
        abort(403)
    data = metier.journal_treso(request.args.get("compte", "").strip(),
                                request.args.get("mois", "").strip())
    doc = metier.doc_adhoc("Journal de tresorerie",
                           mentions="Montants en KMF.",
                           signataire="Le Gestionnaire | Le Directeur de l'EMSP")
    return _page_impression(doc, "table", table=data)


@app.route("/impressions/balance")
def impr_balance():
    if _role_courant().get("financier") != "O":
        abort(403)
    data = metier.balance_comptes()
    doc = metier.doc_adhoc("Situation globale des comptes (balance)",
                           mentions="Montants en KMF.",
                           signataire="Le Gestionnaire | Le Directeur de l'EMSP")
    return _page_impression(doc, "table", table=data)


@app.route("/impressions/etat-poste")
def impr_etat_poste():
    if _role_courant().get("financier") != "O":
        abort(403)
    exercice = request.args.get("exercice", "").strip()
    if exercice:
        # V1.70 : comparatif budget prevu / realise / ecart (annee civile).
        data = metier.etat_poste_budget(exercice)
        doc = metier.doc_adhoc(
            "Budget par poste — prevu / realise / ecart (exercice %s)" % exercice,
            mentions="Montants en KMF. Realise = mouvements de l'annee civile %s." % exercice,
            signataire="Le Gestionnaire | Le Directeur de l'EMSP",
            orientation="paysage")
    else:
        data = metier.etat_par_poste(request.args.get("mois", "").strip(),
                                     request.args.get("compte", "").strip())
        doc = metier.doc_adhoc("Etat des recettes et depenses par poste budgetaire",
                               mentions="Montants en KMF.",
                               signataire="Le Gestionnaire | Le Directeur de l'EMSP")
    return _page_impression(doc, "table", table=data)


@app.route("/impressions/bulletin")
def impr_bulletin():
    if not metier.peut_lire(_role_courant(), "N2_Notes"):
        abort(403)
    mat = request.args.get("matricule", "").strip()
    annee = request.args.get("annee", "").strip()
    semestre = request.args.get("semestre", "").strip()
    # V1.99.50 : edition par session (modele du releve officiel) —
    # '1' = page Premiere session (sans colonnes rattrapage),
    # '2' ou '' = Deuxieme session si des notes de session 2 existent.
    sess = request.args.get("session", "").strip()
    if not (mat and annee):
        flash("Choisissez un etudiant et une annee.", "err")
        return redirect(url_for("impressions"))
    djour = metier.fmt_date(metier._dt.date.today())
    if semestre:
        rel = metier.releve_semestre(mat, annee, semestre,
                                     "1" if sess == "1" else None)
        return render_template("releve_print.html", rel=rel, date_jour=djour,
                               titre_page="Releve de notes")
    bo = metier.bulletin_officiel(mat, annee, sess)
    return render_template("bulletin_officiel.html", bo=bo, date_jour=djour,
                           titre_page="Releve de notes")


@app.route("/impressions/inventaire")
def impr_inventaire():
    # V1.75 : inventaire equipements (M1) regroupe par salle / bailleur / etat.
    axe = request.args.get("axe", "salle").strip().lower()
    data = metier.inventaire_equipements(axe)
    libelle = {"salle": "par salle", "bailleur": "par source de financement",
               "etat": "par etat"}.get(axe, "par salle")
    doc = metier.doc_adhoc("Inventaire des equipements %s" % libelle,
                           mentions="Montants en KMF. Source : module Equipements (M1).",
                           signataire="Le Responsable logistique | Le Directeur de l'EMSP",
                           orientation="paysage")
    return _page_impression(doc, "table", table=data)


@app.route("/impressions/bon-besoin")
def impr_bon_besoin():
    # V1.75 : etat des expressions de besoin (L3), filtrable statut/priorite/salle.
    f = request.args
    data = metier.bon_de_besoin(f.get("statut", ""), f.get("priorite", ""),
                                f.get("salle", ""))
    doc = metier.doc_adhoc("Expression de besoin — equipements et fournitures",
                           mentions="Montants en KMF. Source : module Expression de besoin (L3).",
                           signataire="Le Demandeur | Le Responsable logistique | Le Directeur",
                           orientation="paysage")
    return _page_impression(doc, "table", table=data)


@app.route("/impressions/etat-bailleur")
def impr_etat_bailleur():
    # V1.75 : recettes/depenses par source de financement (F1) — reporting AFD.
    if _role_courant().get("financier") != "O":
        abort(403)
    annee = request.args.get("annee", "").strip()
    data = metier.etat_par_bailleur(annee)
    titre = "Etat des recettes et depenses par source de financement"
    if annee:
        titre += " (annee civile %s)" % annee
    doc = metier.doc_adhoc(titre,
                           mentions="Montants en KMF. Solde = recettes - depenses par bailleur.",
                           signataire="Le Gestionnaire | Le Directeur de l'EMSP")
    return _page_impression(doc, "table", table=data)


@app.route("/api/releve-individuels")
def api_releve_individuels():
    """Enseignants ayant un releve pour une periode (selecteur dependant)."""
    return jsonify(metier.releves_individuels_dispo(request.args.get("mois", "")))


# ===========================================================================
# SAISIE EN GRILLE — REGISTRE DE TRESORERIE (F1) — V1.20
# ===========================================================================
@app.route("/tresorerie")
def tresorerie():
    """Grille de saisie du registre de tresorerie (mode tableur)."""
    _garde_ecriture("F1_Mouvements")
    compte = request.args.get("compte", "").strip()
    comptes = metier.comptes_treso()
    if not compte and comptes:
        compte = comptes[0]["nom"]
    solde_ouverture = metier.solde_courant_compte(compte) if compte else 0.0
    try:
        nb = int(request.args.get("lignes", 12))
    except ValueError:
        nb = 12
    nb = min(max(nb, 1), 100)
    return render_template("tresorerie.html",
                           info=config.TAB_INDEX["F1_Mouvements"],
                           colonnes=metier.treso_grille_colonnes(),
                           comptes=comptes, compte=compte,
                           solde_ouverture=solde_ouverture,
                           nb_lignes=nb, rows_pref=None, erreurs=None,
                           cap=metier.capacite_onglet("F1_Mouvements"),
                           titre_page="Tresorerie — saisie en grille")


@app.route("/tresorerie/enregistrer", methods=["POST"])
def tresorerie_enregistrer():
    _garde_ecriture("F1_Mouvements")
    compte = request.form.get("compte", "").strip()
    colonnes = metier.treso_grille_colonnes()
    # Reconstruire les lignes a partir des colonnes paralleles c0[]..cN[]
    colonnes_vals = [request.form.getlist("c%d[]" % j) for j in range(len(colonnes))]
    nb = max((len(v) for v in colonnes_vals), default=0)
    lignes = []
    for i in range(nb):
        lig = {}
        for j, col in enumerate(colonnes):
            vals = colonnes_vals[j]
            lig[col["libelle"]] = vals[i] if i < len(vals) else ""
        lignes.append(lig)
    n, msg, erreurs = metier.enregistrer_treso_lot(
        compte, lignes, saisi_par=_role_courant().get("login", ""))
    if erreurs or n == 0:
        flash(msg, "err")
        comptes = metier.comptes_treso()
        return render_template("tresorerie.html",
                               info=config.TAB_INDEX["F1_Mouvements"],
                               colonnes=colonnes, comptes=comptes, compte=compte,
                               solde_ouverture=metier.solde_courant_compte(compte),
                               nb_lignes=max(len(lignes), 1),
                               rows_pref=lignes, erreurs=erreurs,
                               cap=metier.capacite_onglet("F1_Mouvements"),
                               titre_page="Tresorerie — saisie en grille")
    flash(msg, "ok")
    auth.journal(session.get("user", ""), "Saisie tresorerie", compte, "%d ligne(s)" % n)
    return redirect(url_for("tresorerie", compte=compte))


# ===========================================================================
# COUCHE REQUETES MULTICRITERES (LECTURE SEULE) — V1.22
# ===========================================================================
@app.route("/requetes")
def requetes():
    vues = [{"id": vid, **metier.VUES[vid]} for vid in metier.VUES_ORDRE]
    return render_template("requetes.html",
                           tables=metier.tables_interrogeables(),
                           vues=vues,
                           info=config.TAB_INDEX["REQ_Hub"],
                           titre_page="Requetes & analyses")


def _parse_filtres(args):
    filtres = []
    for i in range(3):
        col = args.get("f%d_col" % i, "").strip()
        if col:
            filtres.append({"col": col, "op": args.get("f%d_op" % i, "contient"),
                            "val": args.get("f%d_val" % i, "")})
    return filtres


@app.route("/requetes/explorer")
def requetes_explorer():
    args = request.args
    onglet = args.get("onglet", "").strip()
    tables = metier.tables_interrogeables()
    if not onglet and tables:
        onglet = tables[0]["onglet"]
    filtres = _parse_filtres(args)
    colonnes_sel = args.getlist("col")
    tri_col = args.get("tri_col", "").strip()
    tri_sens = args.get("tri_sens", "asc")
    res = metier.explorer(onglet, filtres, colonnes_sel, tri_col, tri_sens)
    if res is None:
        flash("Table inconnue.", "err")
        return redirect(url_for("requetes"))
    if args.get("export") == "xlsx":
        from flask import send_file
        data = metier._xlsx_simple(res["libelle"], res["colonnes"], res["lignes"])
        nom = "export_%s_%s.xlsx" % (onglet, metier._dt.date.today().strftime("%Y%m%d"))
        return send_file(__import__("io").BytesIO(data), as_attachment=True,
                         download_name=metier.nom_export(nom),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return render_template("explorer.html", tables=tables, onglet=onglet, res=res,
                           filtres=filtres, colonnes_sel=colonnes_sel,
                           tri_col=tri_col, tri_sens=tri_sens, operateurs=metier._OPERATEURS,
                           info=config.TAB_INDEX["REQ_Hub"], titre_page="Explorateur")


@app.route("/requetes/vue/<vid>")
def requetes_vue(vid):
    spec = metier.VUES.get(vid)
    if not spec:
        abort(404)
    res = spec["builder"](request.args)
    if request.args.get("export") == "xlsx":
        from flask import send_file
        data = metier._xlsx_simple(res["titre"], res["colonnes"], res["lignes"])
        nom = "vue_%s_%s.xlsx" % (vid, metier._dt.date.today().strftime("%Y%m%d"))
        return send_file(__import__("io").BytesIO(data), as_attachment=True,
                         download_name=metier.nom_export(nom),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    options = {
        "salles": metier.options_liste("Salles (L1)") or [],
        "bailleurs": metier.options_liste("Sources_financement") or [],
        "categories": metier.options_liste("Categories_equipement") or [],
        "mois": metier.mois_dispo(),
        "etudiants": metier.etudiants_dispo(),
    }
    return render_template("vue.html", vid=vid, spec=spec, res=res, options=options,
                           args=request.args, info=config.TAB_INDEX["REQ_Hub"],
                           titre_page=spec["libelle"])


@app.route("/requetes/pivot")
def requetes_pivot():
    args = request.args
    tables = metier.tables_interrogeables()
    onglet = args.get("onglet", "").strip() or (tables[0]["onglet"] if tables else "")
    cols_all = metier.colonnes_table(onglet) if onglet else []
    lig = args.get("lig", "").strip()
    col = args.get("col", "").strip()
    mesure = args.get("mesure", "count").strip()
    mes_col = args.get("mes_col", "").strip()
    res = metier.pivot(onglet, lig, col, mesure, mes_col) if lig else None
    if res and args.get("export") == "xlsx":
        from flask import send_file
        data = metier._xlsx_simple(res["sous_titre"], res["colonnes"], res["lignes"])
        nom = "pivot_%s_%s.xlsx" % (onglet, metier._dt.date.today().strftime("%Y%m%d"))
        return send_file(__import__("io").BytesIO(data), as_attachment=True,
                         download_name=metier.nom_export(nom),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return render_template("pivot.html", tables=tables, onglet=onglet, cols_all=cols_all,
                           lig=lig, col=col, mesure=mesure, mes_col=mes_col, res=res,
                           info=config.TAB_INDEX["REQ_Hub"], titre_page="Tableau croise")


@app.errorhandler(403)
def acces_refuse(e):
    return render_template("404.html", titre_page="Acces refuse",
                           message="Votre role n'autorise pas cette action."), 403


@app.errorhandler(404)
def page_introuvable(e):
    return render_template("404.html", titre_page="Page introuvable"), 404


@app.route("/releve")
def releve():
    role = _role_courant()
    if not metier.peut_lire(role, "N2_Notes"):
        abort(403)
    matricule = _matricule_saisi(request.args.get("matricule", ""))
    annee = request.args.get("annee", "").strip()
    semestre = request.args.get("semestre", "").strip()  # "" = recap annuel
    rel = annuel = None
    if matricule and annee:
        if semestre:
            rel = metier.releve_semestre(matricule, annee, semestre)
        else:
            annuel = metier.releve_annuel(matricule, annee)
    return render_template("releve.html", info=config.TAB_INDEX["REL_Releve"],
                           titre_page="Releve / bulletin", matricule=matricule,
                           annee=annee, semestre=semestre, rel=rel, annuel=annuel,
                           opt_annees=metier.options_liste("Annees_acad"),
                           etudiants=metier.etudiants_dispo())


@app.route("/etudiant")
def etudiant():
    role = _role_courant()
    if not metier.peut_lire(role, "A1_Etudiants"):
        abort(403)
    matricule = _matricule_saisi(request.args.get("matricule", ""))
    fiche = metier.fiche_etudiant(matricule) if matricule else None
    stages = metier.stages_etudiant(matricule) if fiche else []
    annee = fiche.get("Annee acad.", "") if fiche else ""
    droits = metier.droits_inscription(matricule) if fiche else None
    return render_template("etudiant.html", info=config.TAB_INDEX["ETU_Fiche"],
                           titre_page="Fiche etudiant", matricule=matricule,
                           fiche=fiche, stages=stages, annee=annee,
                           droits=droits,
                           peut_encaisser=(_role_courant().get("financier") == "O"),
                           etudiants=metier.recherche_etudiants(),
                           peut_modifier=metier.peut_ecrire(role, "A1_Etudiants"),
                           introuvable=(bool(matricule) and fiche is None))


@app.route("/etudiant/<matricule>/encaisser", methods=["GET", "POST"])
def etudiant_encaisser(matricule):
    """Mini-ecran d'encaissement d'un droit d'inscription (reserve au droit
    financier ; bloque sur poste secondaire par le before_request ; journalise).
    Pre-remplit poste/montant/annee depuis la fiche ; Compte/Mode obligatoires."""
    role = _role_courant()
    _garde_ecriture("F1_Mouvements")
    if role.get("financier") != "O":
        abort(403)
    matricule = str(matricule).strip()
    fiche = metier.fiche_etudiant(matricule)
    if not fiche:
        abort(404)
    droits = metier.droits_inscription(matricule)
    comptes = metier.comptes_treso()
    modes = metier.options_liste("Modes_paiement") or []
    annees = metier.options_liste("Annees_acad") or []
    if request.method == "POST":
        ok, msg = metier.enregistrer_encaissement(
            matricule,
            montant=request.form.get("montant", ""),
            annee=request.form.get("annee", ""),
            compte=request.form.get("compte", ""),
            mode_paiement=request.form.get("mode", ""),
            reference=request.form.get("reference", ""),
            saisi_par=role.get("login", ""))
        if ok:
            flash(msg, "ok")
            auth.journal(session.get("user", ""), "Encaissement inscription",
                         matricule, msg)
            return redirect(url_for("etudiant", matricule=matricule))
        flash(msg, "err")
    return render_template("encaisser.html", info=config.TAB_INDEX["ETU_Fiche"],
                           titre_page="Encaisser un droit d'inscription",
                           fiche=fiche, droits=droits, comptes=comptes,
                           modes=modes, annees=annees)


@app.route("/etudiant/photo/<matricule>")
def etudiant_photo(matricule):
    role = _role_courant()
    if not metier.peut_lire(role, "A1_Etudiants"):
        abort(403)
    chemin, mime = metier.photo_servie(matricule)
    if chemin:
        return send_file(chemin, mimetype=mime)
    # Placeholder portrait (charte #1F4E79) tant qu'aucune photo n'est deposee.
    svg = ('<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 120 120">'
           '<rect width="120" height="120" rx="10" fill="#EAEFF5"/>'
           '<circle cx="60" cy="46" r="21" fill="#1F4E79"/>'
           '<path d="M24 104c0-20 16-31 36-31s36 11 36 31z" fill="#1F4E79"/></svg>')
    return Response(svg, mimetype="image/svg+xml")


@app.route("/etudiant/photo/<matricule>/televerser", methods=["POST"])
def etudiant_photo_televerser(matricule):
    # Action reservee (ecriture A1_Etudiants) ; bloquee sur poste secondaire
    # (lecture seule) par le before_request. N'ecrit que le fichier image.
    _garde_ecriture("A1_Etudiants")
    f = request.files.get("photo")
    maxi = getattr(config, "PHOTO_MAX_OCTETS", 1024 * 1024)
    # Lecture bornee a maxi+1 : une image plus lourde est rejetee sans tout charger.
    donnees = f.read(maxi + 1) if f else b""
    ok, msg = metier.enregistrer_photo(matricule, donnees)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Photo etudiant", matricule, "Televersement")
    return redirect(url_for("etudiant", matricule=matricule))


@app.route("/etudiant/photo/<matricule>/retirer", methods=["POST"])
def etudiant_photo_retirer(matricule):
    _garde_ecriture("A1_Etudiants")
    ok, msg = metier.supprimer_photo(matricule)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Photo etudiant", matricule, "Retrait")
    return redirect(url_for("etudiant", matricule=matricule))


# === Stages : suivi multicritere + tableau de bord (V1.81) =================

def _stages_filtres(src):
    return dict(lieu=src.get("lieu", "").strip(), seance=src.get("seance", "").strip(),
                niveau=src.get("niveau", "").strip(), filiere=src.get("filiere", "").strip(),
                nom=src.get("nom", "").strip(), annee=src.get("annee", "").strip())


@app.route("/stages")
def stages_suivi():
    role = _role_courant()
    if not metier.peut_lire(role, "S1_Stages"):
        abort(403)
    f = _stages_filtres(request.args)
    data = metier.stages_synthese(**f)
    return render_template("stages_suivi.html",
                           info=config.TAB_INDEX["STG_Suivi"],
                           titre_page="Stages — suivi & tableau de bord", **data)


@app.route("/stages/imprimer")
def stages_imprimer():
    role = _role_courant()
    if not metier.peut_lire(role, "S1_Stages"):
        abort(403)
    f = _stages_filtres(request.args)
    data = metier.stages_synthese(**f)
    return render_template("stages_print.html",
                           titre_page="Affectations de stage", **data,
                           date_jour=metier.fmt_date(metier.datetime.date.today()))


# === Stages : affectation automatique (PROPOSITION, lecture seule) (V1.87) ===

def _affect_filtres(src):
    return dict(filiere=src.get("filiere", "").strip(), niveau=src.get("niveau", "").strip(),
                annee=src.get("annee", "").strip(), seance=src.get("seance", "").strip(),
                session=(src.get("session", "").strip() or "Normale"),
                depassement=(src.get("depassement", "") in ("1", "on", "true", "oui")))


@app.route("/stages/affectation")
def stages_affectation():
    role = _role_courant()
    if not metier.peut_lire(role, "S1_Stages"):
        abort(403)
    f = _affect_filtres(request.args)
    donnees = None
    if f["filiere"] and f["niveau"] and f["annee"] and f["seance"]:
        donnees = metier.donnees_affectation_stages(f["filiere"], f["niveau"],
                                                     f["annee"], f["seance"], f["session"])
    return render_template("stages_affectation.html",
                           info=config.TAB_INDEX["STG_Auto"],
                           titre_page="Stages — affectation",
                           filtres=f, donnees=donnees,
                           opt_filieres=metier.options_liste("Filieres"),
                           opt_niveaux=metier.options_liste("Niveaux"),
                           opt_annees=metier.options_liste("Annees_acad"))


@app.route("/stages/affectation/imprimer")
def stages_affectation_imprimer():
    role = _role_courant()
    if not metier.peut_lire(role, "S1_Stages"):
        abort(403)
    f = _affect_filtres(request.args)
    if not (f["filiere"] and f["niveau"] and f["annee"] and f["seance"]):
        flash("Selection incomplete pour l'edition.", "err")
        return redirect(url_for("stages_affectation"))
    d = metier.donnees_affectation_stages(f["filiere"], f["niveau"], f["annee"], f["seance"], f["session"])
    cols = ["N°", "Lieu / service", "Matricule", "Nom", "Prenom", "Debut", "Fin"]
    lignes = []
    i = 0
    for s in d["sections"]:
        for e in s["eleves"]:
            i += 1
            lignes.append([str(i), s["lieu"], e["matricule"], e["nom"], e["prenom"],
                           e.get("debut", ""), e.get("fin", "")])
    for e in d["roster"]:
        if not e["affecte"]:
            i += 1
            lignes.append([str(i), "NON AFFECTE", e["matricule"], e["nom"], e["prenom"], "", ""])
    st = d["stats"]
    leg = ("%s · %s · Annee %s · Seance %s · Effectif %d · Affectes %d · Non affectes %d · "
           "Edite le %s" % (
               f["filiere"], f["niveau"], f["annee"], f["seance"], st["effectif"],
               st["affectes"], st["non_affectes"],
               metier.fmt_date(metier.datetime.date.today())))
    table = {"colonnes": cols, "lignes": lignes, "total": None, "contexte": {"legende": leg}}
    doc = metier.doc_adhoc("Affectation des stages — %s %s (seance %s)"
                           % (f["filiere"], f["niveau"], f["seance"]),
                           mentions="Etat des affectations enregistrees (S1_Stages), "
                                    "groupees par lieu de stage.",
                           signataire="Le responsable des stages | Le Directeur", orientation="paysage")
    return _page_impression(doc, "table", table=table)


# === Resultats par classe : synthese 1 ligne/eleve + impression (V1.85) =====

def _resultats_filtres(src):
    return dict(filiere=src.get("filiere", "").strip(), niveau=src.get("niveau", "").strip(),
                section=src.get("section", "").strip(), annee=src.get("annee", "").strip(),
                periode=src.get("periode", "").strip())


@app.route("/resultats-classe")
def resultats_classe():
    role = _role_courant()
    if not metier.peut_lire(role, "N2_Notes"):
        abort(403)
    f = _resultats_filtres(request.args)
    semestres = (metier.semestres_classe(f["filiere"], f["niveau"])
                 if (f["filiere"] and f["niveau"]) else [])
    data = None
    if f["filiere"] and f["niveau"] and f["annee"] and f["periode"]:
        data = metier.resultats_classe(f["filiere"], f["niveau"], f["section"],
                                       f["annee"], f["periode"])
    delib = (metier.deliberation_classe(f["annee"], f["filiere"], f["niveau"])
             if (f["filiere"] and f["niveau"] and f["annee"]) else None)
    valides = (metier.bulletins_valides(f["annee"], f["filiere"], f["niveau"])
               if (f["filiere"] and f["niveau"] and f["annee"]) else set())
    return render_template("resultats_classe.html",
                           info=config.TAB_INDEX["RES_Classe"],
                           titre_page="Resultats par classe (synthese)",
                           filtres=f, semestres=semestres, data=data,
                           delib=delib, valides=valides,
                           peut_deliberer=metier.peut_ecrire(role, "N2_Notes"),
                           opt_filieres=metier.options_liste("Filieres"),
                           opt_niveaux=metier.options_liste("Niveaux"),
                           opt_sections=metier.options_liste("Sections"),
                           opt_annees=metier.options_liste("Annees_acad"))


@app.route("/resultats-classe/validation", methods=["POST"])
def resultats_validation():
    _garde_ecriture("N2_Notes")
    f = _resultats_filtres(request.form)
    mats = request.form.getlist("valides")
    ok, msg = metier.enregistrer_validations(_role_courant(), f["annee"], f["filiere"],
                                             f["niveau"], mats, session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("resultats_classe", filiere=f["filiere"], niveau=f["niveau"],
                            section=f["section"], annee=f["annee"], periode=f["periode"]))


@app.route("/resultats-classe/deliberation", methods=["POST"])
def resultats_deliberation():
    _garde_ecriture("N2_Notes")
    f = _resultats_filtres(request.form)
    ok, msg = metier.enregistrer_deliberation(_role_courant(), f["annee"], f["filiere"],
                                              f["niveau"], request.form.get("date_deliberation", ""),
                                              session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    return redirect(url_for("resultats_classe", filiere=f["filiere"], niveau=f["niveau"],
                            section=f["section"], annee=f["annee"], periode=f["periode"]))


@app.route("/resultats-classe/attestation/<matricule>")
def resultats_attestation(matricule):
    role = _role_courant()
    if not metier.peut_lire(role, "N2_Notes"):
        abort(403)
    f = _resultats_filtres(request.args)
    ok, msg, octets, nom = metier.generer_attestation(
        matricule, f["annee"], f["filiere"], f["niveau"], session.get("user", ""))
    if not ok:
        flash(msg, "err")
        return redirect(url_for("resultats_classe", filiere=f["filiere"], niveau=f["niveau"],
                                section=f["section"], annee=f["annee"], periode=f["periode"]))
    from flask import send_file
    return send_file(__import__("io").BytesIO(octets), as_attachment=True,
                     download_name=nom, mimetype="application/pdf")


@app.route("/resultats-classe/attestations/lot")
def resultats_attestations_lot():
    role = _role_courant()
    if not metier.peut_lire(role, "N2_Notes"):
        abort(403)
    f = _resultats_filtres(request.args)
    ok, msg, zipo, n = metier.generer_attestations_lot(
        f["filiere"], f["niveau"], f["section"], f["annee"], session.get("user", ""))
    if not ok:
        flash(msg, "err")
        return redirect(url_for("resultats_classe", filiere=f["filiere"], niveau=f["niveau"],
                                section=f["section"], annee=f["annee"], periode=f["periode"]))
    from flask import send_file
    nom = "Attestations_%s_%s_%s.zip" % (f["filiere"], f["niveau"], f["annee"])
    nom = nom.replace(" ", "_").replace("/", "-")
    return send_file(__import__("io").BytesIO(zipo), as_attachment=True,
                     download_name=nom, mimetype="application/zip")


@app.route("/planification/volumes")
def planification_volumes():
    role = _role_courant()
    if not metier.peut_lire(role, "A3_Sessions"):
        abort(403)
    filiere = request.args.get("filiere", "").strip()
    niveau = request.args.get("niveau", "").strip()
    semestre = request.args.get("semestre", "").strip()
    data = (metier.volumes_classe(filiere, niveau, semestre)
            if (filiere and niveau) else None)
    return render_template("volumes_classe.html",
                           info=config.TAB_INDEX["PLN_Volumes"],
                           titre_page="Planification — volumes par classe",
                           filtres={"filiere": filiere, "niveau": niveau, "semestre": semestre},
                           data=data, options=metier.volumes_options())


def _grille_filtres(src):
    return {"filiere": src.get("filiere", "").strip(), "niveau": src.get("niveau", "").strip(),
            "section": src.get("section", "").strip(), "annee": src.get("annee", "").strip(),
            "semestre": src.get("semestre", "").strip()}


@app.route("/planification/grille")
def planification_grille():
    role = _role_courant()
    if not metier.peut_lire(role, "A3_Sessions"):
        abort(403)
    f = _grille_filtres(request.args)
    data = (metier.grille_classe(f["filiere"], f["niveau"], f["section"], f["annee"], f["semestre"])
            if (f["filiere"] and f["niveau"]) else None)
    return render_template("planification_grille.html",
                           info=config.TAB_INDEX["PLN_Grille"],
                           titre_page="Planification — grille hebdomadaire",
                           filtres=f, data=data,
                           vopts=metier.volumes_options(),
                           peut_ecrire=metier.peut_ecrire(role, "A3_Sessions"))


@app.route("/planification/grille/ajouter", methods=["POST"])
def planification_grille_ajouter():
    _garde_ecriture("A3_Sessions")
    role = _role_courant()
    fo = request.form
    ok, msg = metier.creer_seance_grille(
        role, fo.get("annee", ""), fo.get("semestre", ""), fo.get("filiere", ""),
        fo.get("niveau", ""), fo.get("section", ""), fo.get("matiere", ""),
        fo.get("enseignant", ""), fo.get("salle", ""), fo.get("jour", ""),
        fo.get("debut", ""), fo.get("fin", ""), fo.get("vol_prog", ""),
        session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    f = _grille_filtres(fo)
    return redirect(url_for("planification_grille", **f))


@app.route("/planification/grille/supprimer", methods=["POST"])
def planification_grille_supprimer():
    _garde_ecriture("A3_Sessions")
    role = _role_courant()
    ok, msg = metier.supprimer_seance_grille(role, request.form.get("sid", ""),
                                             session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    f = _grille_filtres(request.form)
    return redirect(url_for("planification_grille", **f))


@app.route("/planification/grille/imprimer")
def planification_grille_imprimer():
    """R11 Brique 3 : edition imprimable (PDF A4 paysage) de la grille hebdo
    d'une classe (creneaux x jours Lundi->Samedi, compteur + conflits)."""
    role = _role_courant()
    if not metier.peut_lire(role, "A3_Sessions"):
        abort(403)
    f = _grille_filtres(request.args)
    if not (f["filiere"] and f["niveau"]):
        flash("Choisissez au moins une filiere et un niveau pour imprimer la grille.", "err")
        return redirect(url_for("planification_grille", **f))
    try:
        octets = metier.grille_pdf_bytes(f["filiere"], f["niveau"], f["section"],
                                         f["annee"], f["semestre"])
    except RuntimeError as ex:
        flash(str(ex), "err")
        return redirect(url_for("planification_grille", **f))
    nom = metier.grille_pdf_nom(f["filiere"], f["niveau"], f["section"],
                                f["annee"], f["semestre"])
    return send_file(__import__("io").BytesIO(octets), as_attachment=False,
                     download_name=nom, mimetype="application/pdf")


@app.route("/resultats-classe/imprimer")
def resultats_classe_imprimer():
    role = _role_courant()
    if not metier.peut_lire(role, "N2_Notes"):
        abort(403)
    f = _resultats_filtres(request.args)
    if not (f["filiere"] and f["niveau"] and f["annee"] and f["periode"]):
        flash("Selection incomplete pour les resultats de classe.", "err")
        return redirect(url_for("resultats_classe"))
    d = metier.resultats_classe(f["filiere"], f["niveau"], f["section"], f["annee"], f["periode"])
    annuel = d["stats"]["annuel"]
    cols = ["N°", "Matricule", "Nom", "Prenom", "Moyenne", "Mention", "Decision"]
    if not annuel:
        cols.append("ECTS")
    cols.append("Obs.")
    lignes = []
    for l in d["lignes"]:
        row = [str(l["n"]), l["matricule"], l["nom"], l["prenom"], l["moyenne"],
               l["mention"], l["decision"]]
        if not annuel:
            row.append(l["ects"])
        row.append(l["note"])
        lignes.append(row)
    s2 = d["stats"]
    per = "Annuel" if annuel else ("Semestre %s" % f["periode"])
    leg = ("%s · %s%s · Annee %s · %s · Effectif %d · Admis %d · Ajournes %d · "
           "Moyenne de classe %s · Edite le %s" % (
               f["filiere"], f["niveau"], (" · Section " + f["section"]) if f["section"] else "",
               f["annee"], per, s2["effectif"], s2["admis"], s2["ajourne"],
               s2["moyenne_classe"] or "—", metier.fmt_date(metier.datetime.date.today())))
    table = {"colonnes": cols, "lignes": lignes, "total": None, "contexte": {"legende": leg}}
    doc = metier.doc_adhoc("Resultats de classe — %s %s" % (f["filiere"], f["niveau"]),
                           mentions="Moyennes et decisions PROPOSEES (calcul decret 05-106) ; "
                                    "la deliberation reste validee a la main.",
                           signataire="Le jury | Le Directeur", orientation="paysage")
    return _page_impression(doc, "table", table=table)


@app.route("/stages/affectation/groupe", methods=["POST"])
def stages_affectation_groupe():
    _garde_ecriture("S1_Stages")
    f = _affect_filtres(request.form)
    mats = request.form.getlist("matricules")
    lieu = request.form.get("lieu", "").strip()
    debut = request.form.get("date_debut", "").strip()
    fin = request.form.get("date_fin", "").strip()
    depass = request.form.get("depassement", "") in ("1", "on", "true", "oui")
    if not (f["annee"] and f["seance"]):
        flash("Selection incomplete (annee / seance) pour l'affectation.", "err")
        return redirect(url_for("stages_affectation"))
    ok, msg, n = metier.affecter_groupe_stages(mats, f["annee"], f["seance"],
                                               lieu, debut, fin, depass, f["session"])
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Affectation stages (groupe)",
                     "%s %s seance %s [%s] -> %s" % (f["filiere"], f["niveau"], f["seance"],
                                                     f["session"], lieu),
                     "%d eleve(s)" % n)
    return redirect(url_for("stages_affectation", filiere=f["filiere"], niveau=f["niveau"],
                            annee=f["annee"], seance=f["seance"], session=f["session"]))


# === Sauvegarde horodatee des classeurs (V1.80) ============================

@app.route("/sauvegarde")
def sauvegarde():
    _exige_admin()
    return render_template("sauvegarde.html",
                           info=config.TAB_INDEX["SAV_Sauvegarde"],
                           titre_page="Sauvegarde des donnees",
                           sauvegardes=metier.liste_sauvegardes(),
                           dossier=config.SAUVEGARDES_DIR)


@app.route("/sauvegarde/creer", methods=["POST"])
def sauvegarde_creer():
    _exige_admin()
    ok, msg, nom, copies = metier.sauvegarder_classeurs()
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Sauvegarde classeurs", nom,
                     ", ".join(copies))
    return redirect(url_for("sauvegarde"))


# === Fiche bailleur (V1.99) — referentiel des financeurs (F4_Bailleurs) =====

@app.route("/bailleur")
def bailleur():
    role = _role_courant()
    if not metier.peut_lire(role, "F4_Bailleurs"):
        abort(403)
    idb = (request.args.get("id", "") or "").strip()
    fiche = metier.fiche_bailleur(idb) if idb else None
    return render_template("bailleur.html", info=config.TAB_INDEX["BAIL_Fiche"],
                           titre_page="Fiche bailleur", id_sel=idb, fiche=fiche,
                           bailleurs=metier.liste_bailleurs(),
                           peut_modifier=metier.peut_ecrire(role, "F4_Bailleurs"),
                           documents=(metier.documents_bailleur(idb) if fiche else []),
                           types_document=metier.types_document_bailleur(),
                           tracabilite=(metier.tracabilite_bailleur(idb) if fiche else None),
                           introuvable=(bool(idb) and fiche is None))


@app.route("/bailleur/<id_bailleur>/imprimer")
def bailleur_imprimer(id_bailleur):
    role = _role_courant()
    if not metier.peut_lire(role, "F4_Bailleurs"):
        abort(403)
    fiche = metier.fiche_bailleur(id_bailleur)
    if not fiche:
        abort(404)
    return render_template("bailleur_print.html", info=config.TAB_INDEX["BAIL_Fiche"],
                           fiche=fiche, titre_page="Fiche bailleur - %s" % fiche["id"],
                           documents=metier.documents_bailleur(id_bailleur),
                           tracabilite=metier.tracabilite_bailleur(id_bailleur),
                           date_jour=metier.fmt_date(metier._dt.date.today()))


# --- Documents lies au bailleur (3b) : ajout / statut / telechargement ------
@app.route("/bailleur/<id_bailleur>/document/ajouter", methods=["POST"])
def bailleur_document_ajouter(id_bailleur):
    _garde_ecriture("F4_Bailleurs")
    f = request.files.get("document")
    type_doc = request.form.get("type", "")
    maxi = getattr(config, "DOC_BAILLEUR_MAX_OCTETS", 10 * 1024 * 1024)
    donnees = f.read(maxi + 1) if f else b""
    nom = f.filename if f else ""
    ok, msg, _meta = metier.ajouter_document_bailleur(
        _role_courant(), id_bailleur, donnees, nom, type_doc)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Document bailleur", id_bailleur,
                     "Ajout : " + (nom or ""))
    return redirect(url_for("bailleur", id=id_bailleur))


@app.route("/bailleur/<id_bailleur>/document/statut", methods=["POST"])
def bailleur_document_statut(id_bailleur):
    _garde_ecriture("F4_Bailleurs")
    nom_stocke = request.form.get("nom_stocke", "")
    statut = request.form.get("statut", "")
    ok, msg = metier.marquer_statut_document_bailleur(
        _role_courant(), id_bailleur, nom_stocke, statut)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Document bailleur", id_bailleur,
                     "Statut : " + (statut or "(actif)"))
    return redirect(url_for("bailleur", id=id_bailleur))


@app.route("/bailleur/<id_bailleur>/document/<nom_stocke>")
def bailleur_document_telecharger(id_bailleur, nom_stocke):
    role = _role_courant()
    if not metier.peut_lire(role, "F4_Bailleurs"):
        abort(403)
    chemin = metier.chemin_document_bailleur(id_bailleur, nom_stocke)
    if not chemin:
        abort(404)
    nom_dl = nom_stocke
    for d in metier.documents_bailleur(id_bailleur):
        if d["nom_stocke"] == nom_stocke:
            nom_dl = d["nom_original"] or nom_stocke
            break
    return send_file(chemin, as_attachment=True, download_name=nom_dl)


# === Fiche enseignant (V1.73) — symetrique de la fiche etudiant ============

@app.route("/enseignant")
def enseignant():
    role = _role_courant()
    if not metier.peut_lire(role, "E1_Enseignants"):
        abort(403)
    matricule = _matricule_saisi(request.args.get("matricule", ""))
    fiche = metier.fiche_enseignant(matricule) if matricule else None
    heures = seances = faites = None
    if fiche:
        heures = metier.heures_enseignant(fiche["matricule"])
        seances = metier.seances_enseignant(fiche.get("Nom", ""), fiche.get("Prenom", ""),
                                             fiche["matricule"])
        faites = metier.seances_faites_enseignant(fiche.get("Nom", ""), fiche.get("Prenom", ""),
                                                  fiche["matricule"])
    return render_template("enseignant.html", info=config.TAB_INDEX["ENS_Fiche"],
                           titre_page="Fiche enseignant", matricule=matricule,
                           fiche=fiche, heures=heures, seances=seances, faites=faites,
                           enseignants=metier.recherche_enseignants(),
                           peut_modifier=metier.peut_ecrire(role, "E1_Enseignants"),
                           introuvable=(bool(matricule) and fiche is None))


@app.route("/enseignant/<matricule>/imprimer")
def enseignant_imprimer(matricule):
    role = _role_courant()
    if not metier.peut_lire(role, "E1_Enseignants"):
        abort(403)
    matricule = str(matricule).strip()
    fiche = metier.fiche_enseignant(matricule)
    if not fiche:
        abort(404)
    heures = metier.heures_enseignant(matricule)
    seances = metier.seances_enseignant(fiche.get("Nom", ""), fiche.get("Prenom", ""), matricule)
    faites = metier.seances_faites_enseignant(fiche.get("Nom", ""), fiche.get("Prenom", ""), matricule)
    return render_template("enseignant_print.html",
                           titre_page="Fiche enseignant", fiche=fiche,
                           heures=heures, seances=seances, faites=faites,
                           date_jour=metier.fmt_date(metier.datetime.date.today()))


@app.route("/enseignant/<matricule>/heures/imprimer")
def enseignant_heures_imprimer(matricule):
    """Etat imprimable des heures a payer (Option A) : releve E2 par enseignant.
    Aucun montant KMF — la compta valorise selon les contrats signes."""
    role = _role_courant()
    if not metier.peut_lire(role, "E1_Enseignants"):
        abort(403)
    matricule = str(matricule).strip()
    fiche = metier.fiche_enseignant(matricule)
    if not fiche:
        abort(404)
    heures = metier.heures_enseignant(matricule)
    return render_template("etat_heures_print.html",
                           titre_page="Etat des heures a payer", fiche=fiche,
                           heures=heures,
                           date_jour=metier.fmt_date(metier.datetime.date.today()))


@app.route("/enseignant/photo/<matricule>")
def enseignant_photo(matricule):
    role = _role_courant()
    if not metier.peut_lire(role, "E1_Enseignants"):
        abort(403)
    chemin, mime = metier.photo_servie(matricule)
    if chemin:
        return send_file(chemin, mimetype=mime)
    svg = ('<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 120 120">'
           '<rect width="120" height="120" rx="10" fill="#EAEFF5"/>'
           '<circle cx="60" cy="46" r="21" fill="#1F4E79"/>'
           '<path d="M24 104c0-20 16-31 36-31s36 11 36 31z" fill="#1F4E79"/></svg>')
    return Response(svg, mimetype="image/svg+xml")


@app.route("/enseignant/photo/<matricule>/televerser", methods=["POST"])
def enseignant_photo_televerser(matricule):
    _garde_ecriture("E1_Enseignants")
    f = request.files.get("photo")
    maxi = getattr(config, "PHOTO_MAX_OCTETS", 1024 * 1024)
    donnees = f.read(maxi + 1) if f else b""
    ok, msg = metier.enregistrer_photo_enseignant(matricule, donnees)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Photo enseignant", matricule, "Televersement")
    return redirect(url_for("enseignant", matricule=matricule))


@app.route("/enseignant/photo/<matricule>/retirer", methods=["POST"])
def enseignant_photo_retirer(matricule):
    _garde_ecriture("E1_Enseignants")
    ok, msg = metier.supprimer_photo(matricule)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Photo enseignant", matricule, "Retrait")
    return redirect(url_for("enseignant", matricule=matricule))


@app.route("/etudiant/<matricule>/imprimer")
def etudiant_imprimer(matricule):
    role = _role_courant()
    if not metier.peut_lire(role, "A1_Etudiants"):
        abort(403)
    matricule = str(matricule).strip()
    fiche = metier.fiche_etudiant(matricule)
    if not fiche:
        abort(404)
    stages = metier.stages_etudiant(matricule)
    droits = metier.droits_inscription(matricule)
    return render_template("etudiant_print.html",
                           titre_page="Fiche etudiant", fiche=fiche,
                           stages=stages, droits=droits,
                           date_jour=metier.fmt_date(metier.datetime.date.today()))


@app.route("/releve/export")
def releve_export():
    role = _role_courant()
    if not metier.peut_lire(role, "N2_Notes"):
        abort(403)
    matricule = _matricule_saisi(request.args.get("matricule", ""))
    annee = request.args.get("annee", "").strip()
    semestre = request.args.get("semestre", "").strip()
    if not (matricule and annee):
        abort(404)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    BLEU = "1F4E79"
    ft = Font(name="Calibri", bold=True, size=13, color=BLEU)
    fh = Font(name="Calibri", bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=BLEU)
    wb = Workbook(); ws = wb.active; ws.title = "Releve"
    r = 1
    ws.cell(r, 1, "UNIVERSITE DES COMORES - EMSP - Releve de notes").font = ft; r += 2

    def bloc_semestre(rel, r):
        ws.cell(r, 1, "Etudiant : %s" % rel["nom"]).font = Font(name="Calibri", bold=True); 
        ws.cell(r, 4, "Matricule : %s" % rel["matricule"]); r += 1
        ws.cell(r, 1, "%s %s" % (rel["filiere"], rel["niveau"]))
        ws.cell(r, 4, "Annee : %s" % rel["annee"]); ws.cell(r, 6, "Semestre : %s" % rel["semestre"]); r += 2
        heads = ["UE / Matiere", "Moyenne", "Coef", "ECTS", "Validee"]
        for j, h in enumerate(heads, 1):
            c = ws.cell(r, j, h); c.font = fh; c.fill = fill; c.alignment = Alignment(horizontal="center")
        r += 1
        for ue in rel["ues"]:
            ws.cell(r, 1, "%s  %s" % (ue["num"], ue["intitule"])).font = Font(name="Calibri", bold=True)
            ws.cell(r, 2, ue["moyenne"]); ws.cell(r, 3, ue["coef"]); ws.cell(r, 4, ue["ects"])
            ws.cell(r, 5, "Oui" if ue["validee"] else "Non"); r += 1
            for m in ue["matieres"]:
                ws.cell(r, 1, "   %s" % m["matiere"]); ws.cell(r, 2, m["moyenne"]); r += 1
        ws.cell(r, 1, "Moyenne du semestre").font = Font(name="Calibri", bold=True)
        ws.cell(r, 2, rel["moyenne"]).font = Font(name="Calibri", bold=True); r += 1
        ws.cell(r, 1, "Mention : %s" % rel["mention"]); r += 1
        ws.cell(r, 1, "Proposition : %s" % rel["proposition"]); r += 1
        ws.cell(r, 1, "ECTS acquis : %g / %g" % (rel["ects_acquis"], rel["ects_total"])); r += 2
        return r

    if semestre:
        bloc_semestre(metier.releve_semestre(matricule, annee, semestre), r)
        nomf = "Releve_%s_%s_S%s.xlsx" % (matricule, annee, semestre)
    else:
        an = metier.releve_annuel(matricule, annee)
        for rel in an["semestres"]:
            r = bloc_semestre(rel, r)
        ws.cell(r, 1, "MOYENNE ANNUELLE").font = Font(name="Calibri", bold=True, color=BLEU)
        ws.cell(r, 2, an["moyenne"]).font = Font(name="Calibri", bold=True); r += 1
        ws.cell(r, 1, "Mention : %s   Proposition : %s" % (an["mention"], an["proposition"]))
        nomf = "Releve_annuel_%s_%s.xlsx" % (matricule, annee)
    for col, w in {"A": 42, "B": 10, "C": 8, "D": 8, "E": 9, "F": 16}.items():
        ws.column_dimensions[col].width = w
    metier.bandeau_xlsx(ws, 6)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    auth.journal(role.get("login", ""), "Export releve", matricule, annee)
    return send_file(buf, as_attachment=True, download_name=metier.nom_export(nomf),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/etat-signalements")
def etat_signalements():
    role = _role_courant()
    if not metier.peut_lire(role, "N3_Signalements"):
        abort(403)
    f = {k: request.args.get(k, "").strip() for k in
         ("annee", "semestre", "contexte", "filiere", "niveau", "du", "au")}
    etat = metier.etat_signalements(**f)
    total = sum(e["nb"] for e in etat)
    return render_template("etat_signalements.html", info=config.TAB_INDEX["SIG_Etat"],
                           titre_page="Etat des signalements", etat=etat, total=total, f=f)


@app.route("/etat-signalements/export")
def etat_signalements_export():
    role = _role_courant()
    if not metier.peut_lire(role, "N3_Signalements"):
        abort(403)
    f = {k: request.args.get(k, "").strip() for k in
         ("annee", "semestre", "contexte", "filiere", "niveau", "du", "au")}
    etat = metier.etat_signalements(**f)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    BLEU = "1F4E79"
    wb = Workbook(); ws = wb.active; ws.title = "Signalements"
    ws.cell(1, 1, "EMSP - Etat des signalements / indiscipline (par etudiant)").font = \
        Font(name="Calibri", bold=True, size=13, color=BLEU)
    r = 3
    heads = ["Etudiant / signalement", "Date", "Contexte", "Fonction", "Emetteur", "Motif"]
    for j, h in enumerate(heads, 1):
        c = ws.cell(r, j, h); c.font = Font(name="Calibri", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLEU); c.alignment = Alignment(horizontal="center")
    r += 1
    for e in etat:
        ws.cell(r, 1, "%s  (%s) - %s %s - %d signalement(s)" %
                (e["nom"] or "(inconnu)", e["matricule"], e["filiere"], e["niveau"], e["nb"])
                ).font = Font(name="Calibri", bold=True, color=BLEU)
        r += 1
        for s in e["signalements"]:
            ws.cell(r, 2, s["date"]); ws.cell(r, 3, s["contexte"]); ws.cell(r, 4, s["fonction"])
            ws.cell(r, 5, s["nom"]); ws.cell(r, 6, s["motif"]); r += 1
    for col, w in {"A": 40, "B": 12, "C": 12, "D": 20, "E": 22, "F": 48}.items():
        ws.column_dimensions[col].width = w
    metier.bandeau_xlsx(ws, 6)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    auth.journal(role.get("login", ""), "Export etat signalements", f.get("annee", ""), "")
    return send_file(buf, as_attachment=True,
                     download_name=metier.nom_export("Etat_signalements.xlsx"),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/etat-signalements/imprimer")
def etat_signalements_imprimer():
    """Etat des signalements en edition chartee (kind table generique) — remplace
    l'ancienne impression brute (window.print sur l'ecran)."""
    role = _role_courant()
    if not metier.peut_lire(role, "N3_Signalements"):
        abort(403)
    f = {k: request.args.get(k, "").strip() for k in
         ("annee", "semestre", "contexte", "filiere", "niveau", "du", "au")}
    etat = metier.etat_signalements(**f)
    total = sum(e["nb"] for e in etat)
    colonnes = ["Etudiant / signalement", "Date", "Contexte", "Fonction", "Emetteur", "Motif"]
    lignes = []
    for e in etat:
        lignes.append(["%s (%s) - %s %s - %d signalement(s)" % (
            e["nom"] or "(inconnu)", e["matricule"], e["filiere"], e["niveau"], e["nb"]),
            "", "", "", "", ""])
        for s in e["signalements"]:
            lignes.append(["", s["date"], s["contexte"], s["fonction"], s["nom"], s["motif"]])
    leg = "Etat des signalements / indiscipline (par etudiant) - %d signalement(s) - Edite le %s" % (
        total, metier.fmt_date(metier._dt.date.today()))
    table = {"colonnes": colonnes, "lignes": lignes, "contexte": {"legende": leg}}
    doc = metier.doc_adhoc("Etat des signalements / indiscipline",
                           mentions="Source : signalements (N3). Document interne.",
                           signataire="Le Directeur de l'EMSP", orientation="paysage")
    return _page_impression(doc, "table", table=table)


def _verrou_acquerir(hote, port):
    """Empeche un 2e ECRIVAIN sur le meme dossier de donnees (anti-corruption).
    Verrou = instance/app.lock. Si un serveur repond deja sur l'hote/port enregistre,
    on refuse de demarrer ; sinon le verrou est perime et on le reprend. Retire a la
    sortie propre (atexit) et par la route /quitter."""
    import json as _json, socket as _sock, atexit
    import datetime as _dt
    lock = os.path.join(config.INSTANCE_DIR, "app.lock")
    if os.path.exists(lock):
        try:
            with open(lock, encoding="utf-8") as f:
                d = _json.load(f)
            h = d.get("host", "127.0.0.1")
            cible = "127.0.0.1" if h in ("0.0.0.0", "127.0.0.1") else h
            with _sock.create_connection((cible, int(d.get("port", 0))), timeout=0.7):
                return False                      # un ecrivain repond -> on n'en lance pas un 2e
        except OSError:
            pass                                  # personne ne repond -> verrou perime, on reprend
        except Exception:
            pass
    try:
        os.makedirs(config.INSTANCE_DIR, exist_ok=True)
        with open(lock, "w", encoding="utf-8") as f:
            _json.dump({"host": hote, "port": port, "pid": os.getpid(),
                        "depuis": _dt.datetime.now().strftime("%d/%m/%Y %H:%M:%S")}, f)
        atexit.register(lambda: os.path.exists(lock) and os.remove(lock))
    except Exception:
        pass
    return True


@app.route("/quitter", methods=["POST"])
def quitter():
    """Arret propre du serveur local (poste principal uniquement). Les postes distants
    (lecture seule) ne peuvent pas arreter le serveur."""
    if request.remote_addr not in ("127.0.0.1", "::1", "localhost"):
        abort(403)

    def _stop():
        import time
        time.sleep(0.6)
        try:
            lock = os.path.join(config.INSTANCE_DIR, "app.lock")
            if os.path.exists(lock):
                os.remove(lock)
        except Exception:
            pass
        os._exit(0)
    import threading
    threading.Thread(target=_stop, daemon=True).start()
    return render_template("quitter.html")



# ===========================================================================
# ETATS DE PAIEMENT DES VACATIONS (Bloc 3, V1.99.3)
# Ecran dedie PAIE_Etat : constituer -> arreter -> passer en compta + impression.
# Droits : onglet E4_Etats_paiement (groupe Financier + verrou Acces financier).
# ===========================================================================
@app.template_filter("kmf")
def _filtre_kmf(v):
    """Montant entier formate facon EMSP : '477 000' (sans decimale inutile)."""
    return metier._fmt_kmf(v)


@app.route("/paiement")
def paiement():
    role = _role_courant()
    if not metier.peut_lire(role, "E4_Etats_paiement"):
        abort(403)
    eid = (request.args.get("id", "") or "").strip()
    detail = metier.detail_etat_paiement(eid) if eid else None
    return render_template(
        "paiement.html", info=config.TAB_INDEX["PAIE_Etat"],
        titre_page="États de paiement (vacations)",
        etats=metier.etats_paiement(), detail=detail, id_sel=eid,
        introuvable=(bool(eid) and detail is None),
        mois_dispo=metier.mois_disponibles_paie(),
        annee_defaut=metier._annee_acad_defaut(),
        semestres=config.SEMESTRES_PAIE,
        comptes=[c["nom"] for c in metier.comptes_treso()],
        modes=metier.options_liste("Modes_paiement") or [],
        peut_modifier=metier.peut_ecrire(role, "E4_Etats_paiement"))


@app.route("/paiement/constituer", methods=["POST"])
def paiement_constituer():
    _garde_ecriture("E4_Etats_paiement")
    annee = request.form.get("annee", "").strip()
    semestre = request.form.get("semestre", "").strip()
    mois = request.form.getlist("mois")
    if not annee or not semestre:
        flash("Indiquez l'année académique et le semestre.", "err")
        return redirect(url_for("paiement"))
    if not mois:
        flash("Sélectionnez au moins un mois à inclure.", "err")
        return redirect(url_for("paiement"))
    ok, msg, eid = metier.constituer_etat_paiement(
        annee, semestre, mois, session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Etat de paiement", eid,
                     "Constitution (brouillon) : %d mois" % len(mois))
    return redirect(url_for("paiement", id=eid))


@app.route("/paiement/<eid>/arreter", methods=["POST"])
def paiement_arreter(eid):
    _garde_ecriture("E4_Etats_paiement")
    ok, msg = metier.arreter_etat_paiement(eid, session.get("user", ""))
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Etat de paiement", eid, "Arrete")
    return redirect(url_for("paiement", id=eid))


@app.route("/paiement/<eid>/passer-compta", methods=["POST"])
def paiement_passer_compta(eid):
    _garde_ecriture("E4_Etats_paiement")
    compte = request.form.get("compte", "").strip()
    mode = request.form.get("mode", "").strip()
    poste = request.form.get("poste", "").strip() or None
    ok, msg = metier.passer_etat_en_compta(
        eid, session.get("user", ""), compte, mode, poste)
    flash(msg, "ok" if ok else "err")
    if ok:
        auth.journal(session.get("user", ""), "Etat de paiement", eid,
                     "Passe en compta (compte %s)" % compte)
    return redirect(url_for("paiement", id=eid))


@app.route("/paiement/<eid>/imprimer")
def paiement_imprimer(eid):
    role = _role_courant()
    if not metier.peut_lire(role, "E4_Etats_paiement"):
        abort(403)
    detail = metier.detail_etat_paiement(eid)
    if not detail:
        abort(404)
    return render_template(
        "paiement_print.html", info=config.TAB_INDEX["PAIE_Etat"],
        titre_page="Etat de paiement - %s" % eid, detail=detail,
        date_jour=metier.fmt_date(metier._dt.date.today()))



if __name__ == "__main__":
    # Ecoute LOCALE par defaut (mono-poste). Pour ouvrir au reseau cable interne
    # (postes secondaires en consultation), demarrer avec EMSP_HOST=0.0.0.0
    # (cf. Demarrer_EMSP_RESEAU.bat). Les postes distants sont en lecture seule.
    import sys, time
    hote = os.environ.get("EMSP_HOST", "127.0.0.1")
    port = int(os.environ.get("EMSP_PORT", "5000"))
    if not _verrou_acquerir(hote, port):
        print("")
        print("  EMSP est DEJA en cours d'execution sur ce dossier de donnees.")
        print("  Utilisez la fenetre/onglet deja ouvert (ou le bouton Quitter dedans),")
        print("  puis relancez. Demarrage annule pour eviter tout melange de donnees.")
        print("")
        time.sleep(5)
        sys.exit(0)
    app.run(host=hote, port=port, debug=False)
